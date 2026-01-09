import copy
import hashlib
import html
import io
import json
import math
import os
import re
import shutil
import socket
import sqlite3
import sys
import tempfile
import threading
import time
from collections import defaultdict, Counter, OrderedDict
from datetime import datetime, date, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler

import folium
import unicodedata
from PyQt6.QtCore import Qt, QSize, QPoint, QEvent, QRect, QObject, QTimer, QRectF, QThread, pyqtSignal, QDateTime, \
    QSortFilterProxyModel, QModelIndex, QDate, QAbstractTableModel, QUrl
from PyQt6.QtGui import QFont, QPalette, QColor, QAction, QPixmap, QPainter, QMovie, QRadialGradient, QTextDocument, \
    QImage, QTextCharFormat
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QFileDialog, QStyledItemDelegate, QWidget, QMenu, \
    QComboBox, QMainWindow, QSizePolicy, QFrame, QGraphicsDropShadowEffect, QApplication, QToolTip, QProgressBar, \
    QStackedLayout, QTableWidget, QTableWidgetItem, QHeaderView, QHBoxLayout, QStyle, QStyleOptionViewItem, \
    QStackedWidget, QLineEdit, QAbstractItemView, QSplitter, QDateTimeEdit, QSpinBox, QDoubleSpinBox, QCheckBox, \
    QButtonGroup, QTableView, QTabWidget, QGridLayout, QGroupBox, QMessageBox, QListWidget, QSlider, QScrollArea, \
    QTextEdit, QColorDialog, QListWidgetItem
from branca.element import MacroElement
from jinja2 import Template

from security.security import LicenseManager
from ui.dialog import ModernDialog
from ui.mixins import WatermarkDialogMixin
from utils.constants import HEADER_ALIASES, TABLE_COLUMNS
from utils.helpers import _extract_table_headers_rows, _apply_hidden_cols_to_table_html, _apply_fmt_to_table_html

APP_DIR = os.path.dirname(os.path.abspath("file")) if not getattr(sys, "frozen", False) else sys._MEIPASS

class LicenseGateDialog(QDialog):
    """
    Lisans yoksa/yanlışsa: MainWindow'dan önce bu dialog çıkar.
    Kullanıcı lisans dosyasını seçmezse uygulama kapanır.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Lisans Aktivasyonu")
        self.setModal(True)
        self.setMinimumWidth(520)

        lay = QVBoxLayout(self)

        fp = LicenseManager.device_fingerprint()
        self.lbl = QLabel(
            "Bu uygulama lisans olmadan açılmaz.\n\n"
            f"Cihaz Kodu:\n{fp}\n\n"
            "Satıcıdan bu cihaza özel lisans dosyası (license.json) alıp yükleyin."
        )
        self.lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        lay.addWidget(self.lbl)

        btn_load = QPushButton("Lisans Dosyası Seç (license.json)")
        btn_load.clicked.connect(self._pick_and_install)
        lay.addWidget(btn_load)

        btn_exit = QPushButton("Kapat")
        btn_exit.clicked.connect(self.reject)
        lay.addWidget(btn_exit)

    def _pick_and_install(self):
        path, _ = QFileDialog.getOpenFileName(self, "Lisans Dosyası Seç", "", "JSON (*.json)")
        if not path:
            return

        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)

            # doğrula
            LicenseManager.validate_license(d)

            # install (APPDATA altına yaz)
            target_path = LicenseManager.license_path()
            with open(target_path, "w", encoding="utf-8") as f:
                json.dump(d, f, ensure_ascii=False, indent=2)

            # kaynak dosyayı sil (APPDATA'daki dosyayı ASLA silme)
            try:
                src_abs = os.path.abspath(path)
                dst_abs = os.path.abspath(target_path)
                if src_abs != dst_abs and os.path.exists(src_abs):
                    os.remove(src_abs)
            except Exception as _del_err:
                # Silme başarısız olsa da lisans kurulumu başarılıdır; uygulamayı engellemeyelim.
                pass

            self.accept()

        except Exception as e:
            self.lbl.setText(self.lbl.text() + f"\n\nHata: {e}")


class NoTooltipDelegate(QStyledItemDelegate):
    """Bu delegate tooltip üretimini tamamen engeller."""
    def helpEvent(self, event, view, option, index):
        return False


def _enable_measure_and_balloons(m: folium.Map) -> None:
    """
    Folium haritasına:
    - Mesafe/Alan ölçme aracı (MeasureControl)
    - Çift tık ile modern mesaj baloncuğu ekleme
    - Balon düzenleme (tıklayınca modal)
    - Balon oku/çizgi otomatik yön (üst/alt/sol/sağ)
    - Harita üstü 🗑 ile seçili balon veya seçili ölçüm çizgisini/alanını silme
    - Ölçüm arayüzünü Türkçeleştirme (DOM translate)
    ekler.
    """
    from folium.plugins import MeasureControl
    import folium

    # 1) Mesafe/Alan ölçme aracı
    MeasureControl(
        position="topleft",
        primary_length_unit="meters",
        secondary_length_unit="kilometers",
        primary_area_unit="sqmeters",
        secondary_area_unit="hectares",
        active_color="#e74c3c",
        completed_color="#2c3e50"
    ).add_to(m)

    # 2) CSS: Balon + 4 yön ok + modern modal
    m.get_root().header.add_child(folium.Element("""
    <style>
      .hts-balloon-icon { background: transparent; border: none; }

      /* BALON */
      .hts-balloon{
        background: rgba(255,255,255,0.98);
        border: 3px solid #e74c3c;
        border-radius: 18px;
        padding: 12px 14px;
        font-family: Segoe UI, Arial;
        font-size: 14px;
        font-weight: 700;
        color: #2c3e50;
        box-shadow: 0 6px 18px rgba(0,0,0,0.25);

        white-space: pre-wrap;
        word-break: break-word;
        max-width: 380px;
        min-width: 160px;

        transform: translate(-50%, -120%);
        position: relative;

        /* JS bunları ayarlar */
        --tail-x: 50%;
        --tail-y: 50%;
      }

      /* ok pseudo ortak */
      .hts-balloon:before, .hts-balloon:after{
        content:"";
        position:absolute;
        width:0; height:0;
      }

      /* BOTTOM */
      .hts-balloon[data-tail="bottom"]:after{
        left: var(--tail-x);
        bottom:-12px;
        transform: translateX(-50%);
        border-width: 12px 12px 0 12px;
        border-style: solid;
        border-color: #e74c3c transparent transparent transparent;
      }
      .hts-balloon[data-tail="bottom"]:before{
        left: var(--tail-x);
        bottom:-10px;
        transform: translateX(-50%);
        border-width: 11px 11px 0 11px;
        border-style: solid;
        border-color: rgba(255,255,255,0.98) transparent transparent transparent;
      }

      /* TOP */
      .hts-balloon[data-tail="top"]:after{
        left: var(--tail-x);
        top:-12px;
        transform: translateX(-50%);
        border-width: 0 12px 12px 12px;
        border-style: solid;
        border-color: transparent transparent #e74c3c transparent;
      }
      .hts-balloon[data-tail="top"]:before{
        left: var(--tail-x);
        top:-10px;
        transform: translateX(-50%);
        border-width: 0 11px 11px 11px;
        border-style: solid;
        border-color: transparent transparent rgba(255,255,255,0.98) transparent;
      }

      /* LEFT */
      .hts-balloon[data-tail="left"]:after{
        top: var(--tail-y);
        left:-12px;
        transform: translateY(-50%);
        border-width: 12px 12px 12px 0;
        border-style: solid;
        border-color: transparent #e74c3c transparent transparent;
      }
      .hts-balloon[data-tail="left"]:before{
        top: var(--tail-y);
        left:-10px;
        transform: translateY(-50%);
        border-width: 11px 11px 11px 0;
        border-style: solid;
        border-color: transparent rgba(255,255,255,0.98) transparent transparent;
      }

      /* RIGHT */
      .hts-balloon[data-tail="right"]:after{
        top: var(--tail-y);
        right:-12px;
        transform: translateY(-50%);
        border-width: 12px 0 12px 12px;
        border-style: solid;
        border-color: transparent transparent transparent #e74c3c;
      }
      .hts-balloon[data-tail="right"]:before{
        top: var(--tail-y);
        right:-10px;
        transform: translateY(-50%);
        border-width: 11px 0 11px 11px;
        border-style: solid;
        border-color: transparent transparent transparent rgba(255,255,255,0.98);
      }

      /* Modern mesaj girişi */
      .hts-modal-backdrop{
        position: fixed;
        inset: 0;
        background: rgba(0,0,0,0.35);
        display: flex;
        align-items: center;
        justify-content: center;
        z-index: 999999;
      }
      .hts-modal{
        width: 420px;
        max-width: calc(100vw - 24px);
        background: #ffffff;
        border: 1px solid rgba(0,0,0,0.10);
        border-radius: 14px;
        box-shadow: 0 16px 44px rgba(0,0,0,0.35);
        padding: 14px;
        font-family: Segoe UI, Arial;
        color: #2c3e50;
      }
      .hts-modal h3{
        margin: 0 0 10px 0;
        font-size: 14px;
        font-weight: 800;
        display:flex;
        align-items:center;
        justify-content:space-between;
      }
      .hts-modal .hts-x{
        cursor:pointer;
        border:0;
        background: transparent;
        font-size: 18px;
        line-height: 18px;
        color:#7f8c8d;
      }
      .hts-modal textarea{
        width: 100%;
        height: 110px;
        resize: vertical;
        border-radius: 10px;
        border: 1px solid #d0d7de;
        padding: 10px;
        outline: none;
        font-family: Segoe UI, Arial;
        font-size: 13px;
        font-weight: 600;
        color: #2c3e50;
      }
      .hts-modal textarea:focus{
        border-color: #e74c3c;
        box-shadow: 0 0 0 3px rgba(231,76,60,0.15);
      }
      .hts-modal .hts-row{
        display:flex;
        gap: 10px;
        justify-content:flex-end;
        margin-top: 12px;
      }
      .hts-btn{
        border-radius: 10px;
        border: 0;
        padding: 10px 14px;
        font-family: Segoe UI, Arial;
        font-size: 13px;
        font-weight: 800;
        cursor: pointer;
      }
      .hts-btn-cancel{
        background: #ecf0f1;
        color:#2c3e50;
      }
      .hts-btn-ok{
        background: #e74c3c;
        color: white;
      }
    .leaflet-measure-popup a.leaflet-measure-action,
    .leaflet-measure-popup button.leaflet-measure-action {
      display: none !important;
    }

    /* Sadece sağ üst X (close) kalsın */
    .leaflet-measure-popup .leaflet-popup-close-button {
      display: block !important;
    }
    </style>
    """))

    # 3) JS: TR + balon + 4 yön ok + seç/sil + ölçüm seç/sil
    map_name = m.get_name()
    m.get_root().script.add_child(folium.Element(f"""
    (function() {{
      var MAP_NAME = "{map_name}";

      function esc(s) {{
        return String(s || "")
          .replace(/&/g,"&amp;")
          .replace(/</g,"&lt;")
          .replace(/>/g,"&gt;");
      }}

      function whenMapReady(cb) {{
        var map = window[MAP_NAME];
        if (!map) {{ setTimeout(function(){{ whenMapReady(cb); }}, 30); return; }}
        cb(map);
      }}

      // Modern modal
      function openMessageModal(title, initialText, onOk) {{
        var back = document.createElement("div");
        back.className = "hts-modal-backdrop";

        var box = document.createElement("div");
        box.className = "hts-modal";
        box.innerHTML =
          '<h3>' +
            '<span>' + esc(title) + '</span>' +
            '<button class="hts-x" type="button" title="Kapat">×</button>' +
          '</h3>' +
          '<textarea placeholder="Mesajınızı yazın..."></textarea>' +
          '<div class="hts-row">' +
            '<button class="hts-btn hts-btn-cancel" type="button">İptal</button>' +
            '<button class="hts-btn hts-btn-ok" type="button">Kaydet</button>' +
          '</div>';

        back.appendChild(box);
        document.body.appendChild(back);

        var ta = box.querySelector("textarea");
        var btnX = box.querySelector(".hts-x");
        var btnCancel = box.querySelector(".hts-btn-cancel");
        var btnOk = box.querySelector(".hts-btn-ok");
        ta.value = (initialText || "");

        function close() {{
          try {{ document.body.removeChild(back); }} catch(e){{}}
        }}

        btnX.addEventListener("click", close);
        btnCancel.addEventListener("click", close);
        back.addEventListener("click", function(e) {{ if (e.target === back) close(); }});

        btnOk.addEventListener("click", function() {{
          var t = (ta.value || "").trim();
          if (!t) return;
          close();
          onOk(t);
        }});

        // Enter = kaydet, Shift+Enter = alt satır
        ta.addEventListener("keydown", function(e) {{
          if (e.key === "Enter" && !e.shiftKey) {{
            e.preventDefault();
            btnOk.click();
          }}
        }});

        setTimeout(function() {{ ta.focus(); ta.select(); }}, 0);
      }}

      whenMapReady(function(map) {{
        map.on('popupopen', function(e) {{
          try {{
            trMeasureTexts(e.popup.getElement());
            cleanupMeasurePopup(e.popup.getElement());
          }} catch(err) {{}}
        }});
      map.options.closePopupOnClick = false;

        // -------------------------
        // 1) Measure UI Türkçeleştirme
        // -------------------------
        function trMeasureTexts(root) {{
          if (!root) return;

          var dict = {{
            "Measure distances and areas": "Mesafe ve alan ölç",
            "Start creating a measurement by adding points to the map": "Haritaya nokta ekleyerek ölçüme başla",
            "Cancel": "İptal",
            "Finish measurement": "Ölçümü bitir",
            "Linear measurement": "Mesafe Ölçümü",
            "Area measurement": "Alan Ölçümü",
            "Last point": "Son nokta",
            "Path distance": "Yol mesafesi",
            "Area": "Alan"
          }};

          // SADECE text node değiştir (butonların DOM'u bozulmaz)
          var walker = document.createTreeWalker(root, NodeFilter.SHOW_TEXT, null, false);
          var node;
          while ((node = walker.nextNode())) {{
            var raw = node.nodeValue;
            if (!raw) continue;

            var t = raw.trim();
            if (!t) continue;

            if (dict[t]) {{
              // nodeValue içinde baştaki/sondaki boşlukları koru
              node.nodeValue = raw.replace(t, dict[t]);
            }}
          }}
        }}
        function cleanupMeasurePopup(root) {{
          if (!root) return;

          // Measure popup içindeki linkleri metnine göre gizle
          root.querySelectorAll('a').forEach(function(a) {{
            var t = (a.textContent || '').trim();
            if (t === 'Center on this line' || t === 'Delete') {{
              a.style.display = 'none';
            }}
          }});

          // Aynı linkler bazen button olarak gelebiliyor
          root.querySelectorAll('button').forEach(function(b) {{
            var t = (b.textContent || '').trim();
            if (t === 'Center on this line' || t === 'Delete') {{
              b.style.display = 'none';
            }}
          }});
        }}

        try {{ trMeasureTexts(document.body); }} catch(e){{}}
        try {{ cleanupMeasurePopup(document.body); }} catch(e){{}}

        try {{
          var obs = new MutationObserver(function(muts) {{
            muts.forEach(function(mu) {{
              mu.addedNodes && mu.addedNodes.forEach(function(n) {{
                if (n && n.nodeType === 1) trMeasureTexts(n); cleanupMeasurePopup(n);
              }});
            }});
          }});
          obs.observe(document.body, {{ childList: true, subtree: true }});
        }} catch(e){{}}

        // -------------------------
        // 2) Seçim + Silme (balon/measure)
        // -------------------------
        var balloons = [];
        var selectedBalloon = null;
        var selectedMeasure = null;

        function clearBalloonSelectionVisual() {{
          balloons.forEach(function(b) {{
            var el = b.label.getElement && b.label.getElement();
            if (!el) return;
            var bb = el.querySelector('.hts-balloon');
            if (bb) bb.style.outline = '';
          }});
        }}

        function selectBalloon(b) {{
          selectedBalloon = b;
          selectedMeasure = null;
          clearBalloonSelectionVisual();
          var el = b.label.getElement && b.label.getElement();
          if (el) {{
            var bb = el.querySelector('.hts-balloon');
            if (bb) bb.style.outline = '3px solid rgba(231,76,60,0.75)';
          }}
        }}

        function selectMeasure(layer) {{
          selectedMeasure = layer;
          selectedBalloon = null;
          clearBalloonSelectionVisual();
          try {{
            if (layer.setStyle) layer.setStyle({{ color:'#e74c3c', weight:5, opacity:1 }});
          }} catch(e){{}}
        }}

        function deleteSelected() {{
          if (selectedBalloon) {{
            try {{ map.removeLayer(selectedBalloon.label); }} catch(e){{}}
            try {{ map.removeLayer(selectedBalloon.line); }} catch(e){{}}
            try {{ map.removeLayer(selectedBalloon.anchor); }} catch(e){{}}
            balloons = balloons.filter(function(x) {{ return x !== selectedBalloon; }});
            selectedBalloon = null;
            return;
          }}
          if (selectedMeasure) {{
            try {{ map.removeLayer(selectedMeasure); }} catch(e){{}}
            selectedMeasure = null;
          }}
        }}

        // 🗑 kontrolü (geri geldi)
        var DelCtl = L.Control.extend({{
          options: {{ position: 'topleft' }},
          onAdd: function() {{
            var div = L.DomUtil.create('div', 'leaflet-bar');
            var a = L.DomUtil.create('a', '', div);
            a.href = '#';
            a.title = 'Seçili öğeyi sil';
            a.innerHTML = '🗑';
            a.style.fontSize = '18px';
            a.style.lineHeight = '26px';
            a.style.textAlign = 'center';
            a.style.width = '30px';
            a.style.height = '30px';
            L.DomEvent.on(a, 'click', L.DomEvent.stop)
                     .on(a, 'click', function() {{ deleteSelected(); }});
            return div;
          }}
        }});
        map.addControl(new DelCtl());

        window.addEventListener('keydown', function(ev) {{
          if (ev.key === 'Delete') deleteSelected();
        }});

        // -------------------------
        // 3) Balon oluşturma + 4 yön ok
        // -------------------------
        function makeAnchor(latlng) {{
          var anchorIcon = L.divIcon({{
            className: '',
            html: '<div style="width:12px;height:12px;border-radius:50%;border:2px solid #e74c3c;background:#fff;box-sizing:border-box;"></div>',
            iconSize: [12,12],
            iconAnchor: [6,6]
          }});
          return L.marker(latlng, {{ icon: anchorIcon, draggable: true }});
        }}

        function addBalloon(labelLatLng, anchorLatLng, text) {{
          var icon = L.divIcon({{
            className: 'hts-balloon-icon',
            html: '<div class="hts-balloon" data-tail="bottom">' + esc(text) + '</div>',
            iconSize: [1,1]
          }});

          var label = L.marker(labelLatLng, {{ icon: icon, draggable: true }}).addTo(map);
          var anchor = makeAnchor(anchorLatLng).addTo(map);

          var line = L.polyline([anchor.getLatLng(), label.getLatLng()], {{
            color: '#e74c3c',
            weight: 3,
            opacity: 0.95
          }}).addTo(map);
          line._hts_balloon = true;

          var b = {{ label: label, anchor: anchor, line: line, text: text }};
          balloons.push(b);

          function syncTailAndLine() {{
            var el = label.getElement();
            if (!el) {{
              line.setLatLngs([anchor.getLatLng(), label.getLatLng()]);
              return;
            }}

            var balloon = el.querySelector('.hts-balloon') || el;
            var mapRect = map.getContainer().getBoundingClientRect();
            var r = balloon.getBoundingClientRect();

            var left   = (r.left - mapRect.left);
            var right  = (r.right - mapRect.left);
            var top    = (r.top - mapRect.top);
            var bottom = (r.bottom - mapRect.top);

            var ap = map.latLngToContainerPoint(anchor.getLatLng());
            var cx = (left + right) / 2;
            var cy = (top + bottom) / 2;
            var dx = ap.x - cx;
            var dy = ap.y - cy;

            var side;
            if (Math.abs(dx) > Math.abs(dy)) {{
              side = (dx < 0) ? "left" : "right";
            }} else {{
              side = (dy < 0) ? "top" : "bottom";
            }}

            var pad = 18;
            var tailX = ap.x, tailY = ap.y;

            if (side === "bottom" || side === "top") {{
              tailX = Math.max(left + pad, Math.min(ap.x, right - pad));
              tailY = (side === "bottom") ? bottom : top;
              balloon.style.setProperty("--tail-x", (tailX - left) + "px");
              balloon.style.setProperty("--tail-y", "50%");
            }} else {{
              tailY = Math.max(top + pad, Math.min(ap.y, bottom - pad));
              tailX = (side === "right") ? right : left;
              balloon.style.setProperty("--tail-y", (tailY - top) + "px");
              balloon.style.setProperty("--tail-x", "50%");
            }}

            balloon.setAttribute("data-tail", side);

            var tailLL = map.containerPointToLatLng([tailX, tailY]);
            line.setLatLngs([anchor.getLatLng(), tailLL]);
          }}

          label.on('add', function() {{ setTimeout(syncTailAndLine, 0); }});
          label.on('drag', syncTailAndLine);
          label.on('dragend', syncTailAndLine);
          anchor.on('drag', syncTailAndLine);
          anchor.on('dragend', syncTailAndLine);
          map.on('zoomend moveend', syncTailAndLine);

          setTimeout(syncTailAndLine, 0);

          // seçim
          label.on('click', function() {{ selectBalloon(b); }});
          anchor.on('click', function() {{ selectBalloon(b); }});

          // düzenle (tıklayınca)
          label.on('click', function() {{
            selectBalloon(b);
            openMessageModal("Mesajı Düzenle", b.text, function(newText) {{
              b.text = newText;
              var newIcon = L.divIcon({{
                className: 'hts-balloon-icon',
                html: '<div class="hts-balloon" data-tail="bottom">' + esc(newText) + '</div>',
                iconSize: [1,1]
              }});
              label.setIcon(newIcon);
              setTimeout(syncTailAndLine, 0);
            }});
          }});

          return b;
        }}

        // Qt sağ tık / menü yüzünden dblclick
        if (map.doubleClickZoom) map.doubleClickZoom.disable();

        // Çift tıkla balon ekle
        map.on('dblclick', function(e) {{
          openMessageModal("Mesaj Baloncuğu", "", function(msg) {{
            // label'ı otomatik offset ile yerleştir (anchor = tıklanan yer)
            var p = map.latLngToContainerPoint(e.latlng);
            var p2 = L.point(p.x + 160, p.y - 80);
            var ll2 = map.containerPointToLatLng(p2);
            addBalloon(ll2, e.latlng, msg);
          }});
        }});

        // -------------------------
        // 4) Ölçüm çizgisi/alanı seçimi (tıkla seç)
        // -------------------------
        map.on('layeradd', function(ev) {{
          var layer = ev.layer;
          if (!layer) return;

          // kendi balon çizgimizi hariç tut
          if (layer._hts_balloon) return;

          // measure sonucu polyline/polygon olabilir
          if (layer instanceof L.Polyline || layer instanceof L.Polygon) {{
            layer.on('click', function() {{
              selectMeasure(layer);
            }});
          }}
        }});

      }});
    }})();
    """))


def enforce_normal_table_fonts(root: QWidget):
    from PyQt6.QtWidgets import QTableWidget, QTableView
    normal = QFont("Segoe UI", 10, QFont.Weight.Normal)

    for t in root.findChildren(QTableWidget):
        t.setFont(normal)

    for t in root.findChildren(QTableView):
        t.setFont(normal)


def apply_menu_theme(menu: QMenu):
    menu.setStyleSheet("""
        QMenu {
            background-color: #ffffff;
            color: #2c3e50;
            border: 1px solid #bdc3c7;
            border-radius: 6px;
            padding: 6px;
        }
        QMenu::item {
            background-color: transparent;
            padding: 8px 18px;
            border-radius: 4px;
            color: #2c3e50;
        }
        QMenu::item:selected {
            background-color: #e8f6f3;
            color: #16a085;
        }
        QMenu::separator {
            height: 1px;
            background: #ecf0f1;
            margin: 6px 8px;
        }
    """)
    return menu


def apply_light_combobox_popup(combo: QComboBox):
    """
    Windows tema/palette bazen ComboBox popup view'ını karartıyor.
    Bu fonksiyon her ComboBox için popup view'ı QListView yapıp
    palette + QSS ile %100 beyaz/okunaklı sabitler.
    """
    from PyQt6.QtWidgets import QListView

    view = QListView()
    combo.setView(view)

    view.setStyleSheet("""
        QListView {
            background: #ffffff;
            color: #2c3e50;
            border: 1px solid #bdc3c7;
            outline: none;
        }
        QListView::item {
            padding: 6px 10px;
        }
        QListView::item:selected {
            background: #e8f6f3;
            color: #16a085;
        }
    """)

    pal = view.palette()
    pal.setColor(QPalette.ColorRole.Base, QColor("#ffffff"))
    pal.setColor(QPalette.ColorRole.Text, QColor("#2c3e50"))
    pal.setColor(QPalette.ColorRole.Highlight, QColor("#e8f6f3"))
    pal.setColor(QPalette.ColorRole.HighlightedText, QColor("#16a085"))
    view.setPalette(pal)

    view.setFont(QFont("Segoe UI", 10, QFont.Weight.Normal))
    combo.setFont(QFont("Segoe UI", 10, QFont.Weight.Normal))


class DraggableConnector(MacroElement):
    """
    Baloncuk sürüklenince ve harita zoom/pan olunca,
    sabit raptiye ile baloncuk arasındaki çizgiyi sürekli günceller.
    """
    def __init__(self, fmap, label_marker, connector_line, fixed_lat, fixed_lon):
        super().__init__()
        self._name = "DraggableConnector"
        self.fmap = fmap
        self.label_marker = label_marker
        self.connector_line = connector_line
        self.fixed_lat = fixed_lat
        self.fixed_lon = fixed_lon

        self._template = Template(u"""
        {% macro script(this, kwargs) %}
            var map = {{ this.fmap.get_name() }};

            map.whenReady(function () {
                var label = {{ this.label_marker.get_name() }};
                var line  = {{ this.connector_line.get_name() }};
                var fixed = L.latLng({{ this.fixed_lat }}, {{ this.fixed_lon }});

                // Drag'i garantiye al (DivIcon'da bazen otomatik açılmıyor)
                if (label.dragging) { label.dragging.enable(); }

                function syncLine() {
                    if (!label || !line) return;
                    var p = label.getLatLng();
                    line.setLatLngs([fixed, p]);
                }

                // İlk yükleme sırasında baloncuk sonradan kayarsa yakala
                syncLine();
                var c = 0;
                var t = setInterval(function(){
                    syncLine();
                    c++;
                    if (c >= 10) clearInterval(t);  // ~1 saniye
                }, 100);

                // Drag ile lastik çizgi
                label.on('drag', syncLine);
                label.on('dragend', syncLine);

                // Zoom/pan sonrası da doğru kalsın
                map.on('zoomend moveend', syncLine);
            });
        {% endmacro %}
        """)


class EvidenceWebEngineView(QWebEngineView):
    """Sağ tıklandığında 'Rapora Ekle' menüsü açan özel Web Görüntüleyici."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

    def show_context_menu(self, pos):
        menu = QMenu(self)
        menu = apply_menu_theme(QMenu(self))
        action_add = QAction("📸 Görüntüyü Delil Olarak Rapora Ekle", self)
        action_add.triggered.connect(self.capture_evidence)
        menu.addAction(action_add)

        menu.exec(self.mapToGlobal(pos))

    def capture_evidence(self):
        if not LicenseManager.require_valid_or_exit(self, "Delil ekleme"):
            return
        """
        Akıllı Yakalama: 
        - Harita/Detay pencerelerinde -> TÜM PENCEREYİ (Başlık dahil) alır.
        - Grafik pencerelerinde -> SADECE GRAFİĞİ alır.
        """

        target_widget = self
        window_title = "Görsel Analiz"

        curr = self
        while curr:
            if isinstance(curr, QDialog) or isinstance(curr, QMainWindow):
                window_title = curr.windowTitle()
                c_name = curr.__class__.__name__

                if c_name in [
                    "MapPreviewDialog", "LocationDetailDialog",
                    "MapMultiPointDialog", "ProfileMapDialog",
                    "StalkingAnalysisDialog", "DailyRouteDialog",
                    "ImeiDetailDialog", "ReciprocalDetailDialog"
                ]:
                    target_widget = curr

                break
            curr = curr.parent()

        pixmap = target_widget.grab()

        # ✅ DailyRouteDialog: üst kontrol şeridini (top_frame) kırp, info kartı + harita kalsın
        try:
            if (hasattr(target_widget, "__class__") and target_widget.__class__.__name__ == "DailyRouteDialog"
                    and hasattr(target_widget, "top_frame") and target_widget.top_frame is not None):
                g = target_widget.top_frame.geometry()
                cut_y = max(0, g.bottom() + 1)
                if 0 < cut_y < pixmap.height():
                    pixmap = pixmap.copy(0, cut_y, pixmap.width(), pixmap.height() - cut_y)
        except Exception:
            pass

        # İşlemeye gönder
        self._on_pixmap_ready(pixmap, window_title)

    def _on_pixmap_ready(self, pixmap, title):
        """Ekran görüntüsü alındığında çalışır, AnalysisCenter'a gönderir."""
        parent = self.parent()
        while parent:
            if hasattr(parent, 'main') and hasattr(parent.main.page_analysis, 'add_web_view_evidence'):
                parent.main.page_analysis.add_web_view_evidence(pixmap, title)
                return
            parent = parent.parent()
            if parent is None:
                 break

class WatermarkBackground(QWidget):
    """
    Tüm uygulamanın üstünde tek bir global watermark logo.
    Overlay olarak kullanılır; mouse eventlerini engellemez.
    """
    def __init__(self, logo_path, opacity=0.06, scale=0.85, parent=None):
        super().__init__(parent)
        self.logo_path = logo_path
        self.opacity = opacity
        self.scale = scale
        self._pixmap = QPixmap(self.logo_path)

        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        self.setAttribute(Qt.WidgetAttribute.WA_NoSystemBackground, True)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)

    def resizeEvent(self, event):
        if self.parent():
            self.setGeometry(self.parent().rect())
        super().resizeEvent(event)

    def paintEvent(self, event):
        if self._pixmap.isNull():
            return

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform, True)
        painter.setOpacity(self.opacity)

        w = self.width()
        h = self.height()
        if w <= 0 or h <= 0:
            painter.end()
            return

        target_w = int(w * self.scale)
        target_h = int(target_w * self._pixmap.height() / self._pixmap.width())

        if target_h > int(h * self.scale):
            target_h = int(h * self.scale)
            target_w = int(target_h * self._pixmap.width() / self._pixmap.height())

        x = (w - target_w) // 2
        y = (h - target_h) // 2

        painter.drawPixmap(x, y, target_w, target_h, self._pixmap)
        painter.end()


class AnalysisUtils:
    """Tüm sınıfların ortak kullandığı ağır analiz işlemleri."""

    @staticmethod
    def recalculate_common_analysis_core(project_id):
        try:
            with DB() as conn:
                pid = project_id
                cur = conn.cursor()

                cur.execute("DELETE FROM hts_ortak_imei WHERE ProjeID=?", (pid,))
                cur.execute("DELETE FROM hts_ortak_isim WHERE ProjeID=?", (pid,))
                cur.execute("DELETE FROM hts_ortak_tc   WHERE ProjeID=?", (pid,))
                conn.commit()

                has_any = (
                    cur.execute("SELECT 1 FROM hts_gsm   WHERE ProjeID=? LIMIT 1", (pid,)).fetchone()
                    or cur.execute("SELECT 1 FROM hts_gprs  WHERE ProjeID=? LIMIT 1", (pid,)).fetchone()
                    or cur.execute("SELECT 1 FROM hts_wap   WHERE ProjeID=? LIMIT 1", (pid,)).fetchone()
                    or cur.execute("SELECT 1 FROM hts_rehber WHERE ProjeID=? LIMIT 1", (pid,)).fetchone()
                )
                if not has_any:
                    print("ℹ️ Ortak Analiz (Core): Projede veri kalmadığı için hesaplama pas geçildi.")
                    return

                common_imei_sql = """
                    INSERT INTO hts_ortak_imei (ProjeID, IMEI, KullananSayisi, Numaralar, ToplamKullanim)
                    SELECT ?, IMEI,
                           COUNT(DISTINCT CleanNum) AS KullananSayisi,
                           GROUP_CONCAT(DISTINCT CleanNum) AS Numaralar,
                           COUNT(*) AS ToplamKullanim
                    FROM (
                        SELECT IMEI,
                               substr(replace(replace(replace(NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AS CleanNum
                        FROM hts_gsm
                        WHERE ProjeID=? AND LENGTH(IMEI) > 10

                        UNION ALL

                        SELECT IMEI,
                               substr(replace(replace(replace(NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AS CleanNum
                        FROM hts_gprs
                        WHERE ProjeID=? AND LENGTH(IMEI) > 10

                        UNION ALL

                        SELECT IMEI,
                               substr(replace(replace(replace(NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AS CleanNum
                        FROM hts_wap
                        WHERE ProjeID=? AND LENGTH(IMEI) > 10
                    )
                    GROUP BY IMEI
                    HAVING COUNT(DISTINCT CleanNum) > 1
                """
                cur.execute(common_imei_sql, (pid, pid, pid, pid))

                tc_valid_expr = """
                    TRIM(TC) GLOB '[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]'
                """

                cur.execute(f"""
                    INSERT INTO hts_ortak_isim (ProjeID, AdSoyad, HatSayisi, Numaralar)
                    SELECT ?, Isim,
                           COUNT(DISTINCT KarsiNo) AS HatSayisi,
                           GROUP_CONCAT(DISTINCT KarsiNo) AS Numaralar
                    FROM hts_rehber
                    WHERE ProjeID=?
                      AND LENGTH(Isim) > 1
                      AND TC IS NOT NULL
                      AND {tc_valid_expr}
                    GROUP BY Isim
                    HAVING COUNT(DISTINCT KarsiNo) > 1
                """, (pid, pid))

                rows_tc = cur.execute(f"""
                    SELECT r.TC,
                           COUNT(DISTINCT r.KarsiNo) AS HatSayisi,
                           GROUP_CONCAT(DISTINCT r.KarsiNo) AS Numaralar,
                           (SELECT Isim
                            FROM hts_rehber r2
                            WHERE r2.ProjeID=? AND r2.TC = r.TC
                              AND r2.TC IS NOT NULL
                              AND {tc_valid_expr}
                            LIMIT 1) AS AnyName
                    FROM hts_rehber r
                    WHERE r.ProjeID=?
                      AND r.TC IS NOT NULL
                      AND {tc_valid_expr}
                    GROUP BY r.TC
                    HAVING COUNT(DISTINCT r.KarsiNo) > 1
                    ORDER BY HatSayisi DESC
                """, (pid, pid)).fetchall()

                data_tc = []
                for tc_val, count, nums, any_name in rows_tc:
                    display = f"{tc_val} - {any_name}" if any_name else str(tc_val)
                    data_tc.append([pid, display, count, nums])

                if data_tc:
                    cur.executemany(
                        "INSERT INTO hts_ortak_tc (ProjeID, TC, HatSayisi, Numaralar) VALUES (?,?,?,?)",
                        data_tc
                    )

                conn.commit()
                print("✅ Ortak Analiz (Core) başarıyla tamamlandı.")

        except Exception as e:
            print(f"❌ [recalculate_common_analysis_core] Kritik Hata: {e}")

    @staticmethod
    def delete_gsm_records_core(project_id, gsm_number):
        """
        Belirtilen GSM'e ait tüm verileri siler.
        NOT: ozel_konumlar GSM silmede KESİNLİKLE silinmez (sadece proje silinince silinsin isteği).
        """
        try:
            tabs_gsm = [
                "hts_dosyalari",
                "hts_abone", "hts_gsm",
                "hts_sms", "hts_sabit", "hts_gprs", "hts_wap", "hts_sth",
                "hts_uluslararasi", "hts_ozet", "hts_ozet_iletisim",
                "hts_ozet_baz", "hts_ozet_imei", "hts_rehber", "hts_tum_baz",
            ]

            with DB() as conn:
                for t in tabs_gsm:
                    try:
                        conn.execute(f"DELETE FROM {t} WHERE ProjeID=? AND GSMNo=?", (project_id, gsm_number))
                    except Exception as e:
                        print(f"⚠️ [delete_gsm_records_core] Tablo '{t}' temizlenirken hata: {e}")

                conn.commit()
            return True
        except Exception as e:
            print(f"❌ [delete_gsm_records_core] Genel Hata: {e}")
            return False

    @staticmethod
    def perform_maintenance():
        """
        Veritabanını bakım modunda sıkıştırır.
        WAL modunda güvenli bakım:
        1) CHECKPOINT (WAL -> ana db)
        2) VACUUM
        """
        try:
            mgr = DatabaseManager()
            lock = mgr.lock()
            lock.acquire()
            try:
                conn = mgr.get_connection()
                cur = conn.cursor()

                try:
                    cur.execute("PRAGMA wal_checkpoint(TRUNCATE);")
                except Exception:
                    cur.execute("PRAGMA wal_checkpoint(FULL);")

                cur.execute("VACUUM;")
                # !!! ÖNEMLİ: Burası singleton bağlantı. KAPATMAYIN.
                # conn.close()

                return True
            finally:
                lock.release()

        except Exception as e:
            print(f"❌ [perform_maintenance] Bakım Hatası: {e}")
            return False

    @staticmethod
    def project_has_any_gsm(project_id: int) -> bool:
        if not project_id:
            return False

        try:
            with DB() as conn:
                cur = conn.cursor()
                r = cur.execute("SELECT 1 FROM hts_dosyalari WHERE ProjeID=? LIMIT 1", (project_id,)).fetchone()
                if r:
                    return True
                r = cur.execute("SELECT 1 FROM hts_ozet WHERE ProjeID=? LIMIT 1", (project_id,)).fetchone()
                if r:
                    return True
                r = cur.execute("SELECT 1 FROM hts_gsm WHERE ProjeID=? LIMIT 1", (project_id,)).fetchone()
                return bool(r)
        except Exception as e:
            print(f"project_has_any_gsm hata: {e}")
            return False


class LocalTileServer:
    """
    .mbtiles dosyasını okuyup yerel ağda (localhost) harita karoları sunan sunucu.
    GÜNCELLEME: SQLite bağlantısı Thread Local Storage ile önbelleğe alındı (Performans artışı).
    """
    def __init__(self, mbtiles_path, port=8080):
        self.mbtiles_path = mbtiles_path
        self.port = port
        self.server = None
        self.thread = None

    def start(self):
        """Sunucuyu arka planda başlatır."""
        try:
            handler = self.TileRequestHandler
            handler.db_path = self.mbtiles_path

            self.server = HTTPServer(('localhost', self.port), handler)
            self.thread = threading.Thread(target=self.server.serve_forever)
            self.thread.daemon = True
            self.thread.start()
            print(f"Harita Sunucusu Başlatıldı: http://localhost:{self.port}")
            return True
        except Exception as e:
            print(f"Harita Sunucusu Başlatılamadı: {e}")
            return False

    def stop(self):
        if self.server:
            self.server.shutdown()
            self.server.server_close()

    class TileRequestHandler(BaseHTTPRequestHandler):
        db_path = ""

        _thread_local = threading.local()

        def log_message(self, format, *args):
            return
        def do_GET(self):
            try:
                parts = self.path.strip("/").split("/")
                if len(parts) < 3: return

                zoom = int(parts[0])
                x = int(parts[1])
                y = int(parts[2].split(".")[0])

                tms_y = (2 ** zoom) - 1 - y

                if not hasattr(self._thread_local, "conn"):
                    self._thread_local.conn = sqlite3.connect(self.db_path, check_same_thread=False)

                conn = self._thread_local.conn
                cur = conn.cursor()

                cur.execute("SELECT tile_data FROM tiles WHERE zoom_level=? AND tile_column=? AND tile_row=?", (zoom, x, tms_y))
                row = cur.fetchone()

                if row:
                    self.send_response(200)
                    self.send_header('Content-type', 'image/png')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(row[0])
                else:
                    self.send_error(404)
            except:
                self.send_error(500)


class SleekTooltipPopup(QWidget):
    """Şık, gölgeli ve DIŞARI TIKLAYINCA KAPANAN modern bilgi baloncuğu."""
    def __init__(self, text, parent=None):
        super().__init__(parent)
        self.setWindowFlags(
            Qt.WindowType.ToolTip |
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)

        self.container = QFrame()
        self.container.setObjectName("TooltipContainer")
        container_layout = QVBoxLayout(self.container)
        container_layout.setContentsMargins(12, 10, 12, 10)

        self.label = QLabel(str(text))
        self.label.setObjectName("TooltipLabel")
        self.label.setWordWrap(True)

        container_layout.addWidget(self.label)
        layout.addWidget(self.container)

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 80))
        self.container.setGraphicsEffect(shadow)

        self.setStyleSheet("""
            QFrame#TooltipContainer {
                background: rgba(255, 255, 255, 245);
                border: 1px solid #3498db;
                border-radius: 8px;
            }
            QLabel#TooltipLabel {
                color: #2c3e50;
                font-size: 12px;
            }
        """)

        QApplication.instance().installEventFilter(self)

    def set_text(self, text: str):
        self.label.setText(str(text))
        self.adjustSize()

    def show_at(self, global_pos: QPoint):
        """
        InfoButton.show_popup() bunu çağırıyor.
        Tooltip’i global ekranda düzgün konumlandırıp gösterir.
        """
        self.adjustSize()

        # Ekran dışına taşmayı engelle
        try:
            screen = QApplication.primaryScreen().availableGeometry()
            x, y = global_pos.x(), global_pos.y()

            if x + self.width() > screen.right():
                x = screen.right() - self.width() - 10
            if y + self.height() > screen.bottom():
                y = global_pos.y() - self.height() - 10  # yukarı al
            if x < screen.left():
                x = screen.left() + 10
            if y < screen.top():
                y = screen.top() + 10

            self.move(QPoint(x, y))
        except Exception:
            self.move(global_pos)

        self.show()
        self.raise_()

    def eventFilter(self, source, event):
        if event.type() == QEvent.Type.MouseButtonPress and self.isVisible():
            click_pos = event.globalPosition().toPoint()
            if not self.geometry().contains(click_pos):
                parent_btn = self.parent()
                if parent_btn and isinstance(parent_btn, QWidget):
                    btn_rect = QRect(parent_btn.mapToGlobal(QPoint(0, 0)), parent_btn.size())
                    if btn_rect.contains(click_pos):
                        return super().eventFilter(source, event)
                self.hide()
        return super().eventFilter(source, event)


class InfoButton(QPushButton):
    """Örnekteki gibi minimalist soru işareti butonu ve şık baloncuk tetikleyici."""
    def __init__(self, text_content, parent=None):
        super().__init__("?", parent)
        self.text_content = text_content
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedSize(24, 24)

        self.setObjectName("InfoBtn")

        self.setStyleSheet("""
            QPushButton#InfoBtn {
                background-color: transparent;
                color: #7f8c8d;
                border: 2px solid #bdc3c7;
                border-radius: 12px; /* Tam yuvarlak olması için (24px / 2) */
                font-family: 'Segoe UI', sans-serif;
                font-size: 14px;
                font-weight: bold;
                padding: 0px;
                margin: 0px;
                text-align: center;
            }
            QPushButton#InfoBtn:hover {
                color: #2c3e50;
                border-color: #95a5a6;
                background-color: #ecf0f1;
            }
        """)

        self.setToolTip("")

        self.popup = SleekTooltipPopup(text_content, self)
        self.clicked.connect(self.toggle_popup)

    def enterEvent(self, event):
        super().enterEvent(event)

    def leaveEvent(self, event):
        super().leaveEvent(event)

    def toggle_popup(self):
        if self.popup.isVisible():
            self.popup.hide()
        else:
            self.show_popup()

    def show_popup(self):
        global_pos = self.mapToGlobal(self.rect().bottomLeft())
        self.popup.show_at(global_pos)


class TooltipManager(QObject):
    """
    QToolTip (native) yerine tüm tooltip'leri tek tip SleekTooltipPopup ile gösterir.
    Böylece Windows tema/palet dalgalanması %100 baskılanır.

    Ayrıca QTableView/QTableWidget gibi item-view hücre tooltiplerini de (Qt.ToolTipRole) yakalar.
    """
    def __init__(self, app: QApplication):
        super().__init__(app)
        self.app = app
        self.popup = SleekTooltipPopup("", None)

        try:
            QToolTip.setEnabled(False)
        except Exception:
            pass

        app.installEventFilter(self)

    def _extract_tooltip_text(self, obj, event) -> str:
        """Widget tooltip + ItemView (table/list) hücre tooltip desteği."""
        if isinstance(obj, QWidget):
            t = obj.toolTip()
            if t:
                return t

        try:
            from PyQt6.QtWidgets import QAbstractItemView
            from PyQt6.QtGui import QHelpEvent
            if isinstance(obj, QAbstractItemView) and isinstance(event, QHelpEvent):
                idx = obj.indexAt(event.pos())
                if idx.isValid():
                    data = idx.data(Qt.ItemDataRole.ToolTipRole)
                    if data:
                        return str(data)
        except Exception:
            pass

        return ""

    def eventFilter(self, obj, event):
        et = event.type()

        if et == QEvent.Type.ToolTip:
            text = self._extract_tooltip_text(obj, event)

            if not text:
                return False

            self.popup.set_text(text)

            from PyQt6.QtGui import QCursor
            pos = QCursor.pos() + QPoint(12, 18)
            self.popup.move(pos)
            self.popup.show()
            return True

        if et in (QEvent.Type.Leave, QEvent.Type.HoverLeave):
            self.popup.hide()

        if et == QEvent.Type.MouseMove:
            pass

        return super().eventFilter(obj, event)


class LoadingOverlay(QWidget):
    """
    Ekranı kilitler, YÜKSEK KALİTELİ (Pürüzsüz) büyüteç animasyonu,
    sabitlenmiş metin ve gradyanlı ilerleme çubuğu gösterir.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, False)
        self.setAttribute(Qt.WidgetAttribute.WA_NoSystemBackground, False)

        self.hide()

        main_layout = QVBoxLayout(self)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        container = QWidget()
        container.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(container)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(20)

        self.anim_label = QLabel()
        self.anim_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.anim_label.setFixedSize(120, 120)

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(30)
        shadow.setColor(QColor(255, 255, 255, 100))
        shadow.setOffset(0, 0)
        self.anim_label.setGraphicsEffect(shadow)
        # -------------------------------------------

        gif_path = os.path.join(APP_DIR, "assets", "loading_lens.gif")

        self.movie = None
        if os.path.exists(gif_path):
            self.movie = QMovie(gif_path)
            self.movie.frameChanged.connect(self.update_animation_frame)
        else:
            self.anim_label.setText("🔍")
            self.anim_label.setStyleSheet("font-size: 80px; color: white; background: transparent;")

        layout.addWidget(self.anim_label, 0, Qt.AlignmentFlag.AlignCenter)

        self.text_label = QLabel("Lütfen Bekleyiniz...")
        self.text_label.setAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignTop)
        self.text_label.setWordWrap(True)
        self.text_label.setFixedSize(600, 60)

        self.text_label.setStyleSheet("""
            QLabel {
                color: white;
                font-family: 'Segoe UI';
                font-size: 16px;
                font-weight: bold;
                background-color: transparent;
                padding: 0px; margin: 0px;
            }
        """)
        layout.addWidget(self.text_label, 0, Qt.AlignmentFlag.AlignCenter)

        self.pbar = QProgressBar()
        self.pbar.setFixedWidth(400)
        self.pbar.setFixedHeight(25)
        self.pbar.setTextVisible(True)
        self.pbar.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.pbar.setStyleSheet("""
            QProgressBar {
                border: 2px solid rgba(255, 255, 255, 180);
                border-radius: 12px;
                text-align: center;
                color: white;
                background-color: rgba(30, 30, 30, 200);
                font-weight: bold;
                font-size: 13px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #2ecc71, stop:1 #27ae60);
                border-radius: 10px;
                margin: 2px;
            }
        """)
        self.pbar.hide()

        layout.addWidget(self.pbar, 0, Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(container)

    def update_animation_frame(self):
        """GIF'in o anki karesini alır, pürüzsüzleştirerek yeniden boyutlandırır."""
        pixmap = self.movie.currentPixmap()
        if not pixmap.isNull():
            scaled_pixmap = pixmap.scaled(
                QSize(120, 120),
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            )
            self.anim_label.setPixmap(scaled_pixmap)

    @property
    def text(self):
        return self.text_label.text()

    @text.setter
    def text(self, value):
        self.text_label.setText(str(value))
        QApplication.processEvents()

    def paintEvent(self, event):
        """Arka planı radyal gradyan ile karart (Vignette efekti)."""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        gradient = QRadialGradient(self.width() / 2, self.height() / 2, self.width() / 1.5)
        gradient.setColorAt(0, QColor(40, 40, 40, 180))
        gradient.setColorAt(1, QColor(0, 0, 0, 240))

        painter.fillRect(self.rect(), gradient)

    def start(self, text="İşlem Yapılıyor..."):
        self.text_label.setText(text)
        self.pbar.setValue(0)
        self.pbar.hide()

        if self.parent():
            self.resize(self.parent().size())
        self.raise_()
        self.show()
        QApplication.processEvents()

        if self.movie:
            if self.movie.state() == QMovie.MovieState.Running:
                self.movie.stop()
            self.movie.start()

    def set_progress(self, value):
        if value > 0:
            if self.pbar.isHidden():
                self.pbar.show()
            self.pbar.setValue(value)

    def stop(self):
        if self.movie:
            self.movie.stop()
        self.hide()


class FileDetailPopup(QDialog):
    """Dosya bilgilerini geniş ekranda gösteren detay penceresi. (Overlay logo)"""
    def __init__(self, parent, data, title: str = "Detaylı Dosya Bilgileri", info_text: str | None = None):
        super().__init__(parent)

        self.setWindowTitle(title)
        self.resize(800, 600)
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint)

        central = QWidget(self)
        stack_layout = QStackedLayout(central)
        stack_layout.setStackingMode(QStackedLayout.StackingMode.StackAll)

        content = QWidget(central)
        layout = QVBoxLayout(content)

        default_info = (
            "📄 <b>Dosya Detayları:</b> Aşağıdaki tabloda dosya meta verileri ve hash bilgileri yer almaktadır.\n"
            "Metinleri kopyalamak için hücreye çift tıklayıp seçebilirsiniz."
        )
        if info_text is None:
            info_text = default_info

        self.info_label = QLabel(info_text)
        self.info_label.setWordWrap(True)
        self.info_label.setStyleSheet(
            "background-color: #e8f6f3; padding: 12px; border-radius: 6px; "
            "color: #16a085; font-size: 13px;"
        )
        layout.addWidget(self.info_label)

        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Özellik", "Değer"])
        table.verticalHeader().setVisible(False)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("font-size: 13px; font-family: 'Segoe UI';")

        table.setRowCount(len(data))
        font_bold = QFont("Segoe UI", 10, QFont.Weight.Bold)

        for i, (k, v) in enumerate(data):
            item_k = QTableWidgetItem(k)
            item_k.setFont(font_bold)
            item_k.setBackground(QColor("#f2f2f7"))

            item_v = QTableWidgetItem(v)

            if "MD5" in k or "SHA" in k:
                item_v.setForeground(QColor("#e74c3c"))
                item_v.setFont(QFont("Consolas", 10))

            table.setItem(i, 0, item_k)
            table.setItem(i, 1, item_v)

        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        layout.addWidget(table)

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet(
            "background-color:#34495e; color:white; font-weight:bold; padding:8px;"
        )
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)

        stack_layout.addWidget(content)

        base_dir = APP_DIR
        logo_path = os.path.join(base_dir, "assets", "bg_logo.png")
        if not os.path.exists(logo_path):
            logo_path = os.path.join(base_dir, "assets", "logo.png")

        self.watermark = WatermarkBackground(
            logo_path=logo_path,
            opacity=0.03,
            scale=1.20,
            parent=central
        )
        stack_layout.addWidget(self.watermark)
        self.watermark.raise_()

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(central)


class ElidedItemDelegate(QStyledItemDelegate):
    """
    Metin sığmazsa '.' koyar ve üzerine gelince bizim SleekTooltipPopup ile gösterir.
    (Windows tema bağımsız, %100 sabit görünüm)
    """
    _popup = None

    def paint(self, painter, option, index):
        if not index.isValid():
            return

        painter.save()

        if option.state & QStyle.StateFlag.State_Selected:
            painter.fillRect(option.rect, option.palette.highlight())
            painter.setPen(option.palette.highlightedText().color())
        else:
            painter.setPen(option.palette.text().color())

        text = index.data(Qt.ItemDataRole.DisplayRole)
        if text:
            elided_text = painter.fontMetrics().elidedText(
                str(text),
                Qt.TextElideMode.ElideRight,
                option.rect.width() - 10
            )
            painter.drawText(
                option.rect.adjusted(5, 0, -5, 0),
                Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                elided_text
            )

        painter.restore()

    def helpEvent(self, event, view, option, index):
        if not event or not view:
            return False

        if event.type() == QEvent.Type.ToolTip:
            text = index.data(Qt.ItemDataRole.DisplayRole)
            if not text:
                return False

            if ElidedItemDelegate._popup is None:
                ElidedItemDelegate._popup = SleekTooltipPopup("", None)

            ElidedItemDelegate._popup.set_text(str(text))

            pos = event.globalPos() + QPoint(12, 18)
            ElidedItemDelegate._popup.move(pos)
            ElidedItemDelegate._popup.show()
            return True

        if event.type() in (QEvent.Type.Leave, QEvent.Type.MouseMove):
            if ElidedItemDelegate._popup is not None:
                ElidedItemDelegate._popup.hide()

        return super().helpEvent(event, view, option, index)


class WrapTextDelegate(QStyledItemDelegate):
    """
    Hücre içi metni alt satıra kaydırarak çizer + satır yüksekliğini doğru hesaplar.
    (BAZ kolonlarında kullanılacak)
    """
    def paint(self, painter, option, index):
        if not index.isValid():
            return

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)

        painter.save()

        # Arkaplan (seçili satır vb.) çiz
        style = opt.widget.style() if opt.widget else QApplication.style()
        style.drawPrimitive(QStyle.PrimitiveElement.PE_PanelItemViewItem, opt, painter, opt.widget)

        # Metni wrap'li çiz
        text = opt.text or ""
        doc = QTextDocument()
        doc.setDefaultFont(opt.font)
        doc.setPlainText(text)
        doc.setTextWidth(max(10, opt.rect.width() - 8))

        painter.translate(opt.rect.left() + 4, opt.rect.top() + 2)
        clip = QRectF(0, 0, opt.rect.width() - 8, opt.rect.height() - 4)
        doc.drawContents(painter, clip)

        painter.restore()

    def sizeHint(self, option, index):
        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)

        text = opt.text or ""
        doc = QTextDocument()
        doc.setDefaultFont(opt.font)
        doc.setPlainText(text)

        # view genişliğine göre satır yüksekliği hesapla
        view = opt.widget
        if view is not None and hasattr(view, "columnWidth"):
            w = max(10, view.columnWidth(index.column()) - 8)
        else:
            w = max(10, opt.rect.width() - 8)

        doc.setTextWidth(w)
        h = int(doc.size().height()) + 6
        return QSize(w, max(24, h))


class TipIconDelegate(QStyledItemDelegate):
    """
    'CALL_OUT', 'CALL_IN', 'SMS_OUT', 'SMS_IN' gibi kodlardan
    ortadaki ikon + yön okunu çizer.
    """
    def paint(self, painter, option, index):
        if not index.isValid():
            return

        code = (index.data() or "").upper()

        # ikon + yön
        if code == "CALL_OUT":
            s = "📞➡️"
        elif code == "CALL_IN":
            s = "📞⬅️"
        elif code == "SMS_OUT":
            s = "✉️➡️"
        elif code == "SMS_IN":
            s = "✉️⬅️"
        else:
            s = "•"

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)

        painter.save()
        style = opt.widget.style() if opt.widget else QApplication.style()
        style.drawPrimitive(QStyle.PrimitiveElement.PE_PanelItemViewItem, opt, painter, opt.widget)

        # ortala
        painter.drawText(opt.rect, Qt.AlignmentFlag.AlignCenter, s)
        painter.restore()

    def sizeHint(self, option, index):
        return QSize(44, 24)


class NoElideDelegate(QStyledItemDelegate):
    """Bu kolonda '...' kullanmadan metni çiz (tooltip/ellipsis yok)."""
    def paint(self, painter, option, index):
        opt = QStyleOptionViewItem(option)
        opt.textElideMode = Qt.TextElideMode.ElideNone
        super().paint(painter, opt, index)


def _norm_header(h):
    """Başlıkları standart hale getirir (Türkçe, Boşluk, Kare Karakter Temizliği)."""
    if h is None: return ""
    s = str(h).strip().upper()
    s = s.replace("_X000D_", "").replace("\n", "").replace("\r", "")
    s = s.replace("İ", "I").replace("Ş", "S").replace("Ç", "C").replace("Ö", "O").replace("Ü", "U").replace("Ğ", "G")
    s = re.sub(r'[^A-Z0-9]', '', s)
    return s


def _normalize_msisdn(val: object) -> str:
    digits = re.sub(r"\D", "", "" if val is None else str(val))
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def _extract_gsm_from_filename(dosya_yolu):
    filename = os.path.basename(dosya_yolu)

    match = re.search(r"(5\d{9})", filename)
    if match:
        return _normalize_msisdn(match.group(1))

    match_gen = re.search(r"(\d{10,})", filename)
    return _normalize_msisdn(match_gen.group(1)) if match_gen else "BILINMIYOR"


def _detect_target_gsm(file_path):
    from openpyxl import load_workbook
    found_gsm = None
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
            if not row or len(row) < 2:
                continue

            col_a = str(row[0]).upper().replace("İ", "I") if row[0] else ""

            if "SORGULANAN" in col_a and "NO" in col_a:
                found_gsm = _normalize_msisdn(row[1])
                if found_gsm and found_gsm != "BILINMIYOR" and len(found_gsm) >= 10:
                    break

        wb.close()
    except Exception as e:
        print(f"Excel Okuma Hatası: {e}")

    if not found_gsm or found_gsm == "BILINMIYOR":
        return _extract_gsm_from_filename(file_path)

    return found_gsm


def format_size(num_bytes: int) -> str:
    """Byte cinsinden dosya boyutunu okunur forma çevirir."""
    try:
        n = float(num_bytes)
    except Exception:
        return "-"

    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if n < 1024.0:
            return f"{n:.1f} {unit}"
        n /= 1024.0
    return f"{n:.1f} PB"


def detect_hts_role(path: str) -> tuple[str, str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active

    def norm(x):
        if x is None:
            return ""
        s = str(x).strip().upper()
        s = s.replace("İ","I").replace("Ş","S").replace("Ç","C") \
             .replace("Ö","O").replace("Ü","U").replace("Ğ","G")
        return s

    def norm_space(s: str) -> str:
        s = norm(s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    sorgulanan_no = None

    for row in sheet.iter_rows(min_row=1, max_row=200, values_only=True):
        if not row or len(row) < 2:
            continue
        a = norm(row[0])
        if "SORGULANAN" in a and "NO" in a:
            raw = str(row[1]) if row[1] is not None else ""
            clean = re.sub(r"\D", "", raw)
            if len(clean) >= 10:
                sorgulanan_no = clean
                break

    if not sorgulanan_no:
        wb.close()
        raise Exception("Rol tespiti yapılamadı: 'Sorgulanan No:' bulunamadı.")

    header_row_idx = None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=400, values_only=True), start=1):
        joined = " ".join([norm(c) for c in row if c is not None])
        if "GSM GORUSME SORGU SONUCLARI" in joined:
            header_row_idx = i
            break

    if not header_row_idx:
        wb.close()
        raise Exception("Rol tespiti yapılamadı: 'GSM GÖRÜŞME SORGU SONUÇLARI' başlığı bulunamadı.")

    idx_numara = idx_diger = None
    real_header_idx = None
    for i, row in enumerate(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx+5, values_only=True), start=header_row_idx):
        cells = [norm(c) for c in row]
        if any(("NUMARA" == c or "NUMARA" in c) for c in cells) and any("DIGER" in c for c in cells):
            real_header_idx = i
            for j, c in enumerate(cells):
                if c == "NUMARA":
                    idx_numara = j
                if "DIGER" in c and "NUMARA" in c:
                    idx_diger = j
            break

    if real_header_idx is None or idx_numara is None or idx_diger is None:
        wb.close()
        raise Exception("Rol tespiti yapılamadı: NUMARA / DİĞER NUMARA sütunları bulunamadı.")

    first_data_row = None
    for row in sheet.iter_rows(min_row=real_header_idx+1, max_row=real_header_idx+50, values_only=True):
        if not row:
            continue
        v_num = re.sub(r"\D", "", str(row[idx_numara])) if row[idx_numara] else ""
        v_dig = re.sub(r"\D", "", str(row[idx_diger])) if row[idx_diger] else ""
        if len(v_num) >= 10 or len(v_dig) >= 10:
            first_data_row = row
            break

    if not first_data_row:
        role = None

        hedef_pat = norm_space("İletişimin Tespiti (Arama - Aranma - Mesaj Atma - Mesaj Alma)")
        karsi_pat = norm_space("İletişimin Tespiti (Aranma - Arama - Mesaj Alma - Mesaj Atma)")

        for row in sheet.iter_rows(min_row=1, max_row=120, values_only=True):
            if not row:
                continue

            a0 = str(row[0]) if len(row) > 0 and row[0] is not None else ""
            a1 = str(row[1]) if len(row) > 1 and row[1] is not None else ""
            raw_join = f"{a0} {a1}".strip()

            n = norm_space(raw_join)
            if "TESPIT" in n:
                if hedef_pat in n:
                    role = "HEDEF"
                    break
                if karsi_pat in n:
                    role = "KARSI"
                    break

        wb.close()

        if not role:
            role = "HEDEF"

        return role, sorgulanan_no

    wb.close()

    numara1 = re.sub(r"\D", "", str(first_data_row[idx_numara])) if first_data_row[idx_numara] else ""
    diger1  = re.sub(r"\D", "", str(first_data_row[idx_diger])) if first_data_row[idx_diger] else ""

    if numara1 == sorgulanan_no:
        return "HEDEF", sorgulanan_no
    if diger1 == sorgulanan_no:
        return "KARSI", sorgulanan_no

    return "HEDEF", sorgulanan_no


def ensure_rapor_taslagi_has_id(conn: sqlite3.Connection):
    """rapor_taslagi tablosunda 'id' yoksa ekler/düzeltir."""
    c = conn.cursor()
    tbl = c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='rapor_taslagi'").fetchone()
    if not tbl: return

    cols_info = c.execute("PRAGMA table_info(rapor_taslagi)").fetchall()
    col_names = [r[1] for r in cols_info]
    if "id" in col_names: return

    c.execute("""
        CREATE TABLE IF NOT EXISTS rapor_taslagi_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, Baslik TEXT, Icerik TEXT, 
            Tur TEXT, Tarih TEXT, Sira INTEGER, GenislikYuzde INTEGER DEFAULT 100, YukseklikMm INTEGER,
            Hizalama TEXT DEFAULT 'center', Aciklama TEXT DEFAULT '', HtmlIcerik TEXT, ImagePath TEXT,
            FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE
        )
    """)

    existing = set(col_names)
    copy_cols = [col for col in [
        "ProjeID", "GSMNo", "Baslik", "Icerik", "Tur", "Tarih", "Sira",
        "GenislikYuzde", "YukseklikMm", "Hizalama", "Aciklama",
        "HtmlIcerik", "BaseHtmlIcerik", "HiddenColsJson", "FmtJson",
        "ImagePath"
    ] if col in existing]

    if copy_cols:
        cols_str = ", ".join(copy_cols)
        c.execute(f"INSERT INTO rapor_taslagi_new ({cols_str}) SELECT {cols_str} FROM rapor_taslagi")

    c.execute("DROP TABLE rapor_taslagi")
    c.execute("ALTER TABLE rapor_taslagi_new RENAME TO rapor_taslagi")
    conn.commit()


def ensure_rapor_taslagi_tableprops_columns(conn: sqlite3.Connection):
    cur = conn.cursor()
    tbl = cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='rapor_taslagi'").fetchone()
    if not tbl:
        return

    cols = [r[1] for r in cur.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]

    if "BaseHtmlIcerik" not in cols:
        cur.execute("ALTER TABLE rapor_taslagi ADD COLUMN BaseHtmlIcerik TEXT")
    if "HiddenColsJson" not in cols:
        cur.execute("ALTER TABLE rapor_taslagi ADD COLUMN HiddenColsJson TEXT")
    if "FmtJson" not in cols:
        cur.execute("ALTER TABLE rapor_taslagi ADD COLUMN FmtJson TEXT")

    conn.commit()


def ensure_rapor_meta_ekler_columns(conn):
    cur = conn.cursor()
    cols = [r[1] for r in cur.execute("PRAGMA table_info(rapor_meta_ekler)").fetchall()]

    # Raporda dosya adını gizle
    if "DosyaAdiGizle" not in cols:
        cur.execute("ALTER TABLE rapor_meta_ekler ADD COLUMN DosyaAdiGizle INTEGER DEFAULT 0")

    # Görsel genişliği (%)
    if "GenislikYuzde" not in cols:
        cur.execute("ALTER TABLE rapor_meta_ekler ADD COLUMN GenislikYuzde INTEGER DEFAULT 100")

    conn.commit()


def ensure_hts_dosyalari_meta_columns(conn: sqlite3.Connection):
    """hts_dosyalari tablosuna meta kolonları ekler."""
    cols = [r[1] for r in conn.execute("PRAGMA table_info(hts_dosyalari)").fetchall()]
    for col in ["TalepEdenMakam", "SorguBaslangic", "SorguBitis", "Tespit"]:
        if col not in cols:
            conn.execute(f"ALTER TABLE hts_dosyalari ADD COLUMN {col} TEXT")


def ensure_project_columns(conn: sqlite3.Connection):
    """projeler tablosuna eksik kolonları ekler."""
    cols = [r[1] for r in conn.execute("PRAGMA table_info(projeler)").fetchall()]
    for col in ["suc_tarihi", "gorevlendirme_tarihi", "bilirkisi_adi", "bilirkisi_unvan_sicil"]:
        if col not in cols:
            conn.execute(f"ALTER TABLE projeler ADD COLUMN {col} TEXT")


def ensure_hash_columns(conn: sqlite3.Connection):
    """hts_dosyalari tablosuna MD5 ve SHA256 ekler."""
    cols = [r[1] for r in conn.execute("PRAGMA table_info(hts_dosyalari)").fetchall()]
    if "MD5" not in cols: conn.execute("ALTER TABLE hts_dosyalari ADD COLUMN MD5 TEXT")
    if "SHA256" not in cols: conn.execute("ALTER TABLE hts_dosyalari ADD COLUMN SHA256 TEXT")


def ensure_performance_indexes(conn: sqlite3.Connection):
    """
    Performans indexleri (işlev/mantık değişmez).
    Amaç: En sık filtrelenen/sıralanan alanlarda taramayı azaltmak.
    """
    c = conn.cursor()

    c.execute("CREATE INDEX IF NOT EXISTS idx_gsm_pid_gsmno_tarih  ON hts_gsm  (ProjeID, GSMNo, TARIH)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_gsm_pid_gsmno_baz    ON hts_gsm  (ProjeID, GSMNo, BAZ)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_gsm_pid_gsmno_diger  ON hts_gsm  (ProjeID, GSMNo, DIGER_NUMARA)")

    c.execute("CREATE INDEX IF NOT EXISTS idx_sms_pid_gsmno_tarih  ON hts_sms  (ProjeID, GSMNo, TARIH)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_sms_pid_gsmno_diger  ON hts_sms  (ProjeID, GSMNo, DIGER_NUMARA)")

    c.execute("CREATE INDEX IF NOT EXISTS idx_gprs_pid_gsmno_tarih ON hts_gprs (ProjeID, GSMNo, TARIH)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_gprs_pid_gsmno_baz   ON hts_gprs (ProjeID, GSMNo, BAZ)")

    c.execute("CREATE INDEX IF NOT EXISTS idx_wap_pid_gsmno_tarih  ON hts_wap  (ProjeID, GSMNo, TARIH)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_wap_pid_gsmno_baz    ON hts_wap  (ProjeID, GSMNo, BAZ)")

    c.execute("CREATE INDEX IF NOT EXISTS idx_gsm_pid_numara_tarih  ON hts_gsm  (ProjeID, NUMARA, TARIH)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_gprs_pid_numara_tarih ON hts_gprs (ProjeID, NUMARA, TARIH)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_wap_pid_numara_tarih  ON hts_wap  (ProjeID, NUMARA, TARIH)")

    c.execute("CREATE INDEX IF NOT EXISTS idx_hts_rehber_pid_gsm_adet ON hts_rehber (ProjeID, GSMNo, Adet)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_hts_tum_baz_pid_gsm_sinyal ON hts_tum_baz (ProjeID, GSMNo, Sinyal)")

    c.execute("""
        CREATE INDEX IF NOT EXISTS idx_hts_gsm_detail_lookup
        ON hts_gsm (
            ProjeID,
            substr(replace(replace(replace(NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10),
            substr(replace(replace(replace(DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10),
            (substr(TARIH, 7, 4) || '-' || substr(TARIH, 4, 2) || '-' || substr(TARIH, 1, 2) || substr(TARIH, 11))
        )
    """)

    c.execute("""
        CREATE INDEX IF NOT EXISTS idx_rehber_pid_isim_tc_trim
        ON hts_rehber (ProjeID, Isim, TRIM(TC))
    """)
    c.execute("""
        CREATE INDEX IF NOT EXISTS idx_rehber_pid_tc_trim_karsi
        ON hts_rehber (ProjeID, TRIM(TC), KarsiNo)
    """)

    conn.commit()

    try:
        c.execute("ANALYZE;")
        conn.commit()
    except Exception:
        pass


def _try_open_as_plain_sqlite(db_path: str) -> bool:
    """
    DB plain sqlite mı?
    - plain ise sqlite3 ile açılıp sqlite_master okunabilir.
    - şifreli SQLCipher ise genelde 'file is not a database' vb. hata verir.
    """
    import sqlite3 as _plain_sqlite
    try:
        conn = _plain_sqlite.connect(db_path, timeout=3)
        try:
            conn.execute("SELECT name FROM sqlite_master LIMIT 1;").fetchone()
        finally:
            conn.close()
        return True
    except Exception:
        return False


def _require_free_space_for_migration(db_path: str, multiplier: float = 2.2):
    if not os.path.exists(db_path):
        return
    size = os.path.getsize(db_path)
    drive = os.path.splitdrive(os.path.abspath(db_path))[0] + "\\"
    total, used, free = shutil.disk_usage(drive)
    need = int(size * multiplier)
    if free < need:
        raise RuntimeError(
            f"DB şifreleme için yeterli boş alan yok.\n"
            f"Gereken ~{need/1024/1024/1024:.2f} GB, boş ~{free/1024/1024/1024:.2f} GB."
        )


def migrate_plain_sqlite_to_sqlcipher(db_path: str, key: str, sqlcipher_connect):
    """
    Plain SQLite DB'yi SQLCipher'a çevirir.
    ÖNEMLİ: Plain dosyayı SQLCipher ile doğrudan açmıyoruz (file is not a database hatasına düşmemek için).
    Bunun yerine:
      - encrypted tmp DB oluştur
      - plain DB'yi KEY '' ile attach et
      - sqlcipher_export ile plain -> encrypted kopyala
      - dosyaları rename ile değiştir
    """
    if not os.path.exists(db_path):
        return

    # WAL/SHM varsa önce plain sqlite ile checkpoint dene (zararsız)
    try:
        import sqlite3 as _plain_sqlite
        conn_plain = _plain_sqlite.connect(db_path, isolation_level=None, timeout=30)
        try:
            cur = conn_plain.cursor()
            try:
                cur.execute("PRAGMA wal_checkpoint(TRUNCATE);")
            except Exception:
                cur.execute("PRAGMA wal_checkpoint(FULL);")
        finally:
            conn_plain.close()
    except Exception:
        pass

    _require_free_space_for_migration(db_path)

    tmp_enc = db_path + ".enc_tmp"
    backup_plain = db_path + ".plain_backup"

    if os.path.exists(tmp_enc):
        try:
            os.remove(tmp_enc)
        except Exception:
            pass

    # 1) Encrypted tmp DB'yi SQLCipher ile oluştur
    conn = sqlcipher_connect(tmp_enc, timeout=30, check_same_thread=False)
    try:
        cur = conn.cursor()

        # KEY parametre bağlamayı her driver kabul etmediği için literal veriyoruz.
        # key zaten derive_db_key() ile sha256 hexdigest (0-9a-f) olduğu için güvenli.
        cur.execute(f"PRAGMA key = '{key}';")
        cur.execute("PRAGMA cipher_compatibility = 4;")

        # 2) Plain DB'yi attach et (plain DB'de KEY boş olmalı)
        cur.execute("ATTACH DATABASE ? AS plain KEY '';", (db_path,))

        # 3) Plain -> main(encrypted tmp) export
        cur.execute("SELECT sqlcipher_export('main', 'plain');")

        cur.execute("DETACH DATABASE plain;")
        conn.commit()
    finally:
        conn.close()

    # Eski plain DB'yi yedekle, tmp_enc'yi asıl dosya yap
    if os.path.exists(backup_plain):
        try:
            os.remove(backup_plain)
        except Exception:
            pass

    os.replace(db_path, backup_plain)

    # Plain'in WAL/SHM kalıntılarını temizle
    for ext in (".wal", ".shm"):
        p = db_path + ext
        try:
            if os.path.exists(p):
                os.remove(p)
        except Exception:
            pass

    os.replace(tmp_enc, db_path)


def ensure_encrypted_db(db_path: str, key: str, sqlcipher_connect):
    if not os.path.exists(db_path):
        return

    is_plain = _try_open_as_plain_sqlite(db_path)
    if is_plain:
        migrate_plain_sqlite_to_sqlcipher(db_path, key=key, sqlcipher_connect=sqlcipher_connect)


def run_all_migrations(conn: sqlite3.Connection):
    try:
        ensure_project_columns(conn)
        ensure_hts_dosyalari_meta_columns(conn)
        ensure_hash_columns(conn)
        ensure_rapor_taslagi_has_id(conn)
        ensure_rapor_taslagi_tableprops_columns(conn)
        ensure_performance_indexes(conn)
        ensure_rapor_meta_ekler_columns(conn)


        conn.commit()
    except Exception as e:
        print(f"Migration Hatası: {e}")


def derive_db_key() -> str:
    info = LicenseManager.ensure_valid_or_raise()
    fp = LicenseManager.device_fingerprint()
    raw = f"{info.license_id}|{fp}|HTSMERCEK_DB_KEY_V1"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


class DatabaseManager:
    _instance = None
    _instance_lock = threading.Lock()

    DB_PATH = os.path.join(LicenseManager.appdata_dir(), "htstakip.db")

    try:
        from pysqlcipher3 import dbapi2 as _sqlcipher
        _USING_SQLCIPHER = True
    except Exception:
        try:
            from sqlcipher3 import dbapi2 as _sqlcipher  # sqlcipher3-wheels
            _USING_SQLCIPHER = True
        except Exception:
            _sqlcipher = None
            _USING_SQLCIPHER = False

    def __new__(cls):
        if cls._instance is None:
            with cls._instance_lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if getattr(self, "_initialized", False):
            return

        self._db_lock = threading.RLock()

        if not self._USING_SQLCIPHER or self._sqlcipher is None:
            raise RuntimeError(
                "SQLCipher aktif değil. Şifreli DB için pysqlcipher3 (veya SQLCipher driver) gerekli."
            )
        key = derive_db_key()
        ensure_encrypted_db(self.DB_PATH, key=key, sqlcipher_connect=self._sqlcipher.connect)
        self._connection = self._sqlcipher.connect(
            self.DB_PATH,
            check_same_thread=False,
            timeout=30
        )

        self._configure_db(key)
        self._initialized = True

    def _configure_db(self, key):
        """
        Veritabanı bağlantısını ve şifrelemeyi yapılandırır.
        Eğer şifre anahtarı değişmişse (HWID değişimi vb.) veritabanını sıfırlar.
        """
        with self._db_lock:
            # Mevcut bağlantı üzerinden cursor al
            cur = self._connection.cursor()

            # Anahtarı ayarla
            cur.execute(f"PRAGMA key = '{key}';")
            cur.execute("PRAGMA cipher_compatibility = 4;")

            try:
                # Test sorgusu: Şifre doğru mu?
                cur.execute("SELECT count(*) FROM sqlite_master;")
                cur.fetchone()

                # Başarılıysa performans ayarlarını yap
                cur.execute("PRAGMA journal_mode=WAL;")
                cur.execute("PRAGMA synchronous=NORMAL;")
                cur.execute("PRAGMA foreign_keys=ON;")
                cur.execute("PRAGMA temp_store=MEMORY;")
                cur.execute("PRAGMA busy_timeout=5000;")
                self._connection.commit()

            except Exception as e:
                # Hata yakalama (pysqlcipher veya sqlite3 hataları)
                err_str = str(e)
                err_low = (err_str or "").lower()

                # SADECE gerçekten "anahtar yanlış / db bozuk" durumlarında devreye girsin
                bad_key_markers = (
                    "file is not a database",
                    "not a database",
                    "bad decrypt",
                    "wrong key",
                    "sqlcipher",
                )

                if any(m in err_low for m in bad_key_markers):
                    print("⚠️ Veritabanı anahtarı uyuşmuyor veya DB bozuk. Silmek yerine yedekleniyor...")

                    try:
                        cur.close()
                        self._connection.close()
                    except Exception:
                        pass

                    # Silmek YOK: Yedekle
                    try:
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        base = self.DB_PATH

                        if os.path.exists(base):
                            os.replace(base, base + f".badkey_{ts}.bak")

                        if os.path.exists(base + "-wal"):
                            os.replace(base + "-wal", base + f"-wal.badkey_{ts}.bak")

                        if os.path.exists(base + "-shm"):
                            os.replace(base + "-shm", base + f"-shm.badkey_{ts}.bak")

                    except Exception as bak_err:
                        print(f"DB yedekleme hatası: {bak_err}")

                    # Yeni bağlantı oluştur
                    if self._sqlcipher:
                        self._connection = self._sqlcipher.connect(
                            self.DB_PATH,
                            check_same_thread=False,
                            timeout=30
                        )
                    else:
                        import sqlite3
                        self._connection = sqlite3.connect(
                            self.DB_PATH,
                            check_same_thread=False,
                            timeout=30
                        )

                    cur = self._connection.cursor()
                    cur.execute(f"PRAGMA key = '{key}';")
                    cur.execute("PRAGMA cipher_compatibility = 4;")
                    cur.execute("PRAGMA journal_mode=WAL;")
                    self._connection.commit()

                else:
                    # Başka bir hataysa (örn: disk dolu) durdur
                    raise e

    def lock(self):
        return self._db_lock

    def get_connection(self):
        return self._connection


class DB:
    def __init__(self):
        self.manager = DatabaseManager()
        self.conn = None
        self._lock = self.manager.lock()

    def __enter__(self):
        self._lock.acquire()
        self.conn = self.manager.get_connection()
        return self.conn

    def __exit__(self, exc_type, exc, tb):
        try:
            if not self.conn:
                return False

            if exc_type is None:
                self.conn.commit()
            else:
                self.conn.rollback()
                print(f"⚠️ DB Rollback: {exc}")

            return False
        finally:
            self._lock.release()


def setup_database():
    with DB() as conn:
        c = conn.cursor()

        c.execute(
            "CREATE TABLE IF NOT EXISTS projeler ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, "
            "talep_eden_birim TEXT, dosya_no_tipi TEXT, dosya_no TEXT, suc_bilgisi TEXT, "
            "suc_tarihi TEXT, gorevlendirme_tarihi TEXT, bilirkisi_adi TEXT, "
            "bilirkisi_unvan_sicil TEXT, olusturma_tarihi TEXT)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS taraflar ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, "
            "ProjeID INTEGER REFERENCES projeler(id) ON DELETE CASCADE, "
            "sifat TEXT, ad_soyad TEXT)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_ozet ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, "
            "MinDate TEXT, MaxDate TEXT, UNIQUE(ProjeID, GSMNo))"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_ozet_iletisim ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, "
            "KarsiNo TEXT, Adet INTEGER, Sure INTEGER, Isim TEXT)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_ozet_baz ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, "
            "BazAdi TEXT, Sinyal INTEGER)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_ozet_imei ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, "
            "IMEI TEXT, Adet INTEGER, MinDate TEXT, MaxDate TEXT)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_rehber ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, "
            "KarsiNo TEXT, Adet INTEGER, Sure INTEGER, Isim TEXT, TC TEXT)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_tum_baz ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, "
            "BazAdi TEXT, Sinyal INTEGER)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_ortak_imei ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, IMEI TEXT, "
            "KullananSayisi INTEGER, Numaralar TEXT, ToplamKullanim INTEGER, "
            "UNIQUE(ProjeID, IMEI))"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_ortak_isim ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, AdSoyad TEXT, "
            "HatSayisi INTEGER, Numaralar TEXT, UNIQUE(ProjeID, AdSoyad))"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_ortak_tc ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, TC TEXT, "
            "HatSayisi INTEGER, Numaralar TEXT, UNIQUE(ProjeID, TC))"
        )
        c.execute("CREATE INDEX IF NOT EXISTS idx_rehber ON hts_rehber (ProjeID, GSMNo)")
        c.execute(
            "CREATE TABLE IF NOT EXISTS ozel_konumlar ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, "
            "Lat REAL, Lon REAL, Label TEXT, "
            "FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE)"
        )
        for table_name, columns in TABLE_COLUMNS.items():
            cols_def = ", ".join([f"[{col}] TEXT" for col in columns])
            c.execute(
                f"CREATE TABLE IF NOT EXISTS {table_name} ("
                f"id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, {cols_def}, "
                f"FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE)"
            )
        c.execute(
            "CREATE TABLE IF NOT EXISTS rapor_taslagi ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, Baslik TEXT, Icerik TEXT, "
            "Tur TEXT, Tarih TEXT, Sira INTEGER, GenislikYuzde INTEGER DEFAULT 100, YukseklikMm INTEGER, "
            "Hizalama TEXT DEFAULT 'center', Aciklama TEXT DEFAULT '', HtmlIcerik TEXT, ImagePath TEXT, "
            "FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS hts_dosyalari ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, GSMNo TEXT, Rol TEXT, DosyaAdi TEXT, "
            "DosyaBoyutu INTEGER, DosyaYolu TEXT, TalepEdenMakam TEXT, SorguBaslangic TEXT, SorguBitis TEXT, "
            "Tespit TEXT, MD5 TEXT, SHA256 TEXT, YuklenmeTarihi TEXT DEFAULT CURRENT_TIMESTAMP, "
            "UNIQUE(ProjeID, GSMNo, Rol), FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS rapor_meta ("
            "ProjeID INTEGER PRIMARY KEY, GorevlendirmeMetni TEXT, DosyaHakkindaMetni TEXT, "
            "GenelBilgilendirmeMetni TEXT, DegerlendirmeMetni TEXT, SonucMetni TEXT, "
            "MarginTopMm INTEGER DEFAULT 20, MarginRightMm INTEGER DEFAULT 20, "
            "MarginBottomMm INTEGER DEFAULT 20, MarginLeftMm INTEGER DEFAULT 20, "
            "GuncellemeTarihi TEXT DEFAULT CURRENT_TIMESTAMP, "
            "FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS rapor_meta_ekler ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, "
            "ProjeID INTEGER, "
            "Bolum TEXT, "                      # 'dosya_hakkinda' gibi
            "DosyaAdi TEXT, "
            "DosyaYolu TEXT, "                  # evidence_images altında tutulacak
            "Aciklama TEXT DEFAULT '', "
            "EklemeTarihi TEXT DEFAULT CURRENT_TIMESTAMP, "
            "FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE)"
        )
        c.execute("CREATE INDEX IF NOT EXISTS idx_rapor_meta_ekler_pid ON rapor_meta_ekler (ProjeID, Bolum)")
        c.execute(
            "CREATE TABLE IF NOT EXISTS manuel_numaralar ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, ProjeID INTEGER, Numara TEXT, "
            "Aciklama TEXT DEFAULT 'Manuel Giriş', EklemeTarihi TEXT DEFAULT CURRENT_TIMESTAMP, "
            "FOREIGN KEY(ProjeID) REFERENCES projeler(id) ON DELETE CASCADE)"
        )
        c.execute(
            "CREATE TABLE IF NOT EXISTS baz_kutuphanesi ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, CellID TEXT, BazAdi TEXT, "
            "Lat REAL, Lon REAL, KaynakDosya TEXT, "
            "OgrenmeTarihi TEXT DEFAULT CURRENT_TIMESTAMP, UNIQUE(CellID, BazAdi))"
        )
        c.execute("CREATE INDEX IF NOT EXISTS idx_baz_cell ON baz_kutuphanesi (CellID)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_baz_ad ON baz_kutuphanesi (BazAdi)")

        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_gsm_pid_gsmno ON hts_gsm (ProjeID, GSMNo)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_gsm_pid_gsmno_diger ON hts_gsm (ProjeID, GSMNo, DIGER_NUMARA)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_gsm_pid_imei ON hts_gsm (ProjeID, IMEI)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_gsm_pid_tarih ON hts_gsm (ProjeID, TARIH)")

        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_sms_pid_gsmno ON hts_sms (ProjeID, GSMNo)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_sms_pid_gsmno_diger ON hts_sms (ProjeID, GSMNo, DIGER_NUMARA)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_sms_pid_tarih ON hts_sms (ProjeID, TARIH)")

        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_gprs_pid_gsmno ON hts_gprs (ProjeID, GSMNo)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_gprs_pid_imei ON hts_gprs (ProjeID, IMEI)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_gprs_pid_tarih ON hts_gprs (ProjeID, TARIH)")

        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_wap_pid_gsmno ON hts_wap (ProjeID, GSMNo)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_wap_pid_imei ON hts_wap (ProjeID, IMEI)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_hts_wap_pid_tarih ON hts_wap (ProjeID, TARIH)")

        c.execute("CREATE INDEX IF NOT EXISTS idx_ozet_iletisim_pid_gsm_karsi ON hts_ozet_iletisim (ProjeID, GSMNo, KarsiNo)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_ozet_baz_pid_gsm_baz ON hts_ozet_baz (ProjeID, GSMNo, BazAdi)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_ozet_imei_pid_gsm_imei ON hts_ozet_imei (ProjeID, GSMNo, IMEI)")

        c.execute("CREATE INDEX IF NOT EXISTS idx_manuel_numaralar_pid ON manuel_numaralar (ProjeID)")
        for _t in ["hts_abone", "hts_gsm", "hts_sms", "hts_sabit", "hts_gprs", "hts_wap", "hts_sth", "hts_uluslararasi"]:
            try:
                c.execute(f"ALTER TABLE {_t} ADD COLUMN Rol TEXT")
            except Exception:
                pass
            try:
                c.execute(f"ALTER TABLE {_t} ADD COLUMN DosyaAdi TEXT")
            except Exception:
                pass
        run_all_migrations(conn)


class HtsWorker(QThread):
    progress = pyqtSignal(int)
    log = pyqtSignal(str)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    gsm_detected = pyqtSignal(str)

    def __init__(self, path, pid):
        super().__init__()
        self.path = path
        self.pid = pid
        self.is_running = True
        self.file_name = os.path.basename(path)

    def clean_cell_data(self, value):
        if value is None: return None
        text = str(value).strip()
        text = text.replace('_x000D_', ' ').replace('\n', ' ').replace('\r', ' ')
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def run(self):
        try:
            from openpyxl import load_workbook
            self.log.emit(f"📂 Dosya Analiz Ediliyor:\n{self.file_name}")
            md5_hash = hashlib.md5()
            sha256_hash = hashlib.sha256()

            try:
                with open(self.path, "rb") as f:
                    for byte_block in iter(lambda: f.read(65536), b""):
                        md5_hash.update(byte_block)
                        sha256_hash.update(byte_block)

                file_md5 = md5_hash.hexdigest()
                file_sha256 = sha256_hash.hexdigest()
                self.log.emit("🔒 Dosya İmzaları (Hash) Oluşturuldu.")

            except Exception as hash_err:
                print(f"Hash hatası: {hash_err}")
                file_md5 = "HESAPLANAMADI"
                file_sha256 = "HESAPLANAMADI"
            file_size_mb = os.path.getsize(self.path) / (1024 * 1024)

            wb = load_workbook(self.path, read_only=True, data_only=True)
            sheet = wb.active

            meta_data = {
                "Talep Eden Makam": "",
                "Sorgu Başlangıç Tarihi": "",
                "Sorgu Bitiş Tarihi": "",
                "Tespit": ""
            }

            target_gsm = "BILINMIYOR"

            for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
                if not row or len(row) < 2: continue
                val_a = row[0]; val_b = row[1]
                if not val_a: continue

                key = str(val_a).strip().upper()
                key = key.replace("İ", "I").replace("Ş", "S").replace("Ç", "C").replace("Ö", "O").replace("Ü", "U").replace("Ğ", "G").replace(":", "").strip()

                final_value = ""
                if val_b is not None:
                    if isinstance(val_b, datetime): final_value = val_b.strftime("%d.%m.%Y %H:%M:%S")
                    else: final_value = str(val_b).strip()

                if "TALEP EDEN" in key: meta_data["Talep Eden Makam"] = final_value
                elif "SORGULANAN NO" in key or "GSM NO" in key:
                    clean_gsm = re.sub(r'\D', '', final_value)
                    if len(clean_gsm) >= 10: target_gsm = clean_gsm
                elif "BASLANGIC" in key: meta_data["Sorgu Başlangıç Tarihi"] = final_value
                elif "BITIS" in key: meta_data["Sorgu Bitiş Tarihi"] = final_value
                elif "TESPIT" in key or "KONU" in key: meta_data["Tespit"] = final_value

            if target_gsm == "BILINMIYOR":
                target_gsm = _detect_target_gsm(self.path)

            self.gsm_detected.emit(target_gsm)
            self.log.emit(f"✅ Hedef Numara Tespit Edildi: {target_gsm}\nDosya: {self.file_name}")

            rol, sorgulanan_no = detect_hts_role(self.path)
            self.current_rol = rol

            with DB() as conn:
                conn.execute("""
                    INSERT OR REPLACE INTO hts_dosyalari 
                    (ProjeID, GSMNo, Rol, DosyaAdi, DosyaBoyutu, DosyaYolu, 
                     TalepEdenMakam, SorguBaslangic, SorguBitis, Tespit, MD5, SHA256) 
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (self.pid, target_gsm, rol, self.file_name, int(os.path.getsize(self.path)), self.path,
                      meta_data["Talep Eden Makam"],
                      meta_data["Sorgu Başlangıç Tarihi"],
                      meta_data["Sorgu Bitiş Tarihi"],
                      meta_data["Tespit"],
                      file_md5, file_sha256))

            is_valid_format = False; has_target_no = False
            BLOCK_MAP_KEYS = ["ABONE BILGILERI", "GSM GORUSME", "SABIT TELEFON", "ULUSLARARASI", "MESAJ BILGILERI", "INTERNET BAGLANTI", "STH GORUSME"]

            for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
                row_text = " ".join([str(x).strip().upper() for x in row if x is not None])
                row_text = row_text.replace("İ", "I").replace("Ş", "S").replace("Ç", "C").replace("Ö", "O").replace("Ü", "U").replace("Ğ", "G")
                if "SORGULANAN" in row_text and "NO" in row_text: has_target_no = True
                for key in BLOCK_MAP_KEYS:
                    if key in row_text: is_valid_format = True; break

            if not has_target_no: raise Exception("Hatalı Format: 'Sorgulanan No' satırı bulunamadı.")
            if not is_valid_format: raise Exception("Hatalı Format: Geçerli HTS başlıkları bulunamadı.")

            BLOCK_MAP = {"ABONE BILGILERI": "hts_abone", "GSM GORUSME SORGU SONUCLARI": "hts_gsm", "SABIT TELEFON GORUSME SORGU SONUCLARI": "hts_sabit", "ULUSLARARASI GORUSME SORGU SONUCLARI": "hts_uluslararasi", "MESAJ BILGILERI": "hts_sms", "INTERNET BAGLANTI (GPRS)": "hts_gprs", "INTERNET BAGLANTI (WAP)": "hts_wap", "STH GORUSME SORGU SONUCLARI": "hts_sth"}

            current_table = None; state = "SEARCHING"; column_map = {}; batch_data = []; BATCH_SIZE = 5000
            row_count = sheet.max_row or 50000

            self.log.emit(f"⏳ Veriler Okunuyor...\n{self.file_name}")

            for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if not self.is_running: break

                if r_idx % 5000 == 0:
                    current_prog = int((r_idx / row_count) * 85)
                    self.progress.emit(current_prog)
                    self.log.emit(f"⏳ Okunuyor (%{current_prog})\nDosya: {self.file_name}\nSatır: {r_idx}")
                    QThread.msleep(1)

                if not row: continue

                row_text_raw = " ".join([str(x).strip().upper() for x in row if x is not None])
                row_text = row_text_raw.replace("İ", "I").replace("Ş", "S").replace("Ç", "C").replace("Ö", "O").replace("Ü", "U").replace("Ğ", "G")

                if target_gsm == "BILINMIYOR" and len(row) > 1:
                    col_a = str(row[0]).upper() if row[0] else ""
                    if "SORGULANAN" in col_a and "NO" in col_a:
                         clean = re.sub(r'\D', '', str(row[1]))
                         if clean: target_gsm = clean; self.gsm_detected.emit(target_gsm)

                if "KAYITBULUNAMADI" in str(row[0]).upper().replace(" ", ""):
                    if current_table: self._save_batch(current_table, batch_data, target_gsm); batch_data = []; current_table = None; state = "SEARCHING"
                    continue

                found_new = False
                for key, tbl in BLOCK_MAP.items():
                    if key in row_text:
                        if current_table: self._save_batch(current_table, batch_data, target_gsm); batch_data = []
                        if rol == "KARSI" and tbl == "hts_abone":
                            current_table = None; state = "SEARCHING"; found_new = True; break
                        current_table = tbl; state = "WAITING_HEADER"; found_new = True;
                        break

                if found_new: continue

                if state == "WAITING_HEADER" and current_table:
                    column_map = {}; alias_map = HEADER_ALIASES.get(current_table, {})
                    for c_idx, cell in enumerate(row):
                        if not cell: continue
                        h_clean = _norm_header(str(cell))
                        for ex_head, db_col in alias_map.items():
                            if _norm_header(ex_head) == h_clean: column_map[c_idx] = db_col; break
                    if column_map: state = "READING_DATA"
                    continue

                if state == "READING_DATA" and current_table:
                    sira_col = -1
                    for idx, name in column_map.items():
                        if name == "SIRA_NO": sira_col = idx; break
                    val = row[sira_col] if (sira_col != -1 and sira_col < len(row)) else (row[0] if row else None)

                    if val and str(val).strip().isdigit():
                        entry = {}
                        for c_idx, db_col in column_map.items():
                            if c_idx < len(row):
                                raw_val = row[c_idx]; clean_val = self.clean_cell_data(raw_val)
                                if db_col == "IMEI" and clean_val:
                                    digits_only = re.sub(r'\D', '', clean_val)
                                    if len(digits_only) < 13: clean_val = None
                                entry[db_col] = clean_val
                        if "BAZ" in entry and entry["BAZ"]:
                            baz_raw = str(entry["BAZ"]).strip()

                            coords = re.findall(r"(\d{2}\.\d{4,})", baz_raw)

                            if len(coords) >= 2:
                                try:
                                    lat, lon = float(coords[-2]), float(coords[-1])

                                    final_lat, final_lon = 0, 0
                                    if 35 < lat < 43 and 25 < lon < 46:
                                        final_lat, final_lon = lat, lon
                                    elif 35 < lon < 43 and 25 < lat < 46:
                                        final_lat, final_lon = lon, lat

                                    if final_lat != 0:
                                        cell_id = None

                                        match_par = re.search(r'\((\d{4,})\)', baz_raw)
                                        if match_par:
                                            cell_id = match_par.group(1)
                                        else:
                                            nums = re.findall(r'\d+', baz_raw)
                                            candidates = [n for n in nums if n not in coords and len(n) > 3]
                                            if candidates:
                                                cell_id = candidates[0]
                                        with DB() as db_conn:
                                            db_conn.execute("""
                                                INSERT OR IGNORE INTO baz_kutuphanesi 
                                                (CellID, BazAdi, Lat, Lon, KaynakDosya) 
                                                VALUES (?, ?, ?, ?, ?)
                                            """, (cell_id, baz_raw, final_lat, final_lon, self.file_name))
                                except:
                                    pass
                        entry["GSMNo"] = target_gsm; batch_data.append(entry)

                        if len(batch_data) >= BATCH_SIZE:
                            self.log.emit(f"💾 Veritabanına Yazılıyor...\n{self.file_name}")
                            self._save_batch(current_table, batch_data, target_gsm)
                            batch_data = []

            if current_table and batch_data: self._save_batch(current_table, batch_data, target_gsm)
            self.log.emit(f"📊 İstatistikler ve Özetler Hesaplanıyor...\n{target_gsm}")
            self.progress.emit(86)

            self.calculate_and_save_summary(target_gsm)

            self.progress.emit(100)
            self.finished.emit(f"{target_gsm} - {self.file_name} Tamamlandı.")

        except Exception as e: self.error.emit(str(e))

    def _save_batch(self, table, data, gsm):
        if not data:
            return

        cols = TABLE_COLUMNS[table]
        vals = []

        # Rol + DosyaAdi (kaynak xlsx) bilgisini ham kayıtlara yazacağız
        rol = getattr(self, "current_rol", None) or "HEDEF"
        dosya_adi = getattr(self, "file_name", "") or ""

        for item in data:
            row = [self.pid, gsm, rol, dosya_adi]
            for c in cols:
                row.append(item.get(c, None))
            vals.append(row)

        with DB() as conn:
            ph = ",".join(["?"] * (len(cols) + 4))
            conn.executemany(
                f"INSERT INTO {table} (ProjeID, GSMNo, Rol, DosyaAdi, {','.join(cols)}) VALUES ({ph})",
                vals
            )

    def calculate_and_save_summary(self, gsm):
        try:
            with DB() as conn:
                cur = conn.cursor()
                all_raw_dates = []
                tables_to_check = ["hts_gsm", "hts_sms", "hts_gprs", "hts_wap", "hts_sabit", "hts_sth"]

                for t in tables_to_check:
                    try:
                        res = cur.execute(f"SELECT TARIH FROM {t} WHERE ProjeID=? AND GSMNo=?", (self.pid, gsm)).fetchall()
                        all_raw_dates.extend([r[0] for r in res if r[0]])
                    except:
                        pass
                    QThread.msleep(5)

                valid_dts = []
                for idx, t_str in enumerate(all_raw_dates):
                    if idx % 2000 == 0: QThread.msleep(1)
                    try:
                        cln = str(t_str).strip()
                        fmt = "dd.MM.yyyy HH:mm:ss" if "." in cln else "dd/MM/yyyy HH:mm:ss"
                        if " " not in cln: fmt = fmt.split(" ")[0]

                        dt = QDateTime.fromString(cln, fmt)
                        if not dt.isValid():
                            dt = QDateTime.fromString(cln, "yyyy-MM-dd HH:mm:ss")

                        if dt.isValid(): valid_dts.append(dt)
                    except: pass

                if valid_dts:
                    min_str = min(valid_dts).toString("dd.MM.yyyy HH:mm:ss")
                    max_str = max(valid_dts).toString("dd.MM.yyyy HH:mm:ss")
                    cur.execute("INSERT OR REPLACE INTO hts_ozet (ProjeID, GSMNo, MinDate, MaxDate) VALUES (?, ?, ?, ?)", (self.pid, gsm, min_str, max_str))
                else:
                    cur.execute("INSERT OR REPLACE INTO hts_ozet (ProjeID, GSMNo, MinDate, MaxDate) VALUES (?, ?, ?, ?)", (self.pid, gsm, "", ""))

                self.progress.emit(90)
                QThread.msleep(10)
                cur.execute("DELETE FROM hts_ozet_iletisim WHERE ProjeID=? AND GSMNo=?", (self.pid, gsm))
                cur.execute("DELETE FROM hts_rehber WHERE ProjeID=? AND GSMNo=?", (self.pid, gsm))
                all_contacts = cur.execute("""
                    SELECT DIGER_NUMARA, COUNT(*), SUM(CAST(SURE as INTEGER)), MAX(DIGER_ISIM), MAX(DIGER_TC)
                    FROM hts_gsm 
                    WHERE ProjeID=? 
                      AND GSMNo=? 
                      AND DIGER_NUMARA != ? 
                    GROUP BY DIGER_NUMARA 
                    ORDER BY 2 DESC
                """, (self.pid, gsm, gsm)).fetchall()

                if all_contacts:
                    d_full = [[self.pid, gsm, r[0], r[1], (r[2] if r[2] else 0), r[3], r[4]] for r in all_contacts]
                    cur.executemany("""
                        INSERT INTO hts_rehber (ProjeID, GSMNo, KarsiNo, Adet, Sure, Isim, TC) 
                        VALUES (?,?,?,?,?,?,?)
                    """, d_full)
                    top_20_data = [[self.pid, gsm, r[0], r[1], (r[2] if r[2] else 0), r[3]] for r in all_contacts[:20]]
                    cur.executemany("INSERT INTO hts_ozet_iletisim (ProjeID, GSMNo, KarsiNo, Adet, Sure, Isim) VALUES (?,?,?,?,?,?)", top_20_data)

                self.progress.emit(94)
                QThread.msleep(10)

                cur.execute("DELETE FROM hts_ozet_imei WHERE ProjeID=? AND GSMNo=?", (self.pid, gsm))
                cur.execute("DELETE FROM hts_ozet_baz WHERE ProjeID=? AND GSMNo=?", (self.pid, gsm))
                cur.execute("DELETE FROM hts_tum_baz WHERE ProjeID=? AND GSMNo=?", (self.pid, gsm))

                baz_c = defaultdict(int)
                imei_stats = defaultdict(lambda: {'count': 0, 'dates': []})

                for t in ["hts_gsm", "hts_gprs", "hts_wap"]:
                    try:
                        rows = cur.execute(f"SELECT BAZ, IMEI, TARIH FROM {t} WHERE ProjeID=? AND GSMNo=?", (self.pid, gsm)).fetchall()
                        for idx, r in enumerate(rows):
                            if idx % 2000 == 0: QThread.msleep(1)
                            baz, imei, tarih = r

                            if baz and str(baz).strip():
                                baz_c[str(baz).strip()] += 1

                            if imei and str(imei).strip():
                                clean_imei = str(imei).strip()
                                imei_stats[clean_imei]['count'] += 1
                                if tarih: imei_stats[clean_imei]['dates'].append(tarih)
                    except: pass

                sb = sorted(baz_c.items(), key=lambda x:x[1], reverse=True)
                if sb:
                    all_baz_data = [[self.pid, gsm, k, v] for k,v in sb]
                    cur.executemany("INSERT INTO hts_tum_baz (ProjeID, GSMNo, BazAdi, Sinyal) VALUES (?,?,?,?)", all_baz_data)
                    cur.executemany("INSERT INTO hts_ozet_baz (ProjeID, GSMNo, BazAdi, Sinyal) VALUES (?,?,?,?)", all_baz_data[:20])

                imei_data_to_save = []
                for idx, (imei, info) in enumerate(imei_stats.items()):
                    min_d_str = ""; max_d_str = ""
                    if info['dates']:
                        try:
                            info['dates'].sort(key=lambda x: x.split()[0] if x else "")
                            min_d_str = info['dates'][0]
                            max_d_str = info['dates'][-1]
                        except: pass

                    imei_data_to_save.append([self.pid, gsm, imei, info['count'], min_d_str, max_d_str])

                imei_data_to_save.sort(key=lambda x: x[3], reverse=True) # Adete göre sırala

                if imei_data_to_save:
                    cur.executemany("INSERT INTO hts_ozet_imei (ProjeID, GSMNo, IMEI, Adet, MinDate, MaxDate) VALUES (?,?,?,?,?,?)", imei_data_to_save)

                self.progress.emit(99)
                conn.commit()
                self.recalculate_common_analysis()

        except Exception as e:
            print(f"Özet Hesaplama Hatası: {e}")

    def recalculate_common_analysis(self):
        # Eski uzun kodu silin, sadece bunu yazın:
        AnalysisUtils.recalculate_common_analysis_core(self.pid)


class DateSortFilterProxyModel(QSortFilterProxyModel):
    """Hem Akıllı Metin, Hem Tarih, Hem de SAYISAL SIRALAMA yapan model"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.min_date = None
        self.max_date = None
        self.date_column = -1
        self.date_filter_active = False
        self.search_text = ""

    def setDateRange(self, min_d, max_d):
        self.min_date = min_d; self.max_date = max_d; self.invalidateFilter()

    def setDateColumn(self, col_idx): self.date_column = col_idx
    def setDateFilterActive(self, active): self.date_filter_active = active; self.invalidateFilter()

    def setSearchText(self, text):
        self.search_text = self.normalize_turkish(text)
        self.invalidateFilter()

    def normalize_turkish(self, text):
        if not text: return ""
        text = str(text)
        tr_map = {'İ': 'i', 'I': 'ı', 'ı': 'i', 'Ş': 's', 'ş': 's', 'Ğ': 'g', 'ğ': 'g',
                  'Ü': 'u', 'ü': 'u', 'Ö': 'o', 'ö': 'o', 'Ç': 'c', 'ç': 'c'}
        for k, v in tr_map.items(): text = text.replace(k, v)
        text = unicodedata.normalize('NFD', text.lower())
        text = "".join([c for c in text if not unicodedata.category(c).startswith('M')])
        return text.lower().strip()

    def lessThan(self, left, right):
        left_data = self.sourceModel().data(left, Qt.ItemDataRole.EditRole)
        right_data = self.sourceModel().data(right, Qt.ItemDataRole.EditRole)

        try:
            return float(left_data) < float(right_data)
        except (ValueError, TypeError):
            return str(left_data).lower() < str(right_data).lower()

    def filterAcceptsRow(self, source_row, source_parent):
        if self.search_text:
            row_match = False
            model = self.sourceModel()
            for col in range(model.columnCount(QModelIndex())):
                data = model.data(model.index(source_row, col, source_parent), Qt.ItemDataRole.DisplayRole)
                if data and self.search_text in self.normalize_turkish(data):
                    row_match = True; break
            if not row_match: return False

        if self.date_filter_active and self.date_column != -1:
            date_str = str(self.sourceModel().data(self.sourceModel().index(source_row, self.date_column, source_parent), Qt.ItemDataRole.DisplayRole))

            try:
                clean = date_str.split(" ")[0].strip()

                rd = QDate.fromString(clean, "dd.MM.yyyy")
                if not rd.isValid(): rd = QDate.fromString(clean, "dd/MM/yyyy")
                if not rd.isValid(): rd = QDate.fromString(clean, "yyyy-MM-dd")

                if rd.isValid():
                    if self.min_date and rd < self.min_date: return False
                    if self.max_date and rd > self.max_date: return False
            except:
                pass

        return True


class CustomTableModel(QAbstractTableModel):
    def __init__(self, data=None, headers=None):
        super().__init__()
        self._data = data or []
        self._headers = headers or []

    def data(self, index, role):
        if not index.isValid(): return None

        row_idx = index.row()
        col_idx = index.column()

        if row_idx < 0 or row_idx >= len(self._data): return None
        if col_idx < 0 or col_idx >= len(self._data[row_idx]): return ""

        value = self._data[row_idx][col_idx]

        if role == Qt.ItemDataRole.DisplayRole:
            if value is None: return ""
            return str(value)

        if role == Qt.ItemDataRole.EditRole:
            return value

        if role == Qt.ItemDataRole.TextAlignmentRole:
            if isinstance(value, (int, float)):
                return Qt.AlignmentFlag.AlignCenter
            return Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter

    def rowCount(self, index): return len(self._data)
    def columnCount(self, index): return len(self._headers)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            if section < len(self._headers): return self._headers[section]
            return ""

    def update_data(self, new_data):
        self.beginResetModel()
        self._data = new_data
        self.endResetModel()


class CrossMatchDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, available_numbers):
        super().__init__(parent)

        self.init_watermark(opacity=0.025, scale_ratio=0.90)

        self.project_id = project_id
        self.available_numbers = available_numbers

        self.setWindowTitle("Ortak Temas ve İlişki Analizi")
        self.resize(1300, 850)
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint)

        self.setStyleSheet("""
            /* Dialogun ana zemini temaya göre kalsın */
            CrossMatchDialog { background-color: white; }

            /* Scroll area arka planını şeffaf yap */
            QScrollArea, QScrollArea > QWidget, QScrollArea::viewport {
                background: transparent;
            }

            /* Tablo zemini şeffaf */
            QTableWidget, QTableView {
                background-color: transparent;
                alternate-background-color: transparent;
                gridline-color: rgba(0,0,0,40);
            }

            /* Hücreler: okunabilirlik için hafif beyaz sis */
            QTableWidget::item, QTableView::item {
                background-color: rgba(255,255,255,200);
            }

            /* Headerlar hafif opak beyaz */
            QHeaderView::section {
                background-color: rgba(245,245,245,220);
            }
        """)

        self.layout = QVBoxLayout(self)
        self.stack = QStackedWidget()
        self.layout.addWidget(self.stack)

        self.page_selection = QWidget()
        self.setup_selection_ui()
        self.stack.addWidget(self.page_selection)

        self.page_results = QWidget()
        self.setup_results_ui()
        self.stack.addWidget(self.page_results)

        self.stack.setCurrentIndex(0)

    def setup_selection_ui(self):
        """1. Aşama: İsim Destekli Seçim Ekranı."""
        layout = QVBoxLayout(self.page_selection)

        header_layout = QHBoxLayout()
        lbl_title = QLabel("🎯 Analiz Hedeflerini Belirleyin")
        lbl_title.setStyleSheet("font-weight:bold; font-size:16px; color:#2c3e50;")

        info_btn = InfoButton(
            "<b>Nasıl Kullanılır?</b><br><br>"
            "• <b>Otomatik İsim:</b> Projedeki numaraların sahipleri veritabanından otomatik çekilir.<br>"
            "• <b>Manuel Ekleme:</b> Listede olmayan numaraları ve kime ait olduğunu elle ekleyebilirsiniz.<br>"
            "• <b>Grafik Etiketi:</b> Buradaki 'Hat Sahibi' ismi, analiz grafiğinde baloncuğun içinde görünecektir."
        )

        header_layout.addWidget(lbl_title)
        header_layout.addWidget(info_btn)
        header_layout.addStretch()
        layout.addLayout(header_layout)

        manual_frame = QFrame(); manual_frame.setStyleSheet("background-color: #f4f6f7; border-radius: 8px; padding: 10px; margin-top: 10px;")
        h_manual = QHBoxLayout(manual_frame)

        self.inp_manual = QLineEdit()
        self.inp_manual.setPlaceholderText("Numara (Örn: 555...)")
        self.inp_manual.setFixedWidth(150)

        self.inp_manual_name = QLineEdit()
        self.inp_manual_name.setPlaceholderText("Hat Sahibi / Etiket (Örn: Ahmet YILMAZ)")

        btn_add = QPushButton("➕ Ekle")
        btn_add.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 8px;")
        btn_add.clicked.connect(self.add_manual_number)

        h_manual.addWidget(QLabel("Numara:"))
        h_manual.addWidget(self.inp_manual)
        h_manual.addWidget(QLabel("İsim:"))
        h_manual.addWidget(self.inp_manual_name)
        h_manual.addWidget(btn_add)
        layout.addWidget(manual_frame)

        self.sel_table = QTableWidget()
        self.sel_table.setColumnCount(5)
        self.sel_table.setHorizontalHeaderLabels(["Seç", "Numara", "Hat Sahibi / Etiket", "Kaynak", "ID"])

        self.sel_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.sel_table.setColumnWidth(1, 130)
        self.sel_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

        self.sel_table.setColumnHidden(4, True)
        self.sel_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.sel_table.setStyleSheet("border: 1px solid #bdc3c7;")

        self.sel_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.sel_table.customContextMenuRequested.connect(self.show_context_menu)

        current_row = 0

        with DB() as conn:
            for num in self.available_numbers:
                name_row = conn.execute(
                    "SELECT AD, SOYAD FROM hts_abone WHERE ProjeID=? AND GSMNo=? LIMIT 1",
                    (self.project_id, num)
                ).fetchone()

                if name_row:
                    ad = name_row[0] or ""
                    soyad = name_row[1] or ""
                    display_name = f"{ad} {soyad}".strip()
                    if not display_name: display_name = "BİLİNMİYOR"
                else:
                    display_name = "BİLİNMİYOR"

                self.sel_table.insertRow(current_row)

                chk = QTableWidgetItem(); chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled); chk.setCheckState(Qt.CheckState.Checked)
                self.sel_table.setItem(current_row, 0, chk)
                self.sel_table.setItem(current_row, 1, QTableWidgetItem(str(num)))

                item_name = QTableWidgetItem(display_name)
                item_name.setForeground(QColor("#2c3e50"))
                item_name.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                self.sel_table.setItem(current_row, 2, item_name)

                self.sel_table.setItem(current_row, 3, QTableWidgetItem("Proje Verisi"))
                self.sel_table.setItem(current_row, 4, QTableWidgetItem("AUTO"))
                current_row += 1

            try:
                manual_rows = conn.execute("SELECT id, Numara, Aciklama FROM manuel_numaralar WHERE ProjeID=?", (self.project_id,)).fetchall()
                for db_id, num, name_val in manual_rows:
                    self.sel_table.insertRow(current_row)

                    chk = QTableWidgetItem(); chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled); chk.setCheckState(Qt.CheckState.Checked)
                    self.sel_table.setItem(current_row, 0, chk)

                    item_num = QTableWidgetItem(str(num))
                    item_num.setForeground(QColor("#d35400"))
                    self.sel_table.setItem(current_row, 1, item_num)

                    item_name = QTableWidgetItem(str(name_val))
                    item_name.setForeground(QColor("#d35400"))
                    self.sel_table.setItem(current_row, 2, item_name)

                    self.sel_table.setItem(current_row, 3, QTableWidgetItem("Manuel Giriş"))
                    self.sel_table.setItem(current_row, 4, QTableWidgetItem(str(db_id)))
                    current_row += 1
            except Exception as e:
                print(f"Manuel liste hatası: {e}")

        layout.addWidget(self.sel_table)

        btn_start = QPushButton("🔍 Analizi Başlat")
        btn_start.setStyleSheet("""
            QPushButton { background-color: #2980b9; color: white; font-weight: bold; padding: 12px; font-size: 16px; border-radius: 6px; margin-top: 10px; }
            QPushButton:hover { background-color: #3498db; }
        """)
        btn_start.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_start.clicked.connect(self.start_analysis)
        layout.addWidget(btn_start)

    def setup_results_ui(self):
        """2. Aşama: Sonuç Listesi ve Grafik Ekranı"""
        layout = QVBoxLayout(self.page_results)

        top_bar = QHBoxLayout()
        btn_back = QPushButton("🔙 Seçim Ekranına Dön")
        btn_back.setStyleSheet("background-color: #7f8c8d; color: white; padding: 6px;")
        btn_back.clicked.connect(lambda: self.stack.setCurrentIndex(0))
        top_bar.addWidget(btn_back); top_bar.addStretch()
        layout.addLayout(top_bar)

        splitter = QSplitter(Qt.Orientation.Horizontal)

        left_widget = QWidget(); left_layout = QVBoxLayout(left_widget); left_layout.setContentsMargins(0,0,0,0)

        header_layout = QHBoxLayout()
        header_layout.addWidget(QLabel("🔗 Bağlantı Listesi", styleSheet="font-weight:bold; font-size:14px; color:#2c3e50;"))
        info_cross = InfoButton(
            "<b>🔗 İlişki Ağı ve Bağlantı Analizi:</b><br>"
            "Bu ekran, seçtiğiniz hedeflerin kimlerle ortak görüştüğünü ortaya çıkarır.<br><br>"
            "• <b>Liste (Sol):</b> Ortak temas kurulan 3. şahıslar ve temas sayıları listelenir.<br>"
            "• <b>Grafik (Sağ):</b> İlişkiler görsel bir ağ haritası (Network) olarak çizilir.<br>"
            "• <b>Filtreler:</b> 'Sadece Direkt' butonu ile sadece hedeflerinizin <u>birbirleriyle</u> olan görüşmelerini izole edebilirsiniz."
        )
        header_layout.addWidget(info_cross)
        left_layout.addLayout(header_layout)

        self.res_search = QLineEdit(); self.res_search.setPlaceholderText("Sonuçlarda Ara...")
        self.res_search.textChanged.connect(self.filter_result_table)
        left_layout.addWidget(self.res_search)

        self.res_table = QTableWidget()
        self.res_table.setColumnCount(4)
        self.res_table.setHorizontalHeaderLabels(["Seç", "Numara / Bağlantı", "Temas", "İlişkili Hedefler"])
        self.res_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.res_table.setColumnWidth(0, 40); self.res_table.setColumnWidth(2, 60)
        self.res_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        left_layout.addWidget(self.res_table)

        btn_row = QHBoxLayout()
        btn_all = QPushButton("Tümünü Seç"); btn_all.clicked.connect(lambda: self.toggle_result_selection(True))
        btn_none = QPushButton("Temizle"); btn_none.clicked.connect(lambda: self.toggle_result_selection(False))
        btn_direct = QPushButton("Sadece Direkt"); btn_direct.clicked.connect(self.select_only_direct_results)
        btn_row.addWidget(btn_all); btn_row.addWidget(btn_none); btn_row.addWidget(btn_direct)
        left_layout.addLayout(btn_row)

        self.btn_update_graph = QPushButton("🔄 Grafiği Güncelle")
        self.btn_update_graph.setStyleSheet("background-color: #e67e22; color: white; font-weight: bold; padding: 10px;")
        self.btn_update_graph.clicked.connect(self.draw_graph)
        left_layout.addWidget(self.btn_update_graph)

        self.btn_loc_analysis = QPushButton("📍 İlişki Konumlarını Listele")
        self.btn_loc_analysis.setStyleSheet("""
            QPushButton { background-color: #16a085; color: white; font-weight: bold; padding: 10px; margin-top: 5px; }
            QPushButton:hover { background-color: #1abc9c; }
        """)
        self.btn_loc_analysis.clicked.connect(self.open_location_window)
        left_layout.addWidget(self.btn_loc_analysis)

        splitter.addWidget(left_widget)

        self.browser = EvidenceWebEngineView()
        self.browser.setStyleSheet("background-color: white; border: 1px solid #bdc3c7;")
        splitter.addWidget(self.browser)

        splitter.setSizes([450, 850])
        layout.addWidget(splitter)

    def open_window_safe(self, w: QDialog):
        """Dialog'u NON-MODAL güvenli şekilde aç (GC + fokus/aktif pencere düzeltmesi)."""
        if not hasattr(self, "_open_windows"):
            self._open_windows = []

        self._open_windows.append(w)

        w.setModal(False)
        w.setWindowModality(Qt.WindowModality.NonModal)

        w.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint
        )

        w.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)

        def _cleanup(*_):
            try:
                if w in self._open_windows:
                    self._open_windows.remove(w)
            except Exception:
                pass

            # parent tekrar aktif/fokus gelsin
            QTimer.singleShot(0, self.raise_)
            QTimer.singleShot(0, self.activateWindow)

        if hasattr(w, "finished"):
            w.finished.connect(_cleanup)
        w.destroyed.connect(_cleanup)

        w.show()
        w.raise_()
        w.activateWindow()


    def open_location_window(self):
        """Seçili hedefler VE bulunan ortak kişilerle Konum Analizi penceresini açar."""
        if not hasattr(self, 'selected_targets') or not self.selected_targets:
            ModernDialog.show_warning(self, "Hata", "Lütfen önce analiz edilecek hedefleri seçip 'Başlat'a basın.")
            return

        found_contacts = set()

        for i in range(self.res_table.rowCount()):
            if self.res_table.item(i, 0).checkState() == Qt.CheckState.Checked:
                if i < len(self.data_cache):
                    data = self.data_cache[i]

                    if data['type'] == 'DIRECT':
                        pass
                    else:
                        found_contacts.add(str(data['num']))
        is_any_row_checked = False
        for i in range(self.res_table.rowCount()):
            if self.res_table.item(i, 0).checkState() == Qt.CheckState.Checked:
                is_any_row_checked = True
                break

        if not is_any_row_checked:
            ModernDialog.show_warning(self, "Seçim Yok", "Listeden konumlarını görmek istediğiniz bağlantıları (satırları) seçiniz.")
            return
        dlg = CrossLocationDialog(self, self.project_id, self.selected_targets, list(found_contacts))
        self.open_window_safe(dlg)

    def add_manual_number(self):
        """Listeye ve DB'ye isimli numara ekler."""
        num = self.inp_manual.text().strip()
        clean_num = re.sub(r'\D', '', num)

        name_val = self.inp_manual_name.text().strip()
        if not name_val:
            name_val = "BİLİNMİYOR (Manuel)"

        if not clean_num or len(clean_num) < 3:
            ModernDialog.show_warning(self, "Hata", "Geçerli bir numara giriniz.")
            return

        try:
            new_id = None
            with DB() as conn:
                exists = conn.execute("SELECT 1 FROM manuel_numaralar WHERE ProjeID=? AND Numara=?", (self.project_id, clean_num)).fetchone()
                if exists:
                    ModernDialog.show_warning(self, "Mevcut", "Bu numara zaten listede.")
                    return

                cur = conn.execute("INSERT INTO manuel_numaralar (ProjeID, Numara, Aciklama) VALUES (?,?,?)",
                                  (self.project_id, clean_num, name_val))
                new_id = cur.lastrowid
                conn.commit()

            row = self.sel_table.rowCount()
            self.sel_table.insertRow(row)

            chk = QTableWidgetItem(); chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled); chk.setCheckState(Qt.CheckState.Checked)
            self.sel_table.setItem(row, 0, chk)

            item_num = QTableWidgetItem(clean_num)
            item_num.setForeground(QColor("#d35400"))
            self.sel_table.setItem(row, 1, item_num)

            item_name = QTableWidgetItem(name_val)
            item_name.setForeground(QColor("#d35400"))
            self.sel_table.setItem(row, 2, item_name)

            self.sel_table.setItem(row, 3, QTableWidgetItem("Manuel Giriş"))
            self.sel_table.setItem(row, 4, QTableWidgetItem(str(new_id)))

            self.inp_manual.clear()
            self.inp_manual_name.clear()

        except Exception as e:
            ModernDialog.show_error(self, "Kayıt Hatası", str(e))

    def show_context_menu(self, pos):
        """Manuel eklenenleri silmek için sağ tık menüsü."""
        item = self.sel_table.itemAt(pos)
        if not item:
            return

        row = item.row()

        id_item = self.sel_table.item(row, 4)
        if not id_item:
            return

        db_id_str = id_item.text().strip()

        if db_id_str == "AUTO":
            return

        menu = QMenu(self)
        menu = apply_menu_theme(QMenu(self))
        del_action = QAction("🗑️ Bu Manuel Numarayı Sil", self)
        del_action.triggered.connect(lambda: self.delete_manual_number(row, db_id_str))
        menu.addAction(del_action)
        menu.exec(self.sel_table.mapToGlobal(pos))

    def delete_manual_number(self, row, db_id):
        """Manuel numarayı hem tablodan hem DB'den siler."""
        if not ModernDialog.show_question(self, "Sil", "Bu numarayı listeden kaldırmak istiyor musunuz?"):
            return

        try:
            with DB() as conn:
                conn.execute("DELETE FROM manuel_numaralar WHERE id=?", (db_id,))
                conn.commit()
            self.sel_table.removeRow(row)
        except Exception as e:
            ModernDialog.show_error(self, "Silme Hatası", str(e))

    def start_analysis(self):
        """Seçili numaralarla analizi başlatır."""
        if not LicenseManager.require_valid_or_exit(self, "Ortak temas/ilişki analizi başlat"):
            return
        self.selected_targets = []

        for i in range(self.sel_table.rowCount()):
            if self.sel_table.item(i, 0).checkState() == Qt.CheckState.Checked:
                gsm = self.sel_table.item(i, 1).text()
                self.selected_targets.append(gsm)

        if len(self.selected_targets) < 1:
            ModernDialog.show_warning(self, "Eksik Seçim", "Lütfen en az 1 numara seçiniz.")
            return

        self.run_sql_analysis()
        self.stack.setCurrentIndex(1)

    def run_sql_analysis(self):
        if not LicenseManager.require_valid_or_exit(self, "Ortak temas/ilişki analizi çalıştır"):
            return
        self.data_cache = []

        try:
            with DB() as conn:
                cur = conn.cursor()

                # 1) Direkt temaslar (bunu aynen koruyoruz)
                for i in range(len(self.selected_targets)):
                    for j in range(i + 1, len(self.selected_targets)):
                        gsm1 = self.selected_targets[i]
                        gsm2 = self.selected_targets[j]
                        c1 = cur.execute(
                            "SELECT COUNT(*) FROM hts_gsm WHERE ProjeID=? AND GSMNo=? AND DIGER_NUMARA=?",
                            (self.project_id, gsm1, gsm2)
                        ).fetchone()[0]
                        c2 = cur.execute(
                            "SELECT COUNT(*) FROM hts_gsm WHERE ProjeID=? AND GSMNo=? AND DIGER_NUMARA=?",
                            (self.project_id, gsm2, gsm1)
                        ).fetchone()[0]
                        total = c1 + c2
                        if total > 0:
                            self.data_cache.append({
                                'num': f"{gsm1} <-> {gsm2}",
                                'name': 'Direkt Temas',
                                'count': total,
                                'targets': [gsm1, gsm2],
                                'type': 'DIRECT'
                            })

                # 2) Seçili hedeflerden gerçekten verisi olanları bul
                valid_sources = []
                for gsm in self.selected_targets:
                    if cur.execute(
                        "SELECT 1 FROM hts_gsm WHERE ProjeID=? AND GSMNo=? LIMIT 1",
                        (self.project_id, gsm)
                    ).fetchone():
                        valid_sources.append(gsm)

                # 3) Ortak bağlantılar (BURADA GSM FİLTRESİ EKLİYORUZ)
                if valid_sources:
                    placeholders = ",".join(["?"] * len(valid_sources))

                    sql_common = f"""
                        SELECT DIGER_NUMARA,
                               MAX(DIGER_ISIM),
                               COUNT(DISTINCT GSMNo) as HedefSayisi,
                               COUNT(*) as ToplamGorusme
                        FROM hts_gsm
                        WHERE ProjeID=? AND GSMNo IN ({placeholders})
                        GROUP BY DIGER_NUMARA
                        HAVING HedefSayisi > 1
                        ORDER BY ToplamGorusme DESC
                        LIMIT 500
                    """
                    params = [self.project_id] + valid_sources
                    rows = cur.execute(sql_common, params).fetchall()

                    for r in rows:
                        raw_diger_no = r[0]

                        # ---- GSM filtre: sadece 10 hane ve 5 ile başlayan kalsın ----
                        diger_digits = re.sub(r"\D", "", "" if raw_diger_no is None else str(raw_diger_no))
                        if len(diger_digits) >= 10:
                            diger_digits = diger_digits[-10:]

                        # İSTENEN: sadece GSM numarası (TR GSM: 5xxxxxxxxx)
                        if not (len(diger_digits) == 10 and diger_digits.startswith("5")):
                            continue
                        diger_no = diger_digits
                        # ------------------------------------------------------------

                        diger_isim = r[1] if r[1] else "Bilinmiyor"
                        toplam = r[3]

                        sql_who = f"""
                            SELECT DISTINCT GSMNo
                            FROM hts_gsm
                            WHERE ProjeID=? AND DIGER_NUMARA=? AND GSMNo IN ({placeholders})
                        """
                        t_rows = cur.execute(sql_who, [self.project_id, raw_diger_no] + valid_sources).fetchall()
                        related = [tr[0] for tr in t_rows]

                        self.data_cache.append({
                            'num': diger_no,
                            'name': diger_isim,
                            'count': toplam,
                            'targets': related,
                            'type': 'COMMON'
                        })

        except Exception as e:
            print(f"Analiz Hatası: {e}")

        # Tabloyu yeniden doldur (mevcut davranışı koru)
        self.res_table.setRowCount(len(self.data_cache))
        for i, item in enumerate(self.data_cache):
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            chk.setCheckState(Qt.CheckState.Checked if item['type'] == 'DIRECT' else Qt.CheckState.Unchecked)

            if item['type'] == 'DIRECT':
                self.res_table.setItem(i, 0, chk)
                for c in range(4):
                    if not self.res_table.item(i, c):
                        self.res_table.setItem(i, c, QTableWidgetItem(""))

                display_text = str(item['num'])
                self.res_table.item(i, 1).setText(display_text)
                self.res_table.item(i, 2).setText(str(item['count']))
                self.res_table.item(i, 3).setText(",".join(item.get('targets', [])))

            else:
                self.res_table.setItem(i, 0, chk)

                # Ortak bağlantıda artık 'num' kesin GSM olacak
                num_item = QTableWidgetItem(str(item['num']))
                name_item = QTableWidgetItem(str(item.get('name', '')))
                count_item = QTableWidgetItem(str(item.get('count', 0)))
                targets_item = QTableWidgetItem(",".join(item.get('targets', [])))

                self.res_table.setItem(i, 1, num_item)
                self.res_table.setItem(i, 2, count_item)
                self.res_table.setItem(i, 3, targets_item)

        self.res_table.resizeColumnsToContents()

    def filter_result_table(self, text):
        for i in range(self.res_table.rowCount()):
            match = False
            for j in range(1, 4):
                it = self.res_table.item(i, j)
                if it and text.lower() in it.text().lower(): match = True; break
            self.res_table.setRowHidden(i, not match)

    def toggle_result_selection(self, state):
        st = Qt.CheckState.Checked if state else Qt.CheckState.Unchecked
        for i in range(self.res_table.rowCount()):
            if not self.res_table.isRowHidden(i): self.res_table.item(i, 0).setCheckState(st)

    def select_only_direct_results(self):
        for i in range(self.res_table.rowCount()):
            is_direct = (self.data_cache[i]['type'] == 'DIRECT')
            self.res_table.item(i, 0).setCheckState(Qt.CheckState.Checked if is_direct else Qt.CheckState.Unchecked)

    def draw_graph(self):
        """Vis.js Grafiği (GÜNCELLENDİ: İSİM GÖSTERİMİ)."""
        nodes = []; edges = []; added_nodes = set()
        colors = ['#e74c3c', '#f1c40f', '#3498db', '#9b59b6', '#2ecc71', '#e67e22']

        target_name_map = {}
        for i in range(self.sel_table.rowCount()):
            gsm = self.sel_table.item(i, 1).text()
            name = self.sel_table.item(i, 2).text()
            target_name_map[gsm] = name

        for idx, gsm in enumerate(self.selected_targets):
            color = colors[idx % len(colors)]
            t_id = f"T_{gsm}"
            display_name = target_name_map.get(gsm, "BİLİNMİYOR")
            nodes.append({'id': t_id, 'label': f"{gsm}\n({display_name})", 'color': color, 'size': 45, 'shape': 'box', 'font': {'color': 'white', 'face': 'Segoe UI', 'size': 16}, 'margin': 12})
            added_nodes.add(t_id)

        target_color_map = {gsm: colors[i % len(colors)] for i, gsm in enumerate(self.selected_targets)}

        for i in range(self.res_table.rowCount()):
            if self.res_table.item(i, 0).checkState() == Qt.CheckState.Checked:
                data = self.data_cache[i]

                if data['type'] == 'DIRECT':
                    t1, t2 = data['targets']
                    edges.append({'from': f"T_{t1}", 'to': f"T_{t2}", 'width': 3, 'color': {'color': 'black', 'highlight': 'red'}, 'dashes': True, 'label': f" {data['count']} Temas ", 'font': {'align': 'horizontal', 'background': 'white'}})
                else:
                    c_num = data['num']
                    c_name = data.get('name', 'Bilinmiyor')
                    c_id = f"C_{c_num}"

                    if c_id not in added_nodes:
                        label_text = f"{c_num}\n({data['count']})"
                        if c_name and c_name != "Bilinmiyor":
                            label_text = f"{c_num}\n{c_name}\n({data['count']})"

                        nodes.append({'id': c_id, 'label': label_text, 'color': '#95a5a6', 'size': 25, 'shape': 'dot', 'font': {'background': 'white', 'size': 12}})
                        added_nodes.add(c_id)

                    for t_gsm in data['targets']:
                        line_color = target_color_map.get(t_gsm, '#bdc3c7')
                        edges.append({'from': c_id, 'to': f"T_{t_gsm}", 'color': line_color, 'width': 2})

        nodes_json = json.dumps(nodes); edges_json = json.dumps(edges)

        js_library = ""
        local_js_path = os.path.join(APP_DIR, "assets", "vis-network.min.js")

        if os.path.exists(local_js_path):
            try:
                with open(local_js_path, "r", encoding="utf-8") as f:
                    js_library = f"<script>{f.read()}</script>"
            except: pass

        if not js_library:
            js_library = '<script type="text/javascript" src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>'

        html = f"""<!DOCTYPE html><html><head>{js_library}<style> body {{ margin: 0; padding: 0; overflow: hidden; }} #mynetwork {{ width: 100%; height: 100vh; }} </style></head><body><div id="mynetwork"></div><script type="text/javascript">var nodes = new vis.DataSet({nodes_json});var edges = new vis.DataSet({edges_json});var container = document.getElementById('mynetwork');var data = {{ nodes: nodes, edges: edges }};var options = {{nodes: {{ borderWidth: 2, shadow: true, font: {{ face: 'Segoe UI' }} }},edges: {{ smooth: {{ type: 'continuous', roundness: 0.5 }}, font: {{ align: 'middle' }} }},physics: {{enabled: true,solver: 'repulsion',repulsion: {{ nodeDistance: 350, springLength: 300, damping: 0.09 }},stabilization: {{ enabled: true, iterations: 1000 }}}},interaction: {{ navigationButtons: false, zoomView: true }}}};var network = new vis.Network(container, data, options);network.once("stabilizationIterationsDone", function() {{network.fit({{ animation: {{ duration: 1000, easingFunction: 'easeInOutQuad' }} }});}});</script></body></html>"""
        self.browser.setHtml(html)


class CrossLocationDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, main_targets, found_contacts):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.project_id = project_id

        self.main_targets = [re.sub(r'\D', '', str(t)) for t in main_targets]
        self.found_contacts = [re.sub(r'\D', '', str(t)) for t in found_contacts]

        if not self.found_contacts:
            self.search_mode = "DIRECT_ONLY"
        else:
            self.search_mode = "MIXED"

        self.setWindowTitle("İlişki Konum Analizi (Çift Taraflı Konum Tespiti)")
        self.resize(1500, 800)

        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        top_container = QFrame()
        top_container.setStyleSheet(".QFrame { background-color: #f0f8ff; border: 1px solid #bdc3c7; border-radius: 8px; }")
        v_top = QVBoxLayout(top_container)

        info_label = QLabel(
            "<div style='color:#2c3e50; font-family:Segoe UI; font-size:13px;'>"
            "<b>ℹ️ NASIL KULLANILIR?</b><br>"
            "• Bu ekran, iki tarafın da HTS kaydının bulunduğu ve zaman damgalarının <b>±3 saniye</b> eşleştiği 'Kesin Temas' anlarını gösterir.<br>"
            "• <b>Harita:</b> Satıra <b>ÇİFT TIKLAYARAK</b> iki tarafın konumunu haritada görebilirsiniz.<br>"
            "• <b>Görünüm:</b> Baz adresleri sığmadığında '...' ile biter. <b>Sütun çizgilerini sürükleyerek genişletebilir</b> veya fareyi üzerine getirerek tamamını okuyabilirsiniz."
            "</div>"
        )
        info_label.setWordWrap(True)
        v_top.addWidget(info_label)

        h_filters = QHBoxLayout()
        h_filters.addWidget(QLabel("📅 <b>Tarih Aralığı:</b>", styleSheet="color:#2980b9; font-size:13px;"))

        now = QDateTime.currentDateTime()
        self.dt_start = QDateTimeEdit(now.addYears(-1))
        self.dt_start.setCalendarPopup(True); self.dt_start.setDisplayFormat("dd.MM.yyyy HH:mm")
        self.dt_end = QDateTimeEdit(now)
        self.dt_end.setCalendarPopup(True); self.dt_end.setDisplayFormat("dd.MM.yyyy HH:mm")

        h_filters.addWidget(self.dt_start); h_filters.addWidget(QLabel("➖")); h_filters.addWidget(self.dt_end)

        btn_run = QPushButton("📍 Konumları Getir")
        btn_run.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 6px 15px; border-radius:4px;")
        btn_run.clicked.connect(self.run_location_analysis)
        h_filters.addWidget(btn_run)

        h_filters.addStretch()
        v_top.addLayout(h_filters)
        layout.addWidget(top_container)

        self.table = GenericDatabaseTable(
            ["Tarih/Saat", "Kaynak No", "Hedef No", "Yön", "Süre", "Kaynak Baz", "Hedef Baz", "Kaynak IMEI", "Hedef IMEI"],
            chart_mode='none'
        )
        t = self.table.table
        t.doubleClicked.connect(self.open_map_preview)

        t.setItemDelegate(ElidedItemDelegate(t))
        t.setMouseTracking(True)

        layout.addWidget(self.table)

        QTimer.singleShot(200, self.run_location_analysis)

    def apply_column_settings(self):
        t = self.table.table
        h = t.horizontalHeader()

        t.setWordWrap(False)

        t.setColumnWidth(0, 130) # Tarih
        t.setColumnWidth(1, 110) # Kaynak
        t.setColumnWidth(2, 110) # Hedef
        t.setColumnWidth(3, 70)  # Yön
        t.setColumnWidth(4, 70)  # Süre
        t.setColumnWidth(7, 120) # IMEI 1
        t.setColumnWidth(8, 120) # IMEI 2

        h.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)
        t.setColumnWidth(5, 280)

        h.setSectionResizeMode(6, QHeaderView.ResizeMode.Interactive)
        t.setColumnWidth(6, 280)

        for c in [0, 1, 2, 3, 4, 7, 8]:
            h.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)

    def run_location_analysis(self):
        s_date = self.dt_start.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        e_date = self.dt_end.dateTime().toString("yyyy-MM-dd HH:mm:ss")

        targets_ph = ','.join(['?'] * len(self.main_targets))

        try:
            with DB() as conn:
                params = []
                where_clause = ""

                if self.search_mode == "DIRECT_ONLY":
                    params.extend(self.main_targets + self.main_targets)
                    where_clause = f"AND t1.GSMNo IN ({targets_ph}) AND t1.DIGER_NUMARA IN ({targets_ph})"
                else:
                    contacts_ph = ','.join(['?'] * len(self.found_contacts))
                    params.extend(self.main_targets + self.main_targets)
                    params.extend(self.main_targets + self.found_contacts)
                    params.extend(self.found_contacts + self.main_targets)

                    where_clause = f"""
                        AND (
                            (t1.GSMNo IN ({targets_ph}) AND t1.DIGER_NUMARA IN ({targets_ph}))
                            OR
                            (t1.GSMNo IN ({targets_ph}) AND t1.DIGER_NUMARA IN ({contacts_ph}))
                            OR
                            (t1.GSMNo IN ({contacts_ph}) AND t1.DIGER_NUMARA IN ({targets_ph}))
                        )
                    """

                def to_iso(col):
                    return f"substr({col}, 7, 4) || '-' || substr({col}, 4, 2) || '-' || substr({col}, 1, 2) || substr({col}, 11)"

                sql = f"""
                    SELECT 
                        t1.TARIH, 
                        t1.GSMNo as Kaynak, 
                        t1.DIGER_NUMARA as Hedef, 
                        t1.TIP, 
                        t1.SURE, 
                        t1.BAZ as KaynakBaz,
                        t2.BAZ as HedefBaz, 
                        t1.IMEI as KaynakIMEI,
                        t2.IMEI as HedefIMEI
                    FROM hts_gsm t1
                    JOIN hts_gsm t2 ON 
                        t1.ProjeID = t2.ProjeID AND
                        substr(replace(replace(replace(t2.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = 
                        substr(replace(replace(replace(t1.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND
                        substr(replace(replace(replace(t2.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = 
                        substr(replace(replace(replace(t1.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND
                        datetime({to_iso('t2.TARIH')}) BETWEEN 
                            datetime({to_iso('t1.TARIH')}, '-3 seconds') AND 
                            datetime({to_iso('t1.TARIH')}, '+3 seconds')
                    WHERE t1.ProjeID=?
                      AND ({to_iso('t1.TARIH')}) BETWEEN ? AND ?
                      {where_clause}
                      AND t1.GSMNo != t1.DIGER_NUMARA 
                    ORDER BY 
                        {to_iso('t1.TARIH')} DESC
                """

                final_params = [self.project_id, s_date, e_date] + params
                rows = conn.execute(sql, final_params).fetchall()

                if not rows:
                    self.table.set_data([])
                    ModernDialog.show_info(self, "Sonuç Yok", "Kriterlere uygun ve zamanı eşleşen (±3 sn) kayıt bulunamadı.")
                    return

                display_data = []
                for r in rows:
                    tarih, k, h, tip, s, baz_k, baz_h, imei_k, imei_h = r
                    yon = "Giden ->" if ("Aradı" in str(tip) or "Giden" in str(tip)) else "<- Gelen"
                    display_data.append([
                        tarih, k, h, yon, s,
                        baz_k if baz_k else "",
                        baz_h if baz_h else "",
                        imei_k if imei_k else "",
                        imei_h if imei_h else ""
                    ])

                self.table.set_data(display_data)

                self.apply_column_settings()

        except Exception as e:
            print(f"Konum analizi hatası: {e}")
            ModernDialog.show_error(self, "Hata", str(e))

    def open_map_preview(self, index):
        try:
            row = index.row()
            kaynak_no = self.table.proxy_model.index(row, 1).data()
            hedef_no = self.table.proxy_model.index(row, 2).data()
            kaynak_baz = self.table.proxy_model.index(row, 5).data()
            karsi_baz = self.table.proxy_model.index(row, 6).data()

            if (not kaynak_baz or len(str(kaynak_baz)) < 5) and (not karsi_baz or len(str(karsi_baz)) < 5):
                ModernDialog.show_warning(self, "Konum Yok", "Gösterilecek geçerli bir konum verisi bulunmuyor.")
                return

            lbl1 = f"{kaynak_no} (Kaynak)"
            lbl2 = f"{hedef_no} (Karşı)"

            dlg = MapPreviewDialog(
                self, self.project_id, kaynak_no,
                str(kaynak_baz) if kaynak_baz else "",
                str(karsi_baz) if karsi_baz else "",
                label1=lbl1, label2=lbl2
            )
            dlg.exec()
        except Exception as e:
            print(f"Harita açma hatası: {e}")


class SpeedAnomalyDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, gsm_number):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.setWindowTitle(f"Hız ve Mesafe İhlal Analizi - {gsm_number}")
        self.resize(1150, 650)
        self.project_id = project_id
        self.gsm_number = gsm_number

        self.anomaly_cache = []

        layout = QVBoxLayout(self)

        top_frame = QFrame()
        top_frame.setStyleSheet(
            "background-color: #ffebee; border: 1px solid #ef9a9a; border-radius: 8px;"
        )
        hl = QHBoxLayout(top_frame)

        info_btn = InfoButton(
            "<b>🚀 Hız ve Mesafe İhlali (Impossible Travel):</b><br>"
            "İki sinyal arasında, fiziksel olarak katedilmesi imkansız olan mesafeleri tespit eder.<br><br>"
            "• <b>Mantık:</b> Ardışık iki kayıt arasındaki mesafe ve zaman farkından 'Hız' hesaplanır.<br>"
            "• <b>Mesafe Toleransı (Önemli):</b> Baz istasyonu değişikliklerinden (Drift/Sinyal Sıçraması) kaynaklanan <u>yanlış alarmları</u> önlemek için, belirli bir mesafenin (Örn: 10 km) altındaki değişimler hesaplamaya dahil edilmez.<br>"
            "• <b>Amaç:</b> Klonlanmış hat şüphesi veya hatalı veri tespiti."
        )
        hl.addWidget(info_btn)
        hl.addSpacing(15)

        hl.addWidget(QLabel("🚀 <b>Hız Limiti:</b>"))
        self.spin_speed = QSpinBox()
        self.spin_speed.setRange(50, 2000)
        self.spin_speed.setValue(180)
        self.spin_speed.setSuffix(" km/s")
        hl.addWidget(self.spin_speed)

        hl.addSpacing(20)

        hl.addWidget(QLabel("📏 <b>Mesafe Toleransı (Drift Filtresi):</b>"))
        self.spin_dist = QDoubleSpinBox()
        self.spin_dist.setRange(0.5, 500.0)
        self.spin_dist.setValue(10.0)
        self.spin_dist.setSuffix(" km")
        self.spin_dist.setSingleStep(1.0)
        hl.addWidget(self.spin_dist)

        hl.addStretch()

        btn_run = QPushButton("Analizi Başlat")
        btn_run.setStyleSheet(
            "background-color: #c62828; color: white; font-weight: bold; padding: 8px 15px;"
        )
        btn_run.clicked.connect(self.run_analysis)
        hl.addWidget(btn_run)

        layout.addWidget(top_frame)

        headers = [
            "Başlangıç Zamanı", "Başlangıç Konumu",
            "Bitiş Zamanı", "Bitiş Konumu",
            "Mesafe (km)", "Süre (dk)", "Hız (km/s)"
        ]

        # ✅ Delil ekleme (sağ tık) tamamen GenericDatabaseTable altyapısından gelsin.
        self.table = GenericDatabaseTable(headers=headers, enable_date_filter=False, chart_mode='embedded', info_text=None, enable_evidence_menu=True)
        self.table.project_id = self.project_id
        self.table.gsm_number = self.gsm_number
        self.table.table_title = "Hız/Mesafe İhlali (Impossible Travel)"

        # Görünüm stilleri (önceki QTableWidget görünümüne yakın)
        self.table.table.setStyleSheet('''
            QTableView {
                background: transparent;
                gridline-color: #d0d0d0;
            }
            QTableView::item {
                background-color: rgba(255, 255, 255, 235);
            }
            QHeaderView::section {
                background-color: rgba(240, 240, 240, 245);
                color: #222;
                font-weight: 600;
                border: 1px solid #cfcfcf;
            }
        ''')
        self.table.table.viewport().setAutoFillBackground(False)

        try:
            self.table.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        except Exception:
            pass

        # Çift tık: haritada göster
        self.table.table.doubleClicked.connect(self.open_map_for_row)

        layout.addWidget(self.table)

        self.lbl_status = QLabel("Analiz için 'Başlat' butonuna basınız.")
        layout.addWidget(self.lbl_status)

    def calculate_haversine(self, lat1, lon1, lat2, lon2):
        """İki nokta arasındaki kuş uçuşu mesafeyi hesaplar."""
        R = 6371.0
        dlat = math.radians(lat2 - lat1); dlon = math.radians(lon2 - lon1)
        a = math.sin(dlat / 2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2)**2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return R * c

    def parse_coordinate(self, text):
        if not text: return None
        coords = re.findall(r"(\d{2}\.\d{4,})", str(text))
        if len(coords) >= 2:
            try:
                v1, v2 = float(coords[-2]), float(coords[-1])
                if 35 < v1 < 43 and 25 < v2 < 46: return (v1, v2)
                elif 35 < v2 < 43 and 25 < v1 < 46: return (v2, v1)
                else: return (v1, v2)
            except: return None
        return None

    def clean_gsm(self, val):
        s = re.sub(r'\D', '', str(val))
        return s[-10:] if len(s) >= 10 else s

    def run_analysis(self):
        limit_kmh = self.spin_speed.value()
        limit_dist = self.spin_dist.value()

        self.table.set_data([])
        self.anomaly_cache = []
        self.lbl_status.setText("⏳ Analiz yapılıyor...")
        QApplication.processEvents()

        target_clean = self.clean_gsm(self.gsm_number)
        processed_points = []

        try:
            with DB() as conn:
                sql = """
                    SELECT TARIH, BAZ, NUMARA FROM hts_gsm 
                    WHERE ProjeID=? AND GSMNo=? AND BAZ IS NOT NULL AND length(BAZ) > 5
                    UNION ALL
                    SELECT TARIH, BAZ, NUMARA FROM hts_gprs 
                    WHERE ProjeID=? AND GSMNo=? AND BAZ IS NOT NULL AND length(BAZ) > 5
                    UNION ALL
                    SELECT TARIH, BAZ, NUMARA FROM hts_wap 
                    WHERE ProjeID=? AND GSMNo=? AND BAZ IS NOT NULL AND length(BAZ) > 5
                """
                params = (self.project_id, self.gsm_number) * 3
                rows = conn.execute(sql, params).fetchall()

            if not rows:
                self.lbl_status.setText("Analiz İptal: Veri yok.")
                return

            for r in rows:
                t_str, baz_txt, raw_num = r
                if self.clean_gsm(raw_num) != target_clean: continue

                coord = self.parse_coordinate(baz_txt)
                if not coord: continue

                try:
                    t_str = str(t_str).strip()
                    fmt = "%d.%m.%Y %H:%M:%S" if "." in t_str else "%Y-%m-%d %H:%M:%S"
                    if "/" in t_str: fmt = "%d/%m/%Y %H:%M:%S"

                    dt = datetime.strptime(t_str, fmt)

                    processed_points.append({
                        'dt': dt, 'lat': coord[0], 'lon': coord[1],
                        'baz': baz_txt, 't_str': t_str
                    })
                except: continue

            if len(processed_points) < 2:
                self.lbl_status.setText("Yetersiz koordinatlı veri.")
                return

            processed_points.sort(key=lambda x: x['dt'])
            anomalies = []

            for i in range(len(processed_points) - 1):
                p1 = processed_points[i]
                p2 = processed_points[i+1]

                diff_seconds = (p2['dt'] - p1['dt']).total_seconds()
                if diff_seconds <= 0: continue

                dist_km = self.calculate_haversine(p1['lat'], p1['lon'], p2['lat'], p2['lon'])

                if dist_km < limit_dist:
                    continue

                diff_hours = diff_seconds / 3600.0
                speed = dist_km / diff_hours

                if speed > limit_kmh:
                    time_diff_str = f"{diff_seconds/60:.1f}"
                    dist_str = f"{dist_km:.2f}"

                    row_display = (
                        p1['t_str'], p1['baz'],
                        p2['t_str'], p2['baz'],
                        dist_str,
                        time_diff_str,
                        f"{speed:.0f}"
                    )
                    anomalies.append(row_display)

                    bubble_text = f"📏 {dist_str} km<br>⏱️ {time_diff_str} dk<br>🚀 {speed:.0f} km/s"

                    self.anomaly_cache.append({
                        'p1': p1, 'p2': p2,
                        'info': bubble_text
                    })

            # Tabloya Bas
            self.table.set_data(anomalies)

            self.lbl_status.setText(f"Tespit Edilen: {len(anomalies)} kayıt. (Harita için çift tıklayın)")

        except Exception as e:
            self.lbl_status.setText(f"Hata: {e}")

    def open_map_for_row(self, index):
        """Çift tıklanan satırı haritada baloncuklu olarak gösterir."""
        try:
            row = index.row()
            if row < 0 or row >= len(self.anomaly_cache): return

            payload = self.anomaly_cache[row]
            p1 = payload['p1']; p2 = payload['p2']

            dlg = SpeedAnomalyMapDialog(self, p1, p2, payload['info'])
            dlg.exec()
        except Exception as e:
            print("Harita açma hatası:", e)


class MapDialog(WatermarkDialogMixin, QDialog):
    """Tarih filtreli, toplu seçim, ARAMA, HIGHLIGHT ve KALICI HAFIZA özellikli Harita Analizi."""
    def __init__(self, parent, project_id, gsm_number, init_start_dt, init_end_dt):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.project_id = project_id
        self.gsm_number = gsm_number

        self.searched_location = None
        self.custom_markers = []
        self.temp_search_result = None

        self.setWindowTitle(f"Coğrafi Konum Analizi - {gsm_number}")
        self.resize(1350, 900)
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)
        splitter = QSplitter(Qt.Orientation.Horizontal)

        left_widget = QWidget(); left_widget.setStyleSheet("background-color: #f7f9f9;")
        left_layout = QVBoxLayout(left_widget); left_layout.setContentsMargins(10, 10, 10, 10)

        h_map_title = QHBoxLayout()
        h_map_title.addWidget(QLabel("🌍 Konum Analiz Araçları", styleSheet="font-size:14px; font-weight:bold;"))
        info_map = InfoButton(
            "<b>🌍 Coğrafi Konum ve Baz Analizi:</b><br>"
            "Şüphelinin bulunduğu konumları harita üzerinde keşfedin.<br><br>"
            "• <b>Sinyal Analizi:</b> Listeden baz istasyonlarını seçerek yoğunluk haritası oluşturun.<br>"
            "• <b>Özel Konum Ekleme:</b> 'Ev', 'İş', 'Olay Yeri' gibi kritik adresleri arayıp haritaya <b>Yıldız (⭐)</b> olarak sabitleyebilirsiniz.<br>"
            "• <b>Kalıcı Hafıza:</b> Eklediğiniz özel konumlar projedeki TÜM haritalarda görünmeye devam eder."
        )
        h_map_title.addWidget(info_map)
        left_layout.addLayout(h_map_title)

        search_grp = QFrame(); search_grp.setStyleSheet("background-color: #fff3e0; border-radius: 8px; padding: 5px; border: 1px solid #e67e22;")
        l_search = QVBoxLayout(search_grp)
        l_search.addWidget(QLabel("🔍 Özel Konum Ekle (Kalıcı):", styleSheet="font-weight:bold; color:#d35400"))

        self.inp_search = QLineEdit(); self.inp_search.setPlaceholderText("1. Adım: Adres veya Koordinat Ara...")
        self.inp_search.returnPressed.connect(self.search_location)
        l_search.addWidget(self.inp_search)

        self.inp_desc = QLineEdit()
        self.inp_desc.setPlaceholderText("2. Adım: Konum Açıklaması Girin (Örn: Şüpheli Evi)")
        self.inp_desc.setStyleSheet("background-color: white;")
        l_search.addWidget(self.inp_desc)

        h_s_btns = QHBoxLayout()
        btn_find = QPushButton("Bul"); btn_find.setStyleSheet("background-color: #e67e22; color: white; font-weight: bold; padding: 6px;")
        btn_find.clicked.connect(self.search_location)

        btn_add = QPushButton("Kaydet ve Ekle 💾"); btn_add.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 6px;")
        btn_add.clicked.connect(self.add_custom_marker)

        h_s_btns.addWidget(btn_find); h_s_btns.addWidget(btn_add)
        l_search.addLayout(h_s_btns)

        self.lbl_search_result = QLabel("Konum aranmadı."); self.lbl_search_result.setStyleSheet("color: #7f8c8d; font-size: 11px; margin-top:5px;"); self.lbl_search_result.setWordWrap(True)
        l_search.addWidget(self.lbl_search_result)

        self.loc_table = QTableWidget(); self.loc_table.setColumnCount(1)
        self.loc_table.horizontalHeader().setVisible(False); self.loc_table.verticalHeader().setVisible(False)
        self.loc_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.loc_table.setStyleSheet("background-color: white; border: 1px solid #e67e22;")
        self.loc_table.setFixedHeight(100); self.loc_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.loc_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.loc_table.cellClicked.connect(self.on_custom_loc_clicked)
        l_search.addWidget(self.loc_table)

        btn_del_loc = QPushButton("Seçili Konumu Sil 🗑️"); btn_del_loc.setStyleSheet("background-color: #c0392b; color: white; font-size: 11px; padding: 4px;")
        btn_del_loc.clicked.connect(self.remove_custom_marker)
        l_search.addWidget(btn_del_loc)
        left_layout.addWidget(search_grp)

        grp_date = QFrame(); grp_date.setStyleSheet("background-color: white; border-radius: 8px; padding: 5px; border: 1px solid #bdc3c7; margin-top: 10px;")
        l_date = QVBoxLayout(grp_date)
        l_date.addWidget(QLabel("📅 Analiz Aralığı:", styleSheet="font-weight:bold; color:#2c3e50"))
        self.dt_start = QDateTimeEdit(init_start_dt); self.dt_start.setCalendarPopup(True); self.dt_start.setDisplayFormat("dd.MM.yyyy HH:mm")
        self.dt_end = QDateTimeEdit(init_end_dt); self.dt_end.setCalendarPopup(True); self.dt_end.setDisplayFormat("dd.MM.yyyy HH:mm")
        l_date.addWidget(self.dt_start); l_date.addWidget(self.dt_end)
        btn_filter = QPushButton("Baz İstasyonlarını Getir"); btn_filter.setStyleSheet("background-color: #2980b9; color: white; font-weight: bold; padding: 8px; border-radius: 4px;")
        btn_filter.clicked.connect(self.load_baz_data)
        l_date.addWidget(btn_filter)
        left_layout.addWidget(grp_date)

        left_layout.addWidget(QLabel("📡 Baz İstasyonları (Tıklayıp Odaklanın):", styleSheet="font-weight:bold; margin-top:10px; color:#34495e;"))
        self.table = QTableWidget(); self.table.setColumnCount(3); self.table.setHorizontalHeaderLabels(["Seç", "Konum / Adres", "Sinyal"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 40); self.table.setColumnWidth(2, 70)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet("background-color: white; border: 1px solid #bdc3c7; selection-background-color: #d6eaf8; selection-color: black;")
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.cellClicked.connect(self.on_baz_clicked)
        left_layout.addWidget(self.table)

        h_btns = QHBoxLayout()
        btn_all = QPushButton("✅ Tümünü Seç"); btn_all.setStyleSheet("background-color: #16a085; color: white; padding: 6px;")
        btn_all.clicked.connect(lambda: self.toggle_all(True))
        btn_none = QPushButton("⬜ Seçimi Kaldır"); btn_none.setStyleSheet("background-color: #7f8c8d; color: white; padding: 6px;")
        btn_none.clicked.connect(lambda: self.toggle_all(False))
        h_btns.addWidget(btn_all); h_btns.addWidget(btn_none)
        left_layout.addLayout(h_btns)

        btn_draw = QPushButton("🌍 HARİTAYI GÜNCELLE"); btn_draw.setStyleSheet("background-color: #d35400; color: white; font-weight: bold; padding: 15px; font-size: 14px; border-radius: 5px; margin-top: 10px;")
        btn_draw.clicked.connect(lambda: self.generate_map())
        left_layout.addWidget(btn_draw)
        splitter.addWidget(left_widget)

        right_widget = QWidget(); r_layout = QVBoxLayout(right_widget); r_layout.setContentsMargins(0,0,0,0)
        self.browser = EvidenceWebEngineView(); self.browser.setStyleSheet("background-color: white;")
        r_layout.addWidget(self.browser)
        splitter.addWidget(right_widget); splitter.setSizes([400, 900])
        layout.addWidget(splitter)

        self.load_custom_locations()
        self.load_baz_data()

    def capture_map_evidence(self):
        if hasattr(self.parent(), 'capture_chart_screenshot'):
            self.parent().capture_chart_screenshot(self.browser, self.windowTitle())

    def load_custom_locations(self):
        """Veritabanındaki kayıtlı özel konumları çeker (PROJE BAZLI)."""
        try:
            with DB() as conn:
                rows = conn.execute("SELECT id, Lat, Lon, Label FROM ozel_konumlar WHERE ProjeID=?",
                                   (self.project_id,)).fetchall()

            self.custom_markers = []
            self.loc_table.setRowCount(0)

            for r in rows:
                db_id, lat, lon, label = r
                self.custom_markers.append((lat, lon, label, db_id))

                row_idx = self.loc_table.rowCount()
                self.loc_table.insertRow(row_idx)
                short = label.split(",")[0] if "," in label else label
                self.loc_table.setItem(row_idx, 0, QTableWidgetItem(f"⭐ {short}"))

        except Exception as e:
            print(f"Özel Konum Yükleme Hatası: {e}")

    def on_custom_loc_clicked(self, row, col):
        if row < len(self.custom_markers):
            lat, lon, _, _ = self.custom_markers[row]
            self.generate_map(focus_point=(lat, lon))

    def on_baz_clicked(self, row, col):
        item = self.table.item(row, 1)
        if not item: return

        pt = self.extract_coords(item.text())

        if pt:
            self.generate_map(focus_point=(pt[0], pt[1]))
        else:
            pass

    def search_location(self):
        from geopy.geocoders import Nominatim
        query = self.inp_search.text().strip()
        if not query: return

        def check_net():
            try: socket.create_connection(("8.8.8.8", 53), timeout=1.0); return True
            except: return False

        is_online = check_net()
        lat, lon, name = None, None, ""

        coords = re.findall(r"([-+]?\d{1,2}\.\d+)", query)
        if len(coords) >= 2:
            try:
                lat, lon = float(coords[0]), float(coords[1])
                name = f"Koordinat ({lat}, {lon})"
                self.lbl_search_result.setText(f"✅ Koordinat Bulundu: {lat}, {lon}")
                self.lbl_search_result.setStyleSheet("color: #27ae60;")
                self.temp_search_result = (lat, lon, name)
                self.searched_location = (lat, lon, name)
                self.generate_map(focus_point=(lat, lon))
                return
            except: pass

        if is_online:
            try:
                geolocator = Nominatim(user_agent="htstakip_analiz")
                location = geolocator.geocode(query)
                if location:
                    lat, lon = location.latitude, location.longitude
                    name = location.address
                    self.lbl_search_result.setText(f"✅ Adres Bulundu: {name[:40]}...")
                    self.lbl_search_result.setStyleSheet("color: #27ae60;")
                    self.temp_search_result = (lat, lon, name)
                    self.searched_location = (lat, lon, name)
                    self.generate_map(focus_point=(lat, lon))
                else:
                    self.lbl_search_result.setText("❌ Adres bulunamadı.")
                    self.lbl_search_result.setStyleSheet("color: #c0392b;")
            except Exception as e:
                self.lbl_search_result.setText(f"Hata: {str(e)}")
        else:
            self.lbl_search_result.setText("⚠️ Çevrimdışı moddasınız. Lütfen 'Enlem, Boylam' formatında koordinat giriniz.")
            self.lbl_search_result.setStyleSheet("color: #e67e22; font-weight:bold;")
            ModernDialog.show_warning(self, "Çevrimdışı Arama",
                "İnternet bağlantısı olmadığı için İsim/Adres araması yapılamıyor.\n\n"
                "Lütfen Google Maps vb. bir kaynaktan aldığınız koordinatları şu formatta girin:\n"
                "Örnek: 39.9207, 32.8541")

    def add_custom_marker(self):
        if not self.temp_search_result:
            ModernDialog.show_warning(self, "Hata", "Önce bir konum arayıp bulmalısınız.")
            return

        lat, lon, auto_name = self.temp_search_result

        user_desc = self.inp_desc.text().strip()
        final_label = user_desc if user_desc else auto_name

        try:
            new_id = None
            with DB() as conn:
                cur = conn.execute("INSERT INTO ozel_konumlar (ProjeID, GSMNo, Lat, Lon, Label) VALUES (?,?,?,?,?)",
                                  (self.project_id, self.gsm_number, lat, lon, final_label))
                new_id = cur.lastrowid

            self.custom_markers.append((lat, lon, final_label, new_id))
            row = self.loc_table.rowCount(); self.loc_table.insertRow(row)

            self.loc_table.setItem(row, 0, QTableWidgetItem(f"⭐ {final_label}"))
            self.lbl_search_result.setText("Kaydedildi."); self.inp_search.clear(); self.inp_desc.clear(); self.temp_search_result = None
            self.generate_map(focus_point=(lat, lon))
        except Exception as e: ModernDialog.show_error(self, "Kayıt Hatası", str(e))

    def remove_custom_marker(self):
        row = self.loc_table.currentRow()
        if row >= 0:
            try:
                marker = self.custom_markers[row]
                db_id = marker[3]
                with DB() as conn: conn.execute("DELETE FROM ozel_konumlar WHERE id=?", (db_id,))
            except: pass
            self.loc_table.removeRow(row); del self.custom_markers[row]; self.generate_map()
        else: ModernDialog.show_warning(self, "Seçim Yok", "Listeden silinecek konumu seçiniz.")

    def load_baz_data(self):
        if not getattr(self, "gsm_number", ""):
            return

        start_dt = self.dt_start.dateTime(); end_dt = self.dt_end.dateTime()
        py_start = start_dt.toPyDateTime(); py_end = end_dt.toPyDateTime()
        baz_counter = Counter()
        try:
            with DB() as conn:
                cur = conn.cursor()
                target_gsms = [str(r[0]) for r in cur.execute("SELECT DISTINCT GSMNo FROM hts_dosyalari WHERE ProjeID=? AND (Rol IS NULL OR Rol!='KARSI')", (self.project_id,)).fetchall()]
                if not target_gsms: target_gsms = [self.gsm_number]
                in_placeholders = ",".join(["?"] * len(target_gsms))
                sql = f"""SELECT BAZ, TARIH FROM hts_gsm WHERE ProjeID=? AND GSMNo IN ({in_placeholders}) AND (NUMARA IS NULL OR REPLACE(NUMARA,' ','')=?) UNION ALL SELECT BAZ, TARIH FROM hts_gprs WHERE ProjeID=? AND GSMNo IN ({in_placeholders}) AND (NUMARA IS NULL OR REPLACE(NUMARA,' ','')=?) UNION ALL SELECT BAZ, TARIH FROM hts_wap WHERE ProjeID=? AND GSMNo IN ({in_placeholders}) AND (NUMARA IS NULL OR REPLACE(NUMARA,' ','')=?)"""
                params = ([self.project_id] + target_gsms + [self.gsm_number] + [self.project_id] + target_gsms + [self.gsm_number] + [self.project_id] + target_gsms + [self.gsm_number])
                rows = cur.execute(sql, params).fetchall()
                for r in rows:
                    baz = str(r[0]).strip(); t_str = str(r[1]).strip()
                    if not baz or not t_str: continue
                    try:
                        fmt = "dd.MM.yyyy HH:mm:ss" if "." in t_str else "dd/MM/yyyy HH:mm:ss"
                        if " " not in t_str: fmt = fmt.split(" ")[0]
                        py_fmt = fmt.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y").replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")
                        dt = datetime.strptime(t_str, py_fmt)
                        if py_start <= dt <= py_end: baz_counter[baz] += 1
                    except: pass
            self.table.setRowCount(0); sorted_data = baz_counter.most_common(); self.table.setRowCount(len(sorted_data))
            for i, (baz, count) in enumerate(sorted_data):
                chk = QTableWidgetItem(); chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled); chk.setCheckState(Qt.CheckState.Checked)
                self.table.setItem(i, 0, chk); self.table.setItem(i, 1, QTableWidgetItem(baz));
                item_cnt = QTableWidgetItem(str(count)); item_cnt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(i, 2, item_cnt)
            if not sorted_data: self.browser.setHtml("<h3 style='text-align:center; margin-top:50px;'>Veri bulunamadı.</h3>")
            else: self.generate_map()
        except Exception as e: print(f"Data Error: {e}")

    def toggle_all(self, state):
        st = Qt.CheckState.Checked if state else Qt.CheckState.Unchecked
        for i in range(self.table.rowCount()): self.table.item(i, 0).setCheckState(st)

    def generate_map(self, focus_point=None):
        import folium

        def check_internet_cached(ttl_seconds: int = 10) -> bool:
            try:
                now_ts = datetime.now().timestamp()

                cache = getattr(self, "_net_cache", None)
                if cache and (now_ts - cache["ts"] <= ttl_seconds):
                    return cache["ok"]

                try:
                    socket.create_connection(("8.8.8.8", 53), timeout=1.0)
                    ok = True
                except OSError:
                    ok = False

                self._net_cache = {"ok": ok, "ts": now_ts}
                return ok
            except Exception:
                return False

        is_online = check_internet_cached()
        has_local = os.path.exists(os.path.join(APP_DIR, "turkey.mbtiles"))
        tile_url = "OpenStreetMap" if is_online else ("http://localhost:8080/{z}/{x}/{y}.png" if has_local else "OpenStreetMap")
        attr_info = "OpenStreetMap (Online)" if is_online else ("Çevrimdışı Harita" if has_local else "Kaynak Yok")

        center_lat, center_lon, zoom = 39.0, 35.0, 6
        if focus_point:
            center_lat, center_lon, zoom = focus_point[0], focus_point[1], 15
        elif self.searched_location:
            center_lat, center_lon, zoom = self.searched_location[0], self.searched_location[1], 14

        m = folium.Map(location=[center_lat, center_lon], zoom_start=zoom, tiles=tile_url, attr=attr_info)
        _enable_measure_and_balloons(m)

        try:
            with DB() as conn:
                c_rows = conn.execute(
                    "SELECT Lat, Lon, Label FROM ozel_konumlar WHERE ProjeID=?",
                    (self.project_id,)
                ).fetchall()

                if c_rows:
                    fg_cust = folium.FeatureGroup(name="⭐ Özel Konumlar")
                    for lat, lon, label in c_rows:
                        folium.Marker(
                            [lat, lon],
                            popup=f"<b>ÖZEL KONUM</b><br>{label}",
                            icon=folium.Icon(color="orange", icon="star", prefix='fa'),
                            tooltip=folium.Tooltip(
                                label,
                                permanent=True,
                                style="background-color: #fff3e0; border: 1px solid #e67e22; color: #d35400; font-weight: bold;"
                            )
                        ).add_to(fg_cust)
                    fg_cust.add_to(m)
        except Exception:
            pass

        selected_bazs = []
        max_signal = 1
        for i in range(self.table.rowCount()):
            if self.table.item(i, 0).checkState() == Qt.CheckState.Checked:
                baz = self.table.item(i, 1).text()
                count = int(self.table.item(i, 2).text())
                selected_bazs.append((baz, count))
                max_signal = max(max_signal, count)

        for baz_adi, sinyal in selected_bazs:
            try:
                coords = re.findall(r"(\d{2}\.\d{4,})", baz_adi)
                if len(coords) >= 2:
                    val1, val2 = float(coords[-2]), float(coords[-1])
                    lat, lon = (
                        (val1, val2) if 35 < val1 < 43 and 25 < val2 < 46
                        else ((val2, val1) if 35 < val2 < 43 and 25 < val1 < 46 else (val1, val2))
                    )
                    ratio = sinyal / max_signal
                    color = "red" if ratio > 0.5 else ("orange" if ratio > 0.2 else "blue")
                    folium.Marker(
                        [lat, lon],
                        popup=f"<b>Sinyal:</b> {sinyal}<br>{baz_adi}",
                        icon=folium.Icon(color=color, icon="info-sign")
                    ).add_to(m)
            except Exception:
                pass

        folium.LayerControl().add_to(m)
        data = io.BytesIO()
        m.save(data, close_file=False)
        self.browser.setHtml(data.getvalue().decode())

    def extract_coords(self, text):
        """Kütüphane destekli koordinat bulucu."""
        if not text: return None
        text_str = str(text).strip()

        coords = re.findall(r"(\d{2}\.\d{4,})", text_str)
        if len(coords) >= 2:
            try:
                v1, v2 = float(coords[-2]), float(coords[-1])
                if 35 < v1 < 43 and 25 < v2 < 46: return [v1, v2]
                elif 35 < v2 < 43 and 25 < v1 < 46: return [v2, v1]
                else: return [v1, v2]
            except: pass

        try:
            with DB() as conn:
                cell_id = None
                match = re.search(r'\((\d{4,})\)', text_str)
                if match: cell_id = match.group(1)
                else:
                    nums = re.findall(r'\d+', text_str)
                    cands = [n for n in nums if len(n) > 3]
                    if cands: cell_id = cands[0]

                if cell_id:
                    row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE CellID=?", (cell_id,)).fetchone()
                    if row: return [row[0], row[1]]

                row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE BazAdi=?", (text_str,)).fetchone()
                if row: return [row[0], row[1]]
        except: pass
        return None


class DailyRouteDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, gsm_number, default_datetime):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.project_id = project_id
        self.gsm_number = gsm_number

        self.setWindowTitle(f"Güzergah Analizi (Hareket Dökümü) - {gsm_number}")
        self.resize(1400, 900)
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)

        self.top_frame = QFrame()
        self.top_frame.setStyleSheet("background-color: #f7f9f9; border-bottom: 1px solid #bdc3c7;")
        self.top_frame.setFixedHeight(60)

        hl = QHBoxLayout(self.top_frame)
        hl.setContentsMargins(15, 5, 15, 5)

        info_btn = InfoButton(
            "<b>📅 Güzergah ve Hareket Analizi:</b><br>"
            "Seçilen zaman aralığında şüphelinin baz istasyonu verilerine göre hareket dökümünü harita üzerinde çizer.<br><br>"
            "• <b>Oklar (↝):</b> Hareket yönünü gösterir (Başlangıç ➔ Bitiş).<br>"
            "• <b>Baloncuklar:</b> O konumda bulunulan saati gösterir.<br>"
            "• <b>Amaç:</b> Şüphelinin gün içindeki rotasını, uğradığı noktaları ve bekleme sürelerini kronolojik olarak tespit etmektir."
        )
        hl.addWidget(info_btn)
        hl.addSpacing(15)

        hl.addWidget(QLabel("📅 Başlangıç:", styleSheet="font-weight:bold; color:#2c3e50"))
        self.dt_start = QDateTimeEdit()
        self.dt_start.setCalendarPopup(True)
        self.dt_start.setDisplayFormat("dd.MM.yyyy HH:mm")
        start_def = default_datetime; start_def.setTime(datetime.min.time())
        self.dt_start.setDateTime(start_def)
        hl.addWidget(self.dt_start)

        hl.addSpacing(10)

        hl.addWidget(QLabel("🏁 Bitiş:", styleSheet="font-weight:bold; color:#2c3e50"))
        self.dt_end = QDateTimeEdit()
        self.dt_end.setCalendarPopup(True)
        self.dt_end.setDisplayFormat("dd.MM.yyyy HH:mm")
        end_def = default_datetime; end_def.setTime(datetime.max.time())
        self.dt_end.setDateTime(end_def)
        hl.addWidget(self.dt_end)

        hl.addSpacing(20)

        self.chk_show_lines = QCheckBox("Okları Çiz ↝")
        self.chk_show_lines.setChecked(True)
        self.chk_show_lines.setStyleSheet("font-weight:bold; color:#e67e22;")
        hl.addWidget(self.chk_show_lines)

        hl.addSpacing(10)

        self.chk_show_labels = QCheckBox("Zaman Baloncuklarını Göster 💬")
        self.chk_show_labels.setChecked(True)
        self.chk_show_labels.setStyleSheet("font-weight:bold; color:#27ae60;")
        hl.addWidget(self.chk_show_labels)

        hl.addSpacing(15)

        btn_draw = QPushButton("Analizi Güncelle")
        btn_draw.setStyleSheet("background-color: #2980b9; color: white; font-weight: bold; padding: 6px 15px; border-radius:4px;")
        btn_draw.clicked.connect(self.draw_route)
        hl.addWidget(btn_draw)

        hl.addStretch()
        layout.addWidget(self.top_frame)

        self.info_frame = QFrame()
        self.info_frame.setStyleSheet("background-color: #e3f2fd; border-bottom: 1px solid #90caf9;")
        self.info_frame.setFixedHeight(75)

        h_info = QHBoxLayout(self.info_frame)
        h_info.setContentsMargins(15, 5, 15, 5)

        lbl_icon = QLabel("👤")
        lbl_icon.setStyleSheet("font-size: 28px; border:none; background:transparent;")
        lbl_icon.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        h_info.addWidget(lbl_icon)

        self.lbl_info = QLabel()
        self.lbl_info.setStyleSheet("border:none; background:transparent;")
        self.lbl_info.setWordWrap(True)
        h_info.addWidget(self.lbl_info)

        self.update_info_label()

        h_info.addStretch()
        layout.addWidget(self.info_frame)

        self.browser = EvidenceWebEngineView()
        self.browser.setStyleSheet("background-color: white;")
        layout.addWidget(self.browser, 1)

        QTimer.singleShot(500, self.draw_route)

    def get_subscriber_name(self):
        try:
            with DB() as conn:
                row = conn.execute("SELECT AD, SOYAD FROM hts_abone WHERE ProjeID=? AND GSMNo=? LIMIT 1", (self.project_id, self.gsm_number)).fetchone()
                if row: return f"{row[0] or ''} {row[1] or ''}".strip()
        except: pass
        return "Bilinmeyen Kişi"

    def update_info_label(self):
        abone_name = self.get_subscriber_name()
        s_date = self.dt_start.dateTime().toString("dd.MM.yyyy HH:mm")
        e_date = self.dt_end.dateTime().toString("dd.MM.yyyy HH:mm")

        html_text = (
            f"<div style='font-family: Segoe UI; line-height:1.4; white-space: nowrap;'>"
            f"<span style='font-size:16px; font-weight:bold; color:#1565c0; white-space:nowrap;'>{self.gsm_number}</span> "
            f"<span style='font-size:15px; color:#333; font-weight:600; white-space:nowrap;'> - {abone_name}</span><br>"
            f"<span style='font-size:13px; color:#e65100; font-weight:bold;'>📅 Analiz Aralığı: {s_date} — {e_date}</span>"
            f"</div>"
        )
        self.lbl_info.setText(html_text)

    def calculate_distance(self, lat1, lon1, lat2, lon2):
        try:
            R = 6371000
            phi1 = math.radians(lat1); phi2 = math.radians(lat2)
            dphi = math.radians(lat2 - lat1); dlambda = math.radians(lon2 - lon1)
            a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2) * math.sin(dlambda/2)**2
            c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
            return R * c
        except: return 0

    def resolve_label_collisions(self, anchors, min_sep_m=220, iters=140, pull=0.02, damping=0.88):
        import math, random

        if not anchors:
            return []

        R = 6378137.0

        def ll_to_xy(lat, lon):
            x = math.radians(lon) * R
            y = math.log(math.tan(math.pi/4 + math.radians(lat)/2)) * R
            return x, y

        def xy_to_ll(x, y):
            lon = math.degrees(x / R)
            lat = math.degrees(2 * math.atan(math.exp(y / R)) - math.pi/2)
            return lat, lon

        n = len(anchors)
        fixed_xy = [ll_to_xy(lat, lon) for lat, lon in anchors]

        pos = []
        for (x, y) in fixed_xy:
            jx = random.uniform(-10, 10)
            jy = random.uniform(-10, 10)
            pos.append([x + jx, y + jy])

        vel = [[0.0, 0.0] for _ in range(n)]
        min_sep2 = float(min_sep_m * min_sep_m)

        step_scale = 14.0

        for _ in range(iters):
            forces = [[0.0, 0.0] for _ in range(n)]

            for i in range(n):
                xi, yi = pos[i]
                for j in range(i + 1, n):
                    xj, yj = pos[j]
                    dx = xi - xj
                    dy = yi - yj
                    d2 = dx*dx + dy*dy

                    if d2 < 1e-4:
                        d2 = 1e-4

                    if d2 < min_sep2:
                        d = math.sqrt(d2)
                        ux, uy = dx / d, dy / d

                        overlap = (min_sep_m - d) / min_sep_m
                        mag = overlap * 4.0

                        fx, fy = ux * mag, uy * mag
                        forces[i][0] += fx
                        forces[i][1] += fy
                        forces[j][0] -= fx
                        forces[j][1] -= fy

            for i in range(n):
                bx, by = fixed_xy[i]
                px, py = pos[i]
                forces[i][0] += (bx - px) * pull
                forces[i][1] += (by - py) * pull

            for i in range(n):
                vel[i][0] = (vel[i][0] + forces[i][0]) * damping
                vel[i][1] = (vel[i][1] + forces[i][1]) * damping
                pos[i][0] += vel[i][0] * step_scale
                pos[i][1] += vel[i][1] * step_scale

        return [xy_to_ll(x, y) for x, y in pos]

    def draw_route(self):
        import folium
        from folium.features import DivIcon
        self.update_info_label()
        self.browser.setHtml(
            "<div style='display:flex; justify-content:center; align-items:center; height:100vh; "
            "flex-direction:column; font-family:Segoe UI, sans-serif;'>"
            "<h2 style='color:#3498db;'>Güzergah Hesaplanıyor...</h2>"
            "<p style='color:#7f8c8d;'>Veriler işleniyor...</p></div>"
        )
        QApplication.processEvents()

        spider_data = []

        start_dt_py = self.dt_start.dateTime().toPyDateTime()
        end_dt_py = self.dt_end.dateTime().toPyDateTime()
        s_str = start_dt_py.strftime("%Y-%m-%d %H:%M:%S")
        e_str = end_dt_py.strftime("%Y-%m-%d %H:%M:%S")

        daily_points = defaultdict(list)
        all_coords = []
        total_raw_points = 0

        try:
            with DB() as conn:
                cur = conn.cursor()
                date_filter = (
                    "AND (substr(TARIH, 7, 4) || '-' || substr(TARIH, 4, 2) || '-' || "
                    "substr(TARIH, 1, 2) || substr(TARIH, 11)) BETWEEN ? AND ?"
                )

                sql = f"""
                    SELECT TARIH, BAZ FROM hts_gsm
                    WHERE ProjeID=? AND NUMARA=? AND BAZ IS NOT NULL AND BAZ != '' {date_filter}
                    UNION ALL
                    SELECT TARIH, BAZ FROM hts_gprs
                    WHERE ProjeID=? AND NUMARA=? AND BAZ IS NOT NULL AND BAZ != '' {date_filter}
                    UNION ALL
                    SELECT TARIH, BAZ FROM hts_wap
                    WHERE ProjeID=? AND NUMARA=? AND BAZ IS NOT NULL AND BAZ != '' {date_filter}
                """

                rows = cur.execute(sql, (self.project_id, self.gsm_number, s_str, e_str) * 3).fetchall()
                if len(rows) > 15000:
                    ModernDialog.show_warning(self, "Yoğun Veri", f"Seçili aralıkta {len(rows)} nokta var.")

                for r in rows:
                    t_str = str(r[0]).strip()
                    baz = str(r[1]).strip()
                    if not t_str or not baz:
                        continue
                    try:
                        fmt = "dd.MM.yyyy HH:mm:ss" if "." in t_str else "dd/MM/yyyy HH:mm:ss"
                        if " " not in t_str:
                            fmt = fmt.split(" ")[0]

                        py_fmt = (
                            fmt.replace("dd", "%d")
                               .replace("MM", "%m")
                               .replace("yyyy", "%Y")
                               .replace("HH", "%H")
                               .replace("mm", "%M")
                               .replace("ss", "%S")
                        )
                        dt = datetime.strptime(t_str, py_fmt)

                        coords = re.findall(r"(\d{2}\.\d{4,})", baz)
                        if len(coords) < 2:
                            continue

                        val1, val2 = float(coords[-2]), float(coords[-1])

                        if 35 < val1 < 43 and 25 < val2 < 46:
                            lat, lon = val1, val2
                        elif 35 < val2 < 43 and 25 < val1 < 46:
                            lat, lon = val2, val1
                        else:
                            lat, lon = val1, val2

                        if not (30.0 <= lat <= 46.0 and 20.0 <= lon <= 50.0):
                            continue

                        day_key = dt.strftime("%d.%m.%Y")
                        time_str = dt.strftime("%H:%M")
                        daily_points[day_key].append({
                            'dt': dt, 'lat': lat, 'lon': lon, 'time': time_str, 'addr': baz
                        })
                        total_raw_points += 1
                    except:
                        pass

            if total_raw_points == 0:
                self.browser.setHtml(
                    "<h3 style='text-align:center; margin-top:50px; color:#c0392b'>"
                    "Seçilen tarih/saat aralığında koordinat bulunamadı.</h3>"
                )
                return

            first_day = sorted(daily_points.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y"))[0]
            start_node = daily_points[first_day][0]

            m = folium.Map(
                location=[start_node['lat'], start_node['lon']],
                zoom_start=12,
                tiles='OpenStreetMap',
                attr='OpenStreetMap',
                max_zoom=18
            )
            _enable_measure_and_balloons(m)

            try:
                with DB() as conn:
                    custom_rows = conn.execute(
                        "SELECT Lat, Lon, Label FROM ozel_konumlar WHERE ProjeID=?",
                        (self.project_id,)
                    ).fetchall()

                if custom_rows:
                    fg_custom = folium.FeatureGroup(name="⭐ Kayıtlı Özel Konumlar")
                    for r in custom_rows:
                        clat, clon, clabel = float(r[0]), float(r[1]), str(r[2])
                        folium.Marker(
                            [clat, clon],
                            popup=f"<b>KAYITLI KONUM</b><br>{clabel}<br>Koordinat ({clat}, {clon})",
                            icon=folium.Icon(color="orange", icon="star", prefix='fa'),
                            tooltip=folium.Tooltip(clabel, permanent=True, style="background-color: #fff3e0; border: 1px solid #e67e22; color: #d35400; font-weight: bold;")
                        ).add_to(fg_custom)
                        all_coords.append([clat, clon])
                    fg_custom.add_to(m)
            except Exception as e:
                print(f"Route map custom location error: {e}")

            custom_css = """
            <style>
                .leaflet-tile-pane { z-index: 200 !important; }
                .leaflet-overlay-pane { z-index: 400 !important; }
                .leaflet-marker-pane { z-index: 600 !important; }
                .leaflet-tooltip-pane { z-index: 650 !important; }
                .leaflet-popup-pane { z-index: 700 !important; }

                .custom-map-label {
                    background: rgba(255,255,255,0.98);
                    border: 2px solid #2c3e50;
                    border-radius: 6px;
                    padding: 3px 9px;
                    font-size: 11px;
                    font-weight: 700;
                    font-family: 'Segoe UI', sans-serif;
                    box-shadow: 0 2px 6px rgba(0,0,0,0.35);
                    text-align:center;
                    white-space:nowrap;
                    cursor: grab;
                }
                .custom-map-label:active { cursor: grabbing; }
            </style>
            """
            m.get_root().html.add_child(folium.Element(custom_css))

            day_colors = ['#3498db', '#e74c3c', '#27ae60', '#9b59b6', '#f39c12', '#16a085', '#2c3e50']
            hour_colors = [
                "#e74c3c","#e67e22","#f1c40f","#2ecc71","#1abc9c","#3498db",
                "#9b59b6","#34495e","#16a085","#27ae60","#2980b9","#8e44ad",
                "#c0392b","#d35400","#f39c12","#2ecc71","#1abc9c","#3498db",
                "#9b59b6","#2c3e50","#7f8c8d","#95a5a6","#bdc3c7","#34495e"
            ]

            sorted_days = sorted(daily_points.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y"))
            single_day = (len(sorted_days) == 1)

            draw_labels = self.chk_show_labels.isChecked()
            draw_lines = self.chk_show_lines.isChecked()

            for d_idx, day in enumerate(sorted_days):
                points = daily_points[day]
                points.sort(key=lambda x: x['dt'])
                # ✅ Tekilleştirme: aynı BAZ + aynı saat:dakika -> 1 kayıt kalsın
                # (Her gün için ayrı çalışır)
                seen = set()
                unique_points = []
                for p in points:
                    # baz + dakika + koordinat bazlı tekilleştir
                    # (addr=baz zaten burada baz string'i)
                    k = (p.get('addr', ''), p.get('time', ''), round(float(p.get('lat', 0.0)), 6), round(float(p.get('lon', 0.0)), 6))
                    if k in seen:
                        continue
                    seen.add(k)
                    unique_points.append(p)

                points = unique_points
                filtered_points = []
                if points:
                    filtered_points.append(points[0])
                    for i in range(1, len(points)):
                        prev = filtered_points[-1]
                        curr = points[i]
                        dist_m = self.calculate_distance(prev['lat'], prev['lon'], curr['lat'], curr['lon'])
                        if dist_m > 50 or i == len(points) - 1:
                            filtered_points.append(curr)

                if not filtered_points:
                    continue

                anchors = [(float(pt['lat']), float(pt['lon'])) for pt in filtered_points]
                label_positions = self.resolve_label_collisions(anchors)

                line_coords = []

                for i, pt in enumerate(filtered_points):
                    lat = float(pt['lat'])
                    lon = float(pt['lon'])
                    line_coords.append([lat, lon])
                    all_coords.append([lat, lon])

                    if single_day:
                        h = pt['dt'].hour
                        item_color = hour_colors[h]
                    else:
                        item_color = day_colors[d_idx % len(day_colors)]

                    icon_name = "broadcast-tower"
                    marker_color = "blue"

                    if i == 0:
                        icon_name = "play"
                        marker_color = "green"
                    elif i == len(filtered_points) - 1:
                        icon_name = "stop"
                        marker_color = "red"

                    folium.Marker(
                        location=[lat, lon],
                        icon=folium.Icon(color=marker_color, icon=icon_name, prefix='fa'),
                        popup=f"<b>{pt['time']}</b><br>{pt['addr']}",
                        tooltip=f"{day} - {pt['time']}"
                    ).add_to(m)

                    if draw_labels:
                        lbl_lat, lbl_lon = label_positions[i]

                        label_html = (
                            f"<div class='custom-map-label' "
                            f"style='border-color:{item_color}; color:{item_color};'>"
                            f"{pt['time']}</div>"
                        )

                        label_marker = folium.Marker(
                            location=[lbl_lat, lbl_lon],
                            icon=DivIcon(icon_size=(80, 30), icon_anchor=(40, 15), html=label_html),
                            draggable=True
                        )
                        label_marker.add_to(m)

                        connector = folium.PolyLine(
                            locations=[[lat, lon], [lbl_lat, lbl_lon]],
                            color=item_color,
                            weight=2,
                            opacity=0.85,
                            dash_array="3,3"
                        )
                        connector.add_to(m)

                        m.add_child(DraggableConnector(m, label_marker, connector, lat, lon))

                        spider_data.append({
                            "label": label_marker.get_name(),
                            "line": connector.get_name(),
                            "fixedLat": lat,
                            "fixedLon": lon
                        })

                if draw_lines and len(line_coords) > 1:
                    folium.PolyLine(
                        locations=line_coords,
                        color=item_color if single_day else day_colors[d_idx % len(day_colors)],
                        weight=4,
                        opacity=0.8
                    ).add_to(m)

            if all_coords:
                m.fit_bounds(all_coords)

            if spider_data:
                spider_json = json.dumps(spider_data)

                js = f"""
                <script>
                (function(){{
                    var map = {m.get_name()};
                    var items = {spider_json};

                    // folium objelerini window'dan çek
                    items.forEach(function(it){{
                        it._label = window[it.label];
                        it._line  = window[it.line];
                        it._fixed = L.latLng(it.fixedLat, it.fixedLon);
                        it._pinned = false;
                    }});

                    function dist(a,b) {{
                        var dx=a.x-b.x, dy=a.y-b.y;
                        return Math.sqrt(dx*dx+dy*dy);
                    }}

                    var nearPx = 40;      // cluster eşiği
                    var baseRadius = 70;  // px açılma yarıçapı
                    var refZoom = null;

                    function buildClusters(){{
                        var clusters = [];
                        var used = new Set();

                        items.forEach(function(it, idx){{
                            if (used.has(idx) || it._pinned) return;
                            var cPt = map.latLngToLayerPoint(it._fixed);
                            var cluster = [it];
                            used.add(idx);

                            items.forEach(function(jt, j){{
                                if (used.has(j) || jt._pinned) return;
                                var p = map.latLngToLayerPoint(jt._fixed);
                                if (dist(p, cPt) <= nearPx){{
                                    cluster.push(jt);
                                    used.add(j);
                                }}
                            }});

                            clusters.push(cluster);
                        }});

                        return clusters;
                    }}

                    function zoomFactor(){{
                        var z = map.getZoom();
                        var dz = (refZoom - z);
                        // zoom-out oldukça hızlı büyüsün
                        return Math.max(1.0, Math.pow(1.9, dz));
                    }}

                    function spiderfyCluster(cluster){{
                        if (cluster.length <= 1) {{
                            // tekli ise bazın biraz yanına koy
                            var it = cluster[0];
                            if (it._pinned) return;
                            it._label.setLatLng(it._origLatLng);
                            it._line.setLatLngs([it._fixed, it._origLatLng]);
                            return;
                        }}

                        var N = cluster.length;
                        var k = zoomFactor();
                        var radius = baseRadius * k;

                        // cluster merkezi (bazların ortalaması)
                        var sumX=0, sumY=0;
                        cluster.forEach(function(it){{
                            var p = map.latLngToLayerPoint(it._fixed);
                            sumX += p.x; sumY += p.y;
                        }});
                        var centerPt = L.point(sumX/N, sumY/N);

                        cluster.forEach(function(it, i){{
                            if (it._pinned) return;
                            var ang = (2*Math.PI*i)/N;
                            var dx = radius*Math.cos(ang);
                            var dy = radius*Math.sin(ang);

                            var targetPt = L.point(centerPt.x + dx, centerPt.y + dy);
                            var targetLatLng = map.layerPointToLatLng(targetPt);

                            it._label.setLatLng(targetLatLng);
                            it._line.setLatLngs([it._fixed, targetLatLng]);
                        }});
                    }}

                    function applySpiderfy(){{
                        var clusters = buildClusters();
                        clusters.forEach(spiderfyCluster);
                    }}

                    map.whenReady(function(){{
                        refZoom = map.getZoom();

                        // orijinal konum kayıt
                        items.forEach(function(it){{
                            it._origLatLng = it._label.getLatLng();
                        }});

                        // kullanıcı sürüklerse sabitle
                        items.forEach(function(it){{
                            it._label.on('dragstart', function(){{ it._pinned = true; }});
                            it._label.on('dragend', function(){{
                                it._pinned = true;
                                // yeni orig kabul et
                                it._origLatLng = it._label.getLatLng();
                                it._line.setLatLngs([it._fixed, it._origLatLng]);
                            }});
                        }});

                        map.on('zoomend moveend', applySpiderfy);
                        setTimeout(applySpiderfy, 250);
                    }});
                }})();
                </script>
                """
                m.get_root().html.add_child(folium.Element(js))

            data = io.BytesIO()
            m.save(data, close_file=False)
            self.browser.setHtml(data.getvalue().decode())

        except Exception as e:
            self.browser.setHtml(
                f"<h3 style='text-align:center; color:red'>Hata Oluştu: {str(e)}</h3>"
            )


class HeatmapDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, heatmap_data):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.setWindowTitle("Aktivite Isı Haritası (Yoğunluk Analizi)")
        self.resize(1000, 400)
        self.layout = QVBoxLayout(self)

        lbl_info = QLabel("🔥 Bu grafik, şüphelinin haftanın hangi günlerinde ve günün hangi saatlerinde yoğun iletişim kurduğunu gösterir.")
        lbl_info.setStyleSheet("color: #7f8c8d; font-style: italic; margin-bottom: 10px;")
        self.layout.addWidget(lbl_info)

        self.heatmap_widget = ActivityHeatmapWidget()
        self.heatmap_widget.update_heatmap(heatmap_data)
        self.layout.addWidget(self.heatmap_widget)

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; padding:10px; border-radius:4px;")
        btn_close.clicked.connect(self.accept)
        self.layout.addWidget(btn_close)


class ActivityDetailDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, data_rows, day_name, hour_str):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.setWindowTitle(f"Aktivite Detayı: {day_name} Saat {hour_str}:00 - {hour_str}:59")
        self.resize(900, 500)

        layout = QVBoxLayout(self)

        info = QLabel(f"📅 <b>{day_name}</b> günü, saat <b>{hour_str}:00</b> ile <b>{hour_str}:59</b> arasındaki aktiviteler.")
        info.setStyleSheet("font-size: 14px; color: #2c3e50; background-color: #ecf0f1; padding: 10px; border-radius: 5px;")
        layout.addWidget(info)

        headers = ["Tarih/Saat", "İşlem Türü", "Yön", "Karşı Numara / IP", "Süre / Boyut", "Baz İstasyonu"]

        # ✅ Delil ekleme (sağ tık) tamamen GenericDatabaseTable altyapısından gelsin.
        self.table = GenericDatabaseTable(headers=headers, enable_date_filter=False, chart_mode='embedded', info_text=None, enable_evidence_menu=True)

        # Başlık / GSM bağlama (GenericDatabaseTable.add_selection_to_report buradan okuyacak)
        self.table.project_id = getattr(parent, "current_project_id", None) or getattr(parent, "project_id", None)
        self.table.gsm_number = (
            getattr(parent, "current_gsm_number", None)
            or getattr(parent, "current_gsm", None)
            or getattr(parent, "gsm_number", None)
            or getattr(parent, "gsm_no", None)
            or getattr(parent, "selected_gsm", None)
            or ""
        )
        self.table.table_title = self.windowTitle()

        # Görünüm ayarları (eskisiyle uyumlu)
        try:
            h = self.table.table.horizontalHeader()
            h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Tarih
            h.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)          # Baz
        except Exception:
            pass

        self.table.set_data(data_rows or [])
        layout.addWidget(self.table)

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; padding:10px; border-radius:4px;")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)


class GraphPopupDialog(QDialog):
    """Grafiği geniş bir pencerede açar."""
    def __init__(self, parent=None, title="Analiz Diyagramı"):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(1200, 800)
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)

        self.graph_widget = InteractiveGraphWidget()
        layout.addWidget(self.graph_widget)


class GenericDatabaseTable(QWidget):
    def __init__(self, headers, enable_date_filter=False, chart_mode='embedded', info_text=None, enable_evidence_menu=True):
        super().__init__()
        self.raw_data = []
        self.headers = headers
        self.chart_mode = chart_mode
        self.date_col_index = -1
        self.owner_label = "Merkez"
        self.enable_evidence_menu = bool(enable_evidence_menu)
        self.duration_col_index = -1

        for i, h in enumerate(headers):
            clean_h = str(h).upper().replace("İ", "I").replace("Ğ", "G")
            if "TARIH" in clean_h or "ZAMAN" in clean_h:
                self.date_col_index = i
            if "SÜRE" in clean_h or "DURATION" in clean_h:
                self.duration_col_index = i

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0,0,0,0)

        filter_frame = QFrame()
        filter_frame.setStyleSheet("background-color: #ffffff; border-top-left-radius: 6px; border-top-right-radius: 6px; border-bottom: 1px solid #f3f4f6;")
        h_layout = QHBoxLayout(filter_frame)
        h_layout.setContentsMargins(5, 5, 5, 5)

        if info_text:
            btn_info = InfoButton(info_text)
            h_layout.addWidget(btn_info)

        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("🔍 Tabloda Ara...")
        self.search_bar.textChanged.connect(self.filter_text)
        h_layout.addWidget(self.search_bar)

        if self.chart_mode != 'none':
            self.btn_view_chart = QPushButton("📊 Grafik")
            btn_style = """
                QPushButton { background-color: white; color: #2c3e50; border: 1px solid #bdc3c7; padding: 5px 15px; border-radius: 4px; font-weight: bold; }
                QPushButton:checked { background-color: #3498db; color: white; border: 1px solid #2980b9; }
                QPushButton:hover { background-color: #ecf0f1; }
            """
            self.btn_view_chart.setStyleSheet(btn_style)

            if self.chart_mode == 'embedded':
                self.btn_view_table = QPushButton("📋 Tablo")
                self.btn_view_table.setCheckable(True); self.btn_view_table.setChecked(True)
                self.btn_view_table.setStyleSheet(btn_style)
                self.btn_view_table.clicked.connect(lambda: self.switch_view(0))

                self.btn_view_chart.setCheckable(True)
                self.btn_view_chart.clicked.connect(lambda: self.switch_view(1))

                grp = QButtonGroup(self)
                grp.addButton(self.btn_view_table); grp.addButton(self.btn_view_chart)
                h_layout.addWidget(self.btn_view_table)

            elif self.chart_mode == 'popup':
                self.btn_view_chart.clicked.connect(self.open_chart_popup)

            h_layout.addWidget(self.btn_view_chart)

        self.layout.addWidget(filter_frame)

        self.stack = QStackedWidget()

        self.table_page = QWidget()
        t_layout = QVBoxLayout(self.table_page); t_layout.setContentsMargins(0,0,0,0)
        self.table = QTableView()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        self.table.setItemDelegate(ElidedItemDelegate(self.table))
        self.table.setWordWrap(False)

        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)

        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)

        self.source_model = CustomTableModel([], headers)
        self.model = self.source_model
        self.proxy_model = DateSortFilterProxyModel()
        self.proxy_model.setSourceModel(self.source_model)
        self.proxy_model.setFilterKeyColumn(-1)
        self.table.setModel(self.proxy_model)

        t_layout.addWidget(self.table)
        self.lbl_count = QLabel("Kayıt: 0")
        self.lbl_count.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.lbl_count.setStyleSheet("font-weight: bold; color: #2c3e50; margin-right: 5px;")
        t_layout.addWidget(self.lbl_count)

        self.lbl_duration_sum = QLabel("Toplam Süre: -")
        self.lbl_duration_sum.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.lbl_duration_sum.setStyleSheet("font-weight: bold; color: #16a085; margin-right: 5px;")
        t_layout.addWidget(self.lbl_duration_sum)
        if self.duration_col_index == -1:
            self.lbl_duration_sum.setVisible(False)

        self.stack.addWidget(self.table_page)

        if self.chart_mode == 'embedded':
            self.chart_page = InteractiveGraphWidget()
            self.stack.addWidget(self.chart_page)

        self.layout.addWidget(self.stack)

        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

    def hide_toolbar(self):
        """Üstteki arama çubuğu ve butonları gizler (Sadece tablo kalsın diye)."""
        if self.layout.count() > 0:
            item = self.layout.itemAt(0)
            if item and item.widget():
                item.widget().hide()

    def _reverse_format_duration(self, duration_str):
        """'hh sa mm dk ss sn' formatını saniyeye çevirir."""
        if not duration_str: return 0
        total_seconds = 0

        match = re.search(r'(\d+)\s*sa', duration_str)
        if match: total_seconds += int(match.group(1)) * 3600

        match = re.search(r'(\d+)\s*dk', duration_str)
        if match: total_seconds += int(match.group(1)) * 60

        match = re.search(r'(\d+)\s*sn', duration_str)
        if match: total_seconds += int(match.group(1))

        return total_seconds

    def show_context_menu(self, pos):
        """Sağ tık menüsünü gösterir."""
        if not getattr(self, "enable_evidence_menu", True):
            return

        menu = apply_menu_theme(QMenu())
        add_action = QAction("📑 Seçili Satırları Rapora Ekle", self)
        add_action.triggered.connect(self.add_selection_to_report)
        menu.addAction(add_action)

        menu.exec(self.table.mapToGlobal(pos))

    def add_selection_to_report(self):
        pid = getattr(self, "project_id", None)
        if not pid:
            p = self.parent()
            while p:
                if hasattr(p, "current_project_id") and getattr(p, "current_project_id"):
                    pid = getattr(p, "current_project_id")
                    break
                p = p.parent()
            self.project_id = pid

        if not getattr(self, "project_id", None):
            ModernDialog.show_error(self, "Hata", "Aktif proje (ProjeID) bulunamadı. Önce projeyi açtığınızdan emin olun.")
            return

        try:
            # Bu sınıfta tablo nesnesi "self.table"
            view = getattr(self, "view", None) or getattr(self, "table", None)
            if view is None:
                ModernDialog.show_error(self, "Hata", "Tablo görünümü bulunamadı (self.table/self.view yok).")
                return

            sel_model = view.selectionModel()
            if sel_model is None:
                ModernDialog.show_warning(self, "Seçim Yok", "Tablodan rapora eklenecek hücreleri seçmelisiniz!")
                return

            selected = sel_model.selectedIndexes()
            if not selected:
                ModernDialog.show_warning(self, "Seçim Yok", "Tablodan rapora eklenecek hücreleri seçmelisiniz!")
                return

            include_headers = True

            include_set = set()

            exclude_set = set()

            rows = sorted(set(i.row() for i in selected))
            cols = sorted(set(i.column() for i in selected))

            # Görünür kolonları al + filtre uygula
            visible_cols = [c for c in cols if not view.isColumnHidden(c)]
            if not visible_cols:
                ModernDialog.show_warning(self, "Kolon Yok", "Seçimde görünür kolon bulunamadı.")
                return

            # Model/proxy güvenliği (senin güncel halin)
            proxy = view.model()
            model = proxy.sourceModel() if hasattr(proxy, "sourceModel") else proxy

            # Başlık metinleri (proxy header)
            headers = []
            for c in visible_cols:
                h = proxy.headerData(c, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                headers.append("" if h is None else str(h))

            # Filtre (include doluysa sadece include; yoksa exclude uygula)
            kept_cols = []
            kept_headers = []
            for c, h in zip(visible_cols, headers):
                hn = _norm_header(h)
                if include_set:
                    if hn in include_set:
                        kept_cols.append(c); kept_headers.append(h)
                else:
                    if exclude_set and hn in exclude_set:
                        continue
                    kept_cols.append(c); kept_headers.append(h)

            if not kept_cols:
                ModernDialog.show_warning(self, "Filtre Sonucu Boş", "Kolon filtreleri nedeniyle eklenecek kolon kalmadı.")
                return

            # =========================
            # ✅ Qt icon/pixmap -> HTML <img data:...> gömme
            # =========================
            import base64
            from PyQt6.QtCore import QBuffer, QByteArray, QIODevice
            from PyQt6.QtGui import QPixmap, QImage, QIcon

            def _qimage_to_data_uri(img: QImage) -> str:
                if img is None or img.isNull():
                    return ""
                ba = QByteArray()
                buf = QBuffer(ba)
                buf.open(QIODevice.OpenModeFlag.WriteOnly)
                img.save(buf, "PNG")
                buf.close()
                b64 = base64.b64encode(bytes(ba)).decode("ascii")
                return f"data:image/png;base64,{b64}"

            def _index_to_html_cell(idx) -> str:
                """
                Hücre HTML'i:
                - DisplayRole text
                - varsa DecorationRole ikonunu data-uri PNG olarak ekle
                """
                disp = idx.data(Qt.ItemDataRole.DisplayRole)
                text = "" if disp is None else str(disp)

                deco = idx.data(Qt.ItemDataRole.DecorationRole)
                data_uri = ""

                try:
                    if isinstance(deco, QIcon):
                        pm = deco.pixmap(18, 18)
                        if isinstance(pm, QPixmap) and not pm.isNull():
                            data_uri = _qimage_to_data_uri(pm.toImage())
                    elif isinstance(deco, QPixmap):
                        if not deco.isNull():
                            data_uri = _qimage_to_data_uri(deco.toImage())
                    elif isinstance(deco, QImage):
                        if not deco.isNull():
                            data_uri = _qimage_to_data_uri(deco)
                except Exception:
                    data_uri = ""

                # ikon varsa başa koy (PDF engine kaçırmasın diye inline)
                if data_uri:
                    icon_html = (
                        f"<img src=\"{data_uri}\" "
                        f"style=\"height:14px;width:14px;vertical-align:middle;margin-right:6px;\"/>"
                    )
                    return icon_html + html.escape(text)

                return html.escape(text)

            # HTML tablo üret
            html_table = "<table class='meta-table'>"
            if include_headers:
                html_table += "<tr>" + "".join(f"<th>{html.escape(h)}</th>" for h in kept_headers) + "</tr>"

            for r in rows:
                html_table += "<tr>"
                for c in kept_cols:
                    idx = proxy.index(r, c)
                    html_table += f"<td>{_index_to_html_cell(idx)}</td>"
                html_table += "</tr>"
            html_table += "</table>"

            # Düz metin (ikonlar metne gömülmez; sadece görüntüye gömülür)
            selected_text_lines = []
            for r in rows:
                row_vals = []
                for c in kept_cols:
                    v = proxy.index(r, c).data(Qt.ItemDataRole.DisplayRole)
                    row_vals.append("" if v is None else str(v))
                selected_text_lines.append("\t".join(row_vals))
            selected_text = "\n".join(selected_text_lines)

            # =========================
            # ✅ Başlık üretimi: "GSM - Sekme"
            # =========================
            def _resolve_current_gsm():
                p = self
                while p:
                    gsm = getattr(p, "current_gsm_number", None)
                    if gsm:
                        return str(gsm)
                    p = p.parent()
                return ""

            def _clean_tab_text(t: str) -> str:
                t = (t or "").strip()
                import re
                t = re.sub(r"^[^\wÇĞİÖŞÜçğıöşü0-9]+", "", t, flags=re.UNICODE).strip()
                return t

            def _resolve_tab_name():
                explicit = getattr(self, "report_table_title", None) or getattr(self, "report_title", None)
                if explicit:
                    t = _clean_tab_text(str(explicit))
                    if t:
                        return t

                # en yakın QTabWidget aktif sekmesi
                p = self.parent()
                while p:
                    if isinstance(p, QTabWidget):
                        tw = p
                        try:
                            i = tw.currentIndex()
                            t = _clean_tab_text(tw.tabText(i))
                            if t:
                                return t
                        except Exception:
                            pass
                        break
                    p = p.parent()

                return "Tablo"

            gsm_text = _resolve_current_gsm()
            tab_name = _resolve_tab_name()
            final_title = f"{gsm_text} - {tab_name}" if gsm_text else tab_name

            # ==== ESAS DAVRANIŞ: Rapor merkezi açık olmasa bile DB'ye ekle ====
            from datetime import datetime
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            with DB() as conn:
                c = conn.cursor()

                last_order = c.execute(
                    "SELECT MAX(Sira) FROM rapor_taslagi WHERE ProjeID=?",
                    (self.project_id,)
                ).fetchone()
                new_order = (last_order[0] + 1) if last_order and last_order[0] is not None else 1

                c.execute(
                    """
                    INSERT INTO rapor_taslagi
                    (ProjeID, GSMNo, Baslik, Icerik, Tur, Tarih, Sira, GenislikYuzde, YukseklikMm, Hizalama, Aciklama, HtmlIcerik, ImagePath)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        self.project_id,
                        gsm_text,
                        final_title,
                        selected_text,
                        "TABLE",
                        now,
                        new_order,
                        100,
                        0,
                        "LEFT",
                        "",
                        html_table,
                        None
                    )
                )
                conn.commit()

            # Rapor Merkezi açıksa UI’ya da ekle (varsa)
            rm = getattr(self, "rapor_merkezi", None)
            # ✅ PyQt: kapatılmış (WA_DeleteOnClose) dialog referansı RuntimeError üretebilir.
            alive = False
            try:
                import sip  # type: ignore
                alive = (rm is not None) and (not sip.isdeleted(rm))
            except Exception:
                try:
                    alive = (rm is not None) and rm.isVisible()
                except Exception:
                    alive = False
            if alive and hasattr(rm, "add_block"):
                try:
                    rm.add_block(
                        title=final_title,
                        content=selected_text,
                        block_type="TABLE",
                        html_content=html_table
                    )
                except Exception:
                    pass

            ModernDialog.show_success(self, "Eklendi", "Seçilen tablo rapora delil olarak eklendi!")

        except Exception as e:
            ModernDialog.show_error(self, "Hata", f"Rapora ekleme hatası: {e}")

    def set_owner_info(self, text):
        """Dışarıdan (AnalysisCenter) gelen hat sahibi bilgisini kaydeder."""
        self.owner_label = text

    def prepare_chart_data(self, target_widget):
        """Veriyi hazırlar: İsim ve Adet bilgisini etikete gömer."""
        rows_count = self.proxy_model.rowCount()
        scan_limit = min(rows_count, 2000)

        target_col_idx = -1
        qty_col_idx = -1
        name_col_idx = -1
        center_name = self.owner_label

        for i, header in enumerate(self.headers):
            h_clean = str(header).upper().replace('İ','I').replace('Ğ','G').replace('Ü','U').replace('Ş','S').replace('Ö','O').replace('Ç','C')

            if "BAZ" in h_clean or "KONUM" in h_clean:
                if target_col_idx == -1: target_col_idx = i
            elif "KARSI" in h_clean or "DIGER NUMARA" in h_clean or "ARANAN" in h_clean or "GSM NUMARASI" in h_clean:
                if target_col_idx == -1: target_col_idx = i
            elif "IMEI" in h_clean:
                if target_col_idx == -1: target_col_idx = i

            if "ADET" in h_clean or "SINYAL" in h_clean or "SAYI" in h_clean or "KULLANIM" in h_clean:
                qty_col_idx = i

            if "ISIM" in h_clean or "AD SOYAD" in h_clean or "KISI ADI" in h_clean:
                name_col_idx = i

        if target_col_idx != -1:
            final_data = []

            if qty_col_idx != -1:
                for r in range(scan_limit):
                    idx_target = self.proxy_model.index(r, target_col_idx)
                    val_target = str(self.proxy_model.data(idx_target, Qt.ItemDataRole.EditRole))

                    idx_qty = self.proxy_model.index(r, qty_col_idx)
                    val_qty = self.proxy_model.data(idx_qty, Qt.ItemDataRole.EditRole)
                    try: count = int(val_qty)
                    except: count = 1

                    display_label = val_target

                    if name_col_idx != -1:
                        idx_name = self.proxy_model.index(r, name_col_idx)
                        val_name = str(self.proxy_model.data(idx_name, Qt.ItemDataRole.EditRole))
                        if val_name and val_name not in ["", "None", "Unknown"]:
                            display_label = f"{val_target}\n{val_name}"

                    display_label += f"\n({count} Adet)"

                    final_data.append((display_label, count))

            else:
                raw_items = []
                for r in range(scan_limit):
                    idx = self.proxy_model.index(r, target_col_idx)
                    val = self.proxy_model.data(idx, Qt.ItemDataRole.EditRole)

                    val_name = ""
                    if name_col_idx != -1:
                        idx_n = self.proxy_model.index(r, name_col_idx)
                        val_name = str(self.proxy_model.data(idx_n, Qt.ItemDataRole.EditRole))

                    if val and str(val).strip() not in ['', 'None', '---', 'nan']:
                        if val_name: raw_items.append(f"{val}\n{val_name}")
                        else: raw_items.append(str(val))

                counts = Counter(raw_items).most_common(500)

                # [DÜZELTME] Sayım bittikten sonra etiketlere adet ekle
                for item, cnt in counts:
                    label_with_count = f"{item}\n({cnt} Adet)"
                    final_data.append((label_with_count, cnt))

            if final_data:
                target_widget.load_list_data(center_name, final_data)
            else:
                target_widget.browser.setHtml("<h3 style='text-align:center; margin-top:50px; color:#7f8c8d'>Veri Yok</h3>")
        else:
            target_widget.browser.setHtml("<h3 style='text-align:center; margin-top:50px; color:#e74c3c'>Uygun Sütun Bulunamadı</h3>")

    def switch_view(self, index):
        """Gömülü modda sekme değiştirir."""
        self.stack.setCurrentIndex(index)
        if index == 1 and self.chart_mode == 'embedded':
            self.prepare_chart_data(self.chart_page)

    def open_chart_popup(self):
        """Popup modunda yeni pencere açar."""
        if self.proxy_model.rowCount() == 0:
            ModernDialog.show_warning(self, "Uyarı", "Görselleştirilecek veri yok.")
            return

        dialog = GraphPopupDialog(self, f"Analiz Diyagramı - {self.lbl_count.text()}")
        self.prepare_chart_data(dialog.graph_widget)
        dialog.exec()

    def set_data(self, data):
        self.raw_data = data
        self.source_model.update_data(data)
        self.lbl_count.setText(f"Kayıt: {len(data)}")

        total_duration_seconds = 0
        if self.duration_col_index != -1:
            for row in data:
                try:
                    duration_str = str(row[self.duration_col_index])
                    total_duration_seconds += self._reverse_format_duration(duration_str)
                except:
                    pass
            sec = total_duration_seconds
            m, s = divmod(sec, 60)
            h, m = divmod(m, 60)
            parts = []
            if h: parts.append(f"{int(h)} sa")
            if m: parts.append(f"{int(m)} dk")
            if s or not parts: parts.append(f"{int(s)} sn")
            self.lbl_duration_sum.setText(f"Toplam Süre: {' '.join(parts)}")
        else:
            self.lbl_duration_sum.setText("Toplam Süre: -")

        if len(data) > 0:
            h = self.table.horizontalHeader()

            baz_col_idx = -1
            signal_col_idx = -1

            for i, header in enumerate(self.headers):
                txt = str(header).upper().replace("İ", "I").replace("ı", "I")
                if "BAZ" in txt or "KONUM" in txt:
                    baz_col_idx = i
                if "SINYAL" in txt:
                    signal_col_idx = i

            if baz_col_idx != -1 and signal_col_idx != -1:
                self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

                h.setStretchLastSection(False)

                h.setSectionResizeMode(signal_col_idx, QHeaderView.ResizeMode.Interactive)
                self.table.setColumnWidth(signal_col_idx, 100)
                h.resizeSection(signal_col_idx, 100)
                h.setMaximumSectionSize(160)
                h.setMinimumSectionSize(70)

                h.setSectionResizeMode(baz_col_idx, QHeaderView.ResizeMode.Stretch)

                self.table.setColumnWidth(baz_col_idx, 800)

            else:
                self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
                h.setStretchLastSection(True)
                self.table.resizeColumnsToContents()
                h.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

                for i in range(h.count()):
                    if self.table.columnWidth(i) < 80:
                        self.table.setColumnWidth(i, 80)

        if self.chart_mode == 'embedded' and self.stack.currentIndex() == 1:
            self.switch_view(1)

    def filter_text(self, text):
        self.proxy_model.setSearchText(text)
        self.lbl_count.setText(f"Kayıt: {self.proxy_model.rowCount()}")

    def set_date_range(self, min_dt, max_dt):
        if hasattr(self, 'dt_start') and self.date_col_index != -1:
            self.dt_start.setDateTime(min_dt)
            self.dt_end.setDateTime(max_dt)

    def filter_text(self, text):
        self.proxy_model.setSearchText(text)
        self.lbl_count.setText(f"Kayıt: {self.proxy_model.rowCount()}")

    def apply_date_filter(self):
        if self.date_col_index == -1 or not self.raw_data: return
        min_val = self.dt_start.dateTime(); max_val = self.dt_end.dateTime()
        filtered_data = []
        for row in self.raw_data:
            date_str = str(row[self.date_col_index])
            try:
                fmt = "dd.MM.yyyy HH:mm:ss" if "." in date_str else "dd/MM/yyyy HH:mm:ss"
                if " " not in date_str: fmt = fmt.split(" ")[0]
                row_dt = QDateTime.fromString(date_str, fmt)
                if row_dt.isValid() and min_val <= row_dt <= max_val: filtered_data.append(row)
            except: pass
        self.source_model.update_data(filtered_data)
        self.lbl_count.setText(f"Kayıt: {len(filtered_data)}")


class ActivityHeatmapWidget(QTableWidget):
    cell_clicked_signal = pyqtSignal(int, int)
    def __init__(self, parent=None):
        super().__init__(7, 24, parent)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.verticalHeader().setVisible(True)
        self.horizontalHeader().setVisible(True)
        self.cellClicked.connect(self.on_cell_clicked)
        days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        self.setVerticalHeaderLabels(days)
        self.setHorizontalHeaderLabels([f"{h:02d}" for h in range(24)])

        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.clear_heatmap()

    def on_cell_clicked(self, row, col):
        item = self.item(row, col)
        if item and item.text():
            self.cell_clicked_signal.emit(row, col)

    def clear_heatmap(self):
        for r in range(7):
            for c in range(24):
                self.setItem(r, c, QTableWidgetItem(""))
                self.item(r, c).setBackground(QColor(255, 255, 255))

    def update_heatmap(self, data_matrix):
        max_val = 0
        for row in data_matrix:
            max_val = max(max_val, max(row))

        if max_val == 0: max_val = 1

        for r in range(7):
            for c in range(24):
                val = data_matrix[r][c]
                item = QTableWidgetItem(str(val) if val > 0 else "")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                if val > 0:
                    ratio = val / max_val
                    saturation = int(255 * (1 - ratio))
                    color = QColor(255, saturation, saturation)

                    text_color = Qt.GlobalColor.white if ratio > 0.6 else Qt.GlobalColor.black
                    item.setForeground(text_color)
                    item.setBackground(color)
                    item.setToolTip(f"{self.verticalHeaderItem(r).text()} Saat {c:02d}:00\nToplam Aktivite: {val}")
                else:
                    item.setBackground(QColor(245, 245, 245))

                self.setItem(r, c, item)


class DateLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setPlaceholderText("GG/AA/YYYY (Örn: 25112023)")
        self.setMaxLength(10)
        self.editingFinished.connect(self.validate_and_format)

    def validate_and_format(self):
        text = self.text().strip()
        if not text: return

        clean_text = re.sub(r'\D', '', text)

        if len(clean_text) != 8:
            ModernDialog.show_warning(self, "Format Hatası", "Lütfen tarihi 8 rakam olarak giriniz.\nÖrnek: 25112023")
            self.setFocus()
            self.selectAll()
            return

        try:
            day = int(clean_text[:2])
            month = int(clean_text[2:4])
            year = int(clean_text[4:])

            input_date = date(year, month, day)

            min_date = date(2000, 1, 1)
            today = date.today()

            if input_date < min_date:
                ModernDialog.show_warning(self, "Tarih Hatası", "Tarih 01/01/2000'den küçük olamaz.")
                self.setFocus(); self.selectAll()
                return

            if input_date > today:
                ModernDialog.show_warning(self, "Tarih Hatası", "Tarih bugünden (gelecek) büyük olamaz.")
                self.setFocus(); self.selectAll()
                return

            formatted_date = input_date.strftime("%d/%m/%Y")
            self.setText(formatted_date)

        except ValueError:
            ModernDialog.show_warning(self, "Geçersiz Tarih", "Girdiğiniz tarih takvimde mevcut değil (Örn: 30 Şubat).")
            self.setFocus(); self.selectAll()


class ProjectManager(QWidget):
    sig_upload_started = pyqtSignal(int)
    sig_progress_updated = pyqtSignal(int)
    sig_gsm_detected = pyqtSignal(str)
    sig_file_finished = pyqtSignal()
    sig_queue_finished = pyqtSignal()

    def __init__(self, main_window):
        super().__init__()
        self.main = main_window
        self._ensure_db_columns()

        self.upload_queue = []
        self.total_files_count = 0
        self.processed_files_count = 0
        self.is_uploading = False
        self.success_count = 0

        layout = QHBoxLayout(self)

        self.main_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.main_splitter.setStyleSheet("background-color: transparent;")

        left_widget = QWidget(); left_layout = QVBoxLayout(left_widget)
        left_widget.setStyleSheet("""
            QWidget {
                background-color: #ffffff; 
                border: 1px solid #e5e7eb; 
                border-radius: 8px;
            }
        """)
        left_layout.setContentsMargins(5, 5, 5, 5)

        lbl_p = QLabel("Projeler"); lbl_p.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        left_layout.addWidget(lbl_p)

        self.p_table = GenericDatabaseTable(["ID", "Birim", "Dosya No", "Tip", "Suç"], chart_mode='none', enable_evidence_menu=False)
        self.p_table.table.setColumnHidden(0, True)
        self.p_table.table.clicked.connect(self.on_select)
        left_layout.addWidget(self.p_table)

        btn_layout = QVBoxLayout(); btn_layout.setSpacing(10)
        btn_new = QPushButton("✨ Yeni Proje Oluştur"); btn_new.setStyleSheet("QPushButton { background-color: #27ae60; color: white; font-weight: bold; padding: 10px; border-radius: 5px; font-size: 13px; } QPushButton:hover { background-color: #2ecc71; }")
        btn_new.clicked.connect(self.clear_form)
        btn_layout.addWidget(btn_new)

        h_sub_btns = QHBoxLayout()
        btn_del = QPushButton("🗑️ Projeyi Sil"); btn_del.setStyleSheet("QPushButton { background-color: #c0392b; color: white; font-weight: bold; padding: 10px; border-radius: 5px; } QPushButton:hover { background-color: #e74c3c; }")
        btn_del.clicked.connect(self.delete_project)
        h_sub_btns.addWidget(btn_del)

        btn_vac = QPushButton("🛠️ Veritabanı Bakım"); btn_vac.setStyleSheet("QPushButton { background-color: #7f8c8d; color: white; font-weight: bold; padding: 10px; border-radius: 5px; } QPushButton:hover { background-color: #95a5a6; }")
        btn_vac.clicked.connect(self.vacuum_db)
        h_sub_btns.addWidget(btn_vac)

        btn_layout.addLayout(h_sub_btns)
        left_layout.addLayout(btn_layout)
        self.main_splitter.addWidget(left_widget)

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_widget.setStyleSheet("background-color: transparent;")
        right_layout.setContentsMargins(10, 10, 10, 10)

        # --- Üst başlık satırı (Proje Detayları + Lisans Bilgisi) ---
        header_row = QHBoxLayout()
        header_row.setContentsMargins(0, 0, 0, 8)

        lbl_details = QLabel("Proje Detayları")
        lbl_details.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
            }
        """)

        self.lbl_license_badge = QLabel()
        self.lbl_license_badge.setObjectName("lbl_license_badge")
        self.lbl_license_badge.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.lbl_license_badge.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)

        self.lbl_license_badge.setStyleSheet("""
            QLabel#lbl_license_badge {
                background-color: rgba(15, 118, 110, 0.08);
                border: 1px solid rgba(15, 118, 110, 0.25);
                color: #0f766e;
                padding: 6px 12px;
                border-radius: 12px;
                font-size: 12px;
                font-weight: 600;
            }
        """)

        header_row.addWidget(lbl_details, 0, Qt.AlignmentFlag.AlignLeft)
        header_row.addStretch(1)
        header_row.addWidget(self.lbl_license_badge, 0, Qt.AlignmentFlag.AlignRight)

        right_layout.addLayout(header_row)

        self._refresh_license_badge()
        # --- /Üst başlık satırı ---
        form_frame = QFrame()
        form_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff; 
                border: 1px solid #e5e7eb; 
                border-radius: 8px;
            }
            QLabel { border: none; background: transparent; } 
        """)

        grid = QGridLayout(form_frame)
        grid.setContentsMargins(15, 15, 15, 15)
        grid.setHorizontalSpacing(18)
        grid.setVerticalSpacing(8)

        self.i_birim = QLineEdit()
        self.i_tip = QComboBox(); self.i_tip.addItems(["Esas", "Soruşturma", "Talimat"])
        apply_light_combobox_popup(self.i_tip)
        self.i_dosya = QLineEdit()

        self.i_gorev_tarihi = DateLineEdit()
        self.i_suc_tarihi = DateLineEdit()
        self.i_suc = QLineEdit()

        self.i_bilirkisi_adi = QLineEdit()
        self.i_bilirkisi_unvan = QLineEdit()

        def get_label(text):
            l = QLabel(text)
            l.setStyleSheet("font-weight: bold; color: #34495e; font-size: 13px;")
            return l

        grid.addWidget(get_label("🏛️ Talep Eden Birim:"), 0, 0)
        grid.addWidget(self.i_birim,                    0, 1)

        grid.addWidget(get_label("⚖️ Suç:"),            0, 2)
        grid.addWidget(self.i_suc,                      0, 3)

        grid.addWidget(get_label("📑 Dosya No Tipi:"),   1, 0)
        grid.addWidget(self.i_tip,                      1, 1)

        grid.addWidget(get_label("📅 Suç Tarihi:"),     1, 2)
        grid.addWidget(self.i_suc_tarihi,              1, 3)

        grid.addWidget(get_label("🔢 Dosya No:"),        2, 0)
        grid.addWidget(self.i_dosya,                    2, 1)

        grid.addWidget(get_label("📅 Görevlendirme:"),  2, 2)
        grid.addWidget(self.i_gorev_tarihi,            2, 3)

        grid.addWidget(get_label("👤 Bilirkişi Adı:"),  3, 0)
        grid.addWidget(self.i_bilirkisi_adi,           3, 1)

        grid.addWidget(get_label("🎓 Ünvan/Sicil:"),    3, 2)
        grid.addWidget(self.i_bilirkisi_unvan,         3, 3)

        right_layout.addWidget(form_frame)

        btn_save = QPushButton("💾 Proje Kapak Bilgilerini Kaydet")
        btn_save.setStyleSheet("QPushButton { background-color: #2980b9; color: white; font-weight: bold; padding: 10px; border-radius: 5px; font-size: 14px; margin-top: 5px; } QPushButton:hover { background-color: #3498db; }") # Padding ve margin azaltıldı
        btn_save.clicked.connect(self.save_project)
        right_layout.addWidget(btn_save)

        line = QFrame(); line.setFrameShape(QFrame.Shape.HLine); line.setFrameShadow(QFrame.Shadow.Sunken); line.setStyleSheet("background-color: #ecf0f1; margin-top: 10px; margin-bottom: 10px;") # Marginler 20->10
        right_layout.addWidget(line)

        self.setup_bottom_panels(right_layout)

        self.main_splitter.addWidget(right_widget)

        self.main_splitter.setSizes([650, 650])
        self.main_splitter.setStretchFactor(1, 1)

        layout.addWidget(self.main_splitter)
        self.selected_project_id = None; self.load_projects()

    def _refresh_license_badge(self):
        """
        Proje Merkezi sağ üstte lisans bilgisini gösterir.
        30 gün kala sarı, 7 gün kala kırmızı uyarı verir.
        ID göstermez.
        """
        def apply_style(bg_rgba: str, border_rgba: str, color_hex: str):
            self.lbl_license_badge.setStyleSheet(f"""
                QLabel#lbl_license_badge {{
                    background-color: {bg_rgba};
                    border: 1px solid {border_rgba};
                    color: {color_hex};
                    padding: 6px 12px;
                    border-radius: 12px;
                    font-size: 12px;
                    font-weight: 600;
                }}
            """)

        try:
            d = LicenseManager.load_license_from_disk()
            if not d:
                self.lbl_license_badge.setText("🔐 Lisans: Bulunamadı")
                apply_style("rgba(107, 114, 128, 0.10)", "rgba(107, 114, 128, 0.25)", "#374151")
                return

            info = LicenseManager.validate_license(d)

            # Kalan gün hesabı
            try:
                exp_date = datetime.strptime(info.exp, "%Y-%m-%d").date()
                days_left = (exp_date - date.today()).days
            except Exception:
                days_left = None

            # Metin
            if days_left is None:
                self.lbl_license_badge.setText(f"🔐 {info.customer} • {info.exp} tarihine kadar")
                apply_style("rgba(15, 118, 110, 0.08)", "rgba(15, 118, 110, 0.25)", "#0f766e")
                return

            self.lbl_license_badge.setText(f"🔐 {info.customer} • {info.exp} tarihine kadar")

            # Renkli görsel uyarı
            if days_left <= 7:
                # Kırmızı
                apply_style("rgba(220, 38, 38, 0.10)", "rgba(220, 38, 38, 0.25)", "#b91c1c")
            elif days_left <= 30:
                # Sarı/amber
                apply_style("rgba(245, 158, 11, 0.14)", "rgba(245, 158, 11, 0.30)", "#b45309")
            else:
                # Normal (teal)
                apply_style("rgba(15, 118, 110, 0.08)", "rgba(15, 118, 110, 0.25)", "#0f766e")

        except Exception:
            self.lbl_license_badge.setText("🔐 Lisans: Hatalı")
            apply_style("rgba(220, 38, 38, 0.10)", "rgba(220, 38, 38, 0.25)", "#b91c1c")

    def setup_bottom_panels(self, layout):
        """Alt panelleri (Taraf, GSM, Dosya) oluşturur."""

        lists_layout = QHBoxLayout()
        lists_layout.setSpacing(15)

        taraf_group = QGroupBox("👥 Taraf Bilgileri (Şüpheli / Müşteki)")
        taraf_layout = QVBoxLayout(taraf_group)
        taraf_layout.setSpacing(8)

        input_row = QHBoxLayout()
        self.cmb_sifat = QComboBox(); self.cmb_sifat.addItems(["Şüpheli", "Sanık", "SSÇ", "Müşteki Sanık", "Katılan Sanık", "Müşteki", "Mağdur", "Katılan", "Şikayetçi"]); self.cmb_sifat.setFixedWidth(110)
        apply_light_combobox_popup(self.cmb_sifat)
        self.inp_adsoyad = QLineEdit(); self.inp_adsoyad.setPlaceholderText("Adı Soyadı")
        btn_add_taraf = QPushButton("Ekle"); btn_add_taraf.setStyleSheet("background-color: #16a085; color: white; border:none; padding: 6px;")
        btn_add_taraf.clicked.connect(self.add_taraf)
        input_row.addWidget(self.cmb_sifat); input_row.addWidget(self.inp_adsoyad); input_row.addWidget(btn_add_taraf)
        taraf_layout.addLayout(input_row)

        self.party_table = GenericDatabaseTable(["ID", "Sıfat", "Adı Soyadı"], chart_mode='none', enable_evidence_menu=False)
        self.party_table.table.setColumnHidden(0, True)
        self.party_table.setMinimumHeight(100)

        h_party = self.party_table.table.horizontalHeader()
        h_party.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive); self.party_table.table.setColumnWidth(1, 100)
        h_party.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

        taraf_layout.addWidget(self.party_table)

        btn_del_taraf = QPushButton("Seçili Tarafı Sil 🗑️"); btn_del_taraf.setStyleSheet("background-color: #95a5a6; color: white; padding: 6px; border:none;")
        btn_del_taraf.clicked.connect(self.delete_taraf)
        taraf_layout.addWidget(btn_del_taraf)

        lists_layout.addWidget(taraf_group, 5)

        gsm_group = QGroupBox("📂 Yüklü GSM Numaraları ve Aboneler")
        gsm_layout = QVBoxLayout(gsm_group)
        gsm_layout.setSpacing(8)

        self.pm_gsm_table = GenericDatabaseTable(["NO", "ABONE"], chart_mode='none', enable_evidence_menu=False)
        self.pm_gsm_table.hide_toolbar()
        self.pm_gsm_table.setMinimumHeight(150)

        t = self.pm_gsm_table.table
        h = t.horizontalHeader()

        t.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        t.clicked.connect(self.on_pm_gsm_select)

        t.setTextElideMode(Qt.TextElideMode.ElideNone)
        t.setItemDelegateForColumn(0, NoElideDelegate(t))

        h.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        t.setColumnWidth(0, 300)

        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        t.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        t.setWordWrap(False)

        gsm_layout.addWidget(self.pm_gsm_table)

        gsm_btns = QHBoxLayout()
        btn_del_gsm = QPushButton("Seçili Sil 🗑️"); btn_del_gsm.setStyleSheet("background-color: #c0392b; color: white; padding: 6px; border:none;")
        btn_del_gsm.clicked.connect(self.delete_project_gsm)
        self.btn_upload = QPushButton("📥 HTS Yükle"); self.btn_upload.setStyleSheet("background-color: #d35400; color: white; padding: 6px; border:none;")
        self.btn_upload.clicked.connect(self.upload_excel_pm)
        gsm_btns.addWidget(btn_del_gsm); gsm_btns.addWidget(self.btn_upload)
        gsm_layout.addLayout(gsm_btns)

        lists_layout.addWidget(gsm_group, 5)
        layout.addLayout(lists_layout)

        file_info_group = QGroupBox("📄 Seçili HTS Dosya Bilgileri")
        f_layout = QVBoxLayout(file_info_group)

        self.file_table = QTableWidget()
        self.file_table.setItemDelegate(NoTooltipDelegate(self.file_table))
        self.file_table.setColumnCount(2)
        self.file_table.setHorizontalHeaderLabels(["Özellik", "Bilgiler"])
        self.file_table.verticalHeader().setVisible(False)
        self.file_table.setAlternatingRowColors(True)
        self.file_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.file_table.setMouseTracking(False)
        self.file_table.viewport().setMouseTracking(False)
        self.file_table.setToolTip("")
        self.file_table.viewport().setToolTip("")
        header = self.file_table.horizontalHeader()

        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.file_table.setColumnWidth(0, 300)

        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        self.file_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.file_table.verticalHeader().setDefaultSectionSize(35) # Satır yüksekliği (piksel)
        self.file_table.setWordWrap(False)

        self.file_table.setFixedHeight(200)
        self.file_table.cellDoubleClicked.connect(self.open_file_details_popup)

        f_layout.addWidget(self.file_table)
        layout.addWidget(file_info_group)

        layout.addStretch()
        b_go = QPushButton("🚀 PROJEYİ ANALİZ İÇİN AÇ")
        b_go.setStyleSheet("QPushButton { background-color: #8e44ad; color: white; font-weight: bold; padding: 15px; font-size: 15px; border-radius: 8px; margin-top: 10px; border:none;} QPushButton:hover { background-color: #9b59b6; }")
        b_go.clicked.connect(self.go_analiz)
        layout.addWidget(b_go)

    def open_file_details_popup(self, row, col):
        """Dosya bilgilerini tam ekran popup'ta gösterir."""
        data = []
        for r in range(self.file_table.rowCount()):
            key_item = self.file_table.item(r, 0)
            val_item = self.file_table.item(r, 1)
            if key_item and val_item:
                data.append((key_item.text(), val_item.text()))

        if not data: return

        dlg = FileDetailPopup(self, data)
        dlg.exec()

    def _ensure_db_columns(self):
        """Eğer veritabanında yeni sütunlar yoksa ekler (Otomatik Onarım)."""
        try:
            with DB() as conn:
                columns = [i[1] for i in conn.execute("PRAGMA table_info(projeler)")]

                if "suc_tarihi" not in columns:
                    conn.execute("ALTER TABLE projeler ADD COLUMN suc_tarihi TEXT")

                if "gorevlendirme_tarihi" not in columns:
                    conn.execute("ALTER TABLE projeler ADD COLUMN gorevlendirme_tarihi TEXT")

                if "bilirkisi_adi" not in columns:
                    conn.execute("ALTER TABLE projeler ADD COLUMN bilirkisi_adi TEXT")

                if "bilirkisi_unvan_sicil" not in columns:
                    conn.execute("ALTER TABLE projeler ADD COLUMN bilirkisi_unvan_sicil TEXT")

        except Exception as e:
            print(f"DB Onarım Hatası: {e}")

    def on_pm_gsm_select(self, index):
        """GSM seçildiğinde dosya detaylarını ve HASH bilgilerini gösterir."""
        try:
            if not self.selected_project_id:
                self.file_table.setRowCount(0)
                return

            src = self.pm_gsm_table.proxy_model.mapToSource(index)
            selected_gsm = self.pm_gsm_table.source_model._data[src.row()][0]

            with DB() as conn:
                empty_sql = """
                    SELECT 1 FROM (
                        SELECT 1 FROM hts_gsm          WHERE ProjeID=? AND GSMNo=?
                        UNION ALL
                        SELECT 1 FROM hts_sms          WHERE ProjeID=? AND GSMNo=?
                        UNION ALL
                        SELECT 1 FROM hts_sabit        WHERE ProjeID=? AND GSMNo=?
                        UNION ALL
                        SELECT 1 FROM hts_gprs         WHERE ProjeID=? AND GSMNo=?
                        UNION ALL
                        SELECT 1 FROM hts_wap          WHERE ProjeID=? AND GSMNo=?
                        UNION ALL
                        SELECT 1 FROM hts_sth          WHERE ProjeID=? AND GSMNo=?
                        UNION ALL
                        SELECT 1 FROM hts_uluslararasi WHERE ProjeID=? AND GSMNo=?
                    ) AS t
                    LIMIT 1
                """
                is_empty = conn.execute(
                    empty_sql,
                    (self.selected_project_id, selected_gsm,
                     self.selected_project_id, selected_gsm,
                     self.selected_project_id, selected_gsm,
                     self.selected_project_id, selected_gsm,
                     self.selected_project_id, selected_gsm,
                     self.selected_project_id, selected_gsm,
                     self.selected_project_id, selected_gsm)
                ).fetchone() is None

                if is_empty:
                    ModernDialog.show_warning(
                        self,
                        "HTS Kaydı Boş",
                        f"{selected_gsm} numarasına ait HTS içeriği bulunamadı.\n\n"
                        "Bu kayıt listede tutuldu; ancak görüşme/mesaj verisi olmadığından analiz üretilemez."
                    )

                meta_row = conn.execute("""
                    SELECT TalepEdenMakam, SorguBaslangic, SorguBitis, Tespit
                    FROM hts_dosyalari WHERE ProjeID=? AND GSMNo=?
                    ORDER BY CASE WHEN Rol='HEDEF' THEN 1 ELSE 2 END LIMIT 1
                """, (self.selected_project_id, selected_gsm)).fetchone()

                files = conn.execute("""
                    SELECT Rol, DosyaAdi, DosyaBoyutu, YuklenmeTarihi, MD5, SHA256
                    FROM hts_dosyalari WHERE ProjeID=? AND GSMNo=? ORDER BY Rol
                """, (self.selected_project_id, selected_gsm)).fetchall()

            rows_data = []

            rows_data.append(("📌 Talep Eden Makam", meta_row[0] if meta_row and meta_row[0] else "-", "NORMAL"))
            rows_data.append(("🗓️ Sorgu Başlangıç", meta_row[1] if meta_row and meta_row[1] else "-", "NORMAL"))
            rows_data.append(("🗓️ Sorgu Bitiş", meta_row[2] if meta_row and meta_row[2] else "-", "NORMAL"))
            rows_data.append(("🔎 Tespit", meta_row[3] if meta_row and meta_row[3] else "-", "NORMAL"))
            rows_data.append(("📁 Yüklenen Dosyalar", "", "HEADER"))

            if files:
                for (rol, fname, fsize, ydate, md5_val, sha_val) in files:
                    role_text = "🎯 HEDEF" if str(rol).upper() == "HEDEF" else "🔁 KARŞI"
                    size_text = format_size(int(fsize)) if fsize else "-"
                    date_text = str(ydate) if ydate else "-"

                    rows_data.append((f"{role_text} - {fname}", f"{size_text} | {date_text}", "FILE_MAIN"))

                    md5_val = md5_val if md5_val else "-"
                    sha_val = sha_val if sha_val else "-"
                    rows_data.append(("      └─ 🔒 MD5", md5_val, "FILE_CHILD"))
                    rows_data.append(("      └─ 🔒 SHA-256", sha_val, "FILE_CHILD"))
            else:
                rows_data.append(("Dosya", "Yok", "NORMAL"))

            self.file_table.setRowCount(len(rows_data))
            font_bold = QFont("Segoe UI", 9, QFont.Weight.Bold)
            font_small = QFont("Segoe UI", 8)
            font_small.setItalic(True)

            for i, (key, val, row_type) in enumerate(rows_data):
                item_k = QTableWidgetItem(str(key))
                item_v = QTableWidgetItem(str(val))
                item_v.setToolTip(str(val))

                if row_type == "HEADER":
                    item_k.setFont(font_bold)
                    item_k.setForeground(QColor("#2c3e50"))
                    item_k.setBackground(QColor("#dcdde1"))
                    item_v.setBackground(QColor("#dcdde1"))

                elif row_type == "FILE_MAIN":
                    item_k.setFont(font_bold)
                    item_k.setForeground(QColor("#c0392b"))
                    item_k.setBackground(QColor("#dfe6e9"))
                    item_v.setFont(font_bold)
                    item_v.setForeground(QColor("#c0392b"))
                    item_v.setBackground(QColor("#dfe6e9"))

                elif row_type == "FILE_CHILD":
                    if "🔒" in key:
                        item_v.setFont(font_small)
                        item_v.setForeground(QColor("#7f8c8d"))
                    else:
                        item_v.setForeground(QColor("#333333"))

                self.file_table.setItem(i, 0, item_k)
                self.file_table.setItem(i, 1, item_v)

        except Exception as e:
            print(f"Dosya Detay Hatası: {e}")

    def load_projects(self):
        with DB() as conn:
            rows = conn.execute(
                "SELECT id, talep_eden_birim, dosya_no, dosya_no_tipi, suc_bilgisi, suc_tarihi, gorevlendirme_tarihi "
                "FROM projeler ORDER BY id DESC"
            ).fetchall()

        data = [[r[0], r[1], r[2], r[3], r[4]] for r in rows]
        self.p_table.set_data(data)

        if data:
            try:
                self.p_table.table.selectRow(0)
                first_idx = self.p_table.proxy_model.index(0, 0)
                self.on_select(first_idx)
            except Exception as e:
                print(f"İlk proje otomatik seçim hatası: {e}")
        else:
            self.clear_form()
            if hasattr(self, "gsm_table"):
                self.gsm_table.set_data([])
            if hasattr(self, "file_table"):
                self.file_table.setRowCount(0)

    def on_select(self, idx):
        """Seçili projenin verilerini forma doldurur."""
        try:
            src = self.p_table.proxy_model.mapToSource(idx)
            p_id = self.p_table.source_model._data[src.row()][0]
            self.selected_project_id = p_id

            with DB() as conn:
                row = conn.execute("""
                    SELECT talep_eden_birim, dosya_no_tipi, dosya_no,
                           suc_bilgisi, suc_tarihi, gorevlendirme_tarihi,
                           bilirkisi_adi, bilirkisi_unvan_sicil
                    FROM projeler WHERE id=?
                """, (p_id,)).fetchone()

            if row:
                self.i_birim.setText(row[0] or "")
                self.i_tip.setCurrentText(row[1] or "Esas No")
                self.i_dosya.setText(row[2] or "")
                self.i_suc.setText(row[3] or "")

                self.i_suc_tarihi.setText(row[4] or "")
                self.i_gorev_tarihi.setText(row[5] or "")

                self.i_bilirkisi_adi.setText(row[6] or "")
                self.i_bilirkisi_unvan.setText(row[7] or "")

            self.load_parties()
            self.load_project_gsms()

        except Exception as e:
            print(f"Proje seçim hatası: {e}")

    def load_parties(self):
        if not self.selected_project_id: self.party_table.set_data([]); return
        with DB() as conn: rows = conn.execute("SELECT id, sifat, ad_soyad FROM taraflar WHERE ProjeID=?", (self.selected_project_id,)).fetchall()
        self.party_table.set_data([list(r) for r in rows])

    def clear_form(self):
        self.selected_project_id = None

        self.i_birim.clear()
        self.i_dosya.clear()
        self.i_suc.clear()
        self.i_bilirkisi_adi.clear()
        self.i_bilirkisi_unvan.clear()

        self.i_suc_tarihi.clear()
        self.i_gorev_tarihi.clear()

        self.i_tip.setCurrentIndex(0)

        self.party_table.set_data([])

        if hasattr(self, "pm_gsm_table"):
            self.pm_gsm_table.set_data([])

        if hasattr(self, "file_table"):
            self.file_table.setRowCount(0)

        self.p_table.table.clearSelection()

        sender = self.sender()
        if sender and isinstance(sender, QPushButton) and "Yeni Proje" in sender.text():
            ModernDialog.show_success(self, "Yeni Proje", "Form temizlendi, yeni kayda hazır.")

    def save_project(self):
        """Projeyi kaydeder. Yeni sütunlar (suc_tarihi, gorevlendirme_tarihi, bilirkişi alanları) dahil."""
        birim = self.i_birim.text()
        tip = self.i_tip.currentText()
        dosya = self.i_dosya.text()

        gorev_t = self.i_gorev_tarihi.text()
        suc_t = self.i_suc_tarihi.text()
        suc_bilgi = self.i_suc.text()

        bilirkisi_adi = self.i_bilirkisi_adi.text()
        bilirkisi_unvan = self.i_bilirkisi_unvan.text()

        data = (birim, tip, dosya, suc_bilgi, suc_t, gorev_t, bilirkisi_adi, bilirkisi_unvan)

        try:
            with DB() as conn:
                if self.selected_project_id:
                    conn.execute("""
                        UPDATE projeler SET
                            talep_eden_birim=?,
                            dosya_no_tipi=?,
                            dosya_no=?,
                            suc_bilgisi=?,
                            suc_tarihi=?,
                            gorevlendirme_tarihi=?,
                            bilirkisi_adi=?,
                            bilirkisi_unvan_sicil=?
                        WHERE id=?
                    """, data + (self.selected_project_id,))
                    ModernDialog.show_success(self, "Bilgi", "Proje güncellendi.")
                else:
                    cur = conn.execute("""
                        INSERT INTO projeler
                        (talep_eden_birim, dosya_no_tipi, dosya_no,
                         suc_bilgisi, suc_tarihi, gorevlendirme_tarihi,
                         bilirkisi_adi, bilirkisi_unvan_sicil, olusturma_tarihi)
                        VALUES (?,?,?,?,?,?,?, ?, CURRENT_TIMESTAMP)
                    """, data)
                    self.selected_project_id = cur.lastrowid
                    ModernDialog.show_success(self, "Bilgi", "Yeni proje oluşturuldu.")

                conn.commit()

            self.load_projects()

        except Exception as e:
            ModernDialog.show_error(self, "Kayıt Hatası", str(e))

    def delete_project(self):
        """Projeyi ve TÜM alt verilerini siler (TAM LİSTE GÜNCELLENDİ)."""
        if not self.selected_project_id:
            ModernDialog.show_warning(self, "Uyarı", "Lütfen silinecek projeyi seçin."); return

        if ModernDialog.show_question(self, "Projeyi Sil", "Bu projeyi ve içerisindeki TÜM verileri (HTS, Raporlar, Konumlar) silmek istediğinize emin misiniz?"):

            if hasattr(self.main, 'loader'): self.main.loader.start("Proje Siliniyor...")
            QApplication.processEvents()

            try:
                with DB() as conn:
                    all_tables = [
                        "taraflar", "hts_abone", "hts_gsm", "hts_sms", "hts_sabit",
                        "hts_gprs", "hts_wap", "hts_sth", "hts_uluslararasi",
                        "hts_ozet", "hts_ozet_iletisim", "hts_ozet_baz", "hts_ozet_imei",
                        "hts_rehber", "hts_tum_baz", "ozel_konumlar",
                        "hts_ortak_imei", "hts_ortak_isim", "hts_ortak_tc", "rapor_taslagi"
                    ]

                    for t in all_tables:
                        try: conn.execute(f"DELETE FROM {t} WHERE ProjeID=?", (self.selected_project_id,))
                        except: pass

                    conn.execute("DELETE FROM projeler WHERE id=?", (self.selected_project_id,))
                    conn.commit()

                AnalysisUtils.perform_maintenance()
                self.clear_form()
                self.load_projects()
                self.load_project_gsms()
                ModernDialog.show_success(self, "Silindi", "Proje tamamen silindi ve bakım yapıldı.")

            except Exception as e:
                ModernDialog.show_error(self, "Hata", str(e))
            finally:
                if hasattr(self.main, 'loader'): self.main.loader.stop()

    def vacuum_db(self):
        if ModernDialog.show_question(self, "Bakım", "Veritabanı sıkıştırılacak. Devam edilsin mi?"):
            if AnalysisUtils.perform_maintenance():
                ModernDialog.show_success(self, "Bakım", "Veritabanı sıkıştırıldı.")
            else:
                ModernDialog.show_error(self, "Hata", "Bakım yapılamadı.")

    def add_taraf(self):
        if not self.selected_project_id:
            ModernDialog.show_warning(self, "Uyarı", "Önce projeyi kaydedin veya seçin.")
            return

        ad_soyad = self.inp_adsoyad.text().strip()
        if not ad_soyad: return

        try:
            with DB() as conn:
                conn.execute("INSERT INTO taraflar (ProjeID, sifat, ad_soyad) VALUES (?,?,?)",
                             (self.selected_project_id, self.cmb_sifat.currentText(), ad_soyad))

            self.inp_adsoyad.clear()
            self.load_parties()
            ModernDialog.show_success(self, "Başarılı", "Taraf eklendi.")

        except Exception as e:
            print(f"Ekleme Hatası: {e}")

    def delete_taraf(self):
        """Seçili tarafı siler ve listeyi ANINDA günceller."""
        idx = self.party_table.table.currentIndex()
        if not idx.isValid():
            ModernDialog.show_warning(self, "Seçim Yok", "Lütfen silinecek kişiyi listeden seçiniz.")
            return

        if ModernDialog.show_question(self, "Silme Onayı", "Seçili tarafı listeden silmek istediğinize emin misiniz?"):
            try:
                src = self.party_table.proxy_model.mapToSource(idx)
                rid = self.party_table.source_model._data[src.row()][0]

                with DB() as conn:
                    conn.execute("DELETE FROM taraflar WHERE id=?", (rid,))
                self.load_parties()

                ModernDialog.show_success(self, "Başarılı", "Kayıt silindi.")

            except Exception as e:
                ModernDialog.show_error(self, "Hata", str(e))

    def load_project_gsms(self):
        gsm_table = getattr(self, "pm_gsm_table", None)
        file_table = getattr(self, "file_table", None)

        if not self.selected_project_id:
            if gsm_table is not None:
                gsm_table.set_data([])
            if file_table is not None:
                file_table.setRowCount(0)
            return

        try:
            with DB() as conn:
                rows = conn.execute("""
                    SELECT GSMNo, MAX(YuklenmeTarihi) AS last_upload
                    FROM hts_dosyalari
                    WHERE ProjeID=?
                    GROUP BY GSMNo
                    ORDER BY last_upload DESC
                """, (self.selected_project_id,)).fetchall()

                unique_numbers = [r[0] for r in rows if r and r[0]]

                def has_any_hts_data(gsm: str) -> bool:
                    sql = """
                        SELECT 1 FROM (
                            SELECT 1 FROM hts_gsm          WHERE ProjeID=? AND GSMNo=?
                            UNION ALL
                            SELECT 1 FROM hts_sms          WHERE ProjeID=? AND GSMNo=?
                            UNION ALL
                            SELECT 1 FROM hts_sabit        WHERE ProjeID=? AND GSMNo=?
                            UNION ALL
                            SELECT 1 FROM hts_gprs         WHERE ProjeID=? AND GSMNo=?
                            UNION ALL
                            SELECT 1 FROM hts_wap          WHERE ProjeID=? AND GSMNo=?
                            UNION ALL
                            SELECT 1 FROM hts_sth          WHERE ProjeID=? AND GSMNo=?
                            UNION ALL
                            SELECT 1 FROM hts_uluslararasi WHERE ProjeID=? AND GSMNo=?
                        ) AS t
                        LIMIT 1
                    """
                    params = (
                        self.selected_project_id, gsm,
                        self.selected_project_id, gsm,
                        self.selected_project_id, gsm,
                        self.selected_project_id, gsm,
                        self.selected_project_id, gsm,
                        self.selected_project_id, gsm,
                        self.selected_project_id, gsm,
                    )
                    return conn.execute(sql, params).fetchone() is not None

                tmp = []

                for gsm in unique_numbers:
                    name_rows = conn.execute(
                        "SELECT AD, SOYAD FROM hts_abone WHERE ProjeID=? AND GSMNo=?",
                        (self.selected_project_id, gsm)
                    ).fetchall()

                    unique_names = set()
                    for r in name_rows:
                        ad = str(r[0]).strip() if r and r[0] else ""
                        soyad = str(r[1]).strip() if r and r[1] else ""
                        full_name = f"{ad} {soyad}".strip()
                        if full_name:
                            unique_names.add(full_name)

                    abone_str = " / ".join(sorted(unique_names))
                    empty_flag = 0 if has_any_hts_data(gsm) else 1
                    tmp.append((empty_flag, gsm, abone_str))

            tmp.sort(key=lambda x: x[0])

            display_data = [[gsm, abone_str] for (_, gsm, abone_str) in tmp]

            if gsm_table is not None:
                gsm_table.set_data(display_data)

            if display_data and gsm_table is not None:
                def _select_first_and_load():
                    try:
                        gsm_table.table.selectRow(0)
                        first_idx = gsm_table.proxy_model.index(0, 0)
                        gsm_table.table.setCurrentIndex(first_idx)
                        gsm_table.table.scrollTo(first_idx)

                        if hasattr(self, "on_pm_gsm_select"):
                            self.on_pm_gsm_select(first_idx)

                    except Exception as e:
                        print(f"İlk GSM otomatik seçim hatası: {e}")

                QTimer.singleShot(0, _select_first_and_load)

            else:
                if file_table is not None:
                    file_table.setRowCount(0)

        except Exception as e:
            print(f"Proje GSM yükleme hatası: {e}")
            if gsm_table is not None:
                gsm_table.set_data([])
            if file_table is not None:
                file_table.setRowCount(0)

    def delete_project_gsm(self):
        """Proje ekranından seçili numaraları siler ve ORTAK ANALİZİ GÜNCELLER (LOGLU)."""
        selection = self.pm_gsm_table.table.selectionModel().selectedRows()
        if not selection:
            ModernDialog.show_warning(self, "Uyarı", "Lütfen silinecek numaraları seçin.")
            return

        gsms_to_delete = []
        for idx in selection:
            src_idx = self.pm_gsm_table.proxy_model.mapToSource(idx)
            gsms_to_delete.append(self.pm_gsm_table.source_model._data[src_idx.row()][0])

        if ModernDialog.show_question(self, "Silme Onayı", f"{len(gsms_to_delete)} adet numara silinecek.\nOnaylıyor musunuz?"):

            if hasattr(self.main, 'loader'): self.main.loader.start("Siliniyor...")
            QApplication.processEvents()

            try:
                with DB() as conn:
                    tabs_gsm = [
                        "hts_dosyalari", "hts_abone", "hts_gsm", "hts_sms",
                        "hts_sabit", "hts_gprs", "hts_wap", "hts_sth", "hts_uluslararasi",
                        "hts_ozet", "hts_ozet_iletisim", "hts_ozet_baz", "hts_ozet_imei",
                        "hts_rehber", "hts_tum_baz"
                    ]

                    for gsm in gsms_to_delete:
                        for t in tabs_gsm:
                            try:
                                conn.execute(f"DELETE FROM {t} WHERE ProjeID=? AND GSMNo=?", (self.selected_project_id, gsm))
                            except Exception as e:
                                print(f"⚠️ [PM Delete] '{t}' tablosu silinirken hata: {e}")

                    tabs_common = ["hts_ortak_imei", "hts_ortak_isim", "hts_ortak_tc"]
                    for t in tabs_common:
                        conn.execute(f"DELETE FROM {t} WHERE ProjeID=?", (self.selected_project_id,))

                    conn.commit()

                self.recalculate_common_analysis()
                AnalysisUtils.perform_maintenance()
                self.load_project_gsms()
                self.file_table.setRowCount(0)

                ModernDialog.show_success(self, "Başarılı", "Numaralar silindi.")

            except Exception as e:
                print(f"❌ [PM Delete] Kritik Hata: {e}")
                ModernDialog.show_error(self, "Hata", str(e))
            finally:
                if hasattr(self.main, 'loader'): self.main.loader.stop()

    def recalculate_common_analysis(self):
        if not self.selected_project_id:
            return

        if not AnalysisUtils.project_has_any_gsm(self.selected_project_id):
            return

        AnalysisUtils.recalculate_common_analysis_core(self.selected_project_id)

    def go_analiz(self):
        if not self.selected_project_id: ModernDialog.show_warning(self, "Uyarı", "Lütfen analiz edilecek projeyi seçin."); return
        with DB() as conn:
            has_data = conn.execute("SELECT 1 FROM hts_ozet WHERE ProjeID=? LIMIT 1", (self.selected_project_id,)).fetchone()
            if not has_data: has_data = conn.execute("SELECT 1 FROM hts_gsm WHERE ProjeID=? LIMIT 1", (self.selected_project_id,)).fetchone()
        if has_data: self.main.show_analysis(self.selected_project_id)
        else: ModernDialog.show_error(self, "Veri Yok", "Bu projeye ait yüklenmiş HTS kaydı bulunamadı.")

    def upload_excel_pm(self):
        if not self.selected_project_id:
            ModernDialog.show_warning(self, "Uyarı", "Lütfen önce bir proje seçin veya kaydedin.")
            return

        paths, _ = QFileDialog.getOpenFileNames(self, "Excel Seç", "", "Excel Files (*.xlsx *.xls)")
        if not paths: return

        self.upload_queue_pm = list(paths)
        self.total_count_pm = len(paths)
        self.success_count_pm = 0
        self.is_uploading_pm = True

        if hasattr(self.main, 'loader'):
            self.main.loader.start(f"Yükleniyor... (1/{self.total_count_pm})")

        self.process_next_in_queue_pm()

    def delete_records_for_role(self, gsm: str, rol: str):
        """
        Overwrite seçilince sadece ilgili rolün kayıtlarını temizler.
        DÜZELTME: self.selected_project_id kullanıldı.
        """
        with DB() as conn:
            conn.execute("""
                DELETE FROM hts_dosyalari
                WHERE ProjeID=? AND GSMNo=? AND Rol=?
            """, (self.selected_project_id, gsm, rol))

            if rol == "HEDEF":
                conn.execute("""
                    DELETE FROM hts_dosyalari
                    WHERE ProjeID=? AND GSMNo=?
                """, (self.selected_project_id, gsm))

            conn.commit()

        if rol == "HEDEF":
            self.delete_gsm_database_records(gsm)

        if rol == "KARSI":
            try:
                with DB() as conn:
                    for t in ["hts_karsi_baz", "hts_karsi_imei"]:
                        try:
                            conn.execute(f"DELETE FROM {t} WHERE ProjeID=? AND GSMNo=?", (self.selected_project_id, gsm))
                        except: pass
                    conn.commit()
            except:
                pass

    def process_next_in_queue_pm(self):
        if not hasattr(self, 'upload_queue_pm') or not self.upload_queue_pm:
            self.is_uploading_pm = False
            if hasattr(self.main, 'loader'): self.main.loader.stop()

            if getattr(self, "success_count_pm", 0) > 0:
                ModernDialog.show_success(self, "Yükleme Tamamlandı", f"{self.success_count_pm} / {self.total_count_pm} dosya yüklendi.")
                self.success_count_pm = 0; self.total_count_pm = 0

            self.load_project_gsms()
            if hasattr(self.main, 'analysis_center') and hasattr(self.main.analysis_center, 'load_project_gsms'):
                self.main.analysis_center.load_project_gsms()
            # -----------------------------------
            return

        next_file = self.upload_queue_pm[0]

        current_file_name = os.path.basename(next_file)
        if hasattr(self.main, 'loader'):
            self.main.loader.start(f"Hazırlanıyor:\n{current_file_name}")

        try:
            rol, target_gsm = detect_hts_role(next_file)
        except Exception:
            rol, target_gsm = "HEDEF", _detect_target_gsm(next_file)

        is_same_role_exist = False
        with DB() as conn:
            row = conn.execute("SELECT 1 FROM hts_dosyalari WHERE ProjeID=? AND GSMNo=? AND Rol=? LIMIT 1", (self.selected_project_id, target_gsm, rol)).fetchone()
            is_same_role_exist = row is not None

        if is_same_role_exist:
            if hasattr(self.main, 'loader'): self.main.loader.hide()
            dialog = ModernDialog(self, "Mükerrer Kayıt", f"{target_gsm} ({rol}) zaten yüklü. Üzerine yazılsın mı?", "QUESTION", "Üzerine Yaz", "Atla")
            res = dialog.exec()
            if hasattr(self.main, 'loader'): self.main.loader.show()

            if res != 1:
                self.upload_queue_pm.pop(0)
                QTimer.singleShot(0, self.process_next_in_queue_pm)
                return
            self.delete_records_for_role(target_gsm, rol)

        self.upload_queue_pm.pop(0)
        self.worker = HtsWorker(next_file, self.selected_project_id)

        self.worker.progress.connect(self.on_worker_progress)
        self.worker.log.connect(self.on_worker_log)
        self.worker.finished.connect(self.on_pm_file_finished)
        self.worker.error.connect(self.on_upload_error_pm)

        self.worker.start()

    def on_worker_progress(self, value):
        if hasattr(self.main, 'loader'): self.main.loader.set_progress(value)

    def on_worker_log(self, text):
        """Worker'dan gelen detaylı metni Overlay'e yazar."""
        if hasattr(self.main, 'loader'):
            self.main.loader.text = text
            self.main.loader.update()

    def on_pm_file_finished(self, msg):
        """Tekil dosya bittiğinde çağrılır."""
        self.success_count_pm += 1
        self.process_next_in_queue_pm()

    def on_upload_error_pm(self, err_msg):
        """PM yükleme hatası."""
        if hasattr(self.main, 'loader'): self.main.loader.hide()
        ModernDialog.show_error(self, "Yükleme Hatası", f"Dosya hatası:\n{err_msg}")
        if hasattr(self.main, 'loader'): self.main.loader.show()
        self.process_next_in_queue_pm()

    def check_gsm_exists(self, gsm, rol):
        if not gsm or gsm == "BILINMIYOR":
            return False
        try:
            with DB() as conn:
                q = """SELECT 1 FROM hts_dosyalari
                       WHERE ProjeID=? AND GSMNo=? AND Rol=? LIMIT 1"""
                if conn.execute(q, (self.selected_project_id, gsm, rol)).fetchone():
                    return True
        except:
            return False
        return False

    def delete_gsm_database_records(self, gsm):
        pid = self.selected_project_id if hasattr(self, 'selected_project_id') else self.current_project_id

        if AnalysisUtils.delete_gsm_records_core(pid, gsm):
            with DB() as conn:
                for t in ["hts_ortak_imei", "hts_ortak_isim", "hts_ortak_tc"]:
                    conn.execute(f"DELETE FROM {t} WHERE ProjeID=?", (pid,))

            self.recalculate_common_analysis()
            AnalysisUtils.perform_maintenance()
            return True
        return False

    def clear_file_details(self):
        try:
            self.file_table.clearContents()
            self.file_table.setRowCount(0)
            # Başlıklar dursun istiyorsan clear() kullanma
        except Exception as e:
            print(f"Dosya detay temizleme hatası: {e}")


class InteractiveGraphWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0,0,0,0)

        self.splitter = QSplitter(Qt.Orientation.Horizontal)

        left_widget = QWidget(); left_widget.setStyleSheet("background-color: #f7f9f9;")
        left_layout = QVBoxLayout(left_widget); left_layout.setContentsMargins(10, 10, 10, 10); left_layout.setSpacing(10)

        h_title = QHBoxLayout()
        h_title.addWidget(QLabel("📊 Veri Seçimi", styleSheet="font-size:14px; font-weight:bold; color:#2c3e50;"))
        info_btn = InfoButton(
            "<b>📊 Veri Görselleştirme Aracı:</b><br>"
            "Karmaşık sayısal verileri, anlaşılır ilişki diyagramlarına dönüştürür.<br><br>"
            "1. Sol listeden görselleştirmek istediğiniz verileri (Baz istasyonları, Kişiler vb.) seçin.<br>"
            "2. <b>'Diyagramı Çiz'</b> butonuna basın.<br>"
            "3. Oluşan grafiği fare ile sürükleyip yakınlaştırarak kümeleri inceleyin."
        )
        h_title.addWidget(info_btn); h_title.addStretch()
        left_layout.addLayout(h_title)

        search_frame = QFrame(); search_frame.setStyleSheet("background-color: white; border: 1px solid #bdc3c7; border-radius: 6px;")
        h_search = QHBoxLayout(search_frame); h_search.setContentsMargins(5, 5, 5, 5)
        h_search.addWidget(QLabel("🔍", styleSheet="border:none; font-size:14px;"));
        self.search_bar = QLineEdit(); self.search_bar.setPlaceholderText("Listede Ara..."); self.search_bar.setStyleSheet("border:none; background:transparent; font-size:13px;")
        self.search_bar.textChanged.connect(self.filter_list)
        h_search.addWidget(self.search_bar)
        left_layout.addWidget(search_frame)

        self.table = QTableWidget(); self.table.setColumnCount(3); self.table.setHorizontalHeaderLabels(["", "Veri Adı", "Adet"])
        h = self.table.horizontalHeader(); h.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed); self.table.setColumnWidth(0, 30)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch); h.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed); self.table.setColumnWidth(2, 60)
        self.table.verticalHeader().setVisible(False); self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows); self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setStyleSheet("QTableWidget { background-color: white; border: 1px solid #bdc3c7; border-radius: 4px; } QHeaderView::section { background-color: #ecf0f1; border: none; border-bottom: 1px solid #bdc3c7; padding: 4px; font-weight: bold; color: #2c3e50; }")
        left_layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        btn_all = QPushButton("✅ Tümünü Seç"); btn_all.setStyleSheet("background-color: #16a085; color: white; padding: 6px; border-radius:4px; font-weight:bold; border:none;")
        btn_all.clicked.connect(lambda: self.toggle_all(True))
        btn_none = QPushButton("⬜ Temizle"); btn_none.setStyleSheet("background-color: #7f8c8d; color: white; padding: 6px; border-radius:4px; font-weight:bold; border:none;")
        btn_none.clicked.connect(lambda: self.toggle_all(False))
        btn_row.addWidget(btn_all); btn_row.addWidget(btn_none)
        left_layout.addLayout(btn_row)

        btn_draw = QPushButton("🎨 Diyagramı Çiz"); btn_draw.setStyleSheet("background-color: #d35400; color: white; font-weight: bold; padding: 12px; font-size:13px; border-radius: 5px; margin-top:5px; border:none;")
        btn_draw.clicked.connect(self.draw_selected_graph)
        left_layout.addWidget(btn_draw)

        btn_snapshot = QPushButton("📸 Görüntüyü Rapora Ekle"); btn_snapshot.setStyleSheet("background-color: #8e44ad; color: white; font-weight: bold; padding: 10px; border-radius: 5px; border:none;")
        btn_snapshot.clicked.connect(self.take_snapshot)
        left_layout.addWidget(btn_snapshot)

        self.splitter.addWidget(left_widget)

        right_widget = QWidget(); r_layout = QVBoxLayout(right_widget); r_layout.setContentsMargins(0,0,0,0)
        self.browser = EvidenceWebEngineView(); self.browser.setStyleSheet("background-color: white; border-left: 1px solid #bdc3c7;")
        self.browser.setHtml("<div style='display:flex; justify-content:center; align-items:center; height:100vh; flex-direction:column; font-family:Segoe UI; background-color:#f9f9f9;'><h2 style='color:#7f8c8d;'>👈 Analiz Verisi Seçin</h2><p style='color:#95a5a6;'>Soldaki listeden verileri seçip <b>'Diyagramı Çiz'</b> butonuna basın.</p></div>")
        r_layout.addWidget(self.browser)
        self.splitter.addWidget(right_widget)
        self.splitter.setSizes([380, 900]); self.splitter.setStretchFactor(1, 1)
        self.layout.addWidget(self.splitter)

        self.current_center_name = ""; self.full_data_list = []

    def take_snapshot(self):
        parent = self.parent()
        while parent:
            if hasattr(parent, 'capture_chart_screenshot'): parent.capture_chart_screenshot(self.browser, "Ağ Analizi Grafiği"); break
            if hasattr(parent, 'main') and hasattr(parent.main.page_analysis, 'capture_chart_screenshot'): parent.main.page_analysis.capture_chart_screenshot(self.browser, "Ağ Analizi Grafiği"); break
            parent = parent.parent()

    def load_list_data(self, center_name, data_list):
        self.current_center_name = center_name; self.full_data_list = data_list
        self.table.setRowCount(0); self.table.setRowCount(len(data_list))
        for i, (name, count) in enumerate(data_list):
            chk = QTableWidgetItem(); chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled); chk.setCheckState(Qt.CheckState.Unchecked)
            self.table.setItem(i, 0, chk); self.table.setItem(i, 1, QTableWidgetItem(str(name)))
            cnt = QTableWidgetItem(str(count)); cnt.setTextAlignment(Qt.AlignmentFlag.AlignCenter); self.table.setItem(i, 2, cnt)

    def filter_list(self, text):
        for i in range(self.table.rowCount()): self.table.setRowHidden(i, not (text.lower() in self.table.item(i, 1).text().lower()))

    def toggle_all(self, state):
        st = Qt.CheckState.Checked if state else Qt.CheckState.Unchecked
        for i in range(self.table.rowCount()):
            if not self.table.isRowHidden(i): self.table.item(i, 0).setCheckState(st)

    def draw_selected_graph(self):
        selected_nodes = []
        for i in range(self.table.rowCount()):
            if self.table.item(i, 0).checkState() == Qt.CheckState.Checked:
                raw_name = self.table.item(i, 1).text()
                count = int(self.table.item(i, 2).text())
                selected_nodes.append((raw_name, count))

        if not selected_nodes: ModernDialog.show_warning(self, "Seçim Yapmadınız", "Lütfen listeden en az bir kutucuğu işaretleyin."); return
        self.generate_network_html(self.current_center_name, selected_nodes)

    def generate_network_html(self, center_node_name, connected_nodes):
        """Vis.js HTML kodunu üretir (PROFESYONEL STİL)."""

        nodes_js = f"{{id: 0, label: '{center_node_name}', color: '#e74c3c', size: 50, shape: 'box', font: {{size: 16, face: 'Segoe UI', color: 'white', multi: 'html'}}, margin: 10}},"
        edges_js = ""

        colors = ['#3498db', '#f1c40f', '#9b59b6', '#2ecc71', '#e67e22', '#1abc9c']

        for i, (name, count) in enumerate(connected_nodes, 1):
            size = min(max(30, count * 2), 60)

            label = str(name).replace("'", "").replace("\n", "\\n")

            color = colors[i % len(colors)]

            nodes_js += f"{{id: {i}, label: '{label}', value: {count}, color: '{color}', size: {size}, shape: 'box', font: {{color: 'white', face: 'Segoe UI', size: 14}}, margin: 10}},"

            edges_js += f"{{from: 0, to: {i}, width: 2, color: {{color:'#bdc3c7'}}, label: '{count} Kayıt', font: {{align: 'middle', background: 'white', face: 'Segoe UI'}} }},"

        js_library = ""
        local_js_path = os.path.join(APP_DIR, "assets", "vis-network.min.js")

        if os.path.exists(local_js_path):
            try:
                with open(local_js_path, "r", encoding="utf-8") as f:
                    js_library = f"<script>{f.read()}</script>"
            except: pass

        if not js_library:
            js_library = '<script type="text/javascript" src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>'

        html_content = f"""
        <!DOCTYPE html><html><head>{js_library}
            <style> body {{ margin: 0; padding: 0; overflow: hidden; }} #mynetwork {{ width: 100%; height: 100vh; background-color: #ffffff; }} </style>
        </head><body><div id="mynetwork"></div>
            <script type="text/javascript">
                var nodes = new vis.DataSet([{nodes_js}]);
                var edges = new vis.DataSet([{edges_js}]);
                var container = document.getElementById('mynetwork');
                var data = {{ nodes: nodes, edges: edges }};
                
                var options = {{
                    nodes: {{ borderWidth: 2, shadow: true, font: {{ face: 'Segoe UI' }} }},
                    edges: {{ smooth: {{ type: 'continuous' }} }},
                    physics: {{
                        enabled: true,
                        solver: 'repulsion', 
                        repulsion: {{ nodeDistance: 350, springLength: 300, damping: 0.09 }}, // Geniş yayılım
                        stabilization: {{ enabled: true, iterations: 1000 }}
                    }},
                    interaction: {{ navigationButtons: false, zoomView: true, dragView: true }}
                }};
                
                var network = new vis.Network(container, data, options);
                network.once("stabilizationIterationsDone", function() {{
                    network.fit({{ animation: {{ duration: 1000, easingFunction: 'easeInOutQuad' }} }});
                }});
            </script>
        </body></html>
        """
        self.browser.setHtml(html_content)


class MapPreviewDialog(WatermarkDialogMixin, QDialog):
    """
    İki konumu (A ve B noktası) gösteren harita.
    [GÜNCELLENDİ]: Aynı konumda çakışmayı önleyen (Jitter) mantığı eklendi.
    """
    def __init__(self, parent, project_id, gsm_no, target_baz_text, counter_baz_text, bubble_info=None, label1="Başlangıç", label2="Bitiş"):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.project_id = project_id
        self.gsm_no = gsm_no

        self.setWindowTitle("Konum ve Güzergah Analizi")
        self.resize(1100, 750)

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(5, 5, 5, 5)
        self.layout.setSpacing(5)

        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame { background-color: #e3f2fd; border: 1px solid #90caf9; border-radius: 6px; padding: 2px; }
        """)
        h_info = QHBoxLayout(info_frame)
        h_info.setContentsMargins(4, 2, 4, 2); h_info.setSpacing(10)

        lbl_icon = QLabel("🗺️"); lbl_icon.setStyleSheet("font-size: 20px; border: none;")
        h_info.addWidget(lbl_icon, 0, Qt.AlignmentFlag.AlignTop)

        info_html = (
            f"<div style='font-size:13px; color:#0d47a1; line-height:1.3; font-family:Segoe UI;'>"
            f"🔵 <b>{label1}:</b> {target_baz_text}<br>"
            f"🔴 <b>{label2}:</b> {counter_baz_text}<br>"
            f"<span style='color:#e65100; font-size:11px; font-style:italic;'>(⭐ Yıldızlar: Kayıtlı Özel Konumlar)</span>"
            f"</div>"
        )
        lbl_text = QLabel(info_html); lbl_text.setStyleSheet("border: none;"); lbl_text.setWordWrap(True)
        h_info.addWidget(lbl_text, 1)

        self.layout.addWidget(info_frame, 0)

        self.browser = EvidenceWebEngineView()
        self.browser.setStyleSheet("border: 1px solid #bdc3c7;")
        self.layout.addWidget(self.browser, 1)

        self.draw_dual_map(target_baz_text, counter_baz_text, bubble_info, label1, label2)

    def extract_coords(self, text):
        """Kütüphane destekli koordinat bulucu."""
        if not text: return None
        text_str = str(text).strip()

        coords = re.findall(r"(\d{2}\.\d{4,})", text_str)
        if len(coords) >= 2:
            try:
                v1, v2 = float(coords[-2]), float(coords[-1])
                if 35 < v1 < 43 and 25 < v2 < 46: return [v1, v2]
                elif 35 < v2 < 43 and 25 < v1 < 46: return [v2, v1]
                else: return [v1, v2]
            except: pass

        try:
            with DB() as conn:
                cell_id = None
                match = re.search(r'\((\d{4,})\)', text_str)
                if match: cell_id = match.group(1)
                else:
                    nums = re.findall(r'\d+', text_str)
                    cands = [n for n in nums if len(n) > 3]
                    if cands: cell_id = cands[0]

                if cell_id:
                    row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE CellID=?", (cell_id,)).fetchone()
                    if row: return [row[0], row[1]]

                row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE BazAdi=?", (text_str,)).fetchone()
                if row: return [row[0], row[1]]
        except: pass
        return None

    def draw_dual_map(self, target_text, counter_text, bubble_info=None, lbl1="Başlangıç", lbl2="Bitiş"):
        import folium
        from folium.features import DivIcon
        pt1 = self.extract_coords(target_text) # Mavi
        pt2 = self.extract_coords(counter_text) # Kırmızı

        center = [39.0, 35.0]
        zoom = 6
        if pt1 and pt2:
            center = [(pt1[0] + pt2[0])/2, (pt1[1] + pt2[1])/2]; zoom = 10
        elif pt1: center = pt1; zoom = 13
        elif pt2: center = pt2; zoom = 13

        is_same_location = False
        true_location = None

        if pt1 and pt2:
            if abs(pt1[0] - pt2[0]) < 0.0001 and abs(pt1[1] - pt2[1]) < 0.0001:
                is_same_location = True
                true_location = list(pt1)

                offset = 0.0003
                pt1 = [pt1[0], pt1[1] - offset]
                pt2 = [pt2[0], pt2[1] + offset]

        def check_net():
            try: socket.create_connection(("8.8.8.8", 53), timeout=1.0); return True
            except: return False
        is_online = check_net(); has_local = os.path.exists(os.path.join(APP_DIR, "turkey.mbtiles"))
        tile_url = "OpenStreetMap" if is_online else ("http://localhost:8080/{z}/{x}/{y}.png" if has_local else "OpenStreetMap")
        attr_info = "OpenStreetMap" if is_online else "Offline"

        m = folium.Map(location=center, zoom_start=zoom, tiles=tile_url, attr=attr_info)
        _enable_measure_and_balloons(m)
        if is_same_location and true_location:
            folium.Circle(
                location=true_location,
                radius=40, color="#9b59b6", fill=True, fill_opacity=0.3,
                popup="<b>ORTAK KONUM</b><br>Her iki taraf da bu baz istasyonunda.",
                tooltip="Gerçek Baz İstasyonu"
            ).add_to(m)
            folium.PolyLine(locations=[pt1, pt2], color="#9b59b6", weight=2, dash_array='5, 5', opacity=0.5).add_to(m)

        all_points = []

        if pt1:
            folium.Marker(pt1, popup=f"<b>{lbl1}</b><br>{target_text}", icon=folium.Icon(color="blue", icon="play", prefix='fa'), tooltip=lbl1).add_to(m)
            all_points.append(pt1)

        if pt2:
            folium.Marker(pt2, popup=f"<b>{lbl2}</b><br>{counter_text}", icon=folium.Icon(color="red", icon="stop", prefix='fa'), tooltip=lbl2).add_to(m)
            all_points.append(pt2)

        if pt1 and pt2 and not is_same_location:
            folium.PolyLine(locations=[pt1, pt2], color="black", weight=3, dash_array='10, 10', opacity=0.7).add_to(m)
            if bubble_info:
                mid_lat = (pt1[0] + pt2[0]) / 2; mid_lon = (pt1[1] + pt2[1]) / 2
                bubble_html = f"<div style='background-color:white; border:2px solid #e74c3c; border-radius:15px; padding:3px 8px; font-family:Segoe UI; font-size:11px; font-weight:bold; color:#333; text-align:center; box-shadow:2px 2px 4px rgba(0,0,0,0.3);'>{bubble_info}</div>"
                folium.Marker(location=[mid_lat, mid_lon], icon=DivIcon(icon_size=(150, 36), icon_anchor=(75, 18), html=bubble_html)).add_to(m)

        fg_custom = folium.FeatureGroup(name="⭐ Özel Konumlar")
        try:
            with DB() as conn:
                rows = conn.execute("SELECT Lat, Lon, Label FROM ozel_konumlar WHERE ProjeID=?", (self.project_id,)).fetchall()
                for r in rows:
                    lat, lon, label = r
                    folium.Marker([lat, lon], popup=label, icon=folium.Icon(color="orange", icon="star", prefix='fa'), tooltip=label).add_to(fg_custom)
        except: pass
        fg_custom.add_to(m)

        folium.LayerControl().add_to(m)
        if all_points: m.fit_bounds(all_points, padding=(80, 80))

        data = io.BytesIO()
        m.save(data, close_file=False)
        self.browser.setHtml(data.getvalue().decode())


class StalkingAnalysisDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, owner_gsm, start_dt, end_dt):
        super().__init__(parent)

        self.project_id = project_id
        self.owner_gsm = owner_gsm
        self.start_dt = start_dt
        self.end_dt = end_dt

        self.setWindowTitle(f"Taciz ve Israrlı Takip Analizi - {owner_gsm}")
        self.resize(1200, 700)

        central = QWidget(self)
        central_layout = QVBoxLayout(central); central_layout.setContentsMargins(0, 0, 0, 0); central_layout.setSpacing(8)
        base_dir = APP_DIR if '__file__' in globals() else os.getcwd()
        logo_path = os.path.join(base_dir, "assets", "bg_logo.png")
        if not os.path.exists(logo_path): logo_path = os.path.join(base_dir, "assets", "logo.png")
        self._watermark = WatermarkBackground(logo_path=logo_path, opacity=0.025, scale=1.20, parent=central)
        stack = QStackedLayout(self); stack.setStackingMode(QStackedLayout.StackingMode.StackAll)
        stack.addWidget(central); stack.addWidget(self._watermark); self._watermark.raise_()

        info_frame = QFrame()
        info_frame.setStyleSheet("background-color: #ffebee; border: 1px solid #ef9a9a; border-radius: 8px; padding: 10px;")
        h_info = QHBoxLayout(info_frame)

        lbl_icon = QLabel("🚫"); lbl_icon.setStyleSheet("font-size: 30px;")
        info_text = (
            "<div style='font-size:13px; color:#c62828;'>"
            "<b>ANALİZ MANTIĞI:</b><br>"
            "🔴 <b>Yüksek Risk:</b> Karşı taraf HİÇ aramamış veya aramaların %80'i reddedilmiş (10 sn altı).<br>"
            "🟠 <b>Orta Risk:</b> Giden aramalar, Gelen aramaların 5 katından fazla.<br>"
            "<i>Bu analiz, 'Tek Yönlü Israrlı Arama' şüphesi taşıyan kayıtları filtreler.</i>"
            "</div>"
        )
        h_info.addWidget(lbl_icon); h_info.addWidget(QLabel(info_text)); h_info.addStretch()
        central_layout.addWidget(info_frame)

        headers = ["Risk Durumu", "Hedef Numara", "Kişi Adı", "Giden", "Gelen", "Cevapsız/Kısa (<10sn)", "Toplam Süre"]
        self.table = GenericDatabaseTable(headers, chart_mode='none')
        t = self.table.table
        t.setSortingEnabled(True)
        t.doubleClicked.connect(self.open_details)

        h = t.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        t.setColumnWidth(1, 120); t.setColumnWidth(2, 150)
        h.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)

        central_layout.addWidget(self.table, 1)

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; padding:10px; border-radius:4px;")
        btn_close.clicked.connect(self.accept)
        central_layout.addWidget(btn_close)

        QTimer.singleShot(100, self.run_analysis)

    def run_analysis(self):
        try:
            def clean_gsm(n): return re.sub(r'\D', '', str(n))[-10:]
            short_owner = clean_gsm(self.owner_gsm)

            s_str = self.start_dt.toString("yyyy-MM-dd HH:mm:ss")
            e_str = self.end_dt.toString("yyyy-MM-dd HH:mm:ss")

            sql = f"""
                SELECT 
                    DIGER_NUMARA,
                    MAX(DIGER_ISIM) as Isim,
                    
                    -- GİDEN (Biz Aradık)
                    SUM(CASE 
                        WHEN TIP LIKE '%Aradı%' OR TIP LIKE '%Giden%' OR TIP LIKE '%Gönder%' OR TIP LIKE '%Çıkış%' OR TIP LIKE '%Out%' 
                        THEN 1 ELSE 0 
                    END) as Giden,
                    
                    -- GELEN (O Aradı)
                    SUM(CASE 
                        WHEN TIP LIKE '%Arandı%' OR TIP LIKE '%Gelen%' OR TIP LIKE '%Aldı%' OR TIP LIKE '%Giriş%' OR TIP LIKE '%In%' 
                        THEN 1 ELSE 0 
                    END) as Gelen,
                    
                    -- KISA/REDDEDİLEN (Sadece Giden ve <10sn)
                    SUM(CASE 
                        WHEN (TIP LIKE '%Aradı%' OR TIP LIKE '%Giden%' OR TIP LIKE '%Gönder%') 
                             AND CAST(REPLACE(REPLACE(SURE, ' sn', ''), ' sec', '') as INTEGER) < 10 
                        THEN 1 ELSE 0 
                    END) as Kisa,
                    
                    SUM(CAST(REPLACE(REPLACE(SURE, ' sn', ''), ' sec', '') as INTEGER)) as ToplamSure
                    
                FROM hts_gsm
                WHERE ProjeID=? 
                  AND (substr(TARIH, 7, 4) || '-' || substr(TARIH, 4, 2) || '-' || substr(TARIH, 1, 2) || substr(TARIH, 11)) BETWEEN ? AND ?
                  
                  -- Sadece analiz edilen numaranın kayıtlarını al (Dosya karışıklığını önlemek için)
                  AND substr(replace(replace(replace(NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ?
                  
                GROUP BY DIGER_NUMARA
                HAVING Giden > 5 -- Gürültü önlemek için en az 5 arama
                ORDER BY Giden DESC
            """

            with DB() as conn:
                rows = conn.execute(sql, (self.project_id, s_str, e_str, short_owner)).fetchall()

            analyzed_data = []

            for row in rows:
                num, name, giden, gelen, kisa, total_sure = row

                risk_label = "🟢 Normal"
                risk_val = 0

                denom_gelen = gelen if gelen > 0 else 1
                denom_giden = giden if giden > 0 else 1

                reject_ratio = (kisa / denom_giden) * 100
                imbalance = giden / denom_gelen

                if (gelen == 0 and giden > 5) or (reject_ratio > 80 and giden > 5):
                    risk_label = "🔴 YÜKSEK RİSK"
                    risk_val = 3

                elif imbalance > 5 or reject_ratio > 50:
                    risk_label = "🟡 Şüpheli"
                    risk_val = 2

                if risk_val > 0:
                    m, s = divmod(total_sure, 60)
                    sure_fmt = f"{m} dk {s} sn"

                    analyzed_data.append([
                        risk_label, num, name if name else "Bilinmiyor",
                        giden, gelen, f"{kisa} ({int(reject_ratio)}%)", sure_fmt
                    ])

            analyzed_data.sort(key=lambda x: (x[0], x[3]), reverse=True)

            if not analyzed_data:
                ModernDialog.show_info(self, "Temiz", "Sesli görüşmeler içinde şüpheli/tek yönlü yoğun arama tespit edilemedi.")

            self.table.set_data(analyzed_data)

        except Exception as e:
            print(f"Analiz Hatası: {e}")
            ModernDialog.show_error(self, "Hata", str(e))

    def open_details(self, index):
        try:
            row = index.row()
            target_num = self.table.proxy_model.index(row, 1).data()
            target_name = self.table.proxy_model.index(row, 2).data()

            if target_num:
                dlg = InteractionDetailDialog(
                    self, self.project_id, self.owner_gsm, target_num, target_name,
                    self.start_dt, self.end_dt
                )
                dlg.exec()
        except Exception as e:
            print(f"Detay hatası: {e}")


class ReciprocalDetailDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, target_gsm, other_gsm,
                 other_name="", target_date="", main_target_baz=""):
        super().__init__(parent)

        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.project_id = project_id
        self.target_gsm = target_gsm
        self.other_gsm = other_gsm
        self.other_name = other_name
        self.target_date = target_date
        self.main_target_baz = main_target_baz

        self.setWindowTitle(f"Karşı Taraf Verisi: {other_gsm}")
        self.resize(1300, 800)

        self.layout = QVBoxLayout(self)

        info_frame = QFrame()
        info_frame.setStyleSheet(
            "background-color: #e3f2fd; border-radius: 8px; padding: 10px; "
            "border: 1px solid #90caf9;"
        )
        h_info = QHBoxLayout(info_frame)

        display_date = target_date[:16] if len(target_date) >= 16 else target_date

        lbl_icon = QLabel("📖")
        lbl_icon.setStyleSheet("font-size: 32px; border:none;")

        info_text = (
            f"<div style='font-size:14px; color:#1565c0; font-family: Segoe UI;'>"
            f"<b>📅 Tarih:</b> {display_date}<br>"
            f"<b>📱 Hedef Numara:</b> {target_gsm}<br>"
            f"<b>👤 Karşı Numara:</b> {other_gsm} "
            f"<span style='color:#d35400;'>({other_name})</span><br>"
            f"<i style='color:#c0392b;'>Konum haritası için satıra çift tıklayın.</i>"
            f"</div>"
        )

        h_info.addWidget(lbl_icon)
        h_info.addWidget(QLabel(info_text))
        h_info.addStretch()
        self.layout.addWidget(info_frame)

        headers = [
            "Tarih/Saat", "İşlem Yönü", "İletişim Türü",
            "Süre", "KARŞI IMEI", "KARŞI BAZ"
        ]
        self.table = GenericDatabaseTable(headers, chart_mode='none')
        t = self.table.table

        t.doubleClicked.connect(self.open_map_location)
        t.setWordWrap(True)
        t.setTextElideMode(Qt.TextElideMode.ElideNone)

        # satır yüksekliği içeriğe göre büyüsün
        t.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        # BAZ kolonları satır içine sığdırılsın (wrap için)
        h = t.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Tarih/Saat
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Sahip IMEI
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Karşı IMEI
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)          # Sahip BAZ
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)          # Karşı BAZ

        h = t.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Tarih
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Tür
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Süre
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # IMEI
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        t.setColumnWidth(1, 250)  # İşlem yönü
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)           # Baz

        t.setStyleSheet("""
            QTableWidget, QTableView {
                background-color: rgba(255,255,255,190);
                alternate-background-color: rgba(245,245,245,170);
                gridline-color: rgba(0,0,0,40);
            }
            QTableWidget::item, QTableView::item {
                background-color: transparent;
            }
            QHeaderView::section {
                background-color: rgba(240,240,240,230);
                font-weight: bold;
            }
        """)
        t.viewport().setAutoFillBackground(False)

        self.layout.addWidget(self.table, 1)

        # ✅ YENİ: Karşılıklı temasları açan buton (Kapat'ın üstünde)
        btn_mutual = QPushButton("🔁 Karşılıklı Temasları Göster")
        btn_mutual.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_mutual.setFixedHeight(40)
        btn_mutual.setStyleSheet(
            "background-color:#2c3e50; color:white; font-weight:bold; border-radius:6px;"
        )
        btn_mutual.clicked.connect(self.open_mutual_contacts_match)
        self.layout.addWidget(btn_mutual)

        btn_close = QPushButton("Kapat")
        btn_close.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_close.setFixedHeight(40)
        btn_close.setStyleSheet(
            "background-color:#c0392b; color:white; font-weight:bold; border-radius:6px;"
        )
        btn_close.clicked.connect(self.close)
        self.layout.addWidget(btn_close)

        QTimer.singleShot(
            50,
            lambda: self.load_counterparty_data(
                self.project_id,
                self.target_gsm,
                self.other_gsm,
                self.target_date
            )
        )

    def _open_window_nonmodal(self, w: QDialog):
        """CrossMatchDialog'daki open_window_safe mantığıyla NON-MODAL aç."""
        if not hasattr(self, "_open_windows"):
            self._open_windows = []

        self._open_windows.append(w)

        w.setModal(False)
        w.setWindowModality(Qt.WindowModality.NonModal)

        w.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint
        )

        w.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)

        def _cleanup(*_):
            try:
                if w in self._open_windows:
                    self._open_windows.remove(w)
            except Exception:
                pass

            # parent tekrar aktif/fokus gelsin
            QTimer.singleShot(0, self.raise_)
            QTimer.singleShot(0, self.activateWindow)

        if hasattr(w, "finished"):
            w.finished.connect(_cleanup)
        w.destroyed.connect(_cleanup)

        w.show()
        w.raise_()
        w.activateWindow()

    def open_mutual_contacts_match(self):
        """
        Analysis ekranındaki dt_start/dt_end varsa onu kullanıp
        iki hat arasındaki TÜM temasları karşılıklı eşleştirerek gösterir.
        """
        try:
            parent = self.parent()

            if not parent or not hasattr(parent, "dt_start") or not hasattr(parent, "dt_end"):
                ModernDialog.show_warning(
                    self,
                    "Tarih Aralığı Yok",
                    "Karşılıklı temas detayı için analiz ekranındaki başlangıç/bitiş tarihleri bulunamadı."
                )
                return

            start_dt = parent.dt_start.dateTime()
            end_dt = parent.dt_end.dateTime()

            dlg = MutualContactsMatchDialog(
                self,
                self.project_id,
                self.target_gsm,     # hat sahibi (inceleme yapılan)
                self.other_gsm,      # karşı taraf
                self.other_name,
                start_dt,
                end_dt
            )
            dlg.setWindowFlags(dlg.windowFlags() | Qt.WindowType.WindowStaysOnTopHint)
            self._open_window_nonmodal(dlg)
        except Exception as e:
            print(f"Karşılıklı Temas Açma Hatası: {e}")
            ModernDialog.show_error(self, "Hata", f"Karşılıklı temas ekranı açılamadı: {e}")

    def open_map_location(self, index):
        try:
            row = index.row()
            counter_baz = self.table.proxy_model.index(row, 5).data()
            target_baz = self.main_target_baz
            if not counter_baz or len(str(counter_baz)) < 10:
                ModernDialog.show_warning(self, "Veri Yok", "Karşı tarafın baz bilgisi yok.")
                return

            lbl1 = f"{self.target_gsm} (Hat Sahibi)"
            lbl2 = f"{self.other_gsm} (Karşı Taraf)"

            map_dlg = MapPreviewDialog(
                self,
                self.project_id,
                self.target_gsm,
                str(target_baz),
                str(counter_baz),
                label1=lbl1,
                label2=lbl2
            )
            map_dlg.exec()
        except Exception as e:
            print(f"Hata: {e}")

    def load_counterparty_data(self, pid, my_gsm, other_gsm, full_date):
        try:
            with DB() as conn:
                cur = conn.cursor()
                date_filter = full_date[:16] + '%' if len(full_date) > 16 else full_date + '%'

                def get_last_10(n):
                    s = re.sub(r'\D', '', str(n).strip())
                    return s[-10:] if len(s) >= 10 else s

                my_short = get_last_10(my_gsm)
                other_short = get_last_10(other_gsm)

                sql = """SELECT TARIH,
                                CASE
                                    WHEN TIP LIKE '%Arand%' OR TIP LIKE '%Gelen%' OR TIP LIKE '%Mesaj aldı%'
                                    THEN NUMARA || ' -> ' || DIGER_NUMARA
                                    ELSE DIGER_NUMARA || ' -> ' || NUMARA
                                END as Yon,
                                TIP, SURE, IMEI, BAZ
                         FROM hts_gsm
                         WHERE ProjeID = ?
                           AND substr(replace(replace(replace(NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ?
                           AND substr(replace(replace(replace(DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ?
                           AND TARIH LIKE ?
                         ORDER BY TARIH ASC
                         LIMIT 10001"""
                rows = cur.execute(sql, (pid, other_short, my_short, date_filter)).fetchall()
                if len(rows) > 10000:
                    ModernDialog.show_warning(self, "Limit", "10.000 kayıt sınırı aşıldı.")
                    rows = rows[:10000]
                self.table.set_data([list(r) for r in rows])
        except Exception as e:
            print(f"Veri Hatası: {e}")


class MutualContactsMatchDialog(WatermarkDialogMixin, QDialog):
    """
    İki GSM arasındaki temasları karşılıklı eşleştirir.

    Gösterilecek kolonlar (istenen):
      - Tarih/Saat
      - Hat Sahibi IMEI
      - Hat Sahibi BAZ
      - Tip (ikon)
      - Karşı IMEI
      - Karşı BAZ

    Eşleştirme: ±3 saniye tolerans (InteractionDetailDialog ile aynı mantık).
    """

    def __init__(self, parent, project_id, owner_gsm, other_gsm, other_name, start_dt, end_dt):
        super().__init__(parent)
        self.init_watermark(opacity=0.03, scale_ratio=0.90)

        self.project_id = project_id
        self.owner_gsm = owner_gsm
        self.other_gsm = other_gsm
        self.other_name = other_name

        self.setWindowTitle(f"Karşılıklı Temas Eşleştirme: {owner_gsm} ↔ {other_gsm}")
        self.resize(1400, 800)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # Üst Kart
        info_frame = QFrame()
        info_frame.setStyleSheet(
            "background-color: #e3f2fd; border: 1px solid #90caf9; border-radius: 8px; padding: 10px;"
        )
        h_info = QHBoxLayout(info_frame)

        s_str = start_dt.toString("dd.MM.yyyy HH:mm:ss")
        e_str = end_dt.toString("dd.MM.yyyy HH:mm:ss")

        lbl_icon = QLabel("📇")
        lbl_icon.setStyleSheet("font-size: 30px;")

        info_text = (
            f"<div style='font-size:14px; color:#0d47a1; font-family: Segoe UI;'>"
            f"<b>Tarih Aralığı:</b> {s_str} - {e_str}<br>"
            f"<b>Hat Sahibi:</b> {owner_gsm}<br>"
            f"<b>Karşı Numara:</b> {other_gsm} <span style='color:#ef6c00;'><b>({other_name})</b></span><br>"
            f"<span style='color:#c62828;'>Haritada görmek için satıra çift tıklayın.</span>"
            f"</div>"
        )
        lbl_info = QLabel(info_text)
        lbl_info.setTextFormat(Qt.TextFormat.RichText)
        lbl_info.setWordWrap(True)

        h_info.addWidget(lbl_icon)
        h_info.addWidget(lbl_info, 1)

        layout.addWidget(info_frame)

        headers = [
            "Tarih/Saat",
            "Hat Sahibi IMEI",
            "Hat Sahibi BAZ",
            "Tip",
            "Karşı IMEI",
            "Karşı BAZ",
        ]
        self.table_widget = GenericDatabaseTable(headers, chart_mode='none')
        t = self.table_widget.table

        # ✅ Wrap aktif + elide kapalı
        t.setWordWrap(True)
        try:
            t.setTextElideMode(Qt.TextElideMode.ElideNone)
        except Exception:
            pass

        # ✅ Sadece ilgili kolonlara delegate
        self._wrap_delegate = WrapTextDelegate(t)
        self._tip_delegate = TipIconDelegate(t)
        t.setItemDelegateForColumn(2, self._wrap_delegate)  # Sahip BAZ
        t.setItemDelegateForColumn(5, self._wrap_delegate)  # Karşı BAZ
        t.setItemDelegateForColumn(3, self._tip_delegate)   # Tip ikon

        t.doubleClicked.connect(self.open_map_location)

        h = t.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Tarih
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Sahip IMEI
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)           # Sahip BAZ (wrap)
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)             # Tip
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Karşı IMEI
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)           # Karşı BAZ (wrap)
        t.setColumnWidth(3, 60)

        layout.addWidget(self.table_widget, 1)

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet(
            "background-color:#c0392b; color:white; font-weight:bold; padding:10px; border-radius:6px;"
        )
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)

        self.loader = LoadingOverlay(self)
        self.loader.start("Karşılıklı Kayıtlar Eşleştiriliyor...")

        QTimer.singleShot(100, lambda: self.load_data(start_dt, end_dt))

    def _open_map_nonmodal(self, map_dlg: QDialog):
        """
        Haritayı kesinlikle NON-MODAL açar.
        - Uygulamayı kilitlemez
        - Map kapanınca parent kapanmaz
        - GC map dialogu öldürmesin diye referans tutulur
        """
        if not hasattr(self, "_open_maps"):
            self._open_maps = []

        # ✅ Parent-child ilişkisinde modal kilit oluşmasın
        map_dlg.setModal(False)
        map_dlg.setWindowModality(Qt.WindowModality.NonModal)

        # ✅ Top-level pencere gibi davransın (parent kapanma / minimize zinciri sorunlarını azaltır)
        map_dlg.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint
        )

        # ✅ Kapanınca temizlensin
        map_dlg.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)

        # ✅ Referans tut (GC öldürmesin)
        self._open_maps.append(map_dlg)

        def _cleanup(*_):
            try:
                if map_dlg in self._open_maps:
                    self._open_maps.remove(map_dlg)
            except Exception:
                pass

            # Map kapanınca bu pencereyi tekrar öne al (fokus değil; sadece görünür kalsın diye)
            QTimer.singleShot(0, self.raise_)
            QTimer.singleShot(0, self.activateWindow)

        # finished varsa bağla (QDialog), yoksa destroyed yeter
        if hasattr(map_dlg, "finished"):
            map_dlg.finished.connect(_cleanup)
        map_dlg.destroyed.connect(_cleanup)

        map_dlg.show()
        map_dlg.raise_()
        map_dlg.activateWindow()


    def load_data(self, q_start, q_end):
        try:
            s_date = q_start.toString("yyyy-MM-dd HH:mm:ss")
            e_date = q_end.toString("yyyy-MM-dd HH:mm:ss")

            def clean_gsm(val):
                d = re.sub(r'\D', '', str(val).strip())
                return d[-10:] if len(d) >= 10 else d

            short_owner = clean_gsm(self.owner_gsm)
            short_other = clean_gsm(self.other_gsm)

            with DB() as conn:
                cur = conn.cursor()

                date_filter = (
                    " AND (substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11)) "
                    " BETWEEN ? AND ? "
                )

                sql = f"""
                    SELECT
                        t1.TARIH,
                        t1.IMEI as SahipIMEI,
                        t1.BAZ  as SahipBaz,

                        CASE
                            WHEN (t1.TIP LIKE '%Mesaj%' OR t1.TIP LIKE '%SMS%')
                                 AND (t1.TIP LIKE '%att%' OR t1.TIP LIKE '%gönder%') THEN 'SMS_OUT'
                            WHEN (t1.TIP LIKE '%Mesaj%' OR t1.TIP LIKE '%SMS%')
                                 AND (t1.TIP LIKE '%ald%' OR t1.TIP LIKE '%al%') THEN 'SMS_IN'
                            WHEN (t1.TIP LIKE '%Arad%' OR t1.TIP LIKE '%Aradı%') THEN 'CALL_OUT'
                            WHEN (t1.TIP LIKE '%Arand%' OR t1.TIP LIKE '%Arandı%') THEN 'CALL_IN'
                            ELSE 'OTHER'
                        END as TipKod,

                        t2.IMEI as KarsiIMEI,
                        t2.BAZ  as KarsiBaz
                    FROM hts_gsm t1
                    LEFT JOIN hts_gsm t2 ON
                        t2.ProjeID = t1.ProjeID AND

                        substr(replace(replace(replace(t2.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) =
                        substr(replace(replace(replace(t1.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND

                        substr(replace(replace(replace(t2.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) =
                        substr(replace(replace(replace(t1.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND

                        datetime(
                            substr(t2.TARIH, 7, 4) || '-' || substr(t2.TARIH, 4, 2) || '-' || substr(t2.TARIH, 1, 2) || substr(t2.TARIH, 11)
                        ) BETWEEN
                            datetime(
                                substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11),
                                '-3 seconds'
                            )
                            AND
                            datetime(
                                substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11),
                                '+3 seconds'
                            )

                    WHERE
                        t1.ProjeID = ? AND
                        substr(replace(replace(replace(t1.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ? AND
                        substr(replace(replace(replace(t1.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ?
                        {date_filter}
                    ORDER BY
                        datetime(
                            substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11)
                        ) DESC
                    LIMIT 10001
                """

                params = [self.project_id, short_owner, short_other, s_date, e_date]
                rows = cur.execute(sql, params).fetchall()

            data = []
            for tarih, sahip_imei, sahip_baz, tip_kod, karsi_imei, karsi_baz in rows:
                data.append([
                    tarih or "",
                    sahip_imei or "",
                    sahip_baz or "",
                    tip_kod or "OTHER",
                    karsi_imei or "",
                    karsi_baz or ""
                ])

            if len(data) > 10000:
                ModernDialog.show_warning(self, "Veri Limiti", "Performans için ilk 10.000 kayıt gösteriliyor.")
                data = data[:10000]

            self.table_widget.set_data(data)

            # ✅ Wrap sonrası satır yüksekliğini hesaplat
            self.table_widget.table.resizeRowsToContents()

        except Exception as e:
            print(f"Karşılıklı Eşleştirme Hatası: {e}")
            ModernDialog.show_error(self, "Hata", str(e))
        finally:
            if hasattr(self, "loader"):
                self.loader.stop()

    def open_map_location(self, index):
        try:
            row = index.row()
            # ✅ Yeni kolon düzeni: sahip_baz=2, karsi_baz=5 (Tablo yapınıza göre)
            sahip_baz = self.table_widget.proxy_model.index(row, 2).data()
            karsi_baz = self.table_widget.proxy_model.index(row, 5).data()

            has_sahip = sahip_baz and len(str(sahip_baz)) > 5 and "---" not in str(sahip_baz)
            has_karsi = karsi_baz and len(str(karsi_baz)) > 5 and "---" not in str(karsi_baz)

            if not has_sahip and not has_karsi:
                ModernDialog.show_warning(self, "Konum Yok", "Bu kayıtta her iki taraf için de geçerli baz bilgisi yok.")
                return

            # Tek taraflı durumlarda bile MapPreviewDialog beklediği argümanları alsın diye boş string ver
            if not has_sahip: sahip_baz = ""
            if not has_karsi: karsi_baz = ""

            lbl1 = f"{self.owner_gsm} (Hat Sahibi)"
            lbl2 = f"{self.other_gsm} (Karşı Taraf)"

            map_dlg = MapPreviewDialog(
                self, self.project_id, self.owner_gsm,
                str(sahip_baz), str(karsi_baz),
                label1=lbl1, label2=lbl2
            )

            map_dlg.exec()

        except Exception as e:
            print(f"Harita Hatası: {e}")
            ModernDialog.show_error(self, "Hata", str(e))


class InteractionDetailDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, owner_gsm, target_gsm, target_name, start_dt, end_dt):
        super().__init__(parent)
        self.project_id = project_id
        self.owner_gsm = owner_gsm
        self.target_gsm = target_gsm

        self.setWindowTitle(f"İletişim Detayı: {target_gsm} - {target_name}")
        self.resize(1350, 750)

        central = QWidget(self)
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(0, 0, 0, 0)
        central_layout.setSpacing(0)

        base_dir = os.path.dirname(__file__) if '__file__' in globals() else os.getcwd()
        logo_path = os.path.join(base_dir, "assets", "bg_logo.png")
        if not os.path.exists(logo_path):
            logo_path = os.path.join(base_dir, "assets", "logo.png")

        self._watermark = WatermarkBackground(logo_path=logo_path, opacity=0.03, scale=1.20, parent=central)

        stack = QStackedLayout(self)
        stack.setStackingMode(QStackedLayout.StackingMode.StackAll)
        stack.addWidget(central)
        stack.addWidget(self._watermark)
        self._watermark.raise_()

        self.layout = QVBoxLayout()
        central_layout.addLayout(self.layout)

        info_frame = QFrame()
        info_frame.setStyleSheet("background-color: #e3f2fd; border: 1px solid #90caf9; border-radius: 8px; padding: 10px;")
        h_info = QHBoxLayout(info_frame)
        s_str = start_dt.toString("dd.MM.yyyy")
        e_str = end_dt.toString("dd.MM.yyyy")

        lbl_icon = QLabel("📖")
        lbl_icon.setStyleSheet("font-size: 30px;")
        info_text = (
            f"<div style='font-size:14px; color:#0d47a1; font-family: Segoe UI;'>"
            f"👤 <b>Karşı Taraf:</b> {target_name} ({target_gsm})<br>"
            f"📱 <b>İncelenen Hat:</b> {owner_gsm}<br>"
            f"📅 <b>Analiz Aralığı:</b> {s_str} - {e_str}<br>"
            f"<i>(Çift taraflı konum analizi için satıra çift tıklayın. Karşı baz bilgisi için ±3 saniye tolerans uygulanmıştır.)</i>"
            f"</div>"
        )
        h_info.addWidget(lbl_icon)
        h_info.addWidget(QLabel(info_text))
        h_info.addStretch()
        self.layout.addWidget(info_frame)

        self.headers = [
            "Tarih/Saat",
            "İşlem Yönü",
            "Tür",
            "Süre",
            "Hat Sahibi Konumu (BAZ)",
            "Karşı Taraf Konumu (BAZ)",
            "Karşı IMEI"
        ]
        self.table_widget = GenericDatabaseTable(self.headers, chart_mode='none')
        t = self.table_widget.table
        t.setWordWrap(False)
        t.setItemDelegate(ElidedItemDelegate(t))
        t.doubleClicked.connect(self.open_map_location)

        h = t.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        t.setColumnWidth(1, 180)
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)

        self.layout.addWidget(self.table_widget)

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; padding:10px; border-radius:4px;")
        btn_close.clicked.connect(self.accept)
        self.layout.addWidget(btn_close)

        self.loader = LoadingOverlay(self)
        self.loader.start("Kayıtlar Eşleştiriliyor...")

        QTimer.singleShot(100, lambda: self.load_data_exact(start_dt, end_dt))

    def load_data_exact(self, q_start, q_end):
        try:
            s_date = q_start.toString("yyyy-MM-dd HH:mm:ss")
            e_date = q_end.toString("yyyy-MM-dd HH:mm:ss")

            def clean_gsm(val):
                s = str(val).strip()
                d = re.sub(r'\D', '', s)
                return d[-10:] if len(d) >= 10 else d

            short_owner = clean_gsm(self.owner_gsm)
            short_target = clean_gsm(self.target_gsm)

            with DB() as conn:
                cur = conn.cursor()

                date_filter = " AND (substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11)) BETWEEN ? AND ? "

                sql = f"""
                    SELECT
                        t1.TARIH,
                        t1.NUMARA,
                        t1.DIGER_NUMARA,
                        t1.TIP,
                        t1.SURE,
                        t1.BAZ as SahipBaz,
                        t2.BAZ as KarsiBaz,
                        t2.IMEI as KarsiIMEI
                    FROM hts_gsm t1
                    LEFT JOIN hts_gsm t2 ON
                        t2.ProjeID = t1.ProjeID AND

                        substr(replace(replace(replace(t2.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) =
                        substr(replace(replace(replace(t1.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND

                        substr(replace(replace(replace(t2.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) =
                        substr(replace(replace(replace(t1.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND

                        datetime(
                            substr(t2.TARIH, 7, 4) || '-' || substr(t2.TARIH, 4, 2) || '-' || substr(t2.TARIH, 1, 2) || substr(t2.TARIH, 11)
                        ) BETWEEN
                            datetime(
                                substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11),
                                '-3 seconds'
                            )
                            AND
                            datetime(
                                substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11),
                                '+3 seconds'
                            )

                    WHERE
                        t1.ProjeID = ? AND
                        substr(replace(replace(replace(t1.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ? AND
                        substr(replace(replace(replace(t1.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ?
                        {date_filter}
                    ORDER BY
                        substr(t1.TARIH, 7, 4) || substr(t1.TARIH, 4, 2) || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11) DESC
                """

                params = (self.project_id, short_owner, short_target, s_date, e_date)
                rows = cur.execute(sql, params).fetchall()

            final_data = []
            for tarih_str, numara, diger, tip, sure, sahip_baz, karsi_baz, karsi_imei in rows:
                tip_upper = str(tip).upper()
                if "ARADI" in tip_upper or "GONDER" in tip_upper or "ÇIKIŞ" in tip_upper or "OUT" in tip_upper:
                    yon_str = "Giden (->)"
                elif "ARANDI" in tip_upper or "ALDI" in tip_upper or "GİRİŞ" in tip_upper or "IN" in tip_upper:
                    yon_str = "Gelen (<-)"
                else:
                    yon_str = "Diğer"

                final_data.append([
                    tarih_str,
                    yon_str,
                    tip,
                    f"{sure}" if sure else "",
                    sahip_baz if sahip_baz else "",
                    karsi_baz if karsi_baz else "",
                    karsi_imei if karsi_imei else ""
                ])

            if len(final_data) > 10000:
                ModernDialog.show_warning(self, "Veri Limiti", "Performans için ilk 10.000 kayıt gösteriliyor.")
                final_data = final_data[:10000]

            self.table_widget.set_data(final_data)

        except Exception as e:
            print(f"Veri Yükleme Hatası: {e}")
            ModernDialog.show_error(self, "Hata", str(e))
        finally:
            if hasattr(self, "loader"):
                self.loader.stop()

    def open_map_location(self, index):
        """Çift tıklanan satırdaki İKİ konumu birden gösterir. Veri yoksa UYARIR."""
        try:
            row = index.row()

            kaynak_baz = self.table_widget.proxy_model.index(row, 4).data()
            karsi_baz = self.table_widget.proxy_model.index(row, 5).data()

            has_kaynak = kaynak_baz and len(str(kaynak_baz)) > 5 and "---" not in str(kaynak_baz)
            has_karsi = karsi_baz and len(str(karsi_baz)) > 5 and "---" not in str(karsi_baz)

            if not has_kaynak and not has_karsi:
                ModernDialog.show_warning(self, "Konum Yok", "Bu kayıtta her iki taraf için de geçerli baz istasyonu verisi bulunamadı.")
                return

            if not has_karsi and has_kaynak:
                ModernDialog.show_info(self, "Karşı Veri Eksik",
                    "Seçilen kayıtta <b>Karşı Tarafın</b> baz istasyonu verisi tespit edilemedi.<br>"
                    "<i>(Sebep: Karşı tarafın HTS dosyası yüklenmemiş olabilir veya ±3 saniye içinde eşleşen kayıt bulunamadı.)</i><br>"
                    "Haritada sadece Hat Sahibinin konumu gösterilecektir.")
                karsi_baz = ""

            lbl1 = f"{self.owner_gsm} (İncelenen)"
            lbl2 = f"{self.target_gsm} (Karşı Taraf)"

            dlg = MapPreviewDialog(
                self, self.project_id, self.owner_gsm,
                str(kaynak_baz) if has_kaynak else "",
                str(karsi_baz) if has_karsi else "",
                label1=lbl1,
                label2=lbl2
            )
            dlg.exec()

        except Exception as e:
            print(f"Map Open Error: {e}")
            ModernDialog.show_error(self, "Harita Hatası", str(e))


class LocationDetailDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, owner_gsm, target_baz, start_date=None, end_date=None):
        super().__init__(parent)

        self.setSizeGripEnabled(True)
        self.setMinimumSize(800, 650)

        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowType.WindowMinimizeButtonHint
            | Qt.WindowType.WindowMaximizeButtonHint
        )

        central = QWidget(self)
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(0, 0, 0, 0)
        central_layout.setSpacing(0)

        base_dir = os.path.dirname(__file__)
        logo_path = os.path.join(base_dir, "assets", "bg_logo.png")
        if not os.path.exists(logo_path):
            logo_path = os.path.join(base_dir, "assets", "logo.png")

        self._watermark = WatermarkBackground(
            logo_path=logo_path,
            opacity=0.03,
            scale=1.20,
            parent=central
        )

        stack = QStackedLayout(self)
        stack.setStackingMode(QStackedLayout.StackingMode.StackAll)
        stack.addWidget(central)
        stack.addWidget(self._watermark)
        self._watermark.raise_()
        self.setWindowTitle(f"Konum Analizi: {target_baz}")
        self.resize(1100, 650)

        self.layout = QVBoxLayout()
        central_layout.addLayout(self.layout)

        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #e3f2fd; 
                border: 1px solid #90caf9; 
                border-radius: 8px; 
                padding: 10px;
            }
        """)
        hl = QHBoxLayout(info_frame)

        d_info = "Tüm Zamanlar"
        if start_date and end_date:
            d_info = f"{start_date.toString('dd.MM.yyyy HH:mm')} - {end_date.toString('dd.MM.yyyy HH:mm')}"

        lbl_icon = QLabel("📡")
        lbl_icon.setStyleSheet("font-size: 32px; border:none;")

        info_text = (
            f"<div style='font-size:14px; color:#1565c0; font-family: Segoe UI;'>"
            f"<b>📍 Baz İstasyonu:</b> {target_baz}<br>"
            f"<b>📅 Tarih Aralığı:</b> {d_info}<br>"
            f"<i style='color:#555;'>Bu konumda gerçekleşen tüm GSM ve İnternet trafiği aşağıdadır.</i>"
            f"</div>"
        )

        hl.addWidget(lbl_icon)
        hl.addWidget(QLabel(info_text))
        hl.addStretch()
        self.layout.addWidget(info_frame)

        self.tabs = QTabWidget()
        self.layout.addWidget(self.tabs)

        self.gsm_table = GenericDatabaseTable(
            ["Tarih", "Görüşülen Numara", "Kişi Adı", "TC Kimlik", "Yön", "Süre"],
            enable_date_filter=False,
            chart_mode='none'
        )
        t = self.gsm_table.table
        t.setWordWrap(False)
        t.setItemDelegate(ElidedItemDelegate(t))
        h = t.horizontalHeader()
        h.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tabs.addTab(self.gsm_table, "GSM Görüşmeleri")

        self.net_table = GenericDatabaseTable(
            ["Tür", "Tarih", "Süre", "Kaynak IP", "Hedef IP", "Gönderme", "İndirme"],
            enable_date_filter=False,
            chart_mode='none'
        )
        t2 = self.net_table.table
        t2.setWordWrap(False)
        t2.setItemDelegate(ElidedItemDelegate(t2))
        h2 = t2.horizontalHeader()
        h2.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tabs.addTab(self.net_table, "İnternet (GPRS/WAP)")

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; padding:8px;")
        btn_close.clicked.connect(self.accept)
        self.layout.addWidget(btn_close)

        self.loader = LoadingOverlay(self)
        QTimer.singleShot(
            50,
            lambda: self.load_data(project_id, owner_gsm, target_baz, start_date, end_date)
        )

        def _fix_columns():
            h = self.gsm_table.table.horizontalHeader()
            h.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
            h.setStretchLastSection(True)

            self.gsm_table.table.setColumnWidth(0, 200)
            self.gsm_table.table.setColumnWidth(1, 200)
            self.gsm_table.table.setColumnWidth(2, 200)
            self.gsm_table.table.setColumnWidth(3, 200)
            self.gsm_table.table.setColumnWidth(4, 80)
            self.gsm_table.table.setColumnWidth(5, 90)

            h2 = self.net_table.table.horizontalHeader()
            h2.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
            h2.setStretchLastSection(True)

            self.net_table.table.setColumnWidth(0, 100)
            self.net_table.table.setColumnWidth(1, 200)
            self.net_table.table.setColumnWidth(2, 100)
            self.net_table.table.setColumnWidth(3, 150)
            self.net_table.table.setColumnWidth(4, 100)
            self.net_table.table.setColumnWidth(5, 100)
            self.net_table.table.setColumnWidth(6, 100)

        QTimer.singleShot(200, _fix_columns)

    def load_data(self, pid, owner, baz, min_dt, max_dt):
        self.loader.start("Konum Verileri Taranıyor...")
        QApplication.processEvents()
        try:
            with DB() as conn:
                cur = conn.cursor()
                date_filter_sql = ""
                params_base = []
                if min_dt and max_dt:
                    s_str = min_dt.toString("yyyy-MM-dd HH:mm:ss")
                    e_str = max_dt.toString("yyyy-MM-dd HH:mm:ss")
                    date_filter_sql = " AND (substr(TARIH, 7, 4) || '-' || substr(TARIH, 4, 2) || '-' || substr(TARIH, 1, 2) || substr(TARIH, 11)) BETWEEN ? AND ? "
                    params_base = [s_str, e_str]

                limit_clause = " LIMIT 10001"
                p_std = [pid, owner, baz] + params_base

                gsm_sql = f"SELECT TARIH, DIGER_NUMARA, DIGER_ISIM, DIGER_TC, TIP, SURE FROM hts_gsm WHERE ProjeID=? AND GSMNo=? AND BAZ=? {date_filter_sql} ORDER BY substr(TARIH, 7, 4) || substr(TARIH, 4, 2) || substr(TARIH, 1, 2) || substr(TARIH, 11) DESC {limit_clause}"
                gsm_rows = cur.execute(gsm_sql, p_std).fetchall()

                gprs_sql = f"SELECT 'GPRS', TARIH, SURE, KAYNAK_IP, '', GONDERME, INDIRME FROM hts_gprs WHERE ProjeID=? AND GSMNo=? AND BAZ=? {date_filter_sql} {limit_clause}"
                gprs = cur.execute(gprs_sql, p_std).fetchall()

                wap_sql = f"SELECT 'WAP', TARIH, SURE, KAYNAK_IP, HEDEF_IP, GONDERME, INDIRME FROM hts_wap WHERE ProjeID=? AND GSMNo=? AND BAZ=? {date_filter_sql} {limit_clause}"
                wap = cur.execute(wap_sql, p_std).fetchall()

                net_rows = gprs + wap

                if (len(gsm_rows) + len(net_rows)) > 20000:
                    ModernDialog.show_warning(self, "Limit", "Çok fazla kayıt var, liste kısıtlandı.")

                self.gsm_table.set_data([list(x) for x in gsm_rows])

                def sort_key(row):
                    t = row[1];
                    try: return t[6:10] + t[3:5] + t[0:2] + t[11:]
                    except: return ""
                net_rows.sort(key=sort_key, reverse=True)
                self.net_table.set_data([list(x) for x in net_rows])

                t = self.gsm_table.table
                t.setColumnWidth(0, 140); t.setColumnWidth(1, 110); t.setColumnWidth(2, 150); t.setColumnWidth(3, 100)
                t.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

        except Exception as e:
            print(f"Hata: {e}")
        finally:
            self.loader.stop()


class MapMultiPointDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, gsm_no, focus_text):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.setWindowTitle(f"Konum Analizi: {focus_text[:30]}...")
        self.resize(1000, 700)

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(5, 5, 5, 5)
        self.layout.setSpacing(5)

        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame { background-color: #fff3e0; border: 1px solid #ffcc80; border-radius: 6px; padding: 2px; }
        """)
        h_info = QHBoxLayout(info_frame)
        h_info.setContentsMargins(4, 2, 4, 2); h_info.setSpacing(8)

        lbl_icon = QLabel("📍"); lbl_icon.setStyleSheet("font-size: 18px; border: none;")
        h_info.addWidget(lbl_icon, 0, Qt.AlignmentFlag.AlignTop)

        info_html = (
            f"<div style='font-size:13px; line-height:1.3; font-family:Segoe UI; color:#ef6c00;'>"
            f"<b>📱 İlgili Hat:</b> <span style='color:black; font-weight:bold;'>{gsm_no}</span><br>"
            f"<b>📡 Seçili Baz:</b> {focus_text}<br>"
            f"<i style='color:#7f8c8d; font-size:11px;'>(Diğer tüm özel konumlar da haritada gösterilmektedir.)</i>"
            f"</div>"
        )
        lbl_text = QLabel(info_html); lbl_text.setStyleSheet("border: none;"); lbl_text.setWordWrap(True)
        h_info.addWidget(lbl_text, 1)

        self.layout.addWidget(info_frame, 0)

        self.browser = EvidenceWebEngineView()
        self.browser.setStyleSheet("border: 1px solid #bdc3c7;")
        self.layout.addWidget(self.browser, 1)

        self.draw_smart_map(project_id, gsm_no, focus_text)

    def extract_coords(self, text):
        if not text: return None
        text_str = str(text).strip()

        coords = re.findall(r"(\d{2}\.\d{4,})", text_str)
        if len(coords) >= 2:
            try:
                v1, v2 = float(coords[-2]), float(coords[-1])
                if 35 < v1 < 43 and 25 < v2 < 46: return [v1, v2]
                elif 35 < v2 < 43 and 25 < v1 < 46: return [v2, v1]
                else: return [v1, v2]
            except: pass

        try:
            with DB() as conn:
                cell_id = None
                match = re.search(r'\((\d{4,})\)', text_str)
                if match: cell_id = match.group(1)
                else:
                    nums = re.findall(r'\d+', text_str)
                    cands = [n for n in nums if len(n) > 3]
                    if cands: cell_id = cands[0]

                if cell_id:
                    row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE CellID=?", (cell_id,)).fetchone()
                    if row: return [row[0], row[1]]

                row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE BazAdi=?", (text_str,)).fetchone()
                if row: return [row[0], row[1]]
        except: pass
        return None

    def draw_smart_map(self, pid, gsm, focus_text):
        import folium
        pt_focus = self.extract_coords(focus_text)
        center = pt_focus if pt_focus else [39.0, 35.0]
        zoom = 13 if pt_focus else 6

        def check_internet():
            try: socket.create_connection(("8.8.8.8", 53), timeout=1.5); return True
            except OSError: return False
        is_online = check_internet(); has_local_file = os.path.exists(os.path.join(APP_DIR, "turkey.mbtiles"))
        tile_url = "OpenStreetMap" if is_online else ("http://localhost:8080/{z}/{x}/{y}.png" if has_local_file else "OpenStreetMap")
        attr_info = "OpenStreetMap" if is_online else ("Çevrimdışı Harita" if has_local_file else "Kaynak Yok")

        m = folium.Map(location=center, zoom_start=zoom, tiles=tile_url, attr=attr_info)
        _enable_measure_and_balloons(m)
        fit_points = []
        if pt_focus:
            fit_points.append(pt_focus)

        fg_focus = folium.FeatureGroup(name="🔵 Seçili Baz İstasyonu")
        if pt_focus:
            folium.Marker(
                location=pt_focus,
                popup=f"<div style='width:200px'><b>SEÇİLİ KAYIT</b><br>{focus_text}</div>",
                icon=folium.Icon(color="blue", icon="rss", prefix='fa'),
                tooltip="Seçili Kayıt"
            ).add_to(fg_focus)
            folium.Circle(location=pt_focus, radius=300, color="blue", fill=True, fill_opacity=0.1).add_to(fg_focus)
        fg_focus.add_to(m)

        fg_custom = folium.FeatureGroup(name="⭐ Özel Konumlar")
        try:
            with DB() as conn:
                rows = conn.execute("SELECT Lat, Lon, Label FROM ozel_konumlar WHERE ProjeID=?", (pid,)).fetchall()
                for r in rows:
                    lat, lon, label = r
                    folium.Marker(
                        location=[lat, lon],
                        popup=f"<b>ÖZEL KONUM</b><br>{label}",
                        icon=folium.Icon(color="orange", icon="star", prefix='fa'),
                        tooltip=folium.Tooltip(label, permanent=True, style="background-color: #fff3e0; border: 1px solid #e67e22; color: #d35400; font-weight: bold;")
                    ).add_to(fg_custom)
        except: pass

        fg_custom.add_to(m)
        folium.LayerControl(collapsed=False).add_to(m)

        if pt_focus:
            m.location = pt_focus
        else:
            if fit_points:
                m.fit_bounds(fit_points, padding=(50, 50))
        data = io.BytesIO(); m.save(data, close_file=False); self.browser.setHtml(data.getvalue().decode())


class ImeiSwapTimelineDialog(QDialog):
    def __init__(self, parent=None, imei="", segments=None):
        super().__init__(parent)
        self.setWindowTitle(f"IMEI Swap Timeline - {imei}")
        self.resize(900, 520)

        v = QVBoxLayout(self)

        lbl = QLabel(f"<b>IMEI:</b> {imei}")
        v.addWidget(lbl)

        # segments: [(gsm, first_seen, last_seen, adet)]
        headers = ["GSM", "İlk Görülme", "Son Görülme", "Kayıt"]
        self.table = GenericDatabaseTable(headers=headers, parent=self)
        v.addWidget(self.table)

        rows = []
        for gsm, fs, ls, cnt in (segments or []):
            rows.append([gsm, fs, ls, cnt])
        self.table.set_data(rows)


class ImeiDetailDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, owner_gsm, target_imei, start_date=None, end_date=None, is_common=False):
        super().__init__(parent)

        central = QWidget(self)
        central_layout = QVBoxLayout(central); central_layout.setContentsMargins(0, 0, 0, 0); central_layout.setSpacing(0)
        base_dir = os.path.dirname(__file__) if '__file__' in globals() else os.getcwd()
        logo_path = os.path.join(base_dir, "assets", "bg_logo.png")
        if not os.path.exists(logo_path): logo_path = os.path.join(base_dir, "assets", "logo.png")
        self._watermark = WatermarkBackground(logo_path=logo_path, opacity=0.03, scale=1.20, parent=central)
        stack = QStackedLayout(self); stack.setStackingMode(QStackedLayout.StackingMode.StackAll)
        stack.addWidget(central); stack.addWidget(self._watermark); self._watermark.raise_()

        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.setWindowTitle(f"Cihaz (IMEI) İletişim Analizi: {target_imei}")
        self.resize(1250, 750)

        self.layout = QVBoxLayout()
        central_layout.addLayout(self.layout)

        self.project_id = project_id; self.owner_gsm = owner_gsm; self.is_common = is_common

        owner_display = str(owner_gsm)
        try:
            with DB() as conn:
                row = conn.execute("SELECT AD, SOYAD FROM hts_abone WHERE ProjeID=? AND GSMNo=? LIMIT 1", (project_id, owner_gsm)).fetchone()
                if row:
                    ad = str(row[0]).strip() if row[0] else ""
                    soyad = str(row[1]).strip() if row[1] else ""
                    full_name = f"{ad} {soyad}".strip()
                    if full_name: owner_display = f"{full_name} - {owner_gsm}"
        except: pass

        info_frame = QFrame()
        info_frame.setStyleSheet("background-color: #fff3e0; border: 1px solid #ffcc80; border-radius: 8px; padding: 10px;")
        info_layout = QHBoxLayout(info_frame)

        lbl_icon = QLabel("📱"); lbl_icon.setStyleSheet("font-size: 32px; border:none;")
        info_text = (
            f"<div style='font-size:14px; color:#e65100; font-family: Segoe UI;'>"
            f"<b>Hat Sahibi:</b> {owner_display}<br>"
            f"<b>IMEI No:</b> {target_imei}<br>"
            f"<i>(Çift tıklama ile açılan haritada Hedef ve Karşı taraf ayrımı otomatik yapılır.)</i>"
            f"</div>"
        )
        info_layout.addWidget(lbl_icon); info_layout.addWidget(QLabel(info_text)); info_layout.addStretch()
        self.layout.addWidget(info_frame)

        self.tabs = QTabWidget(); self.layout.addWidget(self.tabs)

        cols = ["Tarih", "Kullanan Numara (IMEI)", "Diğer Numara", "Yön", "Süre", "Kullanan Bazı", "Diğer Baz"]
        self.gsm_table = GenericDatabaseTable(cols, enable_date_filter=False, chart_mode='none')
        t = self.gsm_table.table; t.setWordWrap(False); t.setItemDelegate(ElidedItemDelegate(t))
        t.doubleClicked.connect(self.open_map_location)

        h = t.horizontalHeader()
        h.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        t.setColumnWidth(0, 140); t.setColumnWidth(1, 130); t.setColumnWidth(2, 130)

        self.tabs.addTab(self.gsm_table, "GSM Görüşmeleri")

        self.net_table = GenericDatabaseTable(["Tür", "Tarih", "Süre", "Kullanan", "Kaynak IP", "Hedef IP", "Gönderme", "İndirme", "Baz"], enable_date_filter=False, chart_mode='none')
        t2 = self.net_table.table; t2.setWordWrap(False); t2.setItemDelegate(ElidedItemDelegate(t2))
        t2.doubleClicked.connect(self.open_map_location)
        h2 = t2.horizontalHeader(); h2.setStretchLastSection(True)
        self.tabs.addTab(self.net_table, "İnternet (GPRS/WAP)")

        btn_close = QPushButton("Kapat"); btn_close.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; padding:8px;"); btn_close.clicked.connect(self.accept); self.layout.addWidget(btn_close)

        self.loader = LoadingOverlay(self)
        QTimer.singleShot(50, lambda: self.load_data(project_id, owner_gsm, target_imei, start_date, end_date))

    def load_data(self, pid, owner, imei, min_dt, max_dt):
        self.loader.start("IMEI Verileri İşleniyor...")
        QApplication.processEvents()

        try:
            clean_imei = str(imei).replace(" ", "").strip()

            with DB() as conn:
                cur = conn.cursor()

                date_filter_sql = ""
                params_base = []

                if min_dt and max_dt:
                    s_str = min_dt.toString("yyyy-MM-dd HH:mm:ss")
                    e_str = max_dt.toString("yyyy-MM-dd HH:mm:ss")
                    date_filter_sql = " AND (substr(t1.TARIH, 7, 4) || '-' || substr(t1.TARIH, 4, 2) || '-' || substr(t1.TARIH, 1, 2) || substr(t1.TARIH, 11)) BETWEEN ? AND ? "
                    params_base = [s_str, e_str]

                gsm_sql = f"""
                    SELECT 
                        t1.TARIH, 
                        t1.NUMARA, 
                        t1.DIGER_NUMARA, 
                        t1.TIP, 
                        t1.SURE, 
                        t1.BAZ as KullananBaz, 
                        t2.BAZ as DigerBaz
                    FROM hts_gsm t1
                    LEFT JOIN hts_gsm t2 ON 
                        t2.ProjeID = t1.ProjeID AND
                        -- Çapraz Numara Eşleşmesi
                        substr(replace(replace(replace(t2.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = 
                        substr(replace(replace(replace(t1.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND
                        
                        substr(replace(replace(replace(t2.DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = 
                        substr(replace(replace(replace(t1.NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) AND
                        
                        t2.TARIH = t1.TARIH 
                        
                    WHERE t1.ProjeID=? 
                      AND REPLACE(t1.IMEI, ' ', '')=? 
                      {date_filter_sql}
                    ORDER BY CAST(t1.SIRA_NO as INTEGER)
                    LIMIT 10001
                """

                p_query = [pid, clean_imei] + params_base
                gsm_rows = cur.execute(gsm_sql, p_query).fetchall()

                gprs_sql = f"SELECT 'GPRS', TARIH, SURE, NUMARA, KAYNAK_IP, '', GONDERME, INDIRME, BAZ FROM hts_gprs WHERE ProjeID=? AND REPLACE(IMEI, ' ', '')=? {date_filter_sql.replace('t1.', '')} LIMIT 10001"
                wap_sql = f"SELECT 'WAP', TARIH, SURE, NUMARA, KAYNAK_IP, HEDEF_IP, GONDERME, INDIRME, BAZ FROM hts_wap WHERE ProjeID=? AND REPLACE(IMEI, ' ', '')=? {date_filter_sql.replace('t1.', '')} LIMIT 10001"

                net_rows = cur.execute(gprs_sql, p_query).fetchall() + cur.execute(wap_sql, p_query).fetchall()

                if (len(gsm_rows)+len(net_rows)) >= 10000:
                    ModernDialog.show_warning(self, "Limit", "Liste 10.000 kayıt ile sınırlandı.")

                formatted_gsm = []
                for r in gsm_rows:
                    tarih, num, diger, tip, sure, k_baz, d_baz = r
                    yon = "Giden (->)" if any(x in str(tip).upper() for x in ["ARADI", "GONDER", "ÇIKIŞ", "OUT"]) else "Gelen (<-)"
                    formatted_gsm.append([tarih, num, diger, yon, f"{sure} sn" if sure else "", k_baz if k_baz else "", d_baz if d_baz else ""])

                self.gsm_table.set_data(formatted_gsm)
                self.net_table.set_data([list(r) for r in net_rows])

        except Exception as e:
            ModernDialog.show_error(self, "Hata", str(e))
        finally:
            self.loader.stop()

    def open_map_location(self, index):
        try:
            sender_table = self.sender()
            if not sender_table: return
            row = index.row()

            if sender_table == self.net_table.table:
                baz_text = self.net_table.proxy_model.index(row, 8).data()
                if not baz_text or len(str(baz_text)) < 5:
                    ModernDialog.show_warning(self, "Konum Yok", "Bu kayıtta baz istasyonu bilgisi yok.")
                    return
                dlg = MapMultiPointDialog(self, self.project_id, self.owner_gsm, str(baz_text))
                dlg.exec()
                return

            elif sender_table == self.gsm_table.table:
                kullanan_no = self.gsm_table.proxy_model.index(row, 1).data() # IMEI Sahibi
                diger_no = self.gsm_table.proxy_model.index(row, 2).data()    # Konuşulan Kişi

                kullanan_baz = self.gsm_table.proxy_model.index(row, 5).data()
                diger_baz = self.gsm_table.proxy_model.index(row, 6).data()

                lbl1 = f"{kullanan_no} (IMEI Sahibi)"
                lbl2 = f"{diger_no} (Diğer)"

                try:
                    def clean_gsm(n): return re.sub(r'\D', '', str(n))
                    clean_kullanan = clean_gsm(kullanan_no)

                    with DB() as conn:
                        role_row = conn.execute("SELECT Rol FROM hts_dosyalari WHERE ProjeID=? AND GSMNo=? LIMIT 1", (self.project_id, clean_kullanan)).fetchone()

                        if role_row:
                            role = str(role_row[0]).upper()
                            if role == "HEDEF":
                                lbl1 = f"{kullanan_no} (HEDEF / IMEI Sahibi)"
                                lbl2 = f"{diger_no} (KARŞI TARAF)"
                            elif role == "KARSI" or role == "KARŞI":
                                lbl1 = f"{kullanan_no} (KARŞI / IMEI Sahibi)"
                                lbl2 = f"{diger_no} (HEDEF)"
                except:
                    pass

                has_kullanan = kullanan_baz and len(str(kullanan_baz)) > 5
                has_diger = diger_baz and len(str(diger_baz)) > 5

                if not has_kullanan and not has_diger:
                    ModernDialog.show_warning(self, "Konum Yok", "Görüntülenecek konum verisi bulunamadı.")
                    return

                if has_kullanan and not has_diger:
                    ModernDialog.show_info(self, "Tek Taraflı Veri",
                        f"Diğer tarafın ({diger_no}) konum verisi bulunamadı.<br>Sadece IMEI sahibinin konumu gösteriliyor.")

                dlg = MapPreviewDialog(
                    self, self.project_id, kullanan_no,
                    str(kullanan_baz) if has_kullanan else "",
                    str(diger_baz) if has_diger else "",
                    label1=lbl1, label2=lbl2
                )
                dlg.exec()

        except Exception as e:
            print(f"Harita Hatası: {e}")
            ModernDialog.show_error(self, "Harita Hatası", str(e))


class ProfileCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.home_text = ""
        self.work_text = ""

        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

        self.setStyleSheet("""
            ProfileCard {
                background-color: #ffffff;
                border: 1px solid #e5e7eb;
                border-left: 4px solid #3b82f6; 
                border-radius: 6px;
            }
            ProfileCard:hover {
                border: 1px solid #3b82f6;
                border-left: 4px solid #2563eb;
                background-color: #fcfdff;
            }
            QLabel { border: none; background: transparent; }
            .title { 
                font-weight: 700; color: #9ca3af; 
                font-size: 10px; text-transform: uppercase; letter-spacing: 0.5px;
                margin-bottom: 2px;
            }
            .value { 
                font-weight: 600; color: #1f2937; 
                font-size: 11px; line-height: 14px;
            }
            .icon_box {
                background-color: #eff6ff;
                border-radius: 16px;
                padding: 5px;
            }
        """)

        self.setFixedHeight(65)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(15, 8, 15, 8)
        layout.setSpacing(20)

        self.home_ui = self._create_section("🏠", "Muhtemel Ev (23:00-06:00)", "#dbeafe", "#1e40af")
        layout.addLayout(self.home_ui['layout'], 1)

        line = QFrame()
        line.setFrameShape(QFrame.Shape.VLine)
        line.setStyleSheet("color: #e5e7eb; background-color: #e5e7eb; width: 1px;")
        layout.addWidget(line)

        self.work_ui = self._create_section("💼", "Muhtemel İş (09:00-18:00)", "#fce7f3", "#9d174d")
        layout.addLayout(self.work_ui['layout'], 1)

    def _create_section(self, icon_char, title_text, bg_color, text_color):
        h_layout = QHBoxLayout()
        h_layout.setSpacing(12)
        h_layout.setContentsMargins(0, 0, 0, 0)

        icon_cont = QLabel(icon_char)
        icon_cont.setFixedSize(32, 32)
        icon_cont.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_cont.setStyleSheet(f"background-color: {bg_color}; color: {text_color}; border-radius: 16px; font-size: 16px;")

        v_layout = QVBoxLayout()
        v_layout.setContentsMargins(0, 0, 0, 0)
        v_layout.setSpacing(0)

        title = QLabel(title_text)
        title.setProperty("class", "title")

        value = QLabel("—")
        value.setProperty("class", "value")
        value.setWordWrap(True)
        value.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)

        v_layout.addWidget(title)
        v_layout.addWidget(value)

        h_layout.addWidget(icon_cont)
        h_layout.addLayout(v_layout)

        return {"layout": h_layout, "title": title, "value": value}

    def set_data(self, home_text, work_text):
        self.home_text = home_text
        self.work_text = work_text
        self.home_ui['value'].setText(home_text or "Tespit Edilemedi")
        self.work_ui['value'].setText(work_text or "Tespit Edilemedi")
        self.setToolTip("🖱️ Sol Tık: Haritada Göster\n🖱️ Sağ Tık: Rapora Ekle")

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            target_widget = self.parent()
            while target_widget:
                if hasattr(target_widget, 'open_profile_map_dialog'):
                    target_widget.open_profile_map_dialog(self.home_text, self.work_text)
                    break
                target_widget = target_widget.parent()
        super().mousePressEvent(event)

    def show_context_menu(self, pos):
        menu = QMenu(self)
        menu = apply_menu_theme(QMenu(self))
        action_add = QAction("📸 Profil Adreslerini Rapora Ekle", self)
        action_add.triggered.connect(self.export_to_evidence)
        menu.addAction(action_add)
        menu.exec(self.mapToGlobal(pos))

    def export_to_evidence(self):
        target_widget = self.parent()
        while target_widget:
            if hasattr(target_widget, 'add_evidence_to_report'):
                html = f"""<table style="width:100%; border-collapse:collapse;">
                    <tr><td style="border:1px solid #ddd; padding:8px;"><b>Muhtemel Ev</b></td><td style="border:1px solid #ddd; padding:8px;">{self.home_text}</td></tr>
                    <tr><td style="border:1px solid #ddd; padding:8px;"><b>Muhtemel İş</b></td><td style="border:1px solid #ddd; padding:8px;">{self.work_text}</td></tr></table>"""
                target_widget.add_evidence_to_report("Profil Analizi", html, "TABLE")
                break
            target_widget = target_widget.parent()

    def reset(self):
        try:
            self.set_data("-", "-")
        except Exception:
            for attr in ("lbl_home", "lbl_work", "home_label", "work_label"):
                if hasattr(self, attr):
                    try:
                        getattr(self, attr).setText("-")
                    except:
                        pass


class ProfileMapDialog(WatermarkDialogMixin, QDialog):
    def __init__(self, parent, project_id, gsm_no, home_text, work_text):
        super().__init__(parent)
        self.init_watermark(opacity=0.04, scale_ratio=0.85)
        self.setWindowTitle("Profil ve Konum Analizi Haritası")
        self.resize(1100, 750)

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(5, 5, 5, 5)
        self.layout.setSpacing(5)

        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #f9fbe7; 
                border: 1px solid #cddc39; /* Çerçeve korundu */
                border-radius: 6px; 
                padding: 2px; /* İç boşluk minimuma indi */
            }
        """)

        h_info = QHBoxLayout(info_frame)
        h_info.setContentsMargins(4, 2, 4, 2)
        h_info.setSpacing(10)

        lbl_icon = QLabel("👤")
        lbl_icon.setStyleSheet("font-size: 18px; border: none;")
        lbl_icon.setAlignment(Qt.AlignmentFlag.AlignTop)
        h_info.addWidget(lbl_icon)

        info_html = (
            f"<div style='font-size:13px; color:#33691e; line-height:1.2; font-family:Segoe UI;'>"
            f"<b>🏠 Ev Konumu:</b> {home_text or 'Tespit Edilemedi'}<br>"
            f"<b>💼 İş Konumu:</b> {work_text or 'Tespit Edilemedi'}"
            f" <span style='color:#827717; font-style:italic;'> (Not: Haritada 'Özel Konumlar' da gösterilmektedir.)</span>"
            f"</div>"
        )
        lbl_text = QLabel(info_html)
        lbl_text.setStyleSheet("border: none;")
        lbl_text.setWordWrap(True)
        h_info.addWidget(lbl_text, 1)

        self.layout.addWidget(info_frame, 0)

        self.browser = EvidenceWebEngineView()
        self.browser.setStyleSheet("border: 1px solid #bdc3c7;")
        self.layout.addWidget(self.browser, 1)

        self.draw_map(project_id, gsm_no, home_text, work_text)

    def extract_coords(self, text):
        if not text: return None
        text_str = str(text).strip()

        coords = re.findall(r"(\d{2}\.\d{4,})", text_str)
        if len(coords) >= 2:
            try:
                v1, v2 = float(coords[-2]), float(coords[-1])
                if 35 < v1 < 43 and 25 < v2 < 46: return [v1, v2]
                elif 35 < v2 < 43 and 25 < v1 < 46: return [v2, v1]
                else: return [v1, v2]
            except: pass

        try:
            with DB() as conn:
                cell_id = None
                match = re.search(r'\((\d{4,})\)', text_str)
                if match: cell_id = match.group(1)
                else:
                    nums = re.findall(r'\d+', text_str)
                    cands = [n for n in nums if len(n) > 3]
                    if cands: cell_id = cands[0]

                if cell_id:
                    row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE CellID=?", (cell_id,)).fetchone()
                    if row: return [row[0], row[1]]

                row = conn.execute("SELECT Lat, Lon FROM baz_kutuphanesi WHERE BazAdi=?", (text_str,)).fetchone()
                if row: return [row[0], row[1]]
        except: pass
        return None

    def draw_map(self, pid, gsm, home_txt, work_txt):
        import folium
        pt_home = self.extract_coords(home_txt)
        pt_work = self.extract_coords(work_txt)

        center = [39.0, 35.0]
        if pt_home: center = pt_home
        elif pt_work: center = pt_work

        zoom = 12 if (pt_home or pt_work) else 6

        tile_url = "OpenStreetMap"
        attr_info = "OpenStreetMap"
        try:
            socket.create_connection(("8.8.8.8", 53), timeout=1.5)
        except OSError:
            if os.path.exists(os.path.join(APP_DIR, "turkey.mbtiles")):
                tile_url = "http://localhost:8080/{z}/{x}/{y}.png"
                attr_info = "Local Offline Map"
            else:
                attr_info = "Map Source Not Found"

        m = folium.Map(location=center, zoom_start=zoom, tiles=tile_url, attr=attr_info)
        _enable_measure_and_balloons(m)
        all_points = []

        if pt_home:
            folium.Marker(
                location=pt_home,
                popup=folium.Popup(f"<b>MUHTEMEL EV</b><br>{home_txt}", max_width=300),
                icon=folium.Icon(color="green", icon="home", prefix='fa'),
                tooltip="Muhtemel Ev"
            ).add_to(m)
            all_points.append(pt_home)

        if pt_work:
            folium.Marker(
                location=pt_work,
                popup=folium.Popup(f"<b>MUHTEMEL İŞ</b><br>{work_txt}", max_width=300),
                icon=folium.Icon(color="red", icon="briefcase", prefix='fa'),
                tooltip="Muhtemel İş"
            ).add_to(m)
            all_points.append(pt_work)

        try:
            with DB() as conn:
                rows = conn.execute("SELECT Lat, Lon, Label FROM ozel_konumlar WHERE ProjeID=?", (pid,)).fetchall()

            if rows:
                fg_custom = folium.FeatureGroup(name="Kayıtlı Özel Konumlar")
                for r in rows:
                    lat, lon, label = r
                    folium.Marker(
                        location=[lat, lon],
                        popup=f"<b>KAYITLI KONUM</b><br>{label}",
                        icon=folium.Icon(color="orange", icon="star", prefix='fa'),
                        tooltip=folium.Tooltip(label, permanent=True, style="background-color: #fff3e0; border: 1px solid #e67e22; color: #d35400; font-weight: bold;")
                    ).add_to(fg_custom)
                    all_points.append([lat, lon])
                fg_custom.add_to(m)
        except Exception as e:
            print(f"Özel konum yükleme hatası: {e}")

        folium.LayerControl().add_to(m)

        if all_points:
            m.fit_bounds(all_points, padding=(50, 50))

        data = io.BytesIO()
        m.save(data, close_file=False)
        self.browser.setHtml(data.getvalue().decode())


class LocationSelectorDialog(QDialog):
    """
    Harita görüntüleme ve koordinat seçimi için optimize edilmiş sınıf.
    Beyaz ekran sorunu için 'Load Finished' kontrolü ve yerel dosya protokolü kullanır.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Konum Seçici")
        self.resize(1000, 700)
        self.selected_coords = None

        # Ana Layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0) # Kenar boşluğu yok
        layout.setSpacing(0)

        # Web View (Harita)
        from PyQt6.QtWebEngineWidgets import QWebEngineView
        self.view = QWebEngineView()
        layout.addWidget(self.view)

        # Alt Kontrol Barı
        control_bar = QFrame()
        control_bar.setFixedHeight(50)
        control_bar.setStyleSheet("background-color: #f0f0f0; border-top: 1px solid #ccc;")
        bar_layout = QHBoxLayout(control_bar)
        bar_layout.setContentsMargins(10, 5, 10, 5)

        self.lbl_info = QLabel("Lütfen haritadan bir nokta seçiniz.")
        self.lbl_info.setStyleSheet("color: #333; font-weight: bold;")

        self.btn_confirm = QPushButton("SEÇİMİ ONAYLA")
        self.btn_confirm.setEnabled(False)
        self.btn_confirm.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 5px 15px;")
        self.btn_confirm.clicked.connect(self.accept)

        btn_cancel = QPushButton("İptal")
        btn_cancel.clicked.connect(self.reject)

        bar_layout.addWidget(self.lbl_info)
        bar_layout.addStretch()
        bar_layout.addWidget(self.btn_confirm)
        bar_layout.addWidget(btn_cancel)

        layout.addWidget(control_bar)

        # İletişim kanalı
        self.view.titleChanged.connect(self._handle_title_change)

        # Haritayı UI yüklendikten hemen sonra oluştur (Render hatasını önler)
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(100, self._init_map)

    def _init_map(self):
        import folium
        from folium.plugins import Geocoder
        from PyQt6.QtCore import QUrl

        try:
            # Harita oluştur
            m = folium.Map(location=[39.0, 35.0], zoom_start=6)
            _enable_measure_and_balloons(m)
            # Arama Çubuğu (Geocoder)
            Geocoder(add_marker=False, position='topleft').add_to(m)

            # Tıklama Eventi (JS)
            m.get_root().script.add_child(folium.Element("""
                function onMapClick(e) {
                    var lat = e.latlng.lat;
                    var lng = e.latlng.lng;
                    
                    document.querySelectorAll('.leaflet-marker-icon').forEach(e => e.remove());
                    document.querySelectorAll('.leaflet-marker-shadow').forEach(e => e.remove());

                    L.marker([lat, lng]).addTo(map_""" + m.get_name() + """);
                    document.title = "COORD:" + lat.toFixed(6) + "," + lng.toFixed(6);
                }
                map_""" + m.get_name() + """.on('click', onMapClick);
            """))

            # Dosyayı UTF-8 olarak kaydet
            tmp_dir = tempfile.gettempdir()
            file_path = os.path.join(tmp_dir, "map_select.html")

            m.save(file_path)

            # Dosya yolunu QUrl formatına çevir (Windows/Linux uyumlu)
            local_url = QUrl.fromLocalFile(file_path)
            self.view.load(local_url)

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Harita yüklenemedi:\n{str(e)}")

    def _handle_title_change(self, title):
        if title.startswith("COORD:"):
            try:
                coords = title.split(":")[1].split(",")
                lat, lon = float(coords[0]), float(coords[1])
                self.selected_coords = (lat, lon)
                self.lbl_info.setText(f"Seçilen: {lat:.5f}, {lon:.5f}")
                self.btn_confirm.setEnabled(True)
            except:
                pass

    def get_coordinates(self):
        return self.selected_coords


class EventTabTableWidget(QWidget):
    """
    Olay Merkezli Analiz sekmelerinde (ÖNCE / KRİTİK / SONRA)
    kullanılan tablo + arama çubuğu kapsayıcısı.
    """
    def __init__(self, parent=None, title=""):
        super().__init__(parent)
        self.title = title

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(6, 6, 6, 6)
        main_layout.setSpacing(6)

        # Arama çubuğu
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("🔍 Ara...")
        self.search_bar.setClearButtonEnabled(True)
        self.search_bar.setMinimumHeight(28)
        main_layout.addWidget(self.search_bar)

        # Tablo
        self.table = QTableView()
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.setWordWrap(False)
        self.table.setItemDelegate(ElidedItemDelegate(self.table))

        header = self.table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.verticalHeader().setVisible(False)

        # Header'ları parent'tan al (EventCenteredAnalysisPanel.display_cols)
        headers = []
        try:
            if parent is not None and hasattr(parent, "display_cols"):
                headers = list(parent.display_cols)
        except Exception:
            headers = []

        # Model + Proxy (on_row_double_clicked içinde proxy_model.mapToSource kullanılıyor)
        self.source_model = CustomTableModel([], headers)
        self.proxy_model = DateSortFilterProxyModel(self)
        self.proxy_model.setSourceModel(self.source_model)
        self.proxy_model.setFilterKeyColumn(-1)
        self.table.setModel(self.proxy_model)

        # Filtre bağla
        self.search_bar.textChanged.connect(self.proxy_model.setSearchText)

        main_layout.addWidget(self.table, 1)

    def set_headers(self, headers: list):
        headers = list(headers or [])
        try:
            self.source_model.beginResetModel()
            self.source_model._headers = headers
            self.source_model.endResetModel()
        except Exception:
            # en kötü ihtimal yeniden kur
            self.source_model = CustomTableModel(self.source_model._data if hasattr(self, "source_model") else [], headers)
            self.proxy_model.setSourceModel(self.source_model)

    def set_data(self, data: list, headers: list = None):
        """
        EventCenteredAnalysisPanel.run_analysis() burayı çağırıyor.
        """
        if headers is not None:
            self.set_headers(headers)

        if data is None:
            data = []

        self.source_model.update_data(data)


class EventCenteredAnalysisPanel(QWidget):
    """
    Olay Merkezli Analiz Dashboard - Profesyonel Final Sürüm
    - Görsel Standart: Sabit yükseklikli, kompakt üst bar ve CSS destekli KPI kartları.
    - Gizli Detay: Baz/LAC bilgileri tabloda yer kaplamaz, ÇİFT TIK ile detayda açılır.
    - Zengin Harita: Telefon ikonlu, popup detaylı ve render garantili harita sistemi.
    - Tam Uyumluluk: sync_from_context ve set_owner_info hataları giderildi.
    """
    def __init__(self, db_manager, main_window=None):
        super().__init__()
        self.db = db_manager
        self.main_window = main_window
        self.project_id = None
        self.rapor_merkezi = None
        self.owner_info = None
        self.idx_other = 0

        # T0 otomatik ayar kontrolü (kullanıcı manuel değiştirirse ezmeyelim)
        self._t0_user_edited = False
        self._t0_setting = False

        # Son focus (rapora basmak için)
        self._last_focus_row = None
        self._last_focus_label = None

        # Sütun Yönetimi (Görünür vs Teknik Detay)
        self.raw_cols = TABLE_COLUMNS.get("hts_gsm", [])
        # Tabloyu sade tutmak için BAZ/LAC verilerini eliyoruz + SIRA_NO'yu kaldır
        self.display_cols = ["Delta T"] + [
            c for c in self.raw_cols
            if "BAZ" not in c.upper()
            and "LAC" not in c.upper()
            and c.upper() != "SIRA_NO"
        ]

        # Ham veriyi saklamak için (Çift tık detayı için)
        self.data_store = {"before": [], "crit": [], "after": []}

        self.init_ui()

    def reset(self):
        # varsa kendi iç state/tablolarını sıfırla
        try:
            self.set_context(project_id=None, gsm_number=None, start_qdt=None, end_qdt=None)
        except:
            pass
        try:
            # panel içinde kullandığın tablo/model isimleri neyse onları boşalt
            if hasattr(self, "table"): self.table.set_data([])
            if hasattr(self, "result_table"): self.result_table.set_data([])
        except:
            pass


    # --- MAINWINDOW ZORUNLU ENTEGRASYON METOTLARI ---
    def set_owner_info(self, info):
        self.owner_info = info

    def set_project_id(self, pid: int | None):
        """Bu tablo widget'ına aktif proje id'sini enjekte eder."""
        self.project_id = pid

    def set_report_center(self, rapor_merkezi):
        """Açık olan Rapor Merkezi referansını verir (rapora blok eklemek için)."""
        self.rapor_merkezi = rapor_merkezi

    def sync_from_context(self):
        """MainWindow sekme değişiminde çağrılır."""
        if hasattr(self, 'cmb_gsm') and self.project_id:
            current = self.cmb_gsm.currentText()
            self._load_project_gsms()
            idx = self.cmb_gsm.findText(current)
            if idx >= 0:
                self.cmb_gsm.setCurrentIndex(idx)

        # Sekme değişiminde (kullanıcı T0'ı elle oynamadıysa) ilk HTS tarihine çek
        self._auto_set_t0_to_first_hts_date(force=False)

    def set_context(self, db_manager=None, project_id=None, owner_info=None, gsm_number=None, **kwargs):
        old_pid = getattr(self, 'project_id', None)

        if db_manager:
            self.db = db_manager
        if project_id:
            self.project_id = project_id
        if owner_info:
            self.owner_info = owner_info

        self.sync_from_context()

        if gsm_number:
            idx = self.cmb_gsm.findText(str(gsm_number).strip())
            if idx >= 0:
                self.cmb_gsm.setCurrentIndex(idx)

        # Proje değiştiyse T0'ı kesin olarak o projenin ilk HTS tarihine ayarla
        if project_id and project_id != old_pid:
            self._auto_set_t0_to_first_hts_date(force=True)

    def _on_t0_date_time_changed(self, _qdt):
        """Kullanıcı T0 alanını elle değiştirdiyse bunu işaretle."""
        if getattr(self, "_t0_setting", False):
            return
        self._t0_user_edited = True

    def _auto_set_t0_to_first_hts_date(self, force: bool = False):
        """Olay Anı (T0) alanını HTS kayıtlarının en erken tarihine ayarlar.

        - force=False: kullanıcı T0'ı elle değiştirdiyse alanı ezmez.
        - force=True : proje değişimi gibi durumlarda kesin olarak günceller.
        """
        if not hasattr(self, "dt_t0"):
            return
        if not self.project_id:
            return

        if (not force) and getattr(self, "_t0_user_edited", False):
            return

        iso_sql = (
            "substr(TARIH, 7, 4) || '-' || substr(TARIH, 4, 2) || '-' || "
            "substr(TARIH, 1, 2) || substr(TARIH, 11)"
        )

        sql = f"SELECT MIN({iso_sql}) FROM hts_gsm WHERE ProjeID=? AND TARIH IS NOT NULL AND trim(TARIH)<>''"

        try:
            with DB() as conn:
                r = conn.execute(sql, (self.project_id,)).fetchone()
            min_iso = (r[0] if r else None)
        except Exception:
            min_iso = None

        if not min_iso:
            # Programda yoksa yok: veri yoksa dokunmuyoruz
            return

        min_iso = str(min_iso).strip()

        qdt = QDateTime.fromString(min_iso, "yyyy-MM-dd HH:mm:ss")
        if not qdt.isValid():
            qdt = QDateTime.fromString(min_iso, "yyyy-MM-dd HH:mm")

        if not qdt.isValid():
            return

        try:
            self._t0_setting = True
            self.dt_t0.blockSignals(True)
            self.dt_t0.setDateTime(qdt)
        finally:
            try:
                self.dt_t0.blockSignals(False)
            except Exception:
                pass
            self._t0_setting = False

        if force:
            self._t0_user_edited = False

    # --- UI VE GÖRSEL TASARIM ---
    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(2)

        # =========================================================
        # 1) ÜST KRİTER ALANI
        # =========================================================
        ctrl_group = QGroupBox()
        ctrl_group.setTitle("")
        ctrl_group.setMinimumHeight(98)

        ctrl_v = QVBoxLayout(ctrl_group)
        ctrl_v.setContentsMargins(12, 10, 12, 10)
        ctrl_v.setSpacing(6)

        # Başlık + yardım (info baloncuğu)  ->  ? başlığın hemen sağında
        header_row = QHBoxLayout()
        header_row.setContentsMargins(0, 0, 0, 0)
        header_row.setSpacing(8)

        lbl_ctrl_title = QLabel("📊 Olay Analiz Kriterleri")
        lbl_ctrl_title.setStyleSheet("font-weight: 700; font-size: 12.5pt;")

        btn_help = InfoButton("")
        btn_help.setFixedSize(26, 26)
        btn_help.setToolTip("Bu ekran hakkında yardım")
        btn_help.clicked.connect(self.show_expert_help)

        header_row.addWidget(lbl_ctrl_title)
        header_row.addWidget(btn_help, 0, Qt.AlignmentFlag.AlignVCenter)
        header_row.addStretch(1)
        ctrl_v.addLayout(header_row)

        grid = QGridLayout()
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)
        ctrl_v.addLayout(grid)

        # --- Alanlar
        if not hasattr(self, "dt_t0"):
            self.dt_t0 = QDateTimeEdit(QDateTime.currentDateTime())
            self.dt_t0.setCalendarPopup(True)
            self.dt_t0.setDisplayFormat("dd.MM.yyyy HH:mm:ss")

        # T0 alanı manuel değişim takibi (tek sefer bağla)
        if not hasattr(self, "_t0_signal_connected"):
            self.dt_t0.dateTimeChanged.connect(self._on_t0_date_time_changed)
            self._t0_signal_connected = True
        self.dt_t0.setMinimumWidth(240)
        self.dt_t0.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        if not hasattr(self, "cmb_gsm"):
            self.cmb_gsm = QComboBox()
            self.cmb_gsm.addItem("Tüm Kayıtlar")
        self.cmb_gsm.setMinimumWidth(320)
        self.cmb_gsm.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        if not hasattr(self, "spin_b"):
            self.spin_b = QSpinBox()
            self.spin_b.setRange(1, 720)
            self.spin_b.setValue(24)
            self.spin_b.setSuffix(" sa")

        if not hasattr(self, "spin_c"):
            self.spin_c = QSpinBox()
            self.spin_c.setRange(1, 1440)
            self.spin_c.setValue(30)
            self.spin_c.setSuffix(" dk")

        if not hasattr(self, "spin_a"):
            self.spin_a = QSpinBox()
            self.spin_a.setRange(1, 720)
            self.spin_a.setValue(24)
            self.spin_a.setSuffix(" sa")

        for sp in (self.spin_b, self.spin_c, self.spin_a):
            sp.setMinimumWidth(90)
            sp.setMaximumWidth(110)
            sp.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)

        btn_run = QPushButton("🚀 ANALİZİ BAŞLAT")
        btn_run.setMinimumHeight(34)
        btn_run.setMinimumWidth(180)
        btn_run.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        btn_run.clicked.connect(self.run_analysis)

        btn_add_report = QPushButton("🧾 Rapora Ekle")
        btn_add_report.setMinimumHeight(32)
        btn_add_report.setMinimumWidth(180)
        btn_add_report.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        btn_add_report.clicked.connect(self.add_event_analysis_to_report)

        # --- Satır 0 (tek hizalı)
        grid.addWidget(QLabel("Olay Anı (T0):"), 0, 0)
        grid.addWidget(self.dt_t0, 0, 1)

        grid.addWidget(QLabel("Hedef GSM:"), 0, 2)
        grid.addWidget(self.cmb_gsm, 0, 3)

        # Sağda: üstte 2 buton yan yana, altında Coğrafi Konum Analizi
        btn_geo = QPushButton("🗺 Konum Ekle")
        btn_geo.setMinimumHeight(32)
        btn_geo.setMinimumWidth(368)  # 180 + 180 + aradaki boşluklara yakın
        btn_geo.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        btn_geo.clicked.connect(self.open_geo_location_analysis)  # bu metodu ayrıca ekleyeceksin

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(0, 0, 0, 0)
        btn_row.setSpacing(8)
        btn_row.addWidget(btn_run)
        btn_row.addWidget(btn_add_report)

        btn_stack = QVBoxLayout()
        btn_stack.setContentsMargins(0, 0, 0, 0)
        btn_stack.setSpacing(6)
        btn_stack.addLayout(btn_row)
        btn_stack.addWidget(btn_geo)

        wrap = QHBoxLayout()
        wrap.setContentsMargins(0, 0, 0, 0)
        wrap.addStretch(1)
        wrap.addLayout(btn_stack)

        grid.addLayout(wrap, 0, 4)

        # --- Satır 1 (Pencereler aynı hizalı)
        grid.addWidget(QLabel("Pencereler (Ö/K/S):"), 1, 0)

        win_row = QHBoxLayout()
        win_row.setContentsMargins(0, 0, 0, 0)
        win_row.setSpacing(10)
        win_row.addWidget(QLabel("Önce"))
        win_row.addWidget(self.spin_b)
        win_row.addSpacing(8)
        win_row.addWidget(QLabel("Kritik"))
        win_row.addWidget(self.spin_c)
        win_row.addSpacing(8)
        win_row.addWidget(QLabel("Sonra"))
        win_row.addWidget(self.spin_a)
        win_row.addStretch(1)
        grid.addLayout(win_row, 1, 1, 1, 4)

        grid.setColumnStretch(0, 0)
        grid.setColumnStretch(1, 2)
        grid.setColumnStretch(2, 0)
        grid.setColumnStretch(3, 3)
        grid.setColumnStretch(4, 2)

        main_layout.addWidget(ctrl_group)

        # =========================================================
        # 2) ANA ALAN
        # =========================================================
        root_splitter = QSplitter(Qt.Orientation.Horizontal)
        root_splitter.setHandleWidth(2)

        # --- Sol KPI panel
        if not hasattr(self, "left_panel"):
            self.left_panel = QFrame()

        self.left_panel.setMinimumWidth(230)
        self.left_panel.setMaximumWidth(280)
        self.left_panel.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Expanding)

        if self.left_panel.layout() is None:
            l_lyt = QVBoxLayout(self.left_panel)
            l_lyt.setContentsMargins(10, 10, 10, 10)
            l_lyt.setSpacing(10)

            self.lbl_total = self._add_kpi_card(l_lyt, "TOPLAM İŞLEM", "#3498db")
            self.lbl_crit = self._add_kpi_card(l_lyt, "KRİTİK TEMAS", "#e74c3c")
            self.lbl_burst = self._add_kpi_card(l_lyt, "YOĞUNLUK ARTIŞI", "#f39c12")

            l_lyt.addWidget(QLabel("<b>👤 KRİTİK TOP 5 TEMAS</b>"))
            self.list_top5 = QListWidget()
            l_lyt.addWidget(self.list_top5, 1)
            l_lyt.addStretch(1)

        # --- Sağ taraf (tablo + harita)
        right_splitter = QSplitter(Qt.Orientation.Vertical)
        right_splitter.setHandleWidth(2)

        # Üst: Sekmeler + tablolar
        if not hasattr(self, "tabs"):
            self.tabs = QTabWidget()
            self.tabs.setDocumentMode(True)

            self.t_before = EventTabTableWidget(self, title="ÖNCE")
            self.t_crit = EventTabTableWidget(self, title="KRİTİK")
            self.t_after = EventTabTableWidget(self, title="SONRA")

            # Sekme isimlerine uygun simgeler
            self.tabs.addTab(self.t_before, "⏪ ÖNCE")
            self.tabs.addTab(self.t_crit, "🎯 KRİTİK (Olay Anı)")
            self.tabs.addTab(self.t_after, "⏩ SONRA")

        # Alt: Harita alanı (varsa daha önce oluşturulmuş map_view kullanılır)
        if not hasattr(self, "map_view"):
            self.map_view = EvidenceWebEngineView(self)
            self.map_view.setMinimumHeight(350)

        if not hasattr(self, "map_slider"):
            self.map_slider = QSlider(Qt.Orientation.Horizontal)
            self.map_slider.setRange(0, 0)
            self.map_slider.valueChanged.connect(self.on_slider_moved)

        if not hasattr(self, "lbl_slider_info"):
            self.lbl_slider_info = QLabel("Zaman Akışı için Analiz Başlatın")
            self.lbl_slider_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.lbl_slider_info.setStyleSheet("color: #7f8c8d; font-size: 11px;")

        # Signal bağlama
        if not hasattr(self, "_event_table_signals_bound"):
            self.t_before.table.clicked.connect(self.on_row_single_clicked)
            self.t_crit.table.clicked.connect(self.on_row_single_clicked)
            self.t_after.table.clicked.connect(self.on_row_single_clicked)

            self.t_before.table.doubleClicked.connect(self.on_row_double_clicked)
            self.t_crit.table.doubleClicked.connect(self.on_row_double_clicked)
            self.t_after.table.doubleClicked.connect(self.on_row_double_clicked)

            self.t_before.search_bar.textChanged.connect(self.sync_map_to_filters)
            self.t_crit.search_bar.textChanged.connect(self.sync_map_to_filters)
            self.t_after.search_bar.textChanged.connect(self.sync_map_to_filters)

            self._event_table_signals_bound = True
            self._event_selmodels_bound = set()

        # Harita container
        map_container = QWidget()
        map_v_layout = QVBoxLayout(map_container)
        map_v_layout.setContentsMargins(0, 5, 0, 0)
        map_v_layout.setSpacing(2)
        map_v_layout.addWidget(self.map_view, 1)
        map_v_layout.addWidget(self.lbl_slider_info)
        map_v_layout.addWidget(self.map_slider)

        right_splitter.addWidget(self.tabs)
        right_splitter.addWidget(map_container)

        # Haritayı ekranda daha geniş göstermek istiyorsan burada 2. değeri artır:
        right_splitter.setSizes([520, 640])

        root_splitter.addWidget(self.left_panel)
        root_splitter.addWidget(right_splitter)
        root_splitter.setSizes([250, 1200])

        main_layout.addWidget(root_splitter, 1)


    def _ensure_event_table_focus_hooks(self):
        for t in (self.t_before.table, self.t_crit.table, self.t_after.table):
            sm = t.selectionModel()
            if sm and sm not in self._event_selmodels_bound:
                sm.currentChanged.connect(lambda cur, prev, self=self: self.on_row_single_clicked(cur))
                self._event_selmodels_bound.add(sm)

    # --- RAPORA EKLE ---
    def add_event_analysis_to_report(self):
        """
        Olay Merkezli Analiz çıktısını rapora 'kül halinde' ekler:
        - Kriterler (T0, önce, kritik, sonra)
        - Önce / Kritik / Sonra tabloları
        - Harita (üstünde italik açıklama)
        - Focuslanan an rapora yazılır
        """
        if not self.project_id:
            ModernDialog.show_warning(self, "Uyarı", "Önce bir proje seçiniz.")
            return

        if not hasattr(self, "disp_store") or not isinstance(self.disp_store, dict):
            ModernDialog.show_warning(self, "Uyarı", "Önce 'Analizi Başlat' çalıştırılmalıdır.")
            return

        before_rows = self.disp_store.get("before", [])
        crit_rows = self.disp_store.get("crit", [])
        after_rows = self.disp_store.get("after", [])

        if not (before_rows or crit_rows or after_rows):
            ModernDialog.show_warning(self, "Uyarı", "Rapora eklenecek tablo verisi bulunamadı.")
            return

        # Kriterler
        t0_text = self.dt_t0.dateTime().toString("dd.MM.yyyy HH:mm:ss") if hasattr(self, "dt_t0") else "-"
        b_hours = int(self.spin_b.value()) if hasattr(self, "spin_b") else 0
        c_mins = int(self.spin_c.value()) if hasattr(self, "spin_c") else 0
        a_hours = int(self.spin_a.value()) if hasattr(self, "spin_a") else 0
        gsm_text = self.cmb_gsm.currentText().strip() if hasattr(self, "cmb_gsm") else "Tüm Kayıtlar"

        # Focus bilgisi
        focus_info = "-"
        try:
            if self._last_focus_row:
                t_idx = 1 + self.raw_cols.index("TARIH")
                f_dt = self._last_focus_row[t_idx] if t_idx < len(self._last_focus_row) else "-"
                f_delta = self._last_focus_row[0] if len(self._last_focus_row) > 0 else "-"
                focus_info = f"{f_dt} (Delta: {f_delta})"
        except Exception:
            focus_info = "-"

        # Harita: raporda cluster kapalı (yeşil/sarı kümeler raporda oluşmasın)
        try:
            self._update_rich_map(
                self.data_store.get("before", []),
                self.data_store.get("crit", []),
                self.data_store.get("after", []),
                focus_row=self._last_focus_row,
                focus_label=self._last_focus_label,
                use_cluster=False
            )
        except Exception:
            pass

        # Harita screenshot al
        img_path = ""
        try:
            import os, tempfile
            from datetime import datetime

            out_dir = os.path.join(tempfile.gettempdir(), "hts_mercek_report")
            os.makedirs(out_dir, exist_ok=True)

            img_name = f"event_map_{self.project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            img_path = os.path.join(out_dir, img_name)

            # Mevcut boyutları sakla
            old_size = self.map_view.size()
            old_policy = self.map_view.sizePolicy()
            old_min = self.map_view.minimumSize()
            old_max = self.map_view.maximumSize()
            original_style = self.map_view.styleSheet()

            try:
                # 1. Çerçeveyi kaldır ve arkaplanı ayarla
                self.map_view.setStyleSheet("border: none !important; background: white;")

                # 2. Çözünürlüğü PDF kalitesi için artır
                W, H = 1350, 350
                self.map_view.setFixedSize(W, H) # Geçici olarak sabitle

                # 3. Motorun çizim yapmasına izin ver
                QApplication.processEvents()
                time.sleep(0.8) # Harita katmanlarının (tiles) yüklenmesi için kritik süre

                # 4. Görüntüyü yakala
                pix = self.map_view.grab()

                # 5. KRİTİK EKSİK: Görüntüyü fiziksel olarak kaydet
                if not pix.isNull():
                    pix.save(img_path, "PNG")
                else:
                    img_path = ""

            finally:
                # 6. Arayüzü eski haline döndür (hata üretmesin diye güvenli)
                try:
                    self.map_view.setMinimumSize(old_min)
                    self.map_view.setMaximumSize(old_max)
                    self.map_view.setSizePolicy(old_policy)
                    self.map_view.setStyleSheet(original_style)
                    self.map_view.resize(old_size)
                    self.map_view.update()
                    QApplication.processEvents()
                except Exception:
                    pass

        except Exception as e:
            print(f"Harita yakalama hatası: {e}")
            img_path = ""

        # Tablo HTML (başlık rengi ve kolon genişliği rapor temasıyla uyumlu)
        def _table_html(title: str, headers: list, rows: list, note: str = "") -> str:
            def esc_cell(x):
                s = "" if x is None else str(x)
                return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

            def clean_header(h):
                # "Delta\nT" gibi başlıkları tek satıra çek
                return ("" if h is None else str(h)).replace("\r", " ").replace("\n", " ").strip()

            # Kolon genişlikleri (rapor için)
            col_w = {
                "DELTA T": "10%",   # genişletildi
                "SURE": "6%",
                "TİP": "8%",        # daraltıldı
                "TIP": "8%",
                "NUMARA": "11%",
                "DIGER_NUMARA": "11%",
                "DİGER_NUMARA": "11%",
                "TARIH": "14%",
                "DİGER_ISIM": "16%",
                "DIGER_ISIM": "16%",
                "DİGER_TC": "10%",
                "DIGER_TC": "10%",
                "IMEI": "10%",
            }

            # TH üretimi (meta-table ile aynı tema: #e9eef3)
            ths = ""
            for h in headers:
                h_clean = clean_header(h)
                key = h_clean.upper()
                w = col_w.get(key, None)
                w_style = f"width:{w};" if w else ""
                ths += (
                    f"<th style='border:1px solid #555;padding:4px;background:#e9eef3;"
                    f"text-align:center;font-weight:bold;{w_style}"
                    f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>"
                    f"{esc_cell(h_clean)}</th>"
                )

            def td_style_for_header(hname: str) -> str:
                k = clean_header(hname).upper()
                base = "border:1px solid #555;padding:4px;vertical-align:middle;"

                if k == "BAZ":
                    return base + (
                        "white-space:nowrap;"
                        "overflow:hidden;"
                        "text-overflow:ellipsis;"
                        "max-width:180px;"
                    )

                # boşluksuz/uzun değerler için agresif kır
                hard_break_cols = {"NUMARA", "DIGER_NUMARA", "DİGER_NUMARA", "DIGER_TC", "DİGER_TC", "IMEI", "KAYNAK_IP", "HEDEF_IP"}
                if k in hard_break_cols:
                    return base + (
                        "white-space:normal;"
                        "overflow-wrap:anywhere;"
                        "word-break:break-all;"
                    )

                # normal metinler için daha yumuşak kır
                return base + (
                    "white-space:normal;"
                    "overflow-wrap:anywhere;"
                    "word-break:break-word;"
                )
            body = ""
            for r in rows:
                tds = ""
                for i, v in enumerate(r):
                    hname = headers[i] if i < len(headers) else ""
                    cell = esc_cell(v)
                    tds += f"<td style='{td_style_for_header(hname)}' title='{cell}'>{cell}</td>"
                body += f"<tr>{tds}</tr>"

            note_html = f"<div style='margin:2px 0 6px 0;color:#555;'><i>{esc_cell(note)}</i></div>" if note else ""

            return f"""
            <h3 style="margin:12px 0 4px 0;">{esc_cell(title)}</h3>
            {note_html}
            <table class="meta-table" style="table-layout:fixed; width:100%; max-width:100%;">
                <thead><tr>{ths}</tr></thead>
                <tbody>
                    {body if body else "<tr><td colspan='"+str(len(headers))+"' style='border:1px solid #555;padding:4px;'>-</td></tr>"}
                </tbody>
            </table>
            """

        headers = self.display_cols[:]

        map_legend = (
            "<i style='font-weight:normal;'>"
            "Harita (Kırmızı: kritik zaman penceresi, Mavi: önce penceresi, Mor: sonra penceresi. "
            "Sarı balonlar: aynı/çok yakın noktaların kümelenmiş gösterimi; içindeki sayı, kümedeki kayıt adedi.)"
            "</i>"
        )

        # ✅ Harita yüksekliği sabitlendi (PDF'de taşma/bozulma olmaması için)
        map_block = ""
        if img_path and os.path.exists(img_path):
            map_block = (
                "<div class='map-wrap map-evidence' style='width:100%; margin:10px 0 0 0; text-align:center; page-break-inside:avoid;'>"
                f"<div style='font-size:10px; font-weight:bold; margin:0 0 6px 0; color:#333;'>{map_legend}</div>"
                f"<img src='{img_path}' class='map-img' "
                "style='width:100%; height:auto; display:block; margin:0; border:none;'>"
                "</div>"
            )
        else:
            map_block = "<div style='padding:20px; color:red; text-align:center;'>Harita görüntüsü oluşturulamadı.</div>"

        col_explain = (
            "<div style='margin:6px 0 10px 0; font-size:12px; line-height:1.35; color:#2c3e50;'>"
            "<b>Açıklama:</b> "
            "Delta T: T0'a göre dakika farkı; "
            "Numara: hedef GSM; "
            "Diğer Numara: karşı taraf; "
            "Tip: işlem türü; "
            "Tarih: kayıt zamanı; "
            "Süre: konuşma/işlem süresi; "
            "Diğer İsim/TC/IMEI: abonelik/hat bilgileri."
            "</div>"
        )

        html = f"""
        <div style="margin-bottom:10px;font-size:12px;">
            <div><b>Hedef GSM:</b> {gsm_text}</div>
            <div><b>Olay Anı (T0):</b> {t0_text}</div>
            <div><b>Öncesi:</b> {b_hours} saat</div>
            <div><b>Kritik:</b> {c_mins} dakika önce-sonra</div>
            <div><b>Sonrası:</b> {a_hours} saat</div>
            <div><b>Seçili An (Focus):</b> {focus_info}</div>
        </div>
        {col_explain}
        {_table_html("Önce Tablosu", headers, before_rows)}
        {_table_html("Kritik Tablosu", headers, crit_rows)}
        {_table_html("Sonra Tablosu", headers, after_rows)}
        {map_block}
        """

        # Rapor sırasına ekle (mevcut davranış korunuyor)
        try:
            from datetime import datetime

            with DB() as conn:
                c = conn.cursor()

                last_order = c.execute(
                    "SELECT MAX(Sira) FROM rapor_taslagi WHERE ProjeID=?",
                    (self.project_id,)
                ).fetchone()
                new_order = (last_order[0] + 1) if last_order and last_order[0] is not None else 1

                cols = [r[1] for r in c.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]

                keys = ["ProjeID", "GSMNo", "Baslik", "Icerik", "Tur", "Tarih", "Sira"]
                vals = [
                    self.project_id,
                    gsm_text,
                    f"Olay Merkezli Analiz (Olay Anı: {t0_text})",
                    "",
                    "HTML",
                    datetime.now().strftime("%d.%m.%Y %H:%M"),
                    new_order
                ]
                width_pct = 100
                if "GenislikYuzde" in cols:
                    keys.append("GenislikYuzde"); vals.append(width_pct)
                if "YukseklikMm" in cols:
                    keys.append("YukseklikMm"); vals.append(0)
                if "Hizalama" in cols:
                    keys.append("Hizalama"); vals.append("center")
                if "Aciklama" in cols:
                    keys.append("Aciklama"); vals.append("")
                if "HtmlIcerik" in cols:
                    keys.append("HtmlIcerik"); vals.append(html)

                q = f"INSERT INTO rapor_taslagi ({', '.join(keys)}) VALUES ({', '.join(['?']*len(keys))})"
                c.execute(q, tuple(vals))
                conn.commit()

            ModernDialog.show_info(self, "Başarılı", "Olay Merkezli Analiz çıktısı rapora eklendi.")
        except Exception as e:
            try:
                ModernDialog.show_error(self, "Hata", f"Rapora ekleme sırasında hata oluştu:\n{e}")
            except Exception:
                pass

    # --- FONKSİYONEL MANTIK ---
    def _add_kpi_card(self, layout, title, color):
        card = QFrame()
        card.setStyleSheet(f"background: white; border-left: 6px solid {color}; border-radius: 5px;")
        card.setFixedHeight(65)
        v = QVBoxLayout(card)
        v.setContentsMargins(10, 5, 10, 5)
        t = QLabel(title)
        t.setStyleSheet("color: #7f8c8d; font-size: 10px; font-weight: bold;")
        val = QLabel("-")
        val.setStyleSheet(f"color: {color}; font-size: 20px; font-weight: bold;")
        v.addWidget(t)
        v.addWidget(val)
        layout.addWidget(card)
        return val

    def show_expert_help(self):
        help_html = """
        <div style="font-size:13px; line-height:1.55;">
          <h2>📌 Olay Merkezli Analiz – Bilgi Merkezi</h2>

          <p>
            Bu ekran, belirlenen <b>Olay Anı (T0)</b> merkez alınarak
            öncesi, olay anı ve sonrası HTS kayıtlarının
            <b>zaman, temas ve konum</b> ekseninde analiz edilmesini sağlar.
          </p>

          <h3>🎯 Temel Bileşenler</h3>
          <ul>
            <li><b>Olay Anı (T0):</b> Analizin merkez zamanı</li>
            <li><b>Hedef GSM:</b> Belirli bir numara veya tüm kayıtlar</li>
            <li><b>Pencereler (Ö / K / S):</b>
              <ul>
                <li><b>Önce:</b> Olaydan önceki hazırlık/ön temaslar</li>
                <li><b>Kritik:</b> Olay anına en yakın ve en önemli temaslar</li>
                <li><b>Sonra:</b> Olay sonrası devam/dağılma temasları</li>
              </ul>
            </li>
          </ul>

          <h3>🚀 Analizi Başlat</h3>
          <p>
            Seçilen kriterlere göre verileri işler, üç sekmeyi doldurur,
            harita ve özet göstergeleri üretir.
          </p>

          <h3>🧾 Rapora Ekle</h3>
          <p>
            Yapılan analiz çıktıları delil formatında rapora eklenir.
            Analiz çalıştırılmadan rapora ekleme yapılamaz.
          </p>

          <h3>🗂 Sekmeler</h3>
          <ul>
            <li><b>⏪ ÖNCE:</b> Olay öncesi ilişkiler</li>
            <li><b>🎯 KRİTİK:</b> Olay anına en yakın temaslar</li>
            <li><b>⏩ SONRA:</b> Olay sonrası hareketler</li>
          </ul>

          <p style="margin-top:10px; color:#7f8c8d;">
            İpucu: Tabloda tek tıklama harita odağını, çift tıklama detay penceresini açar.
          </p>
        </div>
        """

        # InfoButton'a uyarla (global baloncuk: SleekTooltipPopup)
        btn = self.sender()
        if isinstance(btn, InfoButton):
            btn.popup.set_text(help_html)   # SleekTooltipPopup.set_text var :contentReference[oaicite:2]{index=2}
            # Eğer InfoButton önce boş içerikle açtıysa, içerik güncellensin diye görünürken de ayarla
            if btn.popup.isVisible():
                btn.popup.adjustSize()
            else:
                btn.show_popup()            # InfoButton.show_popup var :contentReference[oaicite:3]{index=3}
            return

    def on_row_single_clicked(self, index: QModelIndex):
        """Tek tık: haritayı satırın baz noktasına odaklar ve popup'ı açık getirir."""
        tab_idx = self.tabs.currentIndex()
        key = ["before", "crit", "after"][tab_idx]

        if key not in self.data_store:
            return
        if not index.isValid():
            return

        row_i = index.row()
        if row_i < 0 or row_i >= len(self.data_store[key]):
            return

        label = "Kritik" if key == "crit" else ("Önce" if key == "before" else "Sonra")
        self._last_focus_row = self.data_store[key][row_i]
        self._last_focus_label = label
        self._focus_map_to_row(self._last_focus_row, label=label)

    def sync_map_to_filters(self, *_):
        """
        Filtre nasıl tabloyu daraltıyorsa, haritayı da aynı görünen satırlara göre günceller.
        """
        if not hasattr(self, "disp_store"):
            return

        f_before = (self.t_before.search_bar.text() or "").strip().lower()
        f_crit = (self.t_crit.search_bar.text() or "").strip().lower()
        f_after = (self.t_after.search_bar.text() or "").strip().lower()

        def _match(visible_row, ftxt: str) -> bool:
            if not ftxt:
                return True
            s = " ".join([str(x) for x in (visible_row or [])]).lower()
            return ftxt in s

        before_full = [self.data_store["before"][i] for i, vr in enumerate(self.disp_store["before"]) if _match(vr, f_before)]
        crit_full = [self.data_store["crit"][i] for i, vr in enumerate(self.disp_store["crit"]) if _match(vr, f_crit)]
        after_full = [self.data_store["after"][i] for i, vr in enumerate(self.disp_store["after"]) if _match(vr, f_after)]

        self._update_rich_map(before_full, crit_full, after_full, focus_row=None, use_cluster=True)

    def _focus_map_to_row(self, full_row, label="Genel"):
        """Seçilen satıra göre haritayı yeniden üretir ve seçili satır popup'ını açık getirir."""
        before = self.data_store.get("before", [])
        crit = self.data_store.get("crit", [])
        after = self.data_store.get("after", [])
        self._update_rich_map(before, crit, after, focus_row=full_row, focus_label=label, use_cluster=True)

    def on_row_double_clicked(self, index):
        """Tablodaki gizli teknik verileri, dosya detay popup'ı ile aynı UI'de gösterir."""
        try:
            tab_idx = self.tabs.currentIndex()
            key = ["before", "crit", "after"][tab_idx]
            tablew = [self.t_before, self.t_crit, self.t_after][tab_idx]

            src_index = tablew.proxy_model.mapToSource(index)
            src_row = src_index.row()
            if src_row < 0 or src_row >= len(self.data_store.get(key, [])):
                return

            full_row = self.data_store[key][src_row]  # [Delta T] + raw_cols...

            def _fmt(v):
                if v is None:
                    return "-"
                s = str(v).strip()
                return s if s else "-"

            data = []
            data.append(("Olay Anı Farkı", _fmt(full_row[0] if len(full_row) > 0 else None)))

            for i, col in enumerate(self.raw_cols):
                val = full_row[i + 1] if (i + 1) < len(full_row) else None
                data.append((str(col), _fmt(val)))

            # Karşı baz bilgisi analiz tablolarında gösterilmez; sadece detay ekranında üretilir.
            row_map = {c: (full_row[i + 1] if (i + 1) < len(full_row) else None) for i, c in enumerate(self.raw_cols)}
            karsi_baz = self._find_karsi_baz_for_row(row_map)
            data.append(("KARŞI_BAZ", _fmt(karsi_baz)))

            info_text_local = (
                "📊 <b>Görüşme Teknik Detayları:</b> Seçili kaydın ham teknik alanları ve "
                "±3 saniye eşleşmeyle türetilen <b>KARŞI_BAZ</b> bilgisi aşağıdadır."
            )
            dlg = FileDetailPopup(self, data, title="Detaylı Teknik Veri", info_text=info_text_local)
            dlg.exec()
        except Exception as e:
            ModernDialog.show_error(self, "Detay Hatası", str(e))


    def _find_karsi_baz_for_row(self, row_map: dict) -> str:
        t_str = (row_map.get("TARIH") or "").strip()
        num = (row_map.get("NUMARA") or "").strip()
        diger = (row_map.get("DIGER_NUMARA") or "").strip()

        if not t_str or not num or not diger or not self.project_id:
            return ""

        def clean_gsm(val: str) -> str:
            d = re.sub(r"\D", "", str(val))
            return d[-10:] if len(d) >= 10 else d

        n1 = clean_gsm(num)
        n2 = clean_gsm(diger)
        if not n1 or not n2:
            return ""

        dt = None
        for fmt in ("%d.%m.%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(t_str, fmt)
                break
            except Exception:
                continue
        if dt is None:
            return ""

        start_dt = (dt - timedelta(seconds=3)).strftime("%Y-%m-%d %H:%M:%S")
        end_dt = (dt + timedelta(seconds=3)).strftime("%Y-%m-%d %H:%M:%S")

        iso_sql = "substr(TARIH, 7, 4) || '-' || substr(TARIH, 4, 2) || '-' || substr(TARIH, 1, 2) || substr(TARIH, 11)"

        with DB() as conn:
            cur = conn.cursor()
            sql = f"""
                SELECT BAZ
                FROM hts_gsm
                WHERE ProjeID=?
                  AND substr(replace(replace(replace(NUMARA,' ',''),'-',''),'+',''), -10, 10)=?
                  AND substr(replace(replace(replace(DIGER_NUMARA,' ',''),'-',''),'+',''), -10, 10)=?
                  AND ({iso_sql}) BETWEEN ? AND ?
                ORDER BY ({iso_sql}) DESC
                LIMIT 1
            """
            r = cur.execute(sql, (self.project_id, n2, n1, start_dt, end_dt)).fetchone()
            if not r:
                return ""
            return (r[0] or "").strip()

    def open_geo_location_analysis(self):
        if not self.project_id:
            ModernDialog.show_warning(self, "Uyarı", "Önce bir proje seçiniz.")
            return

        selected = self.cmb_gsm.currentText().strip() if hasattr(self, "cmb_gsm") else ""
        is_all = (not selected) or (selected == "Tüm Kayıtlar")
        gsm = "" if is_all else selected

        # 1) HTS yükleme aşamasında hts_ozet tablosuna yazılan analiz aralığını kullan
        init_start_dt, init_end_dt = None, None

        def _parse_qdt(s: str):
            s = (s or "").strip()
            if not s:
                return None
            for fmt in ("dd.MM.yyyy HH:mm:ss", "dd/MM/yyyy HH:mm:ss", "yyyy-MM-dd HH:mm:ss",
                        "dd.MM.yyyy HH:mm", "dd/MM/yyyy HH:mm", "yyyy-MM-dd HH:mm",
                        "dd.MM.yyyy", "dd/MM/yyyy", "yyyy-MM-dd"):
                q = QDateTime.fromString(s, fmt)
                if q.isValid():
                    return q
            return None

        try:
            with DB() as conn:
                cur = conn.cursor()
                if not is_all:
                    row = cur.execute(
                        "SELECT MinDate, MaxDate FROM hts_ozet WHERE ProjeID=? AND GSMNo=?",
                        (self.project_id, gsm)
                    ).fetchone()
                    if row and row[0] and row[1]:
                        q1 = _parse_qdt(str(row[0]))
                        q2 = _parse_qdt(str(row[1]))
                        if q1 and q2:
                            init_start_dt, init_end_dt = q1, q2
                else:
                    # Tüm kayıtlar için projedeki tüm GSM'lerin aralığından min/max türet
                    rows = cur.execute(
                        "SELECT MinDate, MaxDate FROM hts_ozet WHERE ProjeID=?",
                        (self.project_id,)
                    ).fetchall()
                    mins = []
                    maxs = []
                    for r in rows or []:
                        q1 = _parse_qdt(str(r[0]) if r else "")
                        q2 = _parse_qdt(str(r[1]) if r else "")
                        if q1:
                            mins.append(q1)
                        if q2:
                            maxs.append(q2)
                    if mins and maxs:
                        init_start_dt = min(mins, key=lambda x: x.toSecsSinceEpoch())
                        init_end_dt = max(maxs, key=lambda x: x.toSecsSinceEpoch())
        except Exception:
            pass

        # 2) DB'de yoksa mevcut davranış: T0 ± (önce/sonra saat) penceresi
        if init_start_dt is None or init_end_dt is None:
            t0_qt = self.dt_t0.dateTime()  # QDateTime
            before_h = int(self.spin_b.value()) if hasattr(self, "spin_b") else 24
            after_h = int(self.spin_a.value()) if hasattr(self, "spin_a") else 24
            init_start_dt = t0_qt.addSecs(-before_h * 3600)
            init_end_dt = t0_qt.addSecs(after_h * 3600)

        dlg = MapDialog(self, self.project_id, gsm, init_start_dt, init_end_dt)

        # Tüm kayıtlar modunda GSM seçtirmeyelim; baz bilgisi yüklenmesin
        if is_all:
            try:
                dlg.setWindowTitle("Coğrafi Konum Analizi - Tüm Kayıtlar")
            except Exception:
                pass
        else:
            try:
                dlg.load_baz_data()
            except Exception:
                pass

        try:
            dlg.generate_map()
        except Exception:
            pass

        dlg.exec()

    def run_analysis(self):
        if not self.project_id:
            ModernDialog.show_warning(self, "Uyarı", "Önce bir proje seçiniz.")
            return

        t0 = self.dt_t0.dateTime().toPyDateTime()
        c_m = int(self.spin_c.value())
        c_s, c_e = t0 - timedelta(minutes=c_m), t0 + timedelta(minutes=c_m)
        w_s = t0 - timedelta(hours=int(self.spin_b.value()))
        w_e = t0 + timedelta(hours=int(self.spin_a.value()))

        iso_sql = "substr(TARIH, 7, 4) || '-' || substr(TARIH, 4, 2) || '-' || substr(TARIH, 1, 2) || substr(TARIH, 11)"

        params = [self.project_id]
        gsm_filter = ""
        if self.cmb_gsm.currentText() != "Tüm Kayıtlar":
            gsm_filter = " AND GSMNo = ? "
            params.append(self.cmb_gsm.currentText())
        params.extend([w_s.isoformat(sep=" "), w_e.isoformat(sep=" ")])

        select_cols = []
        for c in self.raw_cols:
            if c == "NUMARA":
                select_cols.append("GSMNo as NUMARA")
            else:
                select_cols.append(c)
        select_sql = ", ".join(select_cols)

        query = (
            f"SELECT {select_sql}, {iso_sql} as iso_date "
            f"FROM hts_gsm "
            f"WHERE ProjeID = ?{gsm_filter} "
            f"AND iso_date BETWEEN ? AND ? "
            f"ORDER BY iso_date ASC"
        )

        with DB() as conn:
            rows = conn.execute(query, params).fetchall()

        if not rows:
            ModernDialog.show_info(self, "Bilgi", "Seçilen kriterlere uygun veri bulunamadı.")
            self.lbl_total.setText("0")
            self.lbl_crit.setText("0")
            self.lbl_burst.setText("-")
            self.list_top5.clear()
            self.map_slider.setEnabled(False)
            return

        # ---------------------------------------------------------
        # ✅ DOĞRU FİLTRE: Analiz tablolarında
        # - self kayıt (DIGER_NUMARA==NUMARA) yok
        # - proje içi diğer GSM'lerle olan kayıtlar yok
        # - seçili GSM varsa sadece onun kayıtları var
        # ---------------------------------------------------------
        selected_gsm = self.cmb_gsm.currentText().strip() if hasattr(self, "cmb_gsm") else "Tüm Kayıtlar"

        def _last10(x):
            s = re.sub(r"\D", "", str(x or "")).strip()
            return s[-10:] if len(s) >= 10 else s

        try:
            idx_num = self.raw_cols.index("NUMARA")
        except ValueError:
            idx_num = None

        try:
            idx_other = self.raw_cols.index("DIGER_NUMARA")
        except ValueError:
            idx_other = None

        if idx_num is None or idx_other is None:
            ModernDialog.show_error(self, "Hata", "NUMARA veya DIGER_NUMARA kolonu bulunamadı.")
            return

        # Proje GSM seti: projedeki tüm GSMNo'ların son 10 hanesi
        proj_set = set()
        try:
            with DB() as c:
                for (gsmno,) in c.execute(
                    "SELECT DISTINCT GSMNo FROM hts_gsm WHERE ProjeID=?",
                    (self.project_id,)
                ).fetchall():
                    if gsmno:
                        proj_set.add(_last10(gsmno))
        except Exception:
            proj_set = set()

        target10 = _last10(selected_gsm) if selected_gsm != "Tüm Kayıtlar" else None

        filtered = []
        for r in rows:
            raw = r[:-1]
            if idx_num >= len(raw) or idx_other >= len(raw):
                continue

            num10 = _last10(raw[idx_num])
            other10 = _last10(raw[idx_other])

            # 1) Seçili GSM modunda sadece o numaraya ait kayıtlar
            if target10 and num10 != target10:
                continue

            # 2) Self kayıtlar (DIGER_NUMARA == NUMARA) -> daima çıkar
            if num10 and other10 and num10 == other10:
                continue

            # 3) Proje içi diğer GSM'ler -> daima çıkar
            # (analiz tablolarında "karşı baz/karşı kayıt" görünmesin)
            if proj_set and other10 in proj_set:
                continue

            filtered.append(r)

        rows = filtered

        # Filtre sonrası boş kaldıysa aynı "veri yok" davranışı
        if not rows:
            ModernDialog.show_info(self, "Bilgi", "Seçilen kriterlere uygun veri bulunamadı.")
            self.lbl_total.setText("0")
            self.lbl_crit.setText("0")
            self.lbl_burst.setText("-")
            self.list_top5.clear()
            self.map_slider.setEnabled(False)
            return
        # ---------------------------------------------------------

        self.data_store = {"before": [], "crit": [], "after": []}
        disp_data = {"before": [], "crit": [], "after": []}
        c_nums = []

        # idx_other zaten var ve doğru; tekrar aramaya gerek yok
        for r in rows:
            try:
                dt = datetime.fromisoformat(r[-1])
            except Exception:
                continue

            delta = int((dt - t0).total_seconds() / 60)
            delta_str = f"{'+' if delta >= 0 else ''}{delta} dk"

            full_row = (delta_str,) + r[:-1]

            visible_row = [delta_str]
            for i, col_name in enumerate(self.raw_cols):
                if col_name in self.display_cols:
                    visible_row.append(r[i])

            if dt < c_s:
                self.data_store["before"].append(full_row)
                disp_data["before"].append(visible_row)
            elif c_s <= dt <= c_e:
                self.data_store["crit"].append(full_row)
                disp_data["crit"].append(visible_row)
                try:
                    c_nums.append(r[idx_other])
                except Exception:
                    pass
            else:
                self.data_store["after"].append(full_row)
                disp_data["after"].append(visible_row)

        self.t_before.set_data(disp_data["before"])
        self.t_crit.set_data(disp_data["crit"])
        self.t_after.set_data(disp_data["after"])
        self.disp_store = disp_data

        n_total = len(rows)
        n_crit = len(disp_data["crit"])

        h_before = int(self.spin_b.value())
        h_after = int(self.spin_a.value())
        m_crit = int(self.spin_c.value())

        rate_crit = n_crit / (m_crit / 60.0) if m_crit > 0 else 0
        n_normal = len(disp_data["before"]) + len(disp_data["after"])
        rate_normal = n_normal / (h_before + h_after) if (h_before + h_after) > 0 else 0
        burst_idx = rate_crit / rate_normal if rate_normal > 0 else n_crit

        self.lbl_total.setText(str(n_total))
        self.lbl_crit.setText(str(n_crit))
        burst_text = f"x{burst_idx:.1f} Artış"
        self.lbl_burst.setText(burst_text)
        self.lbl_burst.setStyleSheet(
            f"color: {'#e74c3c' if burst_idx > 2 else '#f39c12'}; font-size: 20px; font-weight: bold;"
        )

        self.list_top5.clear()
        for n, c in Counter(c_nums).most_common(5):
            self.list_top5.addItem(f"📞 {n} ({c} İşlem)")

        self.idx_other = idx_other

        if n_crit > 0:
            self.map_slider.setEnabled(True)
            self.map_slider.setRange(0, n_crit - 1)
            self.map_slider.setValue(0)
            self.lbl_slider_info.setText(f"Kritik pencerede {n_crit} kayıt bulundu. Kaydırarak izleyin.")
        else:
            self.map_slider.setEnabled(False)
            self.lbl_slider_info.setText("Kritik pencerede kayıt yok.")

        self._update_rich_map(
            self.data_store.get("before", []),
            self.data_store.get("crit", []),
            self.data_store.get("after", []),
            focus_row=self._last_focus_row,
            focus_label=self._last_focus_label,
            use_cluster=False
        )
        QApplication.processEvents()

    def on_slider_moved(self, value):
        if not self.data_store.get("crit") or value >= len(self.data_store["crit"]):
            return

        self.tabs.setCurrentIndex(1)
        self.t_crit.table.selectRow(value)

        model_index = self.t_crit.table.model().index(value, 0)
        if model_index.isValid():
            self.t_crit.table.scrollTo(model_index)

        row_data = self.data_store["crit"][value]
        self._last_focus_row = row_data
        self._last_focus_label = "Kritik"
        other_num = row_data[self.idx_other + 1]
        self.lbl_slider_info.setText(f"⌚ Olaydan {row_data[0]} sonra | Hedef: {other_num}")

        # slider focus: popup açık gelecek şekilde yenile
        self._focus_map_to_row(row_data, label="Kritik")

    def _fetch_custom_markers(self):
        """
        Coğrafi Konum Analizi ekranında kaydedilen kullanıcı işaretlerini çeker.
        Tablo: ozel_konumlar (ProjeID, GSMNo, Lat, Lon, Label)
        - Hedef GSM seçiliyse sadece o GSM'nin işaretleri
        - 'Tüm Kayıtlar' ise proje içindeki tüm işaretler
        """
        if not getattr(self, "project_id", None):
            return []

        gsm_filter = None
        try:
            if hasattr(self, "cmb_gsm"):
                t = (self.cmb_gsm.currentText() or "").strip()
                if t and t != "Tüm Kayıtlar":
                    gsm_filter = t
        except Exception:
            gsm_filter = None

        with DB() as conn:
            if gsm_filter:
                rows = conn.execute(
                    "SELECT Lat, Lon, Label, GSMNo FROM ozel_konumlar WHERE ProjeID=? AND GSMNo=? ORDER BY id DESC",
                    (self.project_id, gsm_filter)
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT Lat, Lon, Label, GSMNo FROM ozel_konumlar WHERE ProjeID=? ORDER BY id DESC",
                    (self.project_id,)
                ).fetchall()

        # rows: [(lat, lon, label, gsmno), ...]
        out = []
        for r in rows or []:
            try:
                lat = float(r[0]); lon = float(r[1])
                label = (r[2] or "").strip()
                gsmno = (r[3] or "").strip()
                if label:
                    out.append((lat, lon, label, gsmno))
            except Exception:
                continue
        return out

    def _update_rich_map(self, before, crit, after, focus_row=None, focus_label=None, use_cluster=True):
        import folium
        import io
        from folium.plugins import MarkerCluster

        try:
            b_idx = 1 + self.raw_cols.index("BAZ")
            o_idx = 1 + self.raw_cols.index("DIGER_NUMARA")
            t_idx = 1 + self.raw_cols.index("TARIH")
        except ValueError:
            m = folium.Map(location=[39.0, 35.0], zoom_start=6)
            _enable_measure_and_balloons(m)
            data = io.BytesIO()
            m.save(data, close_file=False)
            self.map_view.setHtml(data.getvalue().decode())
            return

        all_points = []

        def _collect_points(rows):
            for r in rows or []:
                baz_val = r[b_idx] if b_idx < len(r) else None
                pos = self._parse_coords(baz_val)
                if pos:
                    all_points.append(pos)

        _collect_points(before)
        _collect_points(crit)
        _collect_points(after)

        focus_pos = None
        if focus_row:
            baz_val = focus_row[b_idx] if b_idx < len(focus_row) else None
            focus_pos = self._parse_coords(baz_val)

        # ⭐ Kullanıcı işaretlerini de al (ozel_konumlar)
        custom_markers = []
        try:
            custom_markers = self._fetch_custom_markers()
        except Exception:
            pass

        center = focus_pos or (all_points[0] if all_points else [39.0, 35.0])
        m = folium.Map(location=center, zoom_start=14 if focus_pos else (12 if all_points else 6))

        m.get_root().header.add_child(folium.Element(
            '<script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>'
        ))

        m.get_root().header.add_child(folium.Element("""
            <style>
              .leaflet-popup-content { font-family: Segoe UI, Arial; font-size: 12px; line-height: 1.35; }
              .leaflet-popup-content-wrapper { border-radius: 12px; box-shadow: 0 4px 14px rgba(0,0,0,0.25); }
              .hts-popup { min-width: 280px; }
              .hts-popup .title { font-weight: 800; margin-bottom: 6px; font-size: 13px; }
              .hts-popup .row { margin: 2px 0; }
              .hts-popup .k { color:#555; font-weight:700; }
              .hts-popup .v { color:#111; }
            </style>
        """))

        def _is_focus(r):
            if focus_row is None:
                return False
            if r is focus_row:
                return True
            try:
                return (r[0] == focus_row[0]) and (r[o_idx] == focus_row[o_idx]) and (r[t_idx] == focus_row[t_idx])
            except Exception:
                return False

        def _add_markers(rows, color, label):
            layer = MarkerCluster(name=label).add_to(m) if use_cluster else folium.FeatureGroup(name=label).add_to(m)

            for r in rows or []:
                baz_val = r[b_idx] if b_idx < len(r) else None
                pos = self._parse_coords(baz_val)
                if not pos:
                    continue

                other_num = r[o_idx] if o_idx < len(r) else ""
                t_val = r[t_idx] if t_idx < len(r) else ""
                delta_val = r[0] if len(r) > 0 else ""

                popup_html = f"""
                  <div class="hts-popup">
                    <div class="title">{label}</div>
                    <div class="row"><span class="k">Kişi:</span> <span class="v">{other_num}</span></div>
                    <div class="row"><span class="k">Saat:</span> <span class="v">{t_val}</span></div>
                    <div class="row"><span class="k">Delta:</span> <span class="v">{delta_val}</span></div>
                  </div>
                """

                folium.Marker(
                    pos,
                    popup=folium.Popup(
                        popup_html,
                        max_width=420,
                        min_width=280,
                        show=_is_focus(r),
                        auto_close=False,
                        close_button=True
                    ),
                    icon=folium.Icon(color=color, icon="phone", prefix="fa"),
                ).add_to(layer)

        # 3 renk
        _add_markers(crit, "red", "Kritik (Olay anı penceresi)")
        _add_markers(before, "blue", "Önce (pencere)")
        _add_markers(after, "purple", "Sonra (pencere)")

        # ⭐ Kullanıcı işaretleri katmanı
        if custom_markers:
            fg_custom = folium.FeatureGroup(name="⭐ Kullanıcı İşaretleri", show=True).add_to(m)
            for (lat, lon, label, gsmno) in custom_markers:
                popup_html = f"""
                  <div class="hts-popup">
                    <div class="title">⭐ Kullanıcı İşareti</div>
                    <div class="row"><span class="k">Etiket:</span> <span class="v">{label}</span></div>
                    <div class="row"><span class="k">GSM:</span> <span class="v">{gsmno or '-'}</span></div>
                  </div>
                """
                folium.Marker(
                    [lat, lon],
                    popup=folium.Popup(popup_html, max_width=420, min_width=280, auto_close=False, close_button=True),
                    icon=folium.Icon(color="orange", icon="star", prefix="fa"),
                ).add_to(fg_custom)

        if focus_pos:
            m.location = focus_pos
        else:
            if all_points:
                m.fit_bounds(all_points, padding=(40, 40))

        try:
            folium.LayerControl(collapsed=True).add_to(m)
        except Exception:
            pass

        data = io.BytesIO()
        m.save(data, close_file=False)
        self.map_view.setHtml(data.getvalue().decode())

    def _parse_coords(self, text):
        if not text:
            return None
        found = re.findall(r"(\d{2}[\.,]\d+)", str(text))
        if len(found) >= 2:
            try:
                return [float(found[-2].replace(",", ".")), float(found[-1].replace(",", "."))]
            except Exception:
                return None
        return None

    def _load_project_gsms(self):
        if not self.project_id:
            return
        self.cmb_gsm.clear()
        self.cmb_gsm.addItem("Tüm Kayıtlar")
        with DB() as conn:
            rows = conn.execute("SELECT DISTINCT GSMNo FROM hts_gsm WHERE ProjeID=?", (self.project_id,)).fetchall()
            for r in rows:
                self.cmb_gsm.addItem(str(r[0]))


class AnalysisCenter(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main = main_window
        self.current_project_id = None
        self.current_gsm_number = None
        self.is_uploading = False
        self.shutdown_pending = False
        self.upload_queue = []
        self.current_heatmap_data = None
        self.total_count = 0
        self.success_count = 0
        self.flash_state = False
        self.flash_timer = QTimer(); self.flash_timer.timeout.connect(self.toggle_warning_animation)
        self.hide_timer = QTimer(); self.hide_timer.setSingleShot(True); self.hide_timer.timeout.connect(self.stop_warning_animation)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)

        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.left_widget = QWidget()
        self.left_widget.setMinimumWidth(320)
        self.left_widget.setMaximumWidth(400)
        left_layout = QVBoxLayout(self.left_widget)
        left_layout.setContentsMargins(4, 4, 4, 4)

        btn_back = QPushButton("🔙 Projelere Dön")
        btn_back.clicked.connect(self.go_back_to_projects)
        left_layout.addWidget(btn_back)

        self.btn_load = QPushButton("📥 HTS Kaydı Yükle")
        self.btn_load.setStyleSheet("background-color: #d35400; font-weight: bold; padding: 8px; color: white;")
        self.btn_load.clicked.connect(self.upload_excel)
        left_layout.addWidget(self.btn_load)

        self.num_table = GenericDatabaseTable(["No", "Abone"], chart_mode='none', enable_evidence_menu=False)
        self.num_table.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

        self.num_table.table.clicked.connect(self.on_num_select)
        left_layout.addWidget(self.num_table)

        btn_style_base = "QPushButton { color: white; font-weight: bold; padding: 8px; border-radius: 4px; margin-top: 5px; border:none; "

        btn_del_gsm = QPushButton("🗑️ Sil")
        btn_del_gsm.setStyleSheet(
            btn_style_base + "background-color: #c0392b !important; } "
            "QPushButton:hover { background-color: #e74c3c !important; } "
            "QPushButton:pressed { background-color: #b03a2e !important; }"
        )
        btn_del_gsm.clicked.connect(self.delete_current_gsm)
        left_layout.addWidget(btn_del_gsm)

        btn_cross = QPushButton("🔗 Ortak Temas ve İlişki Analizi")
        btn_cross.setStyleSheet(
            btn_style_base + "background-color: #8e44ad !important; } "
            "QPushButton:hover { background-color: #9b59b6 !important; } "
            "QPushButton:pressed { background-color: #7d3c98 !important; }"
        )
        btn_cross.clicked.connect(self.open_cross_match)
        left_layout.addWidget(btn_cross)

        btn_heat = QPushButton("🔥 Aktivite Isı Haritası(Yoğunluk Analizi)")
        btn_heat.setStyleSheet(
            btn_style_base + "background-color: #e67e22 !important; } "
            "QPushButton:hover { background-color: #f39c12 !important; } "
            "QPushButton:pressed { background-color: #d66c1b !important; }"
        )
        btn_heat.clicked.connect(self.open_heatmap_popup)
        left_layout.addWidget(btn_heat)

        btn_speed = QPushButton("🚀 Hız/Mesafe İhlali (Impossible Travel)")
        btn_speed.setStyleSheet(
            btn_style_base + "background-color: #c0392b !important; } "
            "QPushButton:hover { background-color: #e74c3c !important; } "
            "QPushButton:pressed { background-color: #922b21 !important; }"
        )
        btn_speed.clicked.connect(self.open_speed_anomaly)
        left_layout.addWidget(btn_speed)

        btn_stalk = QPushButton("🚫 Taciz / Tek Yönlü Arama Analizi")
        btn_stalk.setStyleSheet(
            btn_style_base + "background-color: #b71c1c !important; } " # Koyu Kırmızı
            "QPushButton:hover { background-color: #c62828 !important; } "
            "QPushButton:pressed { background-color: #880e4f !important; }"
        )
        btn_stalk.clicked.connect(self.open_stalking_analysis)
        left_layout.addWidget(btn_stalk)

        btn_map = QPushButton("🌍 Coğrafi Konum Analizi")
        btn_map.setStyleSheet(
            btn_style_base + "background-color: #2c3e50 !important; } "
            "QPushButton:hover { background-color: #34495e !important; } "
            "QPushButton:pressed { background-color: #1e2b35 !important; }"
        )
        btn_map.clicked.connect(self.open_map_view)
        left_layout.addWidget(btn_map)

        btn_route = QPushButton("📅 Güzergah Analizi(Hareket Dökümü)")
        btn_route.setStyleSheet(
            btn_style_base + "background-color: #16a085 !important; } "
            "QPushButton:hover { background-color: #1abc9c !important; } "
            "QPushButton:pressed { background-color: #117a65 !important; }"
        )
        btn_route.clicked.connect(self.open_daily_route)
        left_layout.addWidget(btn_route)

        btn_report = QPushButton("📝 Rapor ve Tasarım Merkezi")
        btn_report.clicked.connect(self.open_report_center)
        btn_report.setStyleSheet(
            btn_style_base + "background-color: #27ae60 !important; } "
            "QPushButton:hover { background-color: #2ecc71 !important; } "
            "QPushButton:pressed { background-color: #21a65f !important; }"
        )

        left_layout.addWidget(btn_report)

        self.splitter.addWidget(self.left_widget)

        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0,0,0,0)
        right_layout.setSpacing(0)

        top_bar = QFrame()
        top_bar.setFixedHeight(50)
        top_bar.setStyleSheet("background-color: #ecf0f1; border-bottom: 1px solid #bdc3c7;")
        tb_layout = QHBoxLayout(top_bar)
        tb_layout.setContentsMargins(10, 0, 10, 0)

        self.btn_toggle_menu = QPushButton("☰ Menü")
        self.btn_toggle_menu.setCheckable(True)
        self.btn_toggle_menu.setChecked(True)
        self.btn_toggle_menu.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_toggle_menu.setStyleSheet("""
            QPushButton { border: none; font-size: 16px; font-weight: bold; color: #2c3e50; background: transparent; text-align: left; }
            QPushButton:hover { color: #3498db; }
        """)
        self.btn_toggle_menu.clicked.connect(self.toggle_left_panel)
        tb_layout.addWidget(self.btn_toggle_menu)

        self.top_actions_widget = QWidget()
        self.top_actions_widget.setVisible(False)
        top_actions_layout = QHBoxLayout(self.top_actions_widget)
        top_actions_layout.setContentsMargins(10, 0, 0, 0)
        top_actions_layout.setSpacing(5)

        def create_top_btn(text, func, color):
            b = QPushButton(text)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.clicked.connect(func)
            b.setStyleSheet(f"""
                QPushButton {{ background-color: {color}; color: white; font-weight: bold; padding: 6px 12px; border-radius: 4px; border: none; font-size: 12px; }}
                QPushButton:hover {{ opacity: 0.8; }}
            """)
            top_actions_layout.addWidget(b)
            return b

        create_top_btn("🔗 Ortak Temas ve İlişki Analizi", self.open_cross_match, "#8e44ad")

        create_top_btn("🔥 Isı Haritası (Yoğunluk Analizi)", self.open_heatmap_popup, "#e67e22")

        create_top_btn("🚀 Hız/Mesafe İhlali (Imp. Travel)", self.open_speed_anomaly, "#c0392b")

        create_top_btn("🚫 Taciz/Tek Yönlü Arama Analizi", self.open_stalking_analysis, "#b71c1c")

        create_top_btn("🌍 Coğrafi Konum Analizi", self.open_map_view, "#2c3e50")

        create_top_btn("📅 Güzergah Analizi (Hareket Dökümü)", self.open_daily_route, "#16a085")

        create_top_btn("📝 Rapor ve Tasarım Merkezi", self.open_report_center, "#27ae60")

        tb_layout.addWidget(self.top_actions_widget)
        tb_layout.addStretch()

        lbl_title = QLabel("HTSMercek Adli İletişim Kayıtları Analiz Merkezi")
        lbl_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #7f8c8d;")
        tb_layout.addWidget(lbl_title)

        right_layout.addWidget(top_bar)

        self.tabs = QTabWidget()
        self.tab_widgets = {}

        self.overview_widget = QWidget()
        self.overview_widget.setStyleSheet("background-color: transparent;")
        overview_layout = QVBoxLayout(self.overview_widget)
        overview_layout.setContentsMargins(10, 10, 10, 10)
        overview_layout.setSpacing(15)

        v_splitter = QSplitter(Qt.Orientation.Vertical)

        top_container = QWidget()
        top_container.setStyleSheet("background-color: transparent;")
        top_cont_layout = QVBoxLayout(top_container)
        top_cont_layout.setContentsMargins(0, 0, 0, 0)
        top_cont_layout.setSpacing(10)

        self.abone_cols = [
            "NUMARA", "DURUM", "AD", "SOYAD", "ADRES",
            "DOGUM_TARIHI", "DOGUM_YERI", "ILCE", "IL", "TC_KIMLIK_NO",
            "ANNE_ADI", "BABA_ADI", "ABONE_SORGU_ARALIGI",
            "ABONE_BASLANGIC", "ABONE_BITIS", "OPERATOR"
        ]
        self.abone_table = GenericDatabaseTable(self.abone_cols, chart_mode='none', enable_evidence_menu=False)
        self.abone_table.hide_toolbar()
        self.abone_table.table.setColumnHidden(0, True)
        self.abone_table.table.doubleClicked.connect(self.open_abone_detail_popup)
        h_abone = QHBoxLayout()

        lbl_abone = QLabel("👤 Hat Sahibi (Abone) Bilgileri")
        lbl_abone.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")

        info_abone = InfoButton(
            "<b>👤 Abone ve Kimlik Bilgileri:</b><br>"
            "Bu alan, analiz edilen hattın yasal sahibini (Müşteri) gösterir.<br>"
            "• <b>Adresler:</b> Operatörde kayıtlı fatura ve ikamet adresleri burada yer alır.<br>"
            "• <b>Profil Kartı (Alt Kısım):</b> Sinyal verilerine göre sistemin yapay zeka ile tahmin ettiği <b>'Muhtemel Ev'</b> ve <b>'İş'</b> adreslerini gösterir."
        )

        h_abone.addWidget(lbl_abone)
        h_abone.addWidget(info_abone)

        h_abone.addStretch()

        self.abone_table.lbl_count.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        h_abone.addWidget(self.abone_table.lbl_count)

        top_cont_layout.addLayout(h_abone)

        self.abone_table.setFixedHeight(140)
        self.abone_table.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.abone_table.table.verticalHeader().setDefaultSectionSize(35)

        top_cont_layout.addWidget(self.abone_table)

        self.profile_card = ProfileCard()
        top_cont_layout.addWidget(self.profile_card)

        top_container.setMinimumHeight(220)
        top_container.setMaximumHeight(240)

        v_splitter.addWidget(top_container)

        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        filter_frame = QFrame()
        filter_frame.setFixedHeight(45)
        filter_frame.setStyleSheet("background-color: #fff8e1; border: 1px solid #ffcc80; border-radius: 6px; margin-top: 5px;")

        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.setContentsMargins(10, 2, 10, 2)
        filter_layout.setSpacing(10)

        filter_layout.addWidget(QLabel("📅 Analiz Aralığı:", styleSheet="font-weight: bold; color: #d35400;"))

        self.dt_start = QDateTimeEdit(); self.dt_start.setDisplayFormat("dd.MM.yyyy HH:mm:ss"); self.dt_start.setCalendarPopup(True)
        self.dt_end = QDateTimeEdit(); self.dt_end.setDisplayFormat("dd.MM.yyyy HH:mm:ss"); self.dt_end.setCalendarPopup(True)

        filter_layout.addWidget(self.dt_start); filter_layout.addWidget(QLabel("➡")); filter_layout.addWidget(self.dt_end)

        btn_refresh = QPushButton("Analizi Güncelle")
        btn_refresh.setStyleSheet("background-color: #2980b9; color: white; font-weight: bold; padding: 5px 12px; border-radius: 4px;")
        btn_refresh.clicked.connect(self.refresh_all_analysis)
        filter_layout.addWidget(btn_refresh)

        info_refresh = InfoButton(
            "<b>🔄 Dinamik Analiz Güncelleme:</b><br>"
            "Bu buton, sol tarafta seçtiğiniz <b>Tarih Aralığına</b> göre tüm analiz tablolarını yeniden hesaplar.<br><br>"
            "• <b>Ne İşe Yarar?:</b> Tüm veriler yerine sadece olay anına veya suç tarihine odaklanmanızı sağlar.<br>"
            "• <b>Etkilenenler:</b> Tüm sekmeler, 'Rehber', 'En Sık Görüşülenler', 'Baz İstasyonu Yoğunluğu' ve 'IMEI Kullanımı' listeleri bu tarihe göre filtrelenir.<br>"
            "• <b>İpucu:</b> Tarihleri değiştirdikten sonra bu butona basınız."
        )
        filter_layout.addWidget(info_refresh)

        self.lbl_warning = QLabel("⚠️ +1000 KAYIT!")
        self.lbl_warning.setVisible(False)
        filter_layout.addWidget(self.lbl_warning); filter_layout.addStretch()

        bottom_layout.addWidget(filter_frame)

        h_splitter = QSplitter(Qt.Orientation.Horizontal)
        left_v_splitter = QSplitter(Qt.Orientation.Vertical)

        self.comm_tabs = QTabWidget()
        self.contact_table = GenericDatabaseTable(["Karşı Numara", "Temas Sayısı", "Süre", "Kişi Adı"], chart_mode='popup')
        self.contact_table.table.doubleClicked.connect(self.open_interaction_detail)
        self.comm_tabs.addTab(self.contact_table, "📖 Rehber")

        self.top_table = GenericDatabaseTable(["Karşı Numara", "Temas Sayısı", "Süre", "Kişi Adı"], chart_mode='popup')
        self.top_table.table.doubleClicked.connect(self.open_interaction_detail)
        self.comm_tabs.addTab(self.top_table, "🔥 En Sık İletişim Kurulan GSM Numaraları (Top 20)")
        left_v_splitter.addWidget(self.comm_tabs)

        self.sub_tabs = QTabWidget()
        info_name = (
            "<b>👥 Aynı İsimli Numaralar Analizi:</b><br>"
            "Proje veritabanındaki tüm abone ve rehber kayıtlarını (HTS_Rehber, HTS_Abone) tarar.<br><br>"
            "• <b>Mantık:</b> Aynı 'Ad Soyad' bilgisine sahip birden fazla <u>farklı</u> GSM numarası tespit edilirse burada listelenir.<br>"
            "• <b>Amaç:</b> Şüphelinin, farklı numaralar kullansa bile ismen kendini ele verdiği diğer hatlarını tespit etmektir."
        )
        self.common_name_table = GenericDatabaseTable(["Ad Soyad", "Hat Sayısı", "Numaralar"], chart_mode='none', info_text=info_name)
        t = self.common_name_table.table
        h = t.horizontalHeader()
        t.setWordWrap(False)
        t.setTextElideMode(Qt.TextElideMode.ElideNone)
        h.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        h.setStretchLastSection(False)
        self.sub_tabs.addTab(self.common_name_table, "👥 Aynı İsme Kayıtlı GSM Numaraları")
        info_tc = (
            "<b>🆔 TC Kimlik Çapraz Sorgusu:</b><br>"
            "Operatörden gelen yasal abonelik verilerini analiz eder.<br><br>"
            "• <b>Mantık:</b> Aynı TC Kimlik Numarası üzerine kayıtlı olan tüm GSM hatlarını bir araya getirir.<br>"
            "• <b>Kullanım:</b> Şahsın üzerine kayıtlı olup da soruşturma dosyasında bilinmeyen 'Hayalet Hatları' ortaya çıkarmak için kullanılır."
        )
        self.common_tc_table = GenericDatabaseTable(["TC Kimlik - Ad Soyad", "Hat Sayısı", "Numaralar"], chart_mode='none', info_text=info_tc)
        t = self.common_tc_table.table
        h = t.horizontalHeader()
        t.setWordWrap(False)
        t.setTextElideMode(Qt.TextElideMode.ElideNone)
        h.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        h.setStretchLastSection(False)
        self.sub_tabs.addTab(self.common_tc_table, "🆔 Aynı TC Kimlik Numarasına Kayıtlı GSM Numaraları")
        self.common_name_table.table.doubleClicked.connect(self.open_common_subscriber_detail)
        self.common_tc_table.table.doubleClicked.connect(self.open_common_subscriber_detail)
        left_v_splitter.addWidget(self.sub_tabs)

        h_splitter.addWidget(left_v_splitter)

        right_v_splitter = QSplitter(Qt.Orientation.Vertical)
        self.baz_tabs = QTabWidget()
        self.all_baz_table = GenericDatabaseTable(["Baz İstasyonu / Konum", "Sinyal", "Gizli Baz Adı"], chart_mode='popup')
        self.all_baz_table.table.setColumnHidden(2, True)
        self.all_baz_table.table.doubleClicked.connect(self.open_location_detail)
        self.baz_tabs.addTab(self.all_baz_table, "📡 Tüm Baz İstasyonları")

        self.baz_table = GenericDatabaseTable(["Baz İstasyonu / Konum", "Sinyal", "Gizli Baz Adı"], chart_mode='popup')
        self.baz_table.table.setColumnHidden(2, True)
        self.baz_table.table.doubleClicked.connect(self.open_location_detail)
        self.baz_tabs.addTab(self.baz_table, "🔥 En Sık İletişim Kurulan Baz İstasyonları (Top 20)")
        right_v_splitter.addWidget(self.baz_tabs)

        for baz_tbl in (self.all_baz_table.table, self.baz_table.table):
            h = baz_tbl.horizontalHeader()

            baz_tbl.setColumnWidth(0, 800)
            h.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
            h.setMinimumSectionSize(500)
            h.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
            baz_tbl.setColumnWidth(1, 90)
            h.setMinimumSectionSize(60)
            h.setStretchLastSection(False)

        self.imei_tabs = QTabWidget()
        info_imei = (
            "<b>📱 Proje Cihaz Envanteri (Tüm IMEI'ler):</b><br>"
            "Bu projeye yüklenen HTS kayıtlarında (Arama, SMS, İnternet) sinyal vermiş <u>bütün</u> cihazların listesidir.<br><br>"
            "• <b>Kaynak:</b> Hedef şahısların taktığı veya karşı tarafın kullandığı tüm cihazları kapsar.<br>"
            "• <b>Veri:</b> Hangi cihazın (IMEI) toplamda kaç farklı hat ile kullanıldığını gösterir."
        )
        self.imei_table = GenericDatabaseTable(["IMEI No", "Kullanım", "İlk Tarih", "Son Tarih"], chart_mode='none', info_text=info_imei)
        t = self.imei_table.table
        h = t.horizontalHeader()
        t.setWordWrap(False)
        t.setTextElideMode(Qt.TextElideMode.ElideNone)
        h.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        h.setStretchLastSection(False)
        self.imei_table.table.doubleClicked.connect(self.open_imei_detail)
        self.imei_tabs.addTab(self.imei_table, "📱 IMEI Listesi")

        info_common_imei = (
            "<b>⚠️ Ortak Cihaz (Örgütsel Bağ) Analizi:</b><br>"
            "Birden fazla hedef şahsın, SIM kartlarını değiştirerek kullandığı 'Ortak Cihazları' tespit eder.<br><br>"
            "<b>🔍 Tespit Mantığı (Akıllı Filtre):</b><br>"
            "1. Bir cihazı en az 2 farklı hedef numara kullanmış olmalı.<br>"
            "2. <b>Sahiplik Testi:</b> Cihazın 'Ortak' sayılması için, o cihazla <u>Giden Arama</u> veya <u>İnternet</u> kullanımı yapılmış olmalıdır.<br>"
            "<i>(Sadece dışarıdan aranan ve hiç işlem yapmayan cihazlar 'Karşı Taraf' sayılarak elenir.)</i>"
        )
        self.common_imei_table = GenericDatabaseTable(["Ortak IMEI", "Kullanan", "Numaralar", "İşlem"], chart_mode='none', info_text=info_common_imei)
        t = self.common_imei_table.table
        h = t.horizontalHeader()
        t.setWordWrap(False)
        t.setTextElideMode(Qt.TextElideMode.ElideNone)
        h.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        h.setStretchLastSection(False)
        self.common_imei_table.table.doubleClicked.connect(lambda idx: self.open_imei_detail(idx, is_common=True))
        self.imei_tabs.addTab(self.common_imei_table, "⚠️ Ortak IMEI")

        info_swap = (
            "<b>🔁 IMEI Swap Timeline (Cihaz–Hat Zaman Çizelgesi):</b><br>"
            "Bu sekme, <b>Ortak IMEI</b> analizinden gelen cihazların proje verisi içinde "
            "<b>hangi GSM numaraları tarafından</b> ve <b>hangi zaman aralıklarında</b> "
            "kullanıldığını segment segment gösterir.<br><br>"
            "<b>Ne işe yarar?</b><br>"
            "• Aynı IMEI’nin farklı hatlara <b>ne zaman geçtiğini</b> (cihaz el değiştirme / SIM değişimi)<br>"
            "• Bir cihazın <b>birden fazla kişi/hat tarafından</b> kullanılıp kullanılmadığını<br>"
            "• Kısa aralıklı veya tekrarlayan <b>hat rotasyonu</b> paternlerini<br>"
            "• Aynı IMEI’nin farklı GSM’lerde <b>zaman çakışması</b> olup olmadığını ortaya koyar.<br><br>"
            "<b>Tablo sütunları</b><br>"
            "• <b>IMEI</b>: İncelenen cihaza ait IMEI değeri<br>"
            "• <b>Kullanan</b>: Bu IMEI’yi kullanan farklı GSM sayısı<br>"
            "• <b>Zaman Çizelgesi</b>: Her GSM için IMEI’nin ilk ve son görülme aralığı ve kayıt yoğunluğu<br>"
            "• <b>Çakışma</b>: Aynı IMEI’nin farklı GSM’lerde zaman aralığı çakışması varsa <b>VAR</b> olarak işaretlenir.<br><br>"
            "<b>Veri kaynakları</b><br>"
            "Bu analiz, projedeki HTS kayıtlarında yer alan IMEI alanları kullanılarak "
            "<b>GSM, GPRS ve WAP</b> tabloları üzerinden oluşturulur.<br><br>"
            "<b>Kullanım</b><br>"
            "• Bu sekme <b>yükleme sırasında çalışmaz</b>; yalnızca sekmeye girildiğinde (lazy) hesaplanır.<br>"
            "• Bir satıra <b>çift tıklayarak</b>, ilgili IMEI’nin detaylı zaman çizelgesini açabilirsiniz.<br><br>"
            "<b>Analiz notu</b><br>"
            "Zaman çakışması veya sık hat değişimi görülen IMEI’ler, HTS incelemesinde "
            "<b>öncelikli değerlendirilmesi gereken</b> cihazlar olarak ele alınmalıdır."
        )

        self.imei_swap_table = GenericDatabaseTable(
            headers=["IMEI", "Kullanan", "Zaman Çizelgesi", "Çakışma"],
            chart_mode='none',
            info_text=info_swap
        )
        self.imei_tabs.addTab(self.imei_swap_table, "🔁 IMEI Swap Timeline")

        self.imei_tabs.currentChanged.connect(self.on_imei_tabs_changed)
        self.imei_swap_table.table.doubleClicked.connect(self.open_imei_swap_timeline_detail)

        right_v_splitter.addWidget(self.imei_tabs)

        h_splitter.addWidget(right_v_splitter)
        bottom_layout.addWidget(h_splitter)
        v_splitter.addWidget(bottom_widget)

        v_splitter.setStretchFactor(0, 0)
        v_splitter.setStretchFactor(1, 1)
        v_splitter.setSizes([240, 700])

        overview_layout.addWidget(v_splitter)
        icon = self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogContentsView)
        self.tabs.addTab(self.overview_widget, icon, "Genel Bakış")
        self.tab_definitions = {
            "GSM": ["SIRA NO", "NUMARA", "TİP", "DİĞER NUMARA", "TARİH", "SÜRE", "İsim Soyisim", "TC Kimlik No", "IMEI", "BAZ"],
            "SMS": ["SIRA NO", "NUMARA", "TİP", "DİĞER NUMARA", "TARİH", "SÜRE", "İsim Soyisim", "TC Kimlik No", "BOYUT", "İÇERİK TİPİ"],
            "Sabit": ["SIRA NO", "NUMARA", "TİP", "DİĞER NUMARA", "TARİH", "SÜRE", "İsim Soyisim", "TC Kimlik No"],
            "Uluslararası": ["SIRA NO", "NUMARA", "TİP", "DİĞER NUMARA", "TARİH", "SÜRE", "İsim Soyisim", "TC Kimlik No"],
            "İnternet": ["TUR", "SIRA NO", "NUMARA", "TİP", "TARİH", "SÜRE", "IMEI", "KAYNAK IP", "HEDEF IP", "GÖNDERME", "İNDİRME", "BAZ"],
            "STH": ["SIRA NO", "NUMARA", "TİP", "DİĞER NUMARA", "TARİH", "SÜRE", "OPERATÖR", "İsim Soyisim", "TC Kimlik No", "DATA TİP", "DURUM", "PİN NO", "B. GATEWAY", "S. GATEWAY", "B. SANTRAL", "S. SANTRAL"]
        }

        self.tab_descriptions = {
            "GSM": (
                "<b>📞 GSM (Sesli Görüşme) Kayıtları:</b><br>"
                "Hedef numaranın yaptığı tüm gelen ve giden aramaları listeler.<br>"
                "• <b>Analiz İpucu:</b> Sık görüşülen numaraları ve uzun süreli çağrıları inceleyin.<br>"
                "• <b>Konum:</b> 'Baz İstasyonu' sütunu, görüşme anındaki tahmini konumu verir.<br>"
                "• <b>Detay:</b> Satırlara çift tıklayarak harita üzerinde konum analizi yapabilirsiniz."
            ),
            "SMS": (
                "<b>💬 SMS (Kısa Mesaj) Kayıtları:</b><br>"
                "Gönderilen ve alınan SMS trafiğini gösterir.<br>"
                "• <b>Yasal Uyarı:</b> Mesaj içerikleri (metin) operatör kayıtlarında YER ALMAZ.<br>"
                "• Burada sadece kiminle ve ne zaman mesajlaşıldığı bilgisi bulunur."
            ),
            "Sabit": (
                "<b>☎️ Sabit Hat Görüşmeleri:</b><br>"
                "Ev, iş yeri veya kurum telefonları (0212, 0312 vb.) ile yapılan görüşmelerdir.<br>"
                "Şüphelinin irtibatlı olduğu sabit mekanları tespit etmek için kritiktir."
            ),
            "Uluslararası": (
                "<b>✈️ Yurt Dışı (Roaming) Trafiği:</b><br>"
                "Türkiye dışındaki numaralarla yapılan veya yurt dışındayken gerçekleşen iletişimleri gösterir.<br>"
                "Sınır aşan suçlar veya seyahat hareketliliği analizi için kullanılır."
            ),
            "İnternet": (
                "<b>🌐 İnternet (GPRS/WAP) Kayıtları:</b><br>"
                "Telefonun internete bağlandığı anları (Whatsapp, Sosyal Medya vb. arka plan trafiği) gösterir.<br>"
                "• <b>Önemli:</b> İnternet kayıtları çok sık tutulduğu için, şüphelinin <b>fiziki konumunu ve güzergahını</b> en hassas şekilde bu sekmeden takip edebilirsiniz."
            ),
            "STH": (
                "<b>📡 STH (Sabit Telefon Hizmetleri):</b><br>"
                "Alternatif operatörler veya VoIP hizmetleri üzerinden geçen çağrıları içerir."
            )
        }

        self.tab_icons = {
            "GSM": "📞 ", "SMS": "💬 ", "Sabit": "☎️ ",
            "Uluslararası": "✈️ ", "İnternet": "🌐 ", "STH": "📡 ",
            "Olay Merkezli": "🎯 "
        }

        # Olay Merkezli Analiz sekmesi (özet + detay + harita)
        self.tab_descriptions["Olay Merkezli"] = (
            "<b>🎯 Olay Merkezli Analiz:</b><br>"
            "Seçtiğiniz zaman aralığında tek bir olaya odaklanır.<br>"
            "• Özet: yoğunluk / süre / baz dağılımı / koordinat kapsaması<br>"
            "• Detay: satır satır kayıt, koordinat yoksa manuel işaretleme<br>"
            "• Harita: koordinatı olan bazları işaretler, güzergâhı çizer"
        )

        for name, cols in self.tab_definitions.items():
            c_mode = 'embedded' if name == "GSM" else 'none'
            desc = self.tab_descriptions.get(name, "Bu sekme HTS kayıtlarını içerir.")
            wid = GenericDatabaseTable(cols, chart_mode=c_mode, info_text=desc)
            self.tab_widgets[name] = wid

            display_name = f"{self.tab_icons.get(name, '')}{name}"
            self.tabs.addTab(wid, display_name)

        # 🎯 Olay Merkezli Analiz Paneli (GenericDatabaseTable değil, özel widget)
        self.event_panel = EventCenteredAnalysisPanel(self)
        self.tab_widgets["Olay Merkezli"] = self.event_panel
        # Genel Bakış'ın hemen sağına koy
        self.tabs.insertTab(1, self.event_panel, f"{self.tab_icons.get('Olay Merkezli', '')}Olay Merkezli")

        if "GSM" in self.tab_widgets:
            self.tab_widgets["GSM"].table.doubleClicked.connect(self.open_reciprocal_detail)

        right_layout.addWidget(self.tabs)

        self.splitter.addWidget(right_container)
        self.splitter.setSizes([350, 950])

        layout.addWidget(self.splitter)

        self.tabs.currentChanged.connect(self.on_tab_changed)
        self.loaded_tabs = set()
        self.loader = LoadingOverlay(self)
        self.open_detail_windows = []

    def open_abone_detail_popup(self, index):
        """Abone tablosunda çift tıklanan satırı detaylı popup pencerede gösterir."""
        try:
            proxy = self.abone_table.proxy_model
            source_index = proxy.mapToSource(index)
            row = source_index.row()

            raw_data = self.abone_table.source_model._data[row]
            headers = self.abone_cols

            label_map = {
                "NUMARA": "GSM Numarası",
                "DURUM": "Abonelik Durumu",
                "AD": "Ad",
                "SOYAD": "Soyad",
                "ADRES": "Adres",
                "DOGUM_TARIHI": "Doğum Tarihi",
                "DOGUM_YERI": "Doğum Yeri",
                "ILCE": "İlçe",
                "IL": "İl",
                "TC_KIMLIK_NO": "TC Kimlik No",
                "ANNE_ADI": "Anne Adı",
                "BABA_ADI": "Baba Adı",
                "ABONE_SORGU_ARALIGI": "Sorgu Aralığı",
                "ABONE_BASLANGIC": "Abonelik Başlangıç",
                "ABONE_BITIS": "Abonelik Bitiş",
                "OPERATOR": "Operatör"
            }

            display_data = []
            for i, val in enumerate(raw_data):
                if i < len(headers):
                    key_raw = headers[i]
                    key_display = label_map.get(key_raw, key_raw)
                    val_display = str(val) if val is not None else ""
                    display_data.append((key_display, val_display))

            if display_data:
                dlg = FileDetailPopup(self, display_data)
                dlg.setWindowTitle("Hat Sahibi Detayları")
                dlg.exec()

        except Exception as e:
            print(f"Abone detay hatası: {e}")

    def open_report_center(self):
        """Proje bazlı Rapor Merkezi açar (numara seçimi zorunlu değil)."""
        if not self.current_project_id:
            ModernDialog.show_warning(self, "Proje Yok", "Lütfen önce bir proje seçiniz.")
            return

        try:
            # ✅ 1) Zaten açıksa ikinci kez açma
            existing = getattr(self, "rapor_merkezi", None)
            try:
                if existing is not None and existing.isVisible():
                    existing.raise_()
                    existing.activateWindow()
                    return
            except Exception:
                pass

            # ✅ 2) Parent verme (minimize olmasın)
            dlg = ReportCenterDialog(None, self.current_project_id)
            dlg.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)

            self.rapor_merkezi = dlg

            def _clear_report_center_refs(*_):
                # ✅ Dialog kapandıktan sonra child widget'larda kalan referanslar
                # "wrapped C/C++ object ... has been deleted" hatasına neden oluyordu.
                try:
                    if getattr(self, "rapor_merkezi", None) is dlg:
                        setattr(self, "rapor_merkezi", None)
                except Exception:
                    setattr(self, "rapor_merkezi", None)
                try:
                    for w in self.findChildren(QWidget):
                        if getattr(w, "rapor_merkezi", None) is dlg:
                            setattr(w, "rapor_merkezi", None)
                except Exception:
                    pass

            try:
                dlg.finished.connect(_clear_report_center_refs)
            except Exception:
                pass
            try:
                dlg.destroyed.connect(_clear_report_center_refs)
            except Exception:
                pass
            # ✅ 3) Açık olan tüm tablo widget'larına rapor merkezi referansını dağıt
            try:
                for w in self.findChildren(QWidget):
                    if hasattr(w, "add_selection_to_report"):
                        setattr(w, "rapor_merkezi", dlg)
            except Exception:
                pass

            self.open_window_safe(dlg)

        except Exception as e:
            ModernDialog.show_error(self, "Rapor Hatası", str(e))

    def open_stalking_analysis(self):
        """Taciz ve Israrlı Takip analiz penceresini açar."""
        if not self.current_project_id or not self.current_gsm_number:
            ModernDialog.show_warning(self, "Veri Yok", "Lütfen önce bir numara seçin.")
            return

        try:
            s_dt = self.dt_start.dateTime()
            e_dt = self.dt_end.dateTime()

            dlg = StalkingAnalysisDialog(
                self,
                self.current_project_id,
                self.current_gsm_number,
                s_dt,
                e_dt
            )
            self.open_window_safe(dlg)

        except Exception as e:
            ModernDialog.show_error(self, "Hata", f"Analiz başlatılamadı: {e}")

    def open_profile_map_dialog(self, home_txt, work_txt):
        if not self.current_project_id or not self.current_gsm_number:
            ModernDialog.show_warning(self, "Veri Yok", "Lütfen önce bir proje ve numara seçiniz.")
            return

        try:
            dlg = ProfileMapDialog(
                self,
                self.current_project_id,
                self.current_gsm_number,
                home_txt,
                work_txt
            )
            self.open_window_safe(dlg)

        except Exception as e:
            ModernDialog.show_error(self, "Harita Hatası", f"Harita oluşturulurken hata: {e}")

    def open_speed_anomaly(self):
        """Hız ihlal pencresini açar."""
        if not self.current_project_id or not self.current_gsm_number:
            ModernDialog.show_warning(self, "Veri Yok", "Lütfen önce bir numara seçin.")
            return

        try:
            dlg = SpeedAnomalyDialog(self, self.current_project_id, self.current_gsm_number)
            self.open_window_safe(dlg)
        except Exception as e:
            ModernDialog.show_error(self, "Hata", f"Pencere açılamadı: {e}")

    def open_reciprocal_detail(self, index):
        """GSM tablosunda bir satıra çift tıklandığında O TARİHTEKİ karşılıklı analizi açar."""

        try:
            target_table = self.tab_widgets.get("GSM")
            if not target_table: return

            src_idx = target_table.proxy_model.mapToSource(index)
            row_data = target_table.source_model._data[src_idx.row()]

            owner_gsm = str(row_data[1])
            other_gsm = str(row_data[3])
            click_date = str(row_data[4])
            other_name = str(row_data[6]) if len(row_data) > 6 else ""
            main_imei = str(row_data[8]) if len(row_data) > 8 else ""
            main_baz = str(row_data[9]) if len(row_data) > 9 else ""

            def clean_field(text):
                text = str(text).strip()
                if not text or text in ["Kayıt Bulunamadı", "---", "nan", "None", "0"]: return None
                return text

            clean_other_gsm = clean_field(other_gsm)
            clean_imei = clean_field(main_imei)
            clean_baz = clean_field(main_baz)

            if not clean_other_gsm or len(clean_other_gsm) < 3: return

            has_coordinate = False
            if clean_baz:
                coords = re.findall(r"(\d{2}\.\d{4,})", clean_baz)
                if len(coords) >= 2: has_coordinate = True

            if not clean_imei and not has_coordinate:
                msg = (f"Bu kayıt için detay penceresi açılamıyor. "
                       f"Açılış için Ana Hatta ait BAZ VEYA Koordinatlı Baz verisinin olması gerekir.")
                if clean_baz and not has_coordinate:
                    msg = ("Baz istasyonu adı mevcut ancak coğrafi koordinat bilgisi bulunamadı. "
                           "Harita analizi yapılamayacağından detay penceresi açılamamıştır.")
                ModernDialog.show_warning(self, "Detay Verisi Eksik", msg)
                return

            try:
                with DB() as conn:
                    cur = conn.cursor()

                    def get_last_10(n):
                        s = str(n).strip();
                        d = re.sub(r'\D', '', s)
                        return d[-10:] if len(d) >= 10 else d

                    my_short = get_last_10(owner_gsm)
                    other_short = get_last_10(other_gsm)

                    date_filter = click_date
                    if len(click_date) > 16: date_filter = click_date[:16] + '%'
                    else: date_filter = click_date + '%'

                    sql_count = """
                        SELECT COUNT(*)
                        FROM hts_gsm 
                        WHERE ProjeID = ?
                          AND substr(replace(replace(replace(NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ?
                          AND substr(replace(replace(replace(DIGER_NUMARA, ' ', ''), '-', ''), '+', ''), -10, 10) = ?
                          AND TARIH LIKE ?
                    """

                    params = (self.current_project_id, other_short, my_short, date_filter)
                    count = cur.execute(sql_count, params).fetchone()[0]

                    if count == 0:
                         ModernDialog.show_warning(self, "Veri Yok",
                            f"'{other_gsm}' ve '{owner_gsm}' arasında bu saatte kayıtlı trafik bulunamadığından detay penceresi açılamadı."
                                                   )
                         return

            except Exception as e:
                print(f"Ön Kontrol Hatası: {e}")
                ModernDialog.show_error(self, "Hata", "Detay ön kontrolü sırasında veritabanı hatası oluştu.")
                return

            dlg = ReciprocalDetailDialog(
                self,
                self.current_project_id,
                owner_gsm,
                other_gsm,
                other_name,
                click_date,
                main_baz
            )
            self.open_window_safe(dlg)

        except Exception as e:
            print(f"Detay Açma Hatası: {e}")
            ModernDialog.show_error(self, "Hata", f"Detay açılamadı: {e}")

    def open_common_subscriber_detail(self, index):
        """Ortak abone tablosuna tıklanınca detayları açar."""
        sender_view = self.sender()

        try:
            model = sender_view.model()
            row = index.row()

            val_name = model.index(row, 0).data()
            val_nums = model.index(row, 2).data()

            if not val_nums: return

            title_prefix = "TC: " if sender_view == self.common_tc_table.table else ""
            full_name = f"{title_prefix}{val_name}"

            dlg = MultiNumberDetailDialog(
                self,
                self.current_project_id,
                full_name,
                val_nums
            )
            self.open_window_safe(dlg)

        except Exception as e:
            print(f"Ortak Detay Hatası: {e}")

    def add_web_view_evidence(self, pixmap, title="Harita Analizi"):
        """Harita görüntüsünü boyutlandırıp DOSYAYA kaydeder."""
        if pixmap.isNull():
            ModernDialog.show_warning(self, "Hata", "Görüntü oluşturulamadı.")
            return

        gsm_info = self.current_gsm_number if self.current_gsm_number else "Genel"
        final_title = f"{gsm_info} - {title}" if gsm_info not in title else title

        base_dir = os.path.dirname(os.path.abspath(__file__))
        evidence_dir = os.path.join(base_dir, "evidence_images")
        if not os.path.exists(evidence_dir):
            os.makedirs(evidence_dir)

        if pixmap.width() > 2200:
            pixmap = pixmap.scaledToWidth(2200, Qt.TransformationMode.SmoothTransformation)
        if "Coğrafi Konum Analizi" in (title or ""):
            pixmap = self._trim_pixmap_vertical(pixmap, tol=18, pad=4, empty_ratio=0.990, inset=2)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        prefix = "geo_map" if "Coğrafi Konum Analizi" in (title or "") else "event_map"
        filename = f"{prefix}_{timestamp}.png"
        file_path = os.path.join(evidence_dir, filename)

        success = pixmap.save(file_path, "PNG", 90)

        if success:
            self.add_evidence_to_report(final_title, file_path, "IMAGE")
        else:
            ModernDialog.show_error(self, "Hata", "Resim kaydedilemedi.")

    def _trim_pixmap_vertical(self, pixmap, sample_xy=(6, 6), tol=18, pad=6, empty_ratio=0.985, inset=2):
        """
        Üst-alt boşluğu kırpar.
        empty_ratio: satırın bu oranı BG ise satır "boş" kabul edilir.
        inset: kenar çerçevesini hesaba katmamak için x taramasını içeri alır.
        """
        if pixmap.isNull():
            return pixmap

        img = pixmap.toImage().convertToFormat(QImage.Format.Format_RGBA8888)
        w, h = img.width(), img.height()
        if w <= 0 or h <= 0:
            return pixmap

        sx = min(max(sample_xy[0], 0), w - 1)
        sy = min(max(sample_xy[1], 0), h - 1)
        bg = QColor(img.pixel(sx, sy))

        x0 = min(max(inset, 0), w - 1)
        x1 = max(x0 + 1, w - inset)

        step = max(1, (x1 - x0) // 260)

        def row_is_empty(y):
            total = 0
            diff = 0
            for x in range(x0, x1, step):
                total += 1
                c = QColor(img.pixel(x, y))
                if (
                    abs(c.red()   - bg.red())   > tol or
                    abs(c.green() - bg.green()) > tol or
                    abs(c.blue()  - bg.blue())  > tol
                ):
                    diff += 1

            # satırın büyük kısmı BG ise boş say
            if total == 0:
                return True
            return (diff / total) <= (1.0 - empty_ratio)

        # üstten boş satırları kes
        top = 0
        while top < h and row_is_empty(top):
            top += 1

        # alttan boş satırları kes
        bottom = h - 1
        while bottom > top and row_is_empty(bottom):
            bottom -= 1

        # pay
        top = max(0, top - pad)
        bottom = min(h - 1, bottom + pad)

        new_h = bottom - top + 1
        if new_h <= 0:
            return pixmap

        return pixmap.copy(0, top, w, new_h)

    def capture_chart_screenshot(self, widget, title="Grafik Analizi"):
        """Grafiği boyutlandırıp DOSYAYA kaydeder."""

        gsm_info = self.current_gsm_number if self.current_gsm_number else "Genel"
        final_title = f"{gsm_info} - {title}" if gsm_info not in title else title

        pixmap = widget.grab()

        if pixmap.width() > 4000:
            pixmap = pixmap.scaledToWidth(
                4000,
                Qt.TransformationMode.SmoothTransformation
            )

        # ✅ SADECE diagramlar için üst-alt boşluk kırp
        pixmap = self._trim_pixmap_vertical(pixmap, tol=18, pad=6, empty_ratio=0.985, inset=2)

        base_dir = os.path.dirname(os.path.abspath(__file__))
        evidence_dir = os.path.join(base_dir, "evidence_images")
        if not os.path.exists(evidence_dir):
            os.makedirs(evidence_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"chart_{timestamp}.png"
        file_path = os.path.join(evidence_dir, filename)

        success = pixmap.save(file_path, "PNG", 90)

        if success:
            self.add_evidence_to_report(final_title, file_path, "IMAGE")
        else:
            ModernDialog.show_error(self, "Hata", "Grafik kaydedilemedi.")

    def add_evidence_to_report(self, title, content, type_):
        """Delili veritabanına kaydeder (width hatası giderildi)."""
        if not getattr(self, "current_project_id", None):
            return

        try:
            with DB() as conn:
                c = conn.cursor()

                last_order = c.execute(
                    "SELECT MAX(Sira) FROM rapor_taslagi WHERE ProjeID=?",
                    (self.current_project_id,)
                ).fetchone()[0]
                new_order = (last_order or 0) + 1

                gsm_val = self.current_gsm_number or ""

                html_val = content if type_ in ("TABLE", "HTML") else None
                img_val  = content if type_ == "IMAGE" else None

                # ✅ Varsayılan genişlik (akıllı default)
                src_probe = str(content or "")
                title_probe = (title or "")

                # default: %100, harita/diagram: %80
                genislik = 100
                if type_ == "IMAGE":
                    t = title_probe.lower()
                    p = src_probe.lower()  # IMAGE için content=file_path
                    if ("analiz diyagram" in t) or ("coğrafi konum analizi" in t) or ("cografi konum analizi" in t):
                        genislik = 80
                    elif ("chart_" in p) or ("event_map_" in p) or ("geo_map_" in p):
                        genislik = 100
                row_data = [
                    self.current_project_id, gsm_val, title, content, type_,
                    datetime.now().strftime("%d.%m.%Y %H:%M"), new_order,
                    genislik,
                    0, "center", "", html_val, img_val
                ]

                cols = [r[1] for r in c.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]

                keys = ["ProjeID", "GSMNo", "Baslik", "Icerik", "Tur", "Tarih", "Sira"]
                vals = row_data[:7]

                if "GenislikYuzde" in cols: keys.append("GenislikYuzde"); vals.append(row_data[7])
                if "YukseklikMm"   in cols: keys.append("YukseklikMm");   vals.append(row_data[8])
                if "Hizalama"      in cols: keys.append("Hizalama");      vals.append(row_data[9])
                if "Aciklama"      in cols: keys.append("Aciklama");      vals.append(row_data[10])
                if "HtmlIcerik"    in cols: keys.append("HtmlIcerik");    vals.append(row_data[11])
                if "ImagePath"     in cols: keys.append("ImagePath");     vals.append(row_data[12])

                q_marks = ",".join(["?"] * len(keys))
                col_names = ",".join(keys)

                c.execute(f"INSERT INTO rapor_taslagi ({col_names}) VALUES ({q_marks})", vals)
                conn.commit()

            ModernDialog.show_success(self, "Eklendi", "Delil rapora eklendi.")

        except Exception as e:
            ModernDialog.show_error(self, "Rapor Hatası", str(e))

    def capture_table_selection(self, table_widget):
        if not table_widget:
            return

        try:
            selected = table_widget.selectedIndexes()
            if not selected:
                ModernDialog.show_warning(self, "Seçim Yok", "Lütfen tablodan rapora aktarılacak hücreleri seçin.")
                return

            include_headers = True
            include_set = set()
            exclude_set = set()

            rows = sorted(set(i.row() for i in selected))
            cols = sorted(set(i.column() for i in selected))

            # görünür kolonlar
            visible_cols = [c for c in cols if not table_widget.isColumnHidden(c)]
            if not visible_cols:
                ModernDialog.show_warning(self, "Kolon Yok", "Seçimde görünür kolon bulunamadı.")
                return

            headers = []
            for c in visible_cols:
                h = table_widget.horizontalHeaderItem(c)
                headers.append(h.text() if h else f"Kolon {c+1}")

            kept_cols = []
            kept_headers = []
            for c, h in zip(visible_cols, headers):
                hn = _norm_header(h)
                if include_set:
                    if hn in include_set:
                        kept_cols.append(c); kept_headers.append(h)
                else:
                    if exclude_set and hn in exclude_set:
                        continue
                    kept_cols.append(c); kept_headers.append(h)

            if not kept_cols:
                ModernDialog.show_warning(self, "Filtre Sonucu Boş", "Kolon filtreleri nedeniyle eklenecek kolon kalmadı.")
                return

            # HTML tablo üret
            html_table = "<table class='meta-table'>"
            if include_headers:
                html_table += "<tr>" + "".join(f"<th>{html.escape(h)}</th>" for h in kept_headers) + "</tr>"

            for r in rows:
                html_table += "<tr>"
                for c in kept_cols:
                    item = table_widget.item(r, c)
                    val = item.text() if item else ""
                    html_table += f"<td>{html.escape(val)}</td>"
                html_table += "</tr>"
            html_table += "</table>"

            # Plain text
            plain = []
            for r in rows:
                row_vals = []
                for c in kept_cols:
                    item = table_widget.item(r, c)
                    row_vals.append(item.text() if item else "")
                plain.append("\t".join(row_vals))
            selected_text = "\n".join(plain)

            # Rapor Merkezine ekle
            self.report_center.add_block(
                title="Tablo Seçimi",
                content=selected_text,
                block_type="TABLE",
                html_content=html_table
            )
            ModernDialog.show_success(self, "Eklendi", "Seçili tablo rapora eklendi.")
            self.report_center.load_blocks_into_table()
            self.report_center.refresh_preview()

        except Exception as e:
            ModernDialog.show_error(self, "Hata", f"Tablo seçim aktarımı hatası: {e}")

    def go_back_to_projects(self):
        """Projeler ekranına döner ve LİSTELERİ GÜNCELLER."""
        self.close_all_details()

        if hasattr(self.main, 'page_projects'):
            self.main.page_projects.load_projects()
            self.main.page_projects.load_project_gsms()

        self.main.stack.setCurrentIndex(0)

    def open_window_safe(self, window_instance):
        """
        Pencereleri ANA EKRANI KİLİTLEMEDEN (Non-Modal) ve KÜÇÜLTÜLEBİLİR açar.
        """
        try:
            if not hasattr(self, "open_detail_windows"):
                self.open_detail_windows = []

            self.open_detail_windows.append(window_instance)

            if hasattr(window_instance, "finished"):
                window_instance.finished.connect(lambda *_: self.cleanup_closed_window(window_instance))
            else:
                window_instance.destroyed.connect(lambda *_: self.cleanup_closed_window(window_instance))

            window_instance.setWindowFlags(
                Qt.WindowType.Window |
                Qt.WindowType.WindowMinimizeButtonHint |
                Qt.WindowType.WindowMaximizeButtonHint |
                Qt.WindowType.WindowCloseButtonHint
            )

            window_instance.setWindowModality(Qt.WindowModality.NonModal)
            window_instance.show()

            window_instance.raise_()
            window_instance.activateWindow()

        except Exception as e:
            print(f"Pencere açma hatası: {e}")
            ModernDialog.show_error(self, "Pencere Hatası", f"Pencere açılamadı: {e}")

    def cleanup_closed_window(self, dialog):
        # 1) listeden çıkar
        try:
            if hasattr(self, "open_detail_windows") and dialog in self.open_detail_windows:
                self.open_detail_windows.remove(dialog)
        except Exception:
            pass

        # 2) ana pencereyi tekrar aktif hale getir (fokus/klik problemini düzeltir)
        try:
            main_win = getattr(self, "main", None)
            if main_win is not None:
                # bazen dialog kapanınca ana pencere disabled kalabiliyor
                if not main_win.isEnabled():
                    main_win.setEnabled(True)

                # minimize olduysa geri getir
                try:
                    st = main_win.windowState()
                    if st & Qt.WindowState.WindowMinimized:
                        main_win.setWindowState(st & ~Qt.WindowState.WindowMinimized)
                except Exception:
                    pass

                # en üste al + fokus
                main_win.raise_()
                main_win.activateWindow()
                try:
                    QApplication.setActiveWindow(main_win)
                except Exception:
                    pass
        except Exception:
            pass


    def close_all_details(self):
        for w in list(self.open_detail_windows):
            w.close()
        self.open_detail_windows.clear()

    def toggle_left_panel(self):
        """Sol paneli gizler ve üstteki butonları gösterir (Responsive)."""
        is_visible = self.left_widget.isVisible()
        self.left_widget.setVisible(not is_visible)
        self.top_actions_widget.setVisible(is_visible)

        if not is_visible:
            self.btn_toggle_menu.setText("☰ Menü (Kapat)")
        else:
            self.btn_toggle_menu.setText("☰")

    def open_daily_route(self):
        """Günlük güzergah penceresini açar."""
        if not self.current_project_id or not self.current_gsm_number:
            ModernDialog.show_warning(self, "Veri Yok", "Lütfen önce bir numara seçin.")
            return
        default_date = self.dt_start.dateTime()
        dlg = DailyRouteDialog(self, self.current_project_id, self.current_gsm_number, default_date)
        self.open_window_safe(dlg)

    def open_map_view(self):
        """Harita penceresini açar."""
        if not self.current_project_id or not self.current_gsm_number:
            ModernDialog.show_warning(self, "Veri Yok", "Lütfen önce bir numara seçin.")
            return

        try:
            start_dt = self.dt_start.dateTime(); end_dt = self.dt_end.dateTime()
            dlg = MapDialog(self, self.current_project_id, self.current_gsm_number, start_dt, end_dt)
            self.open_window_safe(dlg)
        except Exception as e: ModernDialog.show_error(self, "Hata", f"Harita hatası: {e}")

    def open_cross_match(self):
        numbers = []
        try:
            with DB() as conn:
                rows = conn.execute("SELECT DISTINCT GSMNo FROM hts_ozet WHERE ProjeID=?", (self.current_project_id,)).fetchall()
                numbers = [r[0] for r in rows]
        except: pass

        if not numbers:
            ModernDialog.show_warning(self, "Veri Yok", "Bu projede analiz edilecek hiç numara bulunamadı.")
            return

        dlg = CrossMatchDialog(self, self.current_project_id, numbers)
        self.open_window_safe(dlg)

    def open_heatmap_popup(self):
        """Isı haritasını ANLIK HESAPLAR ve açar."""
        if not self.current_project_id or not self.current_gsm_number:
            ModernDialog.show_warning(self, "Veri Yok", "Lütfen bir numara seçin.")
            return
        data_matrix = [[0 for _ in range(24)] for _ in range(7)]
        try:
            with DB() as conn:
                cur = conn.cursor()
                rows = cur.execute("""
                    SELECT TARIH FROM hts_gsm WHERE ProjeID=? AND GSMNo=?
                    UNION ALL
                    SELECT TARIH FROM hts_sms WHERE ProjeID=? AND GSMNo=?
                """, (self.current_project_id, self.current_gsm_number)*2).fetchall()

                for r in rows:
                    t_str = r[0]
                    if not t_str: continue
                    try:
                        fmt = "dd.MM.yyyy HH:mm:ss" if "." in t_str else "dd/MM/yyyy HH:mm:ss"
                        if " " not in t_str: fmt = fmt.split(" ")[0]
                        py_fmt = fmt.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y").replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")
                        dt = datetime.strptime(t_str, py_fmt)
                        day_idx = dt.weekday()
                        hour_idx = dt.hour
                        data_matrix[day_idx][hour_idx] += 1
                    except: pass

            self.current_heatmap_data = data_matrix

            dlg = HeatmapDialog(self, self.current_heatmap_data)
            dlg.heatmap_widget.cell_clicked_signal.connect(self.open_heatmap_detail)
            self.open_window_safe(dlg)

        except Exception as e:
            ModernDialog.show_error(self, "Hata", f"Isı haritası oluşturulamadı: {e}")

    def open_heatmap_detail(self, day_idx, hour_idx):
        if not self.current_project_id or not self.current_gsm_number: return

        self.loader.start("Kayıtlar Taranıyor...")
        QApplication.processEvents()

        days_tr = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        day_name = days_tr[day_idx]
        hour_str = f"{hour_idx:02d}"

        try:
            with DB() as conn:
                cur = conn.cursor()

                sql_gsm = "SELECT TARIH, 'GSM', TIP, DIGER_NUMARA, SURE || ' sn', BAZ FROM hts_gsm WHERE ProjeID=? AND GSMNo=?"
                sql_sms = "SELECT TARIH, 'SMS', TIP, DIGER_NUMARA, '---', '---' FROM hts_sms WHERE ProjeID=? AND GSMNo=?"
                sql_net = "SELECT TARIH, 'DATA', 'Data', 'İnternet', '---', BAZ FROM hts_gprs WHERE ProjeID=? AND GSMNo=?"

                rows = []
                rows += cur.execute(sql_gsm, (self.current_project_id, self.current_gsm_number)).fetchall()
                rows += cur.execute(sql_sms, (self.current_project_id, self.current_gsm_number)).fetchall()
                rows += cur.execute(sql_net, (self.current_project_id, self.current_gsm_number)).fetchall()

                filtered_rows = []

                for r in rows:
                    t_str = r[0]
                    if not t_str: continue

                    try:
                        fmt = "dd.MM.yyyy HH:mm:ss" if "." in t_str else "dd/MM/yyyy HH:mm:ss"
                        if " " not in t_str: fmt = fmt.split(" ")[0]

                        py_fmt = fmt.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y").replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")

                        dt = datetime.strptime(t_str, py_fmt)

                        if dt.weekday() == day_idx and dt.hour == hour_idx:
                            filtered_rows.append(r)

                    except: pass

                if len(filtered_rows) > 10000:
                    ModernDialog.show_warning(self,
                        "Veri Gösterim Sınırı",
                        f"Bu saat diliminde {len(filtered_rows)} kayıt bulundu. Performans için ilk 10.000 kayıt gösteriliyor."
                                              )
                    filtered_rows = filtered_rows[:10000]

                filtered_rows.sort(key=lambda x: x[0])

            if filtered_rows:
                dialog = ActivityDetailDialog(self, filtered_rows, day_name, hour_str)
                self.loader.stop()
                dialog.exec()
            else:
                self.loader.stop()
                ModernDialog.show_info(self, "Bilgi", "Bu saat diliminde gösterilecek detaylı kayıt bulunamadı.")

        except Exception as e:
            self.loader.stop()
            print(f"Detay Hatası: {e}")
            ModernDialog.show_error(self, "Hata", str(e))

    def recalculate_common_analysis(self):
        if not self.current_project_id:
            return

        if not AnalysisUtils.project_has_any_gsm(self.current_project_id):
            try:
                if hasattr(self, "common_imei_table"): self.common_imei_table.set_data([])
                if hasattr(self, "common_name_table"): self.common_name_table.set_data([])
                if hasattr(self, "common_tc_table"): self.common_tc_table.set_data([])
            except Exception as e:
                print(f"Ortak tablo temizleme hatası: {e}")
            return

        AnalysisUtils.recalculate_common_analysis_core(self.current_project_id)

        self.load_common_analysis_tables()

    def external_upload_started(self, pid):
        """Proje ekranından yükleme başladığında Overlay'i açar."""
        if self.current_project_id != pid: self.set_project(pid)
        self.btn_load.setEnabled(False)

        if hasattr(self.main, 'loader'):
            self.main.loader.start("Yükleme Başlatılıyor...")

    def external_progress_updated(self, value):
        """Proje ekranındaki yüzdeyi Overlay'e yansıtır."""
        if hasattr(self.main, 'loader'):
            self.main.loader.set_progress(value)

    def external_gsm_detected(self, gsm):
        """Proje ekranındaki GSM bilgisini Overlay metnine yansıtır."""
        if hasattr(self.main, 'loader'):
            self.main.loader.text = f"İşleniyor: {gsm}"
            self.main.loader.update()

    def external_queue_finished(self):
        """Yükleme bitince Overlay'i kapatır."""
        self.btn_load.setEnabled(True)
        if hasattr(self.main, 'loader'):
            self.main.loader.stop()
        self.load_numbers()

    def external_file_finished(self):
        """Dosya bitince listeyi yeniler."""
        self.load_numbers()

    def update_tab_visibility(self):
        """Genel Bakış'a dokunmadan diğer sekmeleri günceller (DÜZELTİLDİ)."""
        if not self.current_project_id or not self.current_gsm_number:
            return

        required_tabs = ["Genel Bakış", f"{self.tab_icons.get('Olay Merkezli', '')}Olay Merkezli"]

        table_map = {
            "GSM": "hts_gsm", "SMS": "hts_sms", "Sabit": "hts_sabit",
            "STH": "hts_sth", "Uluslararası": "hts_uluslararasi"
        }

        try:
            with DB() as conn:
                cur = conn.cursor()
                for tab_name in self.tab_definitions.keys():
                    has_data = False

                    if tab_name == "İnternet":
                        # İnternet kayıtları 2 tabloda tutuluyor olabilir
                        c1 = cur.execute(
                            "SELECT 1 FROM hts_gprs WHERE ProjeID=? AND GSMNo=? LIMIT 1",
                            (self.current_project_id, self.current_gsm_number)
                        ).fetchone()
                        c2 = cur.execute(
                            "SELECT 1 FROM hts_wap WHERE ProjeID=? AND GSMNo=? LIMIT 1",
                            (self.current_project_id, self.current_gsm_number)
                        ).fetchone()
                        if c1 or c2:
                            has_data = True

                    elif tab_name in table_map:
                        tbl = table_map[tab_name]
                        if cur.execute(
                            f"SELECT 1 FROM {tbl} WHERE ProjeID=? AND GSMNo=? LIMIT 1",
                            (self.current_project_id, self.current_gsm_number)
                        ).fetchone():
                            has_data = True

                    if has_data:
                        display_name = f"{self.tab_icons.get(tab_name, '')}{tab_name}"
                        required_tabs.append(display_name)

        except Exception as e:
            print(f"Tab Check Error: {e}")

        current_tabs = [self.tabs.tabText(i) for i in range(self.tabs.count())]
        if current_tabs == required_tabs:
            return

        self.tabs.setUpdatesEnabled(False)
        self.tabs.blockSignals(True)

        try:
            for disp_name in required_tabs:
                if disp_name not in current_tabs:
                    raw_name = disp_name
                    for k, v in self.tab_icons.items():
                        if v in disp_name:
                            raw_name = k
                            break

                    if disp_name == "Genel Bakış":
                        self.tabs.insertTab(0, self.overview_widget, disp_name)
                    elif raw_name in self.tab_widgets:
                        self.tabs.addTab(self.tab_widgets[raw_name], disp_name)

            i = 0
            while i < self.tabs.count():
                if self.tabs.tabText(i) not in required_tabs:
                    self.tabs.removeTab(i)
                else:
                    i += 1

        except Exception as e:
            print(f"Sekme Güncelleme Hatası: {e}")

        finally:
            self.tabs.blockSignals(False)
            self.tabs.setUpdatesEnabled(True)

    def set_project(self, pid):
        self.current_project_id = pid
        self.current_gsm_number = None

        self.load_numbers()

        target_table = getattr(self, "num_table", None)
        if target_table is None:
            return

        def _restore_or_select_first():
            try:
                if target_table.source_model.rowCount(QModelIndex()) == 0:
                    if hasattr(self, "loaded_tabs"):
                        self.loaded_tabs.clear()
                    self.clear_all_widgets()
                    return

                target_row = 0
                last_gsm = None

                if hasattr(self, "main") and self.main is not None:
                    last_gsm = getattr(self.main, "last_gsm_number", None)

                if last_gsm:
                    source_data = target_table.source_model._data
                    for r in range(len(source_data)):
                        if source_data[r][0] == last_gsm:
                            target_row = r
                            break

                source_idx = target_table.source_model.index(target_row, 0)
                proxy_idx = target_table.proxy_model.mapFromSource(source_idx)

                if not proxy_idx.isValid():
                    proxy_idx = target_table.proxy_model.index(0, 0)

                target_table.table.blockSignals(True)
                target_table.table.selectRow(proxy_idx.row())
                target_table.table.setCurrentIndex(proxy_idx)
                target_table.table.scrollTo(proxy_idx)
                target_table.table.blockSignals(False)

                self.on_num_select(proxy_idx)

            except Exception as e:
                print(f"Analiz ilk/son GSM otomatik seçim hatası: {e}")

        QTimer.singleShot(10, _restore_or_select_first)

    def load_numbers(self):
        """Sol menüye Numaraları ve Abone İsimlerini (Tekrarsız) yükler."""
        try:
            with DB() as conn:
                rows = conn.execute("""
                    SELECT DISTINCT GSMNo FROM hts_gsm WHERE ProjeID=? AND GSMNo IS NOT NULL AND GSMNo != '' 
                    UNION 
                    SELECT DISTINCT GSMNo FROM hts_abone WHERE ProjeID=? AND GSMNo IS NOT NULL AND GSMNo != ''
                """, (self.current_project_id, self.current_project_id)).fetchall()

                unique_numbers = [r[0] for r in rows]
                display_data = []

                for gsm in unique_numbers:
                    name_rows = conn.execute("""
                        SELECT AD, SOYAD FROM hts_abone 
                        WHERE ProjeID=? AND GSMNo=?
                    """, (self.current_project_id, gsm)).fetchall()

                    unique_names = set()
                    for r in name_rows:
                        ad = str(r[0]).strip() if r[0] else ""
                        soyad = str(r[1]).strip() if r[1] else ""
                        full_name = f"{ad} {soyad}".strip()

                        if full_name:
                            unique_names.add(full_name)

                    abone_str = " / ".join(sorted(list(unique_names)))

                    display_data.append([gsm, abone_str])

            self.num_table.set_data(display_data)

            h = self.num_table.table.horizontalHeader()

            h.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
            self.num_table.table.setColumnWidth(0, 125)

            h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        except Exception as e:
            print(f"Numara Listesi Hatası: {e}")

    def upload_excel(self):
        """Yeni HTS Kaydı Yükle (Sıralı Kuyruk ve Detaylı Bilgi Sistemi ile)"""
        if not self.current_project_id:
            ModernDialog.show_warning(self, "Uyarı", "Lütfen önce bir proje seçiniz.")
            return

        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "HTS Kayıtlarını Seç (Excel)", "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_paths: return

        self.upload_queue = list(file_paths)
        self.total_upload_count = len(file_paths)
        self.success_upload_count = 0

        if hasattr(self.main, 'loader'):
            self.main.loader.start("Yükleme İşlemi Başlatılıyor...")

        self.process_next_upload()

    def process_next_upload(self):
        """Kuyruktaki sıradaki dosyayı işler (Proje Merkezi ile Senkronize)."""
        if not hasattr(self, 'upload_queue') or not self.upload_queue:
            if hasattr(self.main, 'loader'): self.main.loader.stop()

            if self.success_upload_count > 0:
                ModernDialog.show_success(self, "İşlem Tamamlandı", f"{self.success_upload_count} adet dosya başarıyla analiz edildi.")

                if hasattr(self, 'load_project_gsms'):
                    self.load_project_gsms()

                if hasattr(self.main, 'project_manager') and hasattr(self.main.project_manager, 'load_project_gsms'):
                    self.main.project_manager.load_project_gsms()
                # -----------------------------------
            return

        current_file = self.upload_queue[0]
        file_name = os.path.basename(current_file)

        if hasattr(self.main, 'loader'):
            self.main.loader.text = f"Hazırlanıyor:\n{file_name}"
            self.main.loader.update()

        try:
            rol, target_gsm = detect_hts_role(current_file)
        except:
            rol, target_gsm = "HEDEF", _detect_target_gsm(current_file)

        is_exist = False
        with DB() as conn:
            check = conn.execute(
                "SELECT 1 FROM hts_dosyalari WHERE ProjeID=? AND GSMNo=? AND Rol=? LIMIT 1",
                (self.current_project_id, target_gsm, rol)
            ).fetchone()
            if check: is_exist = True

        if is_exist:
            if hasattr(self.main, 'loader'): self.main.loader.hide()

            msg = f"'{file_name}' dosyasındaki numara ({target_gsm}) zaten yüklü.\nÜzerine yazılsın mı?"
            dlg = ModernDialog(self, "Mükerrer Kayıt", msg, "QUESTION", "Üzerine Yaz", "Atla")
            result = dlg.exec()

            if hasattr(self.main, 'loader'): self.main.loader.show()

            if result != 1:
                self.upload_queue.pop(0)
                self.process_next_upload()
                return

            self.delete_existing_records(target_gsm, rol)

        self.worker = HtsWorker(current_file, self.current_project_id)

        self.worker.progress.connect(self.on_upload_progress)
        self.worker.log.connect(self.on_upload_log)
        self.worker.finished.connect(self.on_upload_finished)
        self.worker.error.connect(self.on_upload_error)

        self.worker.start()

    def load_project_gsms(self):
        if not self.current_project_id:
            self.num_table.set_data([])
            self.current_gsm_number = None
            return

        try:
            display_data = []

            with DB() as conn:
                rows = conn.execute("""
                    SELECT GSMNo, MAX(YuklenmeTarihi) AS last_upload
                    FROM hts_dosyalari
                    WHERE ProjeID=?
                    GROUP BY GSMNo
                    ORDER BY datetime(last_upload) DESC, GSMNo DESC
                """, (self.current_project_id,)).fetchall()

                unique_numbers = [r[0] for r in rows if r and r[0]]

                for gsm in unique_numbers:
                    name_rows = conn.execute(
                        "SELECT AD, SOYAD FROM hts_abone WHERE ProjeID=? AND GSMNo=?",
                        (self.current_project_id, gsm)
                    ).fetchall()

                    unique_names = set()
                    for r in name_rows:
                        ad = str(r[0]).strip() if r and r[0] else ""
                        soyad = str(r[1]).strip() if r and r[1] else ""
                        full_name = f"{ad} {soyad}".strip()
                        if full_name:
                            unique_names.add(full_name)

                    abone_str = " / ".join(sorted(unique_names))
                    display_data.append([gsm, abone_str])

            self.num_table.set_data(display_data)

            h = self.num_table.table.horizontalHeader()
            h.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
            self.num_table.table.setColumnWidth(0, 125)
            h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

            if display_data:
                def _select_first_and_load():
                    try:
                        self.num_table.table.selectRow(0)
                        first_idx = self.num_table.proxy_model.index(0, 0)
                        self.num_table.table.setCurrentIndex(first_idx)
                        self.num_table.table.scrollTo(first_idx)

                        if hasattr(self, "on_num_select"):
                            self.on_num_select(first_idx)

                    except Exception as e:
                        print(f"Analiz ilk GSM otomatik seçim hatası: {e}")

                QTimer.singleShot(0, _select_first_and_load)

            else:
                self.current_gsm_number = None
                if hasattr(self, "clear_all_tabs_safe"):
                    self.clear_all_tabs_safe()

        except Exception as e:
            print(f"Numara Listesi Hatası (Analiz): {e}")
            self.num_table.set_data([])
            self.current_gsm_number = None

    def delete_existing_records(self, gsm, rol):
        """Üzerine yazma durumunda eski kayıtları temizler."""
        with DB() as conn:
            conn.execute("DELETE FROM hts_dosyalari WHERE ProjeID=? AND GSMNo=? AND Rol=?", (self.current_project_id, gsm, rol))

    def on_upload_progress(self, val):
        if hasattr(self.main, 'loader'):
            self.main.loader.set_progress(val)

    def on_upload_log(self, text):
        """Worker'dan gelen '📂 Dosya Taranıyor...' gibi detaylı mesajları ekrana basar."""
        if hasattr(self.main, 'loader'):
            self.main.loader.text = text
            self.main.loader.update()

    def on_upload_finished(self, msg):
        """Bir dosya bittiğinde çalışır."""
        self.success_upload_count += 1

        if self.upload_queue:
            self.upload_queue.pop(0)

        self.process_next_upload()

    def on_upload_error(self, err_msg):
        """Hata durumunda çalışır."""
        if hasattr(self.main, 'loader'): self.main.loader.hide()

        ModernDialog.show_error(self, "Hata", f"Yükleme hatası:\n{err_msg}")

        if hasattr(self.main, 'loader'): self.main.loader.show()

        if self.upload_queue:
            self.upload_queue.pop(0)
        self.process_next_upload()

    def request_graceful_close(self):
        if self.is_uploading:
            rem = len(self.upload_queue)
            self.upload_queue.clear()
            self.shutdown_pending = True

            if hasattr(self.main, 'loader'):
                self.main.loader.text = f"KAPATILIYOR... (İptal Edilen: {rem})"
                self.main.loader.update()

            if hasattr(self, 'worker_btn'):
                self.worker_btn.setText("Kapatılıyor...")

    def check_gsm_exists(self, gsm):
        if not gsm or gsm == "BILINMIYOR": return False
        try:
            with DB() as conn:
                if conn.execute("SELECT 1 FROM hts_ozet WHERE ProjeID=? AND GSMNo=?", (self.current_project_id, gsm)).fetchone(): return True
                if conn.execute("SELECT 1 FROM hts_gsm WHERE ProjeID=? AND GSMNo=? LIMIT 1", (self.current_project_id, gsm)).fetchone(): return True
        except: return False
        return False

    def delete_gsm_database_records(self, gsm_no: str):
        """
        Analiz Merkezi: seçilen GSM'ye ait tüm verileri (projede) siler.
        Silme sonrası ortak tabloları aynı projedeki kalan verilerle yeniden üretir.
        """
        pid = getattr(self, "current_project_id", None)
        if not pid:
            ModernDialog.show_warning(self, "Proje Seçilmedi", "Önce bir proje seçmelisiniz.")
            return False

        try:
            ok = AnalysisUtils.delete_gsm_records_core(pid, gsm_no)
            if not ok:
                ModernDialog.show_error(self, "Silme Hatası", "Silme işlemi başarısız.")
                return False

            AnalysisUtils.recalculate_common_analysis_core(pid)

            try:
                if hasattr(self, "load_numbers"):
                    self.load_numbers()
            except Exception as e:
                print(f"⚠️ UI refresh hata: {e}")

            return True

        except Exception as e:
            print(f"❌ [delete_gsm_database_records] Hata: {e}")
            ModernDialog.show_error(self, "Silme Hatası", f"Silme işlemi başarısız:\n{e}")
            return False

    def clear_all_widgets(self):
        """Ekrandaki tüm tabloları ve bilgileri temizler."""
        self.current_gsm_number = None
        self.loaded_tabs.clear()

        self.abone_table.set_data([])
        self.top_table.set_data([])
        self.contact_table.set_data([])
        self.baz_table.set_data([])
        self.all_baz_table.set_data([])
        self.imei_table.set_data([])
        self.common_imei_table.set_data([])
        self.common_name_table.set_data([])
        self.common_tc_table.set_data([])

        for w in self.tab_widgets.values():
            if hasattr(w, "set_data"):
                w.set_data([])
            elif hasattr(w, "reset"):
                w.reset()
            else:
                pass

    def delete_current_gsm(self):
        """Seçili olan BİR veya BİRDEN FAZLA numarayı siler ve EKRANI YENİLER."""

        selection = self.num_table.table.selectionModel().selectedRows()
        gsms_to_delete = []

        if selection:
            for idx in selection:
                src_idx = self.num_table.proxy_model.mapToSource(idx)
                gsm = self.num_table.source_model._data[src_idx.row()][0]
                gsms_to_delete.append(gsm)
        elif self.current_gsm_number:
            gsms_to_delete.append(self.current_gsm_number)

        if not gsms_to_delete:
            ModernDialog.show_warning(self, "Seçim Yok", "Lütfen silinecek numarayı listeden seçiniz.")
            return

        count = len(gsms_to_delete)
        if ModernDialog.show_question(self, "Silme Onayı", f"{count} adet numara silinecek ve analizler güncellenecek.\nOnaylıyor musunuz?"):

            if hasattr(self.main, 'loader'):
                self.main.loader.start(f"{count} Kayıt Siliniyor...")
            QApplication.processEvents()

            try:
                for gsm in gsms_to_delete:
                    self.delete_gsm_database_records(gsm)

                self.clear_all_widgets()

                self.load_numbers()

                if self.num_table.source_model.rowCount(QModelIndex()) > 0:
                    QTimer.singleShot(50, self.select_first_after_delete)
                else:
                    ModernDialog.show_success(self, "Tamamlandı", "Tüm numaralar silindi.")

            except Exception as e:
                ModernDialog.show_error(self, "Hata", str(e))
            finally:
                if hasattr(self.main, 'loader'): self.main.loader.stop()

    def select_first_after_delete(self):
        """Silme işleminden sonra listedeki ilk numarayı seçer."""
        try:
            self.num_table.table.selectRow(0)

            first_index = self.num_table.proxy_model.index(0, 0)
            self.on_num_select(first_index)

            ModernDialog.show_success(self, "Güncellendi", "Silme tamamlandı. Kalan veriler yüklendi.")
        except:
            pass

    def delete_records_for_role(self, gsm: str, rol: str):
        """
        Overwrite seçilince sadece ilgili rolün kayıtlarını temizler.
        DÜZELTME: self.selected_project_id kullanıldı.
        """
        with DB() as conn:
            conn.execute("""
                DELETE FROM hts_dosyalari
                WHERE ProjeID=? AND GSMNo=? AND Rol=?
            """, (self.selected_project_id, gsm, rol))

            if rol == "HEDEF":
                conn.execute("""
                    DELETE FROM hts_dosyalari
                    WHERE ProjeID=? AND GSMNo=?
                """, (self.selected_project_id, gsm))

            conn.commit()

        if rol == "HEDEF":
            self.delete_gsm_database_records(gsm)

        if rol == "KARSI":
            try:
                with DB() as conn:
                    for t in ["hts_karsi_baz", "hts_karsi_imei"]:
                        try:
                            conn.execute(f"DELETE FROM {t} WHERE ProjeID=? AND GSMNo=?", (self.selected_project_id, gsm))
                        except: pass
                    conn.commit()
            except:
                pass

    def on_worker_progress(self, value):
        """Worker'dan gelen ilerleme sinyalini yakalar."""
        if hasattr(self.main, 'loader'):
            self.main.loader.set_progress(value)

    def on_worker_log(self, text):
        """Worker'dan gelen log mesajlarını yakalar ve yükleyici metnini günceller."""
        if hasattr(self.main, 'loader'):
            self.main.loader.text = text
            self.main.loader.update()

    def on_worker_gsm_detected(self, gsm):
        """Worker'dan gelen GSM No bilgisini yakalar ve yükleyici metnini günceller."""
        if hasattr(self.main, 'loader'):
            self.main.loader.text = f"İşleniyor: {gsm}"
            self.main.loader.update()

    def on_single_file_finished(self, msg):
        """Tek dosya yüklemesi bittiğinde çağrılır."""
        self.success_count += 1
        self.load_numbers()
        self.process_next_in_queue()

    def on_num_select(self, index):
        try:
            if not index or not index.isValid():
                return

            target_table = self.num_table
            if target_table is None:
                return

            src_idx = target_table.proxy_model.mapToSource(index)
            row = src_idx.row()

            if row < 0 or row >= len(target_table.source_model._data):
                return

            gsm = target_table.source_model._data[row][0]
            if not gsm:
                return

            self.current_gsm_number = gsm

            if hasattr(self, "main") and self.main is not None:
                self.main.last_gsm_number = gsm

            if hasattr(self, "loaded_tabs"):
                self.loaded_tabs.clear()

            if hasattr(self, "load_overview_tab"):
                self.load_overview_tab()

            if hasattr(self, "load_quick_summary"):
                self.load_quick_summary()

            if hasattr(self, "load_common_analysis_tables"):
                self.load_common_analysis_tables()

            try:
                if hasattr(self, "tabs"):
                    self.tabs.setCurrentIndex(0)
            except:
                pass

        except Exception as e:
            print(f"GSM seçim hatası (Analiz): {e}")

    def on_tab_changed(self, index):
        """Sekme değiştiğinde veriyi yükler (İkon temizliği yapıldı)."""
        if not self.current_gsm_number:
            return

        display_name = self.tabs.tabText(index)

        clean_name = display_name
        if display_name == "Genel Bakış":
            clean_name = "Genel Bakış"
        else:
            for key, icon in self.tab_icons.items():
                if icon in display_name:
                    clean_name = key
                    break

        if clean_name not in self.loaded_tabs:
            if clean_name == "Genel Bakış":
                self.load_overview_tab()

            elif clean_name == "Olay Merkezli":
                # Olay Merkezli panel kendi içinde 'Analiz Et' ile yüklenir.
                if hasattr(self, "event_panel") and self.event_panel is not None:
                    self.event_panel.set_context(
                        project_id=self.current_project_id,
                        gsm_number=self.current_gsm_number,
                        start_qdt=self.dt_start.dateTime(),
                        end_qdt=self.dt_end.dateTime(),
                    )
                    self.event_panel.sync_from_context()

            else:
                self.load_specific_tab(clean_name)

            self.loaded_tabs.add(clean_name)

    def load_overview_tab(self):
        """Genel Bakış sekmesini güvenli bir şekilde yükler."""
        abone_list = []
        owner_label = str(self.current_gsm_number)
        home_txt = "Veri Yok"
        work_txt = "Veri Yok"
        min_dt, max_dt = None, None

        try:
            pid = self.current_project_id
            gsm = self.current_gsm_number

            with DB() as conn:
                cur = conn.cursor()

                cols_sql = ", ".join(self.abone_cols)
                abone_data = cur.execute(f"SELECT {cols_sql} FROM hts_abone WHERE ProjeID=? AND GSMNo=?", (pid, gsm)).fetchall()
                abone_list = [list(r) for r in abone_data]

                ozet = cur.execute("SELECT MinDate, MaxDate FROM hts_ozet WHERE ProjeID=? AND GSMNo=?", (pid, gsm)).fetchone()
                if ozet:
                    min_dt, max_dt = ozet[0], ozet[1]

                home_row = cur.execute("""
                    SELECT BAZ, COUNT(*) as Cnt
                    FROM hts_gsm 
                    WHERE ProjeID=? AND GSMNo=? 
                      AND (CAST(substr(TARIH, 12, 2) AS INT) >= 23 OR CAST(substr(TARIH, 12, 2) AS INT) < 6)
                      AND BAZ IS NOT NULL AND LENGTH(BAZ) > 5
                    GROUP BY BAZ ORDER BY Cnt DESC LIMIT 1
                """, (pid, gsm)).fetchone()

                work_row = cur.execute("""
                    SELECT BAZ, COUNT(*) as Cnt
                    FROM hts_gsm 
                    WHERE ProjeID=? AND GSMNo=? 
                      AND CAST(substr(TARIH, 12, 2) AS INT) BETWEEN 9 AND 18
                      AND BAZ IS NOT NULL AND LENGTH(BAZ) > 5
                    GROUP BY BAZ ORDER BY Cnt DESC LIMIT 1
                """, (pid, gsm)).fetchone()

                if home_row: home_txt = f"{home_row[0]} ({home_row[1]} Sinyal)"
                if work_row: work_txt = f"{work_row[0]} ({work_row[1]} Sinyal)"

            if abone_list:
                try:
                    ad = str(abone_list[0][2]) if abone_list[0][2] else ""
                    soyad = str(abone_list[0][3]) if abone_list[0][3] else ""
                    full_name = f"{ad} {soyad}".strip()
                    if full_name:
                        owner_label = f"{gsm} - {full_name}"
                except: pass

            if min_dt and max_dt:
                try:
                    mn = QDateTime.fromString(min_dt, "dd.MM.yyyy HH:mm:ss")
                    mx = QDateTime.fromString(max_dt, "dd.MM.yyyy HH:mm:ss")
                    if mn.isValid() and mx.isValid():
                        self.dt_start.setDateTime(mn)
                        self.dt_end.setDateTime(mx)
                except: pass
            else:
                now = QDateTime.currentDateTime()
                self.dt_start.setDateTime(now.addYears(-1))
                self.dt_end.setDateTime(now)

            self.abone_table.set_data(abone_list)
            self.profile_card.set_data(home_txt, work_txt)

            for w in [self.contact_table, self.top_table, self.baz_table, self.all_baz_table, self.imei_table]:
                w.set_owner_info(owner_label)
            for w in self.tab_widgets.values():
                w.set_owner_info(owner_label)

            self.load_quick_summary()

        except Exception as e:
            print(f"Overview Yükleme Hatası: {e}")

        finally:
            self.update_tab_visibility()
            self.loaded_tabs.add("Genel Bakış")
            if self.tabs.count() > 0:
                self.tabs.setCurrentIndex(0)

    def load_specific_tab(self, tab_name):
        """Seçili sekmenin verilerini çeker (GSM için NUMARA filtresi aktif)."""
        if hasattr(self.main, 'loader'):
            self.main.loader.start(f"{tab_name} Verileri Yükleniyor...")
        QApplication.processEvents()

        try:
            pid = self.current_project_id
            gsm = self.current_gsm_number

            q_start = self.dt_start.dateTime()
            q_end = self.dt_end.dateTime()
            py_start = q_start.toPyDateTime()
            py_end = q_end.toPyDateTime()

            def parse_fast(t_str):
                if not t_str: return datetime.min
                try:
                    parts = t_str.split(' ')
                    date_part = parts[0]
                    time_part = parts[1] if len(parts) > 1 else "00:00:00"

                    if "." in date_part: d, m, y = map(int, date_part.split('.'))
                    elif "/" in date_part: d, m, y = map(int, date_part.split('/'))
                    elif "-" in date_part: y, m, d = map(int, date_part.split('-'))
                    else: return datetime.min

                    h, mn, s = map(int, time_part.split(':'))
                    return datetime(y, m, d, h, mn, s)
                except: return datetime.min

            with DB() as conn:
                cur = conn.cursor()
                sql = ""
                rows = []

                if tab_name == "GSM":
                    sql = """
                        SELECT SIRA_NO, NUMARA, TIP, DIGER_NUMARA, TARIH, SURE, DIGER_ISIM, DIGER_TC, IMEI, BAZ 
                        FROM hts_gsm 
                        WHERE ProjeID=? AND GSMNo=? AND NUMARA=?
                    """
                    rows = cur.execute(sql, (pid, gsm, gsm)).fetchall()

                elif tab_name == "SMS":
                    sql = "SELECT SIRA_NO, NUMARA, TIP, DIGER_NUMARA, TARIH, SURE, DIGER_ISIM, DIGER_TC, MESAJ_BOYUTU, MESAJ_ICERIK_TIPI FROM hts_sms WHERE ProjeID=? AND GSMNo=?"
                    rows = cur.execute(sql, (pid, gsm)).fetchall()

                elif tab_name == "Sabit":
                    sql = "SELECT SIRA_NO, NUMARA, TIP, DIGER_NUMARA, TARIH, SURE, DIGER_ISIM, DIGER_TC FROM hts_sabit WHERE ProjeID=? AND GSMNo=?"
                    rows = cur.execute(sql, (pid, gsm)).fetchall()

                elif tab_name == "STH":
                    sql = "SELECT SIRA_NO, NUMARA, TIP, DIGER_NUMARA, TARIH, SURE, OPERATOR, DIGER_ISIM, DIGER_TC, DATA_TIP, DURUM, PIN_NO, BASL_GATEWAY, SONL_GATEWAY, BASL_SANTRAL, SONL_SANTRAL FROM hts_sth WHERE ProjeID=? AND GSMNo=?"
                    rows = cur.execute(sql, (pid, gsm)).fetchall()

                elif tab_name == "Uluslararası":
                    sql = "SELECT SIRA_NO, NUMARA, TIP, DIGER_NUMARA, TARIH, SURE, DIGER_ISIM, DIGER_TC FROM hts_uluslararasi WHERE ProjeID=? AND GSMNo=?"
                    rows = cur.execute(sql, (pid, gsm)).fetchall()

                elif tab_name == "İnternet":
                    gprs = cur.execute("SELECT 'GPRS', SIRA_NO, NUMARA, TIP, TARIH, SURE, IMEI, KAYNAK_IP, '', GONDERME, INDIRME, BAZ FROM hts_gprs WHERE ProjeID=? AND GSMNo=?", (pid, gsm)).fetchall()
                    wap = cur.execute("SELECT 'WAP', SIRA_NO, NUMARA, TIP, TARIH, SURE, IMEI, KAYNAK_IP, HEDEF_IP, GONDERME, INDIRME, BAZ FROM hts_wap WHERE ProjeID=? AND GSMNo=?", (pid, gsm)).fetchall()
                    rows = gprs + wap

                if tab_name in self.tab_widgets:
                    w = self.tab_widgets[tab_name]

                    filtered_data = []
                    date_col = w.date_col_index

                    if date_col != -1:
                        temp_rows = []
                        for r in rows:
                            t_str = r[date_col]
                            dt = parse_fast(t_str)
                            if dt and py_start <= dt <= py_end:
                                temp_rows.append((dt, list(r)))

                        temp_rows.sort(key=lambda x: x[0], reverse=False)

                        filtered_data = [x[1] for x in temp_rows]

                    else:
                        filtered_data = [list(r) for r in rows]

                    w.set_data(filtered_data)
                    w.proxy_model.setDateFilterActive(False)

                    if tab_name == "GSM":
                        t = w.table
                        h = t.horizontalHeader()
                        h.setSectionResizeMode(6, QHeaderView.ResizeMode.Interactive)
                        t.setColumnWidth(6, 170)
                        h.setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)
                        t.setColumnWidth(1, 110)
                        t.setColumnWidth(3, 110)
                        t.setColumnWidth(4, 130)

        except Exception as e:
            print(f"Tab Yükleme Hatası ({tab_name}): {e}")
            ModernDialog.show_error(self, "Hata", f"{tab_name} yüklenirken hata: {e}")
        finally:
            if hasattr(self.main, 'loader'):
                self.main.loader.stop()

    def refresh_top_analysis(self):
        """Analiz verilerini SEÇİLİ TARİH ARALIĞINA göre dinamik olarak hesaplar ve yükler."""
        if not self.current_project_id or not self.current_gsm_number: return

        self.stop_warning_animation()
        QApplication.processEvents()

        try:
            pid = self.current_project_id
            gsm = self.current_gsm_number

            q_start = self.dt_start.dateTime()
            q_end = self.dt_end.dateTime()
            py_start = q_start.toPyDateTime()
            py_end = q_end.toPyDateTime()

            def is_date_in_range(t_str):
                if not t_str: return None
                try:
                    fmt = "dd.MM.yyyy HH:mm:ss" if "." in t_str else "dd/MM/yyyy HH:mm:ss"
                    if " " not in t_str: fmt = fmt.split(" ")[0]
                    py_fmt = fmt.replace("dd", "%d").replace("MM", "%m").replace("yyyy", "%Y").replace("HH", "%H").replace("mm", "%M").replace("ss", "%S")
                    dt = datetime.strptime(t_str, py_fmt)

                    return dt if py_start <= dt <= py_end else None
                except:
                    return None

            with DB() as conn:
                cur = conn.cursor()
                raw_contacts = cur.execute("""
                    SELECT DIGER_NUMARA, TARIH, SURE, DIGER_ISIM, DIGER_TC 
                    FROM hts_gsm 
                    WHERE ProjeID=? AND GSMNo=? AND DIGER_NUMARA != ?
                """, (pid, gsm, gsm)).fetchall()

                raw_locations = []
                for t in ["hts_gsm", "hts_gprs", "hts_wap"]:
                    try:
                        raw_locations.extend(cur.execute(f"SELECT BAZ, IMEI, TARIH FROM {t} WHERE ProjeID=? AND GSMNo=?", (pid, gsm)).fetchall())
                    except: pass

            full_contact_list = []
            contact_map = defaultdict(lambda: {'count': 0, 'duration': 0, 'name': None, 'tc': None})
            total_contacts_count = 0

            for r in raw_contacts:
                diger_no, tarih, sure, isim, tc = r
                dt = is_date_in_range(tarih)

                if dt:
                    contact_map[diger_no]['count'] += 1

                    try:
                        clean_sure = re.sub(r'[^\d\.]', '', str(sure or '0'))
                        contact_map[diger_no]['duration'] += int(float(clean_sure))
                    except:
                        pass

                    if isim and not contact_map[diger_no]['name']: contact_map[diger_no]['name'] = isim
                    if tc and not contact_map[diger_no]['tc']: contact_map[diger_no]['tc'] = tc
                    total_contacts_count += 1

            full_contact_list_temp = []
            for k, v in contact_map.items():
                fmt_sure = self.format_seconds(v['duration'])
                full_contact_list_temp.append([k, v['count'], fmt_sure, v['name'] or ''])

            full_contact_list_temp.sort(key=lambda x: x[1], reverse=True)
            full_contact_list = full_contact_list_temp

            baz_counter = Counter()
            imei_stats = defaultdict(lambda: {'count': 0, 'min_dt': None, 'max_dt': None})

            for r in raw_locations:
                baz, imei, tarih = r
                dt = is_date_in_range(tarih)

                if dt:
                    if baz and str(baz).strip(): baz_counter[str(baz).strip()] += 1

                    if imei and str(imei).strip() and len(str(imei).strip()) >= 13:
                        clean_imei = str(imei).strip()
                        imei_stats[clean_imei]['count'] += 1

                        if imei_stats[clean_imei]['min_dt'] is None or dt < imei_stats[clean_imei]['min_dt']:
                            imei_stats[clean_imei]['min_dt'] = dt
                        if imei_stats[clean_imei]['max_dt'] is None or dt > imei_stats[clean_imei]['max_dt']:
                            imei_stats[clean_imei]['max_dt'] = dt

            full_baz_list_temp = baz_counter.most_common()
            full_baz_list = []
            for baz_adi, sinyal in full_baz_list_temp:
                if len(baz_adi) > 50:
                    display_baz = baz_adi[:125] + "..."
                else:
                    display_baz = baz_adi
                full_baz_list.append([display_baz, sinyal, baz_adi])

            imei_list = []
            for imei, stats in imei_stats.items():
                min_s = stats['min_dt'].strftime("%d.%m.%Y %H:%M:%S") if stats['min_dt'] else ""
                max_s = stats['max_dt'].strftime("%d.%m.%Y %H:%M:%S") if stats['max_dt'] else ""
                imei_list.append([imei, stats['count'], min_s, max_s])

            imei_list.sort(key=lambda x: x[1], reverse=True)

            self.contact_table.set_data(full_contact_list)
            self.top_table.set_data(full_contact_list[:20])

            self.all_baz_table.set_data(full_baz_list)
            self.baz_table.set_data(full_baz_list[:20])

            self.imei_table.set_data(imei_list)

            self.load_common_analysis_tables()

            if total_contacts_count > 10000:
                self.lbl_warning.setText(f"⚠️ DİKKAT: +{total_contacts_count} KAYIT! Analiz Tarihini Daraltın")
                self.flash_timer.start(500)
                self.lbl_warning.setVisible(True)
            else:
                self.stop_warning_animation()

            # Tabloları tekrar sıralat
            self.contact_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)
            self.all_baz_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)
            self.imei_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)

        except Exception as e:
            if hasattr(self.main, 'loader'): self.main.loader.stop()
            ModernDialog.show_error(self, "Analiz Güncelleme Hatası", f"Veri dinamik olarak hesaplanamadı: {e}")
            print(f"Hata: {e}")
        finally:
            if hasattr(self.main, 'loader'):
                self.main.loader.stop()

    def on_imei_tabs_changed(self, idx: int):
        try:
            # Swap tab indexini isimle kontrol (daha güvenli)
            tab_text = self.imei_tabs.tabText(idx) if hasattr(self, "imei_tabs") else ""
            if "Swap Timeline" not in tab_text:
                return

            # Proje yoksa çık
            pid = getattr(self, "current_project_id", None)
            if not pid:
                self.imei_swap_table.set_data([])
                return

            # Daha önce yüklendiyse tekrar hesaplama (istersen kaldır)
            if getattr(self, "_imei_swap_loaded_pid", None) == pid:
                return

            self.refresh_imei_swap_timeline(pid)
            self._imei_swap_loaded_pid = pid

        except Exception as e:
            print(f"IMEI Swap Timeline tab changed hata: {e}")


    def refresh_imei_swap_timeline(self, pid: int):
        """
        LAZY: Kullanıcı Swap sekmesine girince çalışır.
        Yükleme/parse akışını KESİNLİKLE etkilemez.
        """
        try:
            with DB() as conn:
                cur = conn.cursor()

                # Ortak IMEI listesini mevcut tablodan değil DB’den çekmek daha net:
                # hts_ortak_imei zaten var ve ortak analiz sonrası dolu olmalı.
                imei_rows = cur.execute("""
                    SELECT IMEI, KullananSayisi, Numaralar
                    FROM hts_ortak_imei
                    WHERE ProjeID=?
                    ORDER BY KullananSayisi DESC
                    LIMIT 2000
                """, (pid,)).fetchall()

                out_rows = []
                for imei, ksay, nums in imei_rows:
                    imei = (str(imei).strip() if imei is not None else "")
                    if not imei:
                        continue

                    num_list = []
                    if nums:
                        # Numaralar genelde "05xx..., 05xx..." biçiminde
                        for x in str(nums).replace(";", ",").split(","):
                            x = x.strip()
                            if x:
                                num_list.append(x)

                    segments = self._query_imei_segments(conn, pid, imei, num_list)

                    # Timeline metni (kısa)
                    tl_parts = []
                    overlap = False
                    last_end = None

                    for gsm, first_seen, last_seen, cnt in segments:
                        tl_parts.append(f"{gsm} [{first_seen} - {last_seen}] ({cnt})")

                        # Overlap kontrolü (kaba ama hızlı)
                        if last_end is not None and first_seen <= last_end:
                            overlap = True
                        if last_end is None or last_seen > last_end:
                            last_end = last_seen

                    timeline_txt = " | ".join(tl_parts) if tl_parts else "-"
                    out_rows.append([imei, int(ksay or 0), timeline_txt, "VAR" if overlap else "YOK"])

                self.imei_swap_table.set_data(out_rows)

        except Exception as e:
            print(f"refresh_imei_swap_timeline hata: {e}")
            self.imei_swap_table.set_data([])


    def _query_imei_segments(self, conn, pid: int, imei: str, num_list: list):
        """
        Bir IMEI için, hangi GSM’lerde hangi zaman aralığında kullanılmış?
        Kaynak: GSM + GPRS + WAP (mevcut DB tabloların)
        Dönen: [(gsm, first_seen, last_seen, count)] -> first/last 'YYYYMMDDHHMMSS' formatında string
        """
        cur = conn.cursor()

        # num_list boşsa IN filtresi uygulamayalım (ortak imei tablosu zaten daraltılmış olabilir)
        in_sql = ""
        params = [pid, imei, pid, imei, pid, imei]
        if num_list:
            placeholders = ",".join(["?"] * len(num_list))
            in_sql = f" AND GSMNo IN ({placeholders}) "
            params += num_list * 3

        # TARIH formatı dd.mm.yyyy HH:MM:SS varsayımıyla normalize (senin sorguların bu mantığı zaten kullanıyor)
        # 'YYYYMMDDHHMMSS' üretelim ki MIN/MAX düzgün çalışsın
        sql = f"""
            SELECT GSMNo,
                   MIN(k) as first_k,
                   MAX(k) as last_k,
                   COUNT(*) as cnt
            FROM (
                SELECT GSMNo,
                       (substr(TARIH, 7, 4) || substr(TARIH, 4, 2) || substr(TARIH, 1, 2) ||
                        replace(replace(substr(TARIH, 12), ':',''), ' ', '')) as k
                FROM hts_gsm
                WHERE ProjeID=? AND IMEI=? {in_sql}

                UNION ALL

                SELECT GSMNo,
                       (substr(TARIH, 7, 4) || substr(TARIH, 4, 2) || substr(TARIH, 1, 2) ||
                        replace(replace(substr(TARIH, 12), ':',''), ' ', '')) as k
                FROM hts_gprs
                WHERE ProjeID=? AND IMEI=? {in_sql}

                UNION ALL

                SELECT GSMNo,
                       (substr(TARIH, 7, 4) || substr(TARIH, 4, 2) || substr(TARIH, 1, 2) ||
                        replace(replace(substr(TARIH, 12), ':',''), ' ', '')) as k
                FROM hts_wap
                WHERE ProjeID=? AND IMEI=? {in_sql}
            )
            GROUP BY GSMNo
            ORDER BY first_k ASC
        """

        rows = cur.execute(sql, params).fetchall()

        # k -> tekrar dd.mm.yyyy HH:MM:SS gibi okunur üretmek istersen burada formatlayabiliriz,
        # ama hızlı olsun diye k’yı olduğu gibi kullanıyoruz.
        out = []
        for gsm, first_k, last_k, cnt in rows:
            gsm = str(gsm) if gsm is not None else ""
            out.append((gsm, str(first_k or ""), str(last_k or ""), int(cnt or 0)))
        return out

    def open_imei_swap_timeline_detail(self, model_index):
        try:
            r = model_index.row()
            imei_item = self.imei_swap_table.table.item(r, 0)
            imei = imei_item.text().strip() if imei_item else ""
            if not imei:
                return

            # Proje ID için sağlam fallback
            pid = (
                getattr(self, "current_project_id", None)
                or getattr(self, "current_pid", None)
                or getattr(self, "pid", None)
            )
            if not pid:
                print("IMEI Swap Timeline: ProjeID bulunamadı (current_project_id/current_pid/pid yok).")
                return

            with DB() as conn:
                cur = conn.cursor()

                row = cur.execute("""
                    SELECT Numaralar
                    FROM hts_ortak_imei
                    WHERE ProjeID=? AND IMEI=? LIMIT 1
                """, (pid, str(imei))).fetchone()

                num_list = []
                if row and row[0]:
                    for x in str(row[0]).replace(";", ",").split(","):
                        x = x.strip()
                        if x:
                            num_list.append(x)

                segments = self._query_imei_segments(conn, pid, str(imei), num_list)

            dlg = ImeiSwapTimelineDialog(self, imei=str(imei), segments=segments)
            self.open_window_safe(dlg)

        except Exception as e:
            print(f"open_imei_swap_timeline_detail hata: {e}")

    def load_common_analysis_tables(self):
        """Ortak analiz tablolarını (Tarihten bağımsız) yükler."""
        if not self.current_project_id: return

        try:
            with DB() as conn:
                cur = conn.cursor()
                pid = self.current_project_id

                common_rows = cur.execute("""
                    SELECT IMEI, KullananSayisi, Numaralar, ToplamKullanim 
                    FROM hts_ortak_imei 
                    WHERE ProjeID=? 
                    ORDER BY KullananSayisi DESC
                """, (pid,)).fetchall()
                self.common_imei_table.set_data([list(r) for r in common_rows])

                if common_rows:
                    self.imei_tabs.tabBar().setTabTextColor(1, QColor("red"))
                    self.imei_tabs.setTabText(1, f"⚠️ Ortak IMEI ({len(common_rows)})")
                else:
                    self.imei_tabs.tabBar().setTabTextColor(1, QColor("black"))
                    self.imei_tabs.setTabText(1, "Ortak IMEI (Yok)")

                name_rows = cur.execute("""
                    SELECT AdSoyad, HatSayisi, Numaralar 
                    FROM hts_ortak_isim 
                    WHERE ProjeID=? 
                    ORDER BY HatSayisi DESC
                """, (pid,)).fetchall()
                self.common_name_table.set_data([list(r) for r in name_rows])

                h_name = self.common_name_table.table.horizontalHeader()
                h_name.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
                self.common_name_table.table.setColumnWidth(0, 180)
                h_name.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
                self.common_name_table.table.setColumnWidth(1, 180)
                h_name.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

                tc_rows = cur.execute("""
                    SELECT TC, HatSayisi, Numaralar 
                    FROM hts_ortak_tc 
                    WHERE ProjeID=? 
                    ORDER BY HatSayisi DESC
                """, (pid,)).fetchall()
                self.common_tc_table.set_data([list(r) for r in tc_rows])

                h_tc = self.common_tc_table.table.horizontalHeader()
                h_tc.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
                self.common_tc_table.table.setColumnWidth(0, 180)
                h_tc.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
                self.common_tc_table.table.setColumnWidth(1, 180)
                h_tc.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

                if name_rows: self.sub_tabs.tabBar().setTabTextColor(0, QColor("#d35400"))
                if tc_rows: self.sub_tabs.tabBar().setTabTextColor(1, QColor("#d35400"))

                self.common_imei_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)
                self.common_name_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)
                self.common_tc_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)

        except Exception as e:
            print(f"Ortak Tablo Yükleme Hatası: {e}")

    def format_seconds(self, seconds):
        if not seconds: return "0 sn"
        try:
            sec = int(seconds)
        except:
            return "0 sn"
        m, s = divmod(sec, 60)
        h, m = divmod(m, 60)
        parts = []
        if h: parts.append(f"{int(h)} sa")
        if m: parts.append(f"{int(m)} dk")
        if s or not parts: parts.append(f"{int(s)} sn")
        return " ".join(parts)

    def load_quick_summary(self):
        if not self.current_project_id or not self.current_gsm_number: return
        if not LicenseManager.require_valid_or_exit(self, "GSM değiştirildi"):
            return

        self.stop_warning_animation()

        try:
            pid = self.current_project_id
            gsm = self.current_gsm_number

            with DB() as conn:
                cur = conn.cursor()

                contact_rows = cur.execute("""
                    SELECT KarsiNo, Adet, Sure, Isim 
                    FROM hts_rehber 
                    WHERE ProjeID=? AND GSMNo=?
                      AND TC IS NOT NULL
                      AND TRIM(TC) GLOB '[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]'
                    ORDER BY Adet DESC
                """, (pid, gsm)).fetchall()

                full_contact_list = []
                for r in contact_rows:
                    karsi_no, adet, sure, isim = r
                    fmt_sure = self.format_seconds(sure)
                    full_contact_list.append([karsi_no, adet, fmt_sure, isim])

                self.contact_table.set_data(full_contact_list)
                self.top_table.set_data(full_contact_list[:20])

                b_rows = cur.execute("""
                    SELECT BazAdi, Sinyal 
                    FROM hts_tum_baz 
                    WHERE ProjeID=? AND GSMNo=? 
                    ORDER BY Sinyal DESC
                """, (pid, gsm)).fetchall()

                full_baz_list = []
                for baz_adi, sinyal in b_rows:
                    if len(baz_adi) > 50:
                        display_baz = baz_adi[:125] + "..."
                    else:
                        display_baz = baz_adi
                    full_baz_list.append([display_baz, sinyal, baz_adi])

                self.all_baz_table.set_data(full_baz_list)
                self.baz_table.set_data(full_baz_list[:20])

                i_rows = cur.execute("""
                    SELECT TRIM(IMEI), MAX(Adet), MIN(MinDate), MAX(MaxDate) 
                    FROM hts_ozet_imei 
                    WHERE ProjeID=? AND GSMNo=? 
                    GROUP BY TRIM(IMEI) 
                    ORDER BY MAX(Adet) DESC
                """, (pid, gsm)).fetchall()
                self.imei_table.set_data([list(r) for r in i_rows])

                h_imei = self.imei_table.table.horizontalHeader()
                h_imei.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
                self.imei_table.table.setColumnWidth(0, 145)
                h_imei.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
                self.imei_table.table.setColumnWidth(1, 145)
                h_imei.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
                self.imei_table.table.setColumnWidth(2, 150)
                h_imei.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)

                self.load_common_analysis_tables()

                self.contact_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)
                self.all_baz_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)
                self.imei_table.table.sortByColumn(1, Qt.SortOrder.DescendingOrder)

        except Exception as e:
            print(f"Hızlı Yükleme Hatası: {e}")

    def refresh_all_analysis(self):
        """Tarih aralığı değişince tüm sekmeleri yeniler ve UI geri bildirimi sağlar."""

        btn = self.sender()
        original_text = btn.text() if btn else "Analizi Güncelle"

        if btn:
            btn.setEnabled(False)
            btn.setText("YÜKLENİYOR...")

        QApplication.processEvents()

        if hasattr(self.main, 'loader'):
            self.main.loader.start("Analizler Güncelleniyor...")

        try:
            self.refresh_top_analysis()

            self.loaded_tabs.clear()
            self.loaded_tabs.add("Genel Bakış")

            current_index = self.tabs.currentIndex()
            display_name = self.tabs.tabText(current_index)

            clean_name = display_name
            if display_name != "Genel Bakış":
                for key, icon in self.tab_icons.items():
                    if icon in display_name:
                        clean_name = key
                        break

            if clean_name != "Genel Bakış":
                self.load_specific_tab(clean_name)
                self.loaded_tabs.add(clean_name)

        except Exception as e:
            print(f"Analiz Güncelleme Hatası: {e}")
            ModernDialog.show_error(self, "Hata", f"Güncelleme sırasında hata: {e}")

        finally:
            if btn:
                btn.setText(original_text)
                btn.setEnabled(True)

            if hasattr(self.main, 'loader'):
                self.main.loader.stop()

    def toggle_warning_animation(self):
        """Uyarı etiketini yanıp söndürür."""
        self.flash_state = not self.flash_state
        if self.flash_state:
            self.lbl_warning.setStyleSheet("font-weight: bold; color: white; background-color: #e74c3c; padding: 5px; border-radius: 4px; margin-left: 10px;")
        else:
            self.lbl_warning.setStyleSheet("font-weight: bold; color: black; background-color: #f1c40f; padding: 5px; border-radius: 4px; margin-left: 10px;")

    def stop_warning_animation(self):
        """Animasyonu durdurur ve etiketi gizler."""
        self.flash_timer.stop()
        self.lbl_warning.setVisible(False)

    def _check_dt(self, t_str, start, end):
        if not t_str: return False
        try:
            fmt = "dd.MM.yyyy HH:mm:ss" if "." in t_str else "dd/MM/yyyy HH:mm:ss"
            if " " not in t_str: fmt = fmt.split(" ")[0]
            dt = QDateTime.fromString(t_str, fmt)
            return dt.isValid() and start <= dt <= end
        except: return False

    def open_interaction_detail(self, index):
        """İletişim detaylarını açar (Tablo Kaynağı Düzeltildi)."""
        try:
            sender_widget = self.sender()

            target_table = None

            if sender_widget == self.contact_table.table:
                target_table = self.contact_table
            elif sender_widget == self.top_table.table:
                target_table = self.top_table

            if target_table is None:
                target_table = self.top_table

            src = target_table.proxy_model.mapToSource(index)
            d = target_table.source_model._data[src.row()]

            target_gsm = d[0]
            target_name = d[3]

            dlg = InteractionDetailDialog(
                self,
                self.current_project_id,
                self.current_gsm_number,
                target_gsm,
                target_name,
                self.dt_start.dateTime(),
                self.dt_end.dateTime()
            )
            self.open_window_safe(dlg)

        except Exception as e:
            print(f"Detay Açma Hatası: {e}")

    def open_location_detail(self, index):
        if not self.current_project_id or not self.current_gsm_number: return

        try:
            proxy_model = self.sender().model()
            if not proxy_model: return

            row = index.row()

            full_baz_adi_index = proxy_model.index(row, 2)
            full_baz_adi = full_baz_adi_index.data()

            if not full_baz_adi:
                QMessageBox.critical(self, "Hata", "Baz İstasyonu tam adı okunamadı. Lütfen tablo verilerini kontrol edin.")
                return

            pid = self.current_project_id
            gsm = self.current_gsm_number
            min_dt, max_dt = self.get_date_range()

            dialog = LocationDetailDialog(
                self,
                pid,
                gsm,
                full_baz_adi,
                min_dt,
                max_dt
            )
            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Konum Detayları yüklenirken hata oluştu: {e}")

    def get_date_range(self):
        """UI'dan seçili başlangıç ve bitiş tarihlerini (QDateTime) döndürür."""
        return self.dt_start.dateTime(), self.dt_end.dateTime()

    def open_imei_detail(self, index, is_common=False):
        """IMEI detay penceresini açar. Ortak veya Kişisel moda göre davranır."""
        try:
            if is_common:
                src = self.common_imei_table.proxy_model.mapToSource(index)
                d = self.common_imei_table.source_model._data[src.row()]
                target_imei = d[0]
            else:
                src = self.imei_table.proxy_model.mapToSource(index)
                d = self.imei_table.source_model._data[src.row()]
                target_imei = d[0]

            dlg = ImeiDetailDialog(
                self,
                self.current_project_id,
                self.current_gsm_number,
                target_imei,
                self.dt_start.dateTime(),
                self.dt_end.dateTime(),
                is_common=is_common
            )
            self.open_window_safe(dlg)

        except Exception as e:
            print(f"IMEI Detay Hatası: {e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("HTSMercek - Adli İletişim Kayıtları Analiz Merkezi")
        self.setGeometry(100, 100, 1200, 800)

        setup_database()

        central = QWidget(self)
        self.setCentralWidget(central)

        central_layout = QStackedLayout(central)
        central_layout.setStackingMode(QStackedLayout.StackingMode.StackAll)

        self.stack = QStackedWidget(central)
        central_layout.addWidget(self.stack)

        base_dir = os.path.dirname(__file__)
        logo_path = os.path.join(base_dir, "assets", "bg_logo.png")
        if not os.path.exists(logo_path):
            logo_path = os.path.join(base_dir, "assets", "logo.png")

        self.watermark = WatermarkBackground(
            logo_path=logo_path,
            opacity=0.04,
            scale=1.2,
            parent=central
        )
        central_layout.addWidget(self.watermark)
        self.watermark.raise_()

        self.last_gsm_number = None

        self.page_projects = ProjectManager(self)
        self.page_analysis = AnalysisCenter(self)

        self.stack.addWidget(self.page_projects)
        self.stack.addWidget(self.page_analysis)
        self.page_projects.sig_upload_started.connect(self.page_analysis.external_upload_started)
        self.page_projects.sig_progress_updated.connect(self.page_analysis.external_progress_updated)
        self.page_projects.sig_gsm_detected.connect(self.page_analysis.external_gsm_detected)
        self.page_projects.sig_file_finished.connect(self.page_analysis.external_file_finished)
        self.page_projects.sig_queue_finished.connect(self.page_analysis.external_queue_finished)

        self.loader = LoadingOverlay(self)
        mbtiles_file = os.path.join(APP_DIR, "turkey.mbtiles")
        if os.path.exists(mbtiles_file):
            self.map_server = LocalTileServer(mbtiles_file)
            self.map_server.start()
        else:
            print("UYARI: Harita dosyası (turkey.mbtiles) bulunamadı. Haritalar online çalışacak.")
            self.map_server = None
        enforce_normal_table_fonts(self)

        self.license_timer = QTimer(self)
        self.license_timer.timeout.connect(self._perform_security_check)
        self.license_timer.start(300000)

    def _perform_security_check(self):
        """Arka planda lisans kontrolü yapar."""
        try:
            LicenseManager.ensure_valid_or_raise()
        except Exception as e:
            try:
                self.license_timer.stop()
            except Exception:
                pass

            try:
                if "ModernDialog" in globals():
                    ModernDialog.show_error(
                        self,
                        "Lisans Hatası",
                        f"Lisans doğrulaması başarısız oldu.\nProgram kapatılacak.\n\nDetay: {e}"
                    )
                else:
                    QMessageBox.critical(
                        self,
                        "Lisans Hatası",
                        f"Lisans doğrulaması başarısız oldu.\nProgram kapatılacak.\n\nDetay: {e}"
                    )
            finally:
                # sys.exit(0) yerine güvenli çıkış
                if "_quit_app" in globals():
                    _quit_app()
                else:
                    try:
                        QApplication.quit()
                    except Exception:
                        os._exit(0)

    def show_analysis(self, pid):
        self.page_analysis.set_project(pid)
        self.stack.setCurrentIndex(1)

    def closeEvent(self, event):
        """Kullanıcı X tuşuna bastığında çalışır."""

        if self.page_analysis.is_uploading and not self.page_analysis.shutdown_pending:

            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setWindowTitle("Yükleme Devam Ediyor")
            msg.setText("Şu an dosya yükleme işlemi devam ediyor.")
            msg.setInformativeText("Veritabanı güvenliği için mevcut dosyanın bitmesi beklenmelidir.\n\nProgram, mevcut dosya biter bitmez OTOMATİK KAPANACAKTIR.\nOnaylıyor musunuz?")
            msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            msg.setDefaultButton(QMessageBox.StandardButton.No)

            reply = msg.exec()

            if reply == QMessageBox.StandardButton.Yes:
                self.page_analysis.request_graceful_close()
                event.ignore()
            else:
                event.ignore()

        else:
            event.accept()

    def resizeEvent(self, event):
            if hasattr(self, 'loader'):
                self.loader.resize(self.width(), self.height())
            super().resizeEvent(event)


class MultiNumberDetailDialog(QDialog):
    """Ortak isme/TC'ye ait numaralarla PROJEDEKİ TÜM HEDEFLER arasındaki ilişkiyi gösterir."""
    def __init__(self, parent, project_id, owner_name, number_list_str):
        super().__init__(parent)

        central = QWidget(self)
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(0, 0, 0, 0)
        central_layout.setSpacing(0)

        base_dir = os.path.dirname(__file__)
        logo_path = os.path.join(base_dir, "assets", "bg_logo.png")
        if not os.path.exists(logo_path):
            logo_path = os.path.join(base_dir, "assets", "logo.png")

        self._watermark = WatermarkBackground(
            logo_path=logo_path,
            opacity=0.03,
            scale=1.20,
            parent=central
        )

        stack = QStackedLayout(self)
        stack.setStackingMode(QStackedLayout.StackingMode.StackAll)
        stack.addWidget(central)
        stack.addWidget(self._watermark)
        self._watermark.raise_()

        self.setWindowTitle(f"İlişki Analizi: {owner_name}")
        self.resize(1300, 750)

        layout = QVBoxLayout()
        central_layout.addLayout(layout)

        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #e1f5fe; 
                border: 1px solid #81d4fa; 
                border-radius: 8px; 
                padding: 10px;
            }
        """)
        hl = QHBoxLayout(info_frame)

        lbl_icon = QLabel("🔗")
        lbl_icon.setStyleSheet("font-size: 32px; border:none;")

        info_text = (
            f"<div style='font-size:14px; color:#0277bd; font-family: Segoe UI;'>"
            f"<b>👥 Ortak Kişi:</b> {owner_name}<br>"
            f"<b>ℹ️ Bilgi:</b> Bu kişinin numaraları ile <u>Projedeki Tüm Hedefler</u> arasındaki trafik aşağıdadır."
            f"</div>"
        )

        hl.addWidget(lbl_icon)
        hl.addWidget(QLabel(info_text))
        hl.addStretch()
        layout.addWidget(info_frame)

        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        target_numbers = [n.strip() for n in number_list_str.split(',') if n.strip()]

        for other_gsm in target_numbers:
            page = QWidget()
            l = QVBoxLayout(page)

            cols = ["Hat Sahibi (Hedef No)", "Tarih/Saat", "İletişim Türü", "Yön", "Süre / Mesaj", "Baz İstasyonu"]
            table = GenericDatabaseTable(cols, chart_mode='none')

            h = table.table.horizontalHeader()
            h.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

            try:
                with DB() as conn:
                    cur = conn.cursor()
                    limit_clause = " LIMIT 10001"
                    sql = f"""
                        SELECT * FROM (
                            SELECT t1.GSMNo || ' - ' || COALESCE((SELECT AD || ' ' || SOYAD FROM hts_abone WHERE ProjeID=t1.ProjeID AND GSMNo=t1.GSMNo LIMIT 1), ''),
                                   t1.TARIH, 'ARAMA', t1.TIP, t1.SURE || ' sn', t1.BAZ 
                            FROM hts_gsm t1 WHERE t1.ProjeID=? AND t1.DIGER_NUMARA=?
                            UNION ALL
                            SELECT t2.GSMNo || ' - ' || COALESCE((SELECT AD || ' ' || SOYAD FROM hts_abone WHERE ProjeID=t2.ProjeID AND GSMNo=t2.GSMNo LIMIT 1), ''),
                                   t2.TARIH, 'SMS', t2.TIP, t2.MESAJ_BOYUTU || ' byte', '---'
                            FROM hts_sms t2 WHERE t2.ProjeID=? AND t2.DIGER_NUMARA=?
                        ) ORDER BY 2 DESC {limit_clause}
                    """
                    params = (project_id, other_gsm, project_id, other_gsm)
                    rows = cur.execute(sql, params).fetchall()

                    if len(rows) > 10000:
                        ModernDialog.show_warning(self, "Limit", f"'{other_gsm}' için liste 10.000 kayıtla sınırlandı.")

                    table.set_data([list(r) for r in rows])
                    t = table.table
                    t.setColumnWidth(0, 230); t.setColumnWidth(1, 140); t.setColumnWidth(2, 80)
                    t.setColumnWidth(3, 70); t.setColumnWidth(4, 100)
                    h.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)

                    tab_title = f"✅ {other_gsm} ({len(rows)})" if rows else f"{other_gsm} (0)"

            except Exception as e:
                print(f"Hata: {e}")
                table.set_data([])
                tab_title = f"{other_gsm} (Hata)"

            l.addWidget(table)
            self.tabs.addTab(page, tab_title)

        btn_close = QPushButton("Kapat")
        btn_close.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; padding:8px;")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)


class ReportHtmlBuilder:
    def __init__(self, project_id: int):
        self.project_id = project_id

    def _ensure_meta_row(self):
        """
        rapor_meta satırı yoksa default metinlerle oluşturur.
        Satır varsa ama eski sürümden kalma şekilde bütün metin kolonları NULL ise
        (yani hiç initialize edilmemişse) yine defaultları yazar.
        Kullanıcı bilerek metni silip "" yaptıysa dokunmaz.
        """
        pid = self.project_id
        try:
            pid = int(pid)
        except Exception:
            pid = str(pid).strip()

        with DB() as conn:
            row = conn.execute("""
                SELECT
                    GorevlendirmeMetni,
                    DosyaHakkindaMetni,
                    GenelBilgilendirmeMetni,
                    DegerlendirmeMetni,
                    SonucMetni
                FROM rapor_meta
                WHERE ProjeID=?
            """, (pid,)).fetchone()

            # hiç satır yoksa: defaultlarla INSERT
            if not row:
                gorev = self.build_default_gorevlendirme()
                genel = self.build_default_genel_bilgi()
                conn.execute("""
                    INSERT INTO rapor_meta (
                        ProjeID,
                        GorevlendirmeMetni,
                        DosyaHakkindaMetni,
                        GenelBilgilendirmeMetni,
                        DegerlendirmeMetni,
                        SonucMetni,
                        GuncellemeTarihi
                    ) VALUES (?,?,?,?,?,?,CURRENT_TIMESTAMP)
                """, (pid, gorev, "", genel, "", ""))
                conn.commit()
                return

            # satır var ama eski bozuk sürümden kalma: tüm metinler NULL ise initialize et
            gorev_db, dosya_db, genel_db, deg_db, sonuc_db = row
            if (gorev_db is None and dosya_db is None and genel_db is None and deg_db is None and sonuc_db is None):
                gorev = self.build_default_gorevlendirme()
                genel = self.build_default_genel_bilgi()
                conn.execute("""
                    UPDATE rapor_meta SET
                        GorevlendirmeMetni=?,
                        DosyaHakkindaMetni=?,
                        GenelBilgilendirmeMetni=?,
                        DegerlendirmeMetni=?,
                        SonucMetni=?,
                        GuncellemeTarihi=CURRENT_TIMESTAMP
                    WHERE ProjeID=?
                """, (gorev, "", genel, "", "", pid))
                conn.commit()

    def _meta_desc_to_text(self, acik) -> str:
        import re, html as _html
        if not acik:
            return ""
        s = str(acik).strip()
        if not s:
            return ""

        # RichDescriptionDialog -> toHtml() çıktısı gelirse düz metne indir.
        if "<" in s and ">" in s:
            s = re.sub(r"(?is)<!doctype.*?>", "", s)
            s = re.sub(r"(?is)<head.*?>.*?</head>", "", s)
            s = re.sub(r"(?is)<(script|style).*?>.*?</\1>", "", s)
            s = re.sub(r"(?is)</?html.*?>", "", s)
            s = re.sub(r"(?is)</?body.*?>", "", s)
            s = re.sub(r"(?is)<br\s*/?>", "\n", s)
            s = re.sub(r"(?is)</p\s*>", "\n", s)
            s = re.sub(r"(?is)<[^>]+>", "", s)
            s = _html.unescape(s)

        return s.strip()

    def _sanitize_rich_html(self, s: str) -> str:
        """
        RichDescriptionDialog.toHtml() gibi tam HTML dönen içeriklerde:
        - DOCTYPE/html/head/body dış kabuğunu temizle
        - sadece body içini al (varsa)
        - <style>/<script>/<meta> temizle
        """
        if not s:
            return ""
        txt = str(s)

        # DOCTYPE temizle
        txt = re.sub(r"<!DOCTYPE[^>]*>", "", txt, flags=re.IGNORECASE | re.DOTALL)

        # body içini çek (varsa)
        m = re.search(r"<body[^>]*>(.*?)</body>", txt, flags=re.IGNORECASE | re.DOTALL)
        if m:
            txt = m.group(1)

        # style/script/meta/head/html temizle
        txt = re.sub(r"<style[^>]*>.*?</style>", "", txt, flags=re.IGNORECASE | re.DOTALL)
        txt = re.sub(r"<script[^>]*>.*?</script>", "", txt, flags=re.IGNORECASE | re.DOTALL)
        txt = re.sub(r"<meta[^>]*>", "", txt, flags=re.IGNORECASE | re.DOTALL)
        txt = re.sub(r"</?head[^>]*>", "", txt, flags=re.IGNORECASE | re.DOTALL)
        txt = re.sub(r"</?html[^>]*>", "", txt, flags=re.IGNORECASE | re.DOTALL)

        return txt.strip()

    def _strip_tags_len(self, html_text: str) -> int:
        if not html_text:
            return 0
        t = re.sub(r"<[^>]+>", "", str(html_text), flags=re.DOTALL)
        t = re.sub(r"\s+", " ", t).strip()
        return len(t)

    def _render_meta_ekler_html(self, bolum: str, baslik: str) -> str:
        """
        Dosya eklerini a., b., c. şeklinde listeler.
        HATA ÇÖZÜMÜ: 'ek_no' değişkeni kaldırıldı, enumerate kullanıldı.
        """
        rows = self._fetch_meta_ekler(bolum)
        if str(bolum) == "delil_cekmece":
            return ""

        parts = [f"<h2>{html.escape(baslik)}</h2>"]

        # enumerate(rows) ile otomatik sayaç: 0, 1, 2...
        for i, (db_id, dosya_adi, dosya_yolu, aciklama, dosya_adi_gizle, gen_pct) in enumerate(rows):

            # 0 -> a, 1 -> b...
            harf = chr(97 + i)

            fname = (dosya_adi or (os.path.basename(dosya_yolu or "") if dosya_yolu else "")).strip()
            gen_pct = int(gen_pct or 100)
            gen_pct = max(10, min(100, gen_pct))

            safe_desc_html = self._sanitize_rich_html(aciklama or "")
            desc_len = self._strip_tags_len(safe_desc_html)
            desc_is_long = desc_len >= 700

            # İsim satırı (a. DosyaAdı)
            if int(dosya_adi_gizle or 0) == 0:
                name_line = (
                    f"<div class='metaek-name'>"
                    f"<span class='metaek-no'>{harf}.</span> "
                    f"<span class='metaek-fname'>{html.escape(fname)}</span>"
                    f"</div>"
                )
            else:
                name_line = f"<div class='metaek-name'><span class='metaek-no'>{harf}.</span></div>"

            # Resim
            if dosya_yolu and os.path.exists(dosya_yolu):
                src = QUrl.fromLocalFile(dosya_yolu).toString()
                img_html = (
                    f"<div class='metaek-imgwrap'>"
                    f" <img class='metaek-img' "
                    f" src='{src}' "
                    f" style='width:{gen_pct}%;' "
                    f" onclick=\"openImgModal('{src}')\" />"
                    f"</div>"
                )
            else:
                img_html = "<div class='metaek-missing'>[Dosya bulunamadı]</div>"

            desc_html = ""
            if safe_desc_html.strip():
                desc_html = f"<div class='metaek-desc'>{safe_desc_html}</div>"

            # DİKKAT: ID olarak veritabanı ID'sini (db_id) kullanıyoruz
            parts.append(
                f"<div class='metaek-block' id='metaek-{db_id}'>"
                f" {name_line}"
                f" {img_html}"
                f" {desc_html}"
                f"</div>"
            )

        return "\n".join(parts)

    def _fetch_meta_ekler_files(self, bolum: str):
        """
        rapor_meta_ekler tablosundan Dosya Hakkında / Delil Çekmecesi eklerini çeker.
        Kolonlar eski sürümlerde yoksa (DosyaAdiGizle, GenislikYuzde) fallback uygular.
        Dönen: [(id, DosyaAdi, DosyaYolu, Aciklama, DosyaAdiGizle, GenislikYuzde), ...]
        """
        with DB() as conn:
            cols = {r[1] for r in conn.execute("PRAGMA table_info(rapor_meta_ekler)").fetchall()}

            id_expr = "id" if "id" in cols else "rowid AS id"
            name_expr = "DosyaAdi" if "DosyaAdi" in cols else "NULL AS DosyaAdi"
            path_expr = "DosyaYolu" if "DosyaYolu" in cols else "NULL AS DosyaYolu"
            desc_expr = "Aciklama" if "Aciklama" in cols else "'' AS Aciklama"
            hide_expr = "DosyaAdiGizle" if "DosyaAdiGizle" in cols else "0 AS DosyaAdiGizle"
            w_expr = "GenislikYuzde" if "GenislikYuzde" in cols else "100 AS GenislikYuzde"

            rows = conn.execute(
                f"""
                SELECT {id_expr},
                       {name_expr},
                       {path_expr},
                       {desc_expr},
                       {hide_expr},
                       {w_expr}
                FROM rapor_meta_ekler
                WHERE ProjeID=? AND Bolum=?
                ORDER BY id ASC
                """,
                (self.project_id, bolum)
            ).fetchall()

        return rows

    def _fetch_meta_ekler(self, bolum: str):
        """
        rapor_meta_ekler:
          id, DosyaAdi, DosyaYolu, Aciklama, DosyaAdiGizle, GenislikYuzde
        """
        # ✅ ZORUNLU: sqlcipher param bağlama hatasını engeller
        pid = self.project_id
        try:
            pid = int(pid)
        except Exception:
            # yine de güvenli fallback
            pid = str(pid)

        bolum = str(bolum)

        with DB() as conn:
            rows = conn.execute(
                "SELECT id, DosyaAdi, DosyaYolu, Aciklama, "
                "COALESCE(DosyaAdiGizle,0) as DosyaAdiGizle, "
                "COALESCE(GenislikYuzde,100) as GenislikYuzde "
                "FROM rapor_meta_ekler "
                "WHERE ProjeID=? AND Bolum=? "
                "ORDER BY id ASC",
                (pid, bolum),
            ).fetchall()
        return rows

    def _fetch_meta(self):
        pid = self.project_id
        try:
            pid = int(pid)
        except Exception:
            pid = str(pid)
        with DB() as conn:
            return conn.execute("""
                SELECT ProjeID AS id,
                       GorevlendirmeMetni,
                       DosyaHakkindaMetni,
                       GenelBilgilendirmeMetni,
                       DegerlendirmeMetni,
                       SonucMetni,
                       MarginTopMm,
                       MarginRightMm,
                       MarginBottomMm,
                       MarginLeftMm
                FROM rapor_meta WHERE ProjeID=?
            """, (pid,)).fetchone()

    def _fetch_blocks(self):
        with DB() as conn:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]

            if "id" in cols:
                id_expr = "id"
            else:
                id_expr = "rowid AS id"

            gen_expr  = "GenislikYuzde" if "GenislikYuzde" in cols else "100 AS GenislikYuzde"
            acik_expr = "Aciklama"      if "Aciklama"      in cols else "'' AS Aciklama"
            html_expr = "HtmlIcerik"    if "HtmlIcerik"    in cols else "NULL AS HtmlIcerik"
            base_expr = "BaseHtmlIcerik" if "BaseHtmlIcerik" in cols else "NULL AS BaseHtmlIcerik"
            hid_expr  = "HiddenColsJson" if "HiddenColsJson" in cols else "NULL AS HiddenColsJson"
            fmt_expr  = "FmtJson"        if "FmtJson"        in cols else "NULL AS FmtJson"
            img_expr  = "ImagePath"     if "ImagePath"     in cols else "NULL AS ImagePath"

            # [DÜZELTME] Icerik sütununu da çekiyoruz (Eski kayıtlar veya fallback için)
            raw_icerik_expr = "Icerik" if "Icerik" in cols else "NULL AS Icerik"

            return conn.execute(f"""
                SELECT {id_expr},
                       Sira,
                       Baslik,
                       Tur,
                       {gen_expr},
                       {acik_expr},
                       {html_expr},
                       {base_expr},
                       {hid_expr},
                       {fmt_expr},
                       {img_expr},
                       {raw_icerik_expr}
                FROM rapor_taslagi
                WHERE ProjeID=?
                ORDER BY Sira ASC, id ASC
            """, (self.project_id,)).fetchall()

    def _fetch_project_info(self):
        """
        Proje tablosundan:
        birim, dosya no tipi, dosya no, suç, suç tarihi, görevlendirme tarihi,
        bilirkişi adı, bilirkişi ünvan/sicil

        ayrıca taraflar:
        [(sifat, adsoyad), ...]
        """
        with DB() as conn:
            p = conn.execute("""
                SELECT talep_eden_birim,
                       dosya_no_tipi,
                       dosya_no,
                       suc_bilgisi,
                       suc_tarihi,
                       gorevlendirme_tarihi,
                       bilirkisi_adi,
                       bilirkisi_unvan_sicil
                FROM projeler WHERE id=?
            """, (self.project_id,)).fetchone()

            parties = conn.execute("""
                SELECT sifat, ad_soyad
                FROM taraflar
                WHERE ProjeID=?
                ORDER BY id ASC
            """, (self.project_id,)).fetchall()

        return p, parties

    def _wrap_as_block(self, title: str, inner_html: str, genislik_yuzde: int = 100, aciklama: str = "") -> str:
        """
        Delil çekmecesi blok temasıyla aynı görünecek şekilde HTML'i sarar.
        DB'ye yazmaz; sadece raporda kullanılır.
        """
        w = int(genislik_yuzde or 100)
        t = html.escape(title or "")
        a = html.escape(aciklama or "")

        out = []
        out.append(f"<div class='block' style='width:{w}% !important;'>")  # ✅ genişlik kesin uygula
        out.append(f"<div class='block-title'>{t}</div>")
        out.append("<div class='block-inner'>")
        out.append(inner_html or "")
        out.append("</div>")  # block-inner
        if aciklama:
            out.append(f"<div class='block-desc'>Delil Açıklaması: {a}</div>")
        out.append("</div>")
        return "\n".join(out)

    # -------------------------
    # Defaults
    # -------------------------
    def build_default_gorevlendirme(self) -> str:
        """
        Yeni şablon:
        ("Mahkeme Adı" "Dosya No" "Dosya Tipi" sayılı dosyasından "Talep Tarihi"
         tarihli yazısı ile talep edilen, talebe istinaden BTK tarafından sunulan
         HTS kayıtlarını incelemek ve ..... bilirkişi raporunu bildirmek üzere
         bilirkişi olarak tayin edildim.
         İş bu raporla yukarıda belirtilen talep gereği incelemeler yapılmış olup;
         gerekli açıklamalar ilgili bölümlerdedir.)

        Talep Tarihi kullanıcı girişi olacak -> placeholder bırakıyoruz.
        Diğer alanlar projeler tablosundan geliyor.
        """
        p, _ = self._fetch_project_info()
        birim, tip, dosya_no, suc, suc_t, gorev_t, bilirkisi_adi, bilirkisi_unvan = (
            p or ("","","","","","","","")
        )

        dosya_tip_ve_no = f"{dosya_no} {tip}".strip()
        mahkeme_adi = (birim or "").strip()

        talep_tarihi_placeholder = "../../20.."

        return (
            f'{mahkeme_adi} {dosya_tip_ve_no} sayılı dosyasından '
            f'{talep_tarihi_placeholder} tarihli yazısı ile talep edilen, '
            f'talebe istinaden Bilgi Teknolojileri ve İletişim Kurumu tarafından sunulan '
            f'HTS kayıtlarını incelemek ve ilgili hususlarda bilirkişi raporunu bildirmek üzere '
            f'bilirkişi olarak tayin edildim.'
            f'<br><br>'
            f'<p>İş bu raporla yukarıda belirtilen talep gereği incelemeler yapılmış olup; '
            f'gerekli açıklamalar ilgili bölümlerdedir.</p>'
        )

    def build_auto_hts_dosya_abone_details(self) -> str:
        """
        HTS DOSYA VE ABONE BİLGİLERİ altında,
        delil çekmecesi temasıyla aynı blok görünümünde otomatik detay basar.
        - Daha kompakt tablo
        - MD5 / SHA-256 değerleri italik
        - GSM başlığı + bloklar aynı sayfada kalmaya zorlanır
        """
        pid = self.project_id

        label_map = {
            "GSMNo": "Numara",
            "Numara": "Numara",
            "Durum": "Durum",
            "Ad": "Ad",
            "Soyad": "Soyad",
            "Adres": "Adres",
            "DogumTarihi": "Doğum Tarihi",
            "DogumYeri": "Doğum Yeri",
            "Ilce": "İlçe",
            "Il": "İl",
            "TCKimlikNo": "TC Kimlik No",
            "AnneAdi": "Anne Adı",
            "BabaAdi": "Baba Adı",
            "AboneSorguAraligi": "Abone Sorgu Aralığı",
            "AboneBaslangic": "Abone Başlangıç",
            "AboneBitis": "Abone Bitiş",
            "Operator": "Operatör",
            "dogum_tarihi": "Doğum Tarihi",
            "dogum_yeri": "Doğum Yeri",
            "tc_kimlik_no": "TC Kimlik No",
            "anne_adi": "Anne Adı",
            "baba_adi": "Baba Adı",
            "abone_sorgu_araligi": "Abone Sorgu Aralığı",
            "abone_baslangic": "Abone Başlangıç",
            "abone_bitis": "Abone Bitiş",
            "operator": "Operatör",
            "ilce": "İlçe",
            "il": "İl",
            "ad": "Ad",
            "soyad": "Soyad",
            "adres": "Adres",
            "durum": "Durum",
        }
        abone_skip = {
            "id", "ID", "projeid", "ProjeID", "gsmno", "GSMNo",
            "SiraNo", "SIRA_NO", "sira_no",
            # EKLE:
            "Rol", "rol",
            "DosyaAdi", "dosyaadi", "dosya_adi", "DosyaAdı"
        }

        dosya_label = {
            "DosyaAdi": "Dosya Adı",
            "DosyaBoyutu": "Boyut",
            "TalepEdenMakam": "Talep Eden Makam",
            "SorguBaslangic": "Sorgu Başlangıç",
            "SorguBitis": "Sorgu Bitiş",
            "Tespit": "Tespit",
            "YuklemeTarihi": "Yükleme",
            "MD5": "MD5",
            "SHA256": "SHA-256",
        }

        def meta_table(rows):
            """
            rows: [(label, value), ...]
            MD5/SHA değerlerini italik basar.
            Font ve punto CSS'ten (table_head/table_body) gelir.
            """
            out = []
            out.append("<table class='meta-table' style='table-layout:fixed; width:100%;'>")
            out.append("<colgroup><col style='width:30%;'><col style='width:70%;'></colgroup>")
            out.append("<tr><th>ÖZELLİK</th><th>DEĞER</th></tr>")
            for k, v in rows:
                key_txt = html.escape(str(k))
                val_txt = html.escape(str(v or ""))

                if str(k).upper() in ("MD5", "SHA256", "SHA-256", "SHA_256"):
                    val_txt = f"<i>{val_txt}</i>"

                out.append(
                    "<tr>"
                    f"<td class='key'>{key_txt}</td>"
                    f"<td>{val_txt}</td>"
                    "</tr>"
                )
            out.append("</table>")
            return "\n".join(out)

        html_parts = []

        with DB() as conn:
            gsms = []
            try:
                rows = conn.execute(
                    "SELECT DISTINCT GSMNo FROM hts_dosyalari WHERE ProjeID=? ORDER BY GSMNo",
                    (pid,)
                ).fetchall()
                gsms += [r[0] for r in rows if r and r[0]]
            except Exception:
                pass
            try:
                rows = conn.execute(
                    "SELECT DISTINCT GSMNo FROM hts_ozet WHERE ProjeID=? ORDER BY GSMNo",
                    (pid,)
                ).fetchall()
                gsms += [r[0] for r in rows if r and r[0]]
            except Exception:
                pass
            seen = set()
            gsms = [g for g in gsms if not (g in seen or seen.add(g))]

            dosya_cols = []
            try:
                dosya_cols = [r[1] for r in conn.execute("PRAGMA table_info(hts_dosyalari)").fetchall()]
            except Exception:
                dosya_cols = []

            def col_exists(c): return c in dosya_cols

            select_cols = ["Rol", "DosyaAdi", "DosyaBoyutu", "TalepEdenMakam", "SorguBaslangic", "SorguBitis"]
            if col_exists("Tespit"): select_cols.append("Tespit")
            if col_exists("YuklemeTarihi"): select_cols.append("YuklemeTarihi")
            if col_exists("MD5"): select_cols.append("MD5")
            if col_exists("SHA256"): select_cols.append("SHA256")

            for gsm in gsms:
                gsm_esc = html.escape(gsm)

                section_buf = []
                section_buf.append(
                    f"<h3 style='margin:0 0 2mm 0; text-decoration:underline; text-underline-offset:0.8mm; font-size:13pt; font-weight:700;'> "
                    f"GSM Hattı: {gsm_esc}</h3>"
                )

                file_rows = []
                try:
                    col_sql = ", ".join(select_cols)
                    file_rows = conn.execute(
                        f"SELECT {col_sql} FROM hts_dosyalari "
                        "WHERE ProjeID=? AND GSMNo=? ORDER BY id",
                        (pid, gsm)
                    ).fetchall()
                except Exception:
                    file_rows = []

                if file_rows:
                    grouped = {}
                    for row in file_rows:
                        rol = (row[0] or "").strip().upper() or "DIGER"
                        grouped.setdefault(rol, []).append(row)

                    rol_order = ["HEDEF", "KARSI", "KARŞI", "DIGER"]
                    def rk(r):
                        try: return rol_order.index(r)
                        except ValueError: return 99

                    for rol in sorted(grouped.keys(), key=rk):
                        rol_title = "HEDEF HTS DOSYASI" if rol == "HEDEF" else \
                                    "KARŞI HTS DOSYASI" if rol in ("KARSI", "KARŞI") else \
                                    f"{rol} HTS DOSYASI"

                        for idx_f, row in enumerate(grouped[rol], start=1):
                            kv_rows = []
                            for i, col in enumerate(select_cols[1:], start=1):
                                val = row[i] if i < len(row) else ""
                                if col.lower() == "dosyaboyutu" and val not in (None, ""):
                                    try:
                                        b = int(val)
                                        if b >= 1024*1024:
                                            val = f"{round(b/(1024*1024),2)} MB"
                                        elif b >= 1024:
                                            val = f"{round(b/1024,2)} KB"
                                        else:
                                            val = f"{b} B"
                                    except Exception:
                                        pass
                                kv_rows.append((dosya_label.get(col, col), val))

                            inner = meta_table(kv_rows)
                            title = rol_title if len(grouped[rol]) == 1 else f"{rol_title} ({idx_f})"
                            section_buf.append(self._wrap_as_block(title, inner, 100))
                else:
                    section_buf.append(self._wrap_as_block(
                        "Dosya Bilgileri",
                        "<p><i>Bu GSM hattı için kayıtlı dosya bilgisi bulunamadı.</i></p>"
                    ))

                abone_dict = {}
                try:
                    abone_cols = [r[1] for r in conn.execute("PRAGMA table_info(hts_abone)").fetchall()]
                    row = conn.execute(
                        "SELECT * FROM hts_abone WHERE ProjeID=? AND GSMNo=? LIMIT 1",
                        (pid, gsm)
                    ).fetchone()
                    if abone_cols and row:
                        abone_dict = {abone_cols[i]: row[i] for i in range(min(len(abone_cols), len(row)))}
                except Exception:
                    abone_dict = {}

                if abone_dict:
                    kv_rows = []
                    for k, v in abone_dict.items():
                        if (k or "") in abone_skip:
                            continue
                        label = label_map.get(k, k)
                        kv_rows.append((label, v))
                    inner = meta_table(kv_rows) if kv_rows else "<p><i>Abone bilgisi bulunamadı.</i></p>"
                    section_buf.append(self._wrap_as_block("ABONE BİLGİLERİ", inner, 100))
                else:
                    section_buf.append(self._wrap_as_block(
                        "ABONE BİLGİLERİ",
                        "<p><i>Bu GSM hattı için abone bilgisi bulunamadı.</i></p>"
                    ))

                html_parts.append("<div class='gsm-section'>")
                html_parts.append("\n".join(section_buf))
                html_parts.append("</div><br>")

        return "\n".join(html_parts)

    def build_default_genel_bilgi(self) -> str:
        """
        Rapor Merkezi "Genel Bilgilendirme" metni.
        HEDEF/KARSI ayrımı: ham kayıtlarda Rol + DosyaAdi tutuluyorsa doğru sayım yapılır.
        """
        pid = self.project_id
        all_totals: list[int] = []

        def fmt_int(v):
            try:
                return f"{int(v):,}".replace(",", ".")
            except Exception:
                return str(v or "0")

        def _table_count(cur, table: str, pid_: int, gsm_: str, rol_: str, dosya_adi_: str) -> int:
            """
            Önce Rol+DosyaAdi ile dener.
            Eski DB'lerde Rol/DosyaAdi kolonu yoksa fallback: sadece ProjeID+GSMNo sayar (ayırt edemez).
            """
            # 1) Rol + DosyaAdi
            try:
                row = cur.execute(
                    f"SELECT COUNT(*) FROM {table} WHERE ProjeID=? AND GSMNo=? AND Rol=? AND DosyaAdi=?",
                    (pid_, gsm_, rol_, dosya_adi_)
                ).fetchone()
                return int(row[0] or 0) if row else 0
            except Exception:
                pass

            # 2) Rol
            try:
                row = cur.execute(
                    f"SELECT COUNT(*) FROM {table} WHERE ProjeID=? AND GSMNo=? AND Rol=?",
                    (pid_, gsm_, rol_)
                ).fetchone()
                return int(row[0] or 0) if row else 0
            except Exception:
                pass

            # 3) Fallback (ayırt edemez)
            try:
                row = cur.execute(
                    f"SELECT COUNT(*) FROM {table} WHERE ProjeID=? AND GSMNo=?",
                    (pid_, gsm_)
                ).fetchone()
                return int(row[0] or 0) if row else 0
            except Exception:
                return 0

        def line(css_class: str, indent_level: int, bullet: str, html_text: str) -> str:
            """
            Qt tarafı mm'yi stabil uygulamadığı için px kullanıyoruz.
            indent_level eski kodda mm gibi düşünülüyordu.
            1mm ≈ 3.78px -> pratikte 4px çarpanı yeterli ve stabil.
            """
            b = html.escape(str(bullet or ""))
            px = int(round(int(indent_level) * 4))  # mm -> px yaklaşık dönüşüm

            return (
                f"<div class='gb-line {html.escape(css_class)}' style='margin-left:{px}px;'>"
                f"<span class='gb-bullet'>{b}</span>"
                f"<span class='gb-text'>{html_text}</span>"
                f"</div>"
            )

        with DB() as conn:
            cur = conn.cursor()

            gsm_rows = cur.execute(
                "SELECT DISTINCT GSMNo FROM hts_dosyalari WHERE ProjeID=? ORDER BY GSMNo",
                (pid,)
            ).fetchall()
            gsms = [r[0] for r in gsm_rows if r and r[0]]

            out = []
            out.append("<div class='gb-block'>")

            out.append(line("gb-intro", 0, "", "Bilgi Teknolojileri ve İletişim Kurumu tarafından gönderilen yazı ekinde,"))
            out.append("<br>")

            for gsm in gsms:
                file_rows = cur.execute(
                    "SELECT Rol, DosyaAdi FROM hts_dosyalari WHERE ProjeID=? AND GSMNo=?",
                    (pid, gsm)
                ).fetchall()

                hedef_files = [r[1] for r in file_rows if r and (r[0] == "HEDEF") and r[1]]
                karsi_files = [r[1] for r in file_rows if r and (r[0] == "KARSI") and r[1]]

                # (sende tek dosya varsayımı vardı; aynen korudum)
                hedef_file = hedef_files[0] if hedef_files else None
                karsi_file = karsi_files[0] if karsi_files else None

                # GSM satırı
                out.append(line(
                    "gb-gsm", 0, "•",
                    f'<b>"{html.escape(gsm)}"</b> GSM numaralı hatta ait olduğu belirtilen,'
                ))

                def get_counts(rol, dosya):
                    if not dosya:
                        return None
                    gsm_c = _table_count(cur, "hts_gsm", pid, gsm, rol, dosya)
                    sms_c = _table_count(cur, "hts_sms", pid, gsm, rol, dosya)
                    gprs_c = _table_count(cur, "hts_gprs", pid, gsm, rol, dosya)
                    wap_c = _table_count(cur, "hts_wap", pid, gsm, rol, dosya)
                    sth_c = _table_count(cur, "hts_sth", pid, gsm, rol, dosya)
                    ulus_c = _table_count(cur, "hts_uluslararasi", pid, gsm, rol, dosya)
                    sabit_c = _table_count(cur, "hts_sabit", pid, gsm, rol, dosya)

                    internet_c = int(gprs_c or 0) + int(wap_c or 0)
                    total = int(gsm_c or 0) + int(sms_c or 0) + int(internet_c or 0) + int(sth_c or 0) + int(ulus_c or 0) + int(sabit_c or 0)

                    return {
                        "gsm": gsm_c,
                        "sms": sms_c,
                        "internet": internet_c,
                        "sth": sth_c,
                        "ulus": ulus_c,
                        "sabit": sabit_c,
                        "total": total
                    }

                hedef_counts = get_counts("HEDEF", hedef_file) if hedef_file else None
                karsi_counts = get_counts("KARSI", karsi_file) if karsi_file else None

                hedef_total = int(hedef_counts["total"]) if hedef_counts else 0
                karsi_total = int(karsi_counts["total"]) if karsi_counts else 0

                if hedef_file and karsi_file and hedef_total == 0 and karsi_total == 0:
                    out.append(line(
                        "gb-file", 6, "○",
                        f'<b>"{html.escape(hedef_file)}"</b>, <b>"{html.escape(karsi_file)}"</b> isimli dosyalar içerisinde kayıt bulunamamıştır.'
                    ))
                    out.append("<br>")
                    continue

                def emit_file_block(file_name: str, counts: dict | None):
                    if not file_name:
                        return
                    total = int((counts or {}).get("total", 0) or 0)
                    all_totals.append(total)

                    if total == 0:
                        out.append(line(
                            "gb-file", 12, "○",
                            f'<b>"{html.escape(file_name)}"</b> isimli dosya içerisinde kayıt bulunamamıştır.'
                        ))
                        return

                    # dosya satırı
                    out.append(line(
                        "gb-file", 6, "○",
                        f'<b>"{html.escape(file_name)}"</b> isimli dosya içerisinde,'
                    ))

                    def add_stat(key, label):
                        val = int((counts or {}).get(key, 0) or 0)
                        if val > 0:
                            out.append(line("gb-stat", 24, "▪", f'"{fmt_int(val)}" adet {label}'))

                    add_stat("gsm", "GSM görüşme")
                    add_stat("sms", "SMS gönderme")
                    add_stat("internet", "İnternet bağlantısı")
                    add_stat("sth", "STH görüşmesi")
                    add_stat("ulus", "Uluslararası görüşme")
                    add_stat("sabit", "Sabit Hat görüşmesi")

                    out.append(line(
                        "gb-total", 0, "",
                        f'olmak üzere toplamda "<b>{fmt_int(total)}</b>" kayıt olduğu görülmüştür.'
                    ))

                # HEDEF
                emit_file_block(hedef_file, hedef_counts)

                # KARSI (varsa)
                emit_file_block(karsi_file, karsi_counts)

                out.append("<br>")

            grand_total = sum(all_totals) if all_totals else 0
            out.append(
                line(
                    "gb-grand-total",
                    0,
                    "",
                    f'Bu haliyle toplam "<b>{fmt_int(grand_total)}</b>" kayıt incelenerek rapora devam olunmuştur.'
                )
            )
            out.append("<br>")

            out.append("</div>")
            return "\n".join(out)

    def _file_uri(self, path: str) -> str:
        """Windows/local path -> file:/// URI"""
        if not path:
            return ""
        p = path.replace("\\", "/")
        if re.match(r"^[A-Za-z]:/", p):
            return "file:///" + p
        return p
    def _postprocess_report_tables(self, html_content: str, title: str) -> str:
        """
        - Chromium PDF'de tablo bölünmelerini azaltmak için THEAD'i garanti eder
        - Delil başlığına göre istenmeyen sütunları kaldırır
        - Sütunları daraltan colgroup/width enjekte etmez (auto layout)
        - Sadece çok uzun tek-parça değerlerde (IMEI/ID/IP/URL) kırılmayı açar
        """

        if title and any(k in title.upper() for k in [
            "HEDEF HTS",
            "KARŞI HTS",
            "ABONE BİLGİLERİ"
        ]):
            try:
                soup = BeautifulSoup(html, "html.parser")

                table = soup.find("table")
                if table:
                    thead = table.find("thead")
                    tbody = table.find("tbody")

                    if thead and tbody:
                        ths = thead.find_all("th")
                        oz_idx = None
                        for i, th in enumerate(ths):
                            if "ÖZELLİK" in th.get_text().upper():
                                oz_idx = i
                                break

                        if oz_idx is not None:
                            for tr in tbody.find_all("tr"):
                                tds = tr.find_all("td")
                                if len(tds) > oz_idx:
                                    tds[oz_idx]["style"] = (
                                        tds[oz_idx].get("style", "")
                                        + " font-weight:700 !important;"
                                    )

                html_content = str(soup)
            except Exception:
                pass
        try:
            from bs4 import BeautifulSoup
        except Exception:
            return html_content

        def tr_norm(s: str) -> str:
            s = (s or "").strip()
            s = s.casefold()
            # Türkçe karakterleri sadeleştir (eşleşme kaçmasın)
            s = (s.replace("ı", "i").replace("İ", "i")
                   .replace("ş", "s").replace("ğ", "g").replace("ç", "c")
                   .replace("ö", "o").replace("ü", "u"))
            s = " ".join(s.split())
            return s

        soup = BeautifulSoup(html_content, "html.parser")

        t = tr_norm(title or "")

        hide_cols = set()
        # GSM: "SIRA NO" görünmesin
        if "gsm" in t:
            hide_cols |= {tr_norm("SIRA NO")}

        # Internet / MutualContactsMatchDialog: belirtilen kolonlar görünmesin
        if ("internet" in t) or ("mutualcontactsmatchdialog" in t) or ("mutual contacts" in t):
            hide_cols |= {
                tr_norm("TUR"),
                tr_norm("SIRA NO"),
                tr_norm("HEDEF IP"),
                tr_norm("GÖNDERME"),
                tr_norm("İNDİRME"),
            }

        for table in soup.find_all("table"):
            # varsa önceki colgroup'ları sök (daraltmayı bitir)
            cg = table.find("colgroup")
            if cg:
                cg.decompose()

            # sınıf ekle
            existing_cls = table.get("class", [])
            if "report-table" not in existing_cls:
                table["class"] = existing_cls + ["report-table"]

            # THEAD yoksa: ilk tr'deki th'leri THEAD'e al
            thead = table.find("thead")
            if not thead:
                first_tr = table.find("tr")
                if first_tr and first_tr.find("th"):
                    thead = soup.new_tag("thead")
                    tr_new = soup.new_tag("tr")
                    for th in first_tr.find_all("th", recursive=False):
                        tr_new.append(th.extract())
                    thead.append(tr_new)
                    table.insert(0, thead)

            # Başlık hücrelerinden index bul
            header_tr = table.find("thead").find("tr") if table.find("thead") else None
            headers = header_tr.find_all(["th", "td"], recursive=False) if header_tr else []
            header_names = [tr_norm(h.get_text(" ", strip=True)) for h in headers]

            drop_idx = set()
            if hide_cols and header_names:
                for i, hn in enumerate(header_names):
                    if not hn:  # "" ise
                        drop_idx.add(i)

            # İstenen kolonları kaldır
            if drop_idx:
                # header
                if header_tr:
                    for i in sorted(drop_idx, reverse=True):
                        cells = header_tr.find_all(["th", "td"], recursive=False)
                        if i < len(cells):
                            cells[i].decompose()

                # body rows
                for tr in table.find_all("tr"):
                    tds = tr.find_all(["td", "th"], recursive=False)
                    for i in sorted(drop_idx, reverse=True):
                        if i < len(tds):
                            tds[i].decompose()

            # Çok uzun tek parça token kırılabilsin (geri kalan kelimeler bölünmesin)
            for td in table.find_all("td"):
                text = td.get_text(" ", strip=True) or ""
                # en uzun “boşluksuz” parça
                max_token = max((len(x) for x in text.split()), default=0)
                if max_token >= 26:
                    cls = td.get("class", [])
                    if "cell-break-anywhere" not in cls:
                        td["class"] = cls + ["cell-break-anywhere"]

        return str(soup)

    def build_html(self, disabled_sections=None) -> str:
        """
        SQL Ayarlı ve Güvenli HTML Oluşturucu
        - Tüm metinler Segoe UI
        - Görevlendirme / Dosya Hakkında / HTS Dosya+Abone / Değerlendirme / Sonuç: iki yana yaslı
        - Otomatik üretilen tablolar justify içine sokulmaz (HTML bozulmasını engeller)
        - Kapak watermark HTML ile basılmaz, PDF post-process ile basılır
        """
        # ✅ disabled_sections her durumda set olsun (işlev aynı)
        disabled_sections = set(disabled_sections or [])
        self._ensure_meta_row()

        s = StyleConfig.load()
        if not isinstance(s, dict):
            s = StyleConfig.DEFAULTS.copy()

        margins = s.get("margins", StyleConfig.DEFAULTS["margins"])

        def css_gen(key: str) -> str:
            val = s.get(key)
            if not isinstance(val, dict):
                return ""
            css = f"font-size: {val.get('size', 11)}pt !important;"
            if val.get("bold"):
                css += " font-weight: bold !important;"
            if val.get("italic"):
                css += " font-style: italic !important;"
            if val.get("under"):
                css += " text-decoration: underline !important;"
            return css

        body_cfg = s.get('body', {})
        if not isinstance(body_cfg, dict):
            body_cfg = {}

        justify_size = body_cfg.get('size', 12)
        justify_bold = "bold" if body_cfg.get('bold') else "normal"
        justify_italic = "italic" if body_cfg.get('italic') else "normal"
        justify_under = "underline" if body_cfg.get('under') else "none"

        p, parties = self._fetch_project_info()
        meta = self._fetch_meta()
        blocks = self._fetch_blocks()

        (birim, tip, dosya_no, suc, suc_t, gorev_t, bilirkisi_adi, bilirkisi_unvan) = p or ("", "", "", "", "", "", "", "")
        dosya_numarasi = f"{dosya_no} {tip}".strip()

        def mget(idx, default=""):
            # ✅ sınır + tip güvenliği (işlev aynı; sadece patlamayı engeller)
            if not meta or idx >= len(meta):
                val = default
            else:
                val = meta[idx] if meta[idx] is not None else default

            if not isinstance(val, str):
                val = "" if val is None else str(val)

            val = val.replace('align="left"', '').replace("align='left'", '')
            val = val.replace('text-align:left', '').replace('text-align: left', '')
            val = val.replace('text-align:start', '').replace('text-align: start', '')
            val = val.replace('white-space:nowrap', '').replace('white-space: pre', '')
            return val

        gorev_metin = mget(1, self.build_default_gorevlendirme())

        dosya_ekler_html = self._render_meta_ekler_html("dosya_hakkinda", "")
        dosya_metin = (mget(2, "") or "") + (dosya_ekler_html or "")

        genel_metin = mget(3, self.build_default_genel_bilgi())
        deger_metin = mget(4, "")
        sonuc_metin = mget(5, "")

        cover_parties_rows = ""
        if "taraflar" not in disabled_sections and parties:
            grouped = OrderedDict()  # sıralamayı korur

            for s_val, n_val in parties:
                sifat = (s_val or "").strip()
                adsoyad = (n_val or "").strip()
                if not sifat or not adsoyad:
                    continue
                grouped.setdefault(sifat, []).append(adsoyad)

            r_buf = []
            for sifat, names in grouped.items():
                joined = ", ".join(html.escape(x) for x in names)
                r_buf.append(
                    f"<tr>"
                    f"<td class='k'>{html.escape(sifat)}</td>"
                    f"<td class='c'>:</td>"
                    f"<td class='v party-v'>{joined}</td>"
                    f"</tr>"
                )
            cover_parties_rows = "\n".join(r_buf)

        mt = margins.get("top", 25)
        mr = margins.get("right", 20)
        mb = margins.get("bottom", 20)
        ml = margins.get("left", 25)

        html_out = [f"""<!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
    @page {{ size: A4; margin: {mt}mm {mr}mm {mb}mm {ml}mm; }}
    * {{ box-sizing: border-box; }}

    body {{
        font-family: "Segoe UI", "DejaVu Sans", Arial, sans-serif;
        color: #111;
        {css_gen('body')}
        width: 100% !important;
        margin: 0 !important;
        padding: 0 !important;
    }}

    h2 {{ margin: 6mm 0 3mm 0; {css_gen('section_title')} }}
    .center {{ text-align:center; }}

    .page-break{{
      page-break-after: always;
      height: 0 !important;
      line-height: 0 !important;
      margin: 0 !important;
      padding: 0 !important;
      border: 0 !important;
    }}

    .kv-table {{ width: 100%; border-collapse: collapse; margin-top: 4mm; }}
    .kv-table td {{ padding: 1.2mm 0; vertical-align: top; {css_gen('kv_values')} }}
    .kv-table .k {{ width: 30%; text-align: left; padding-right: 2mm; {css_gen('kv_keys')} }}
    .kv-table .c {{ width: 2%; text-align: left; font-weight:bold; }}
    .kv-table .v {{ width: 68%; text-align: left; }}

    table {{
        width: 100% !important;
        border-collapse: collapse !important;
        table-layout: auto !important;
        border-spacing: 0;
    }}

    .party-v {{
      text-align: justify;
      text-align-last: left;
      white-space: normal;
      word-break: break-word;
    }}

    /* -------------------------------------------------------
       ✅ TABLO STANDARDI
       - HTS/Abone tabloları (.meta-table)
       - Delil tabloları (.block-inner table)
       Hepsi StyleConfig table_head + table_body ile aynı olsun
    ------------------------------------------------------- */

    /* Genel tablo hücreleri */
    .meta-table th{{
        white-space: nowrap !important;
    }}
    .block-inner th {{
        background: #e9eef3 !important;
        color: #111 !important;
        text-align: center !important;
        border: 1px solid #555 !important;
        white-space: normal !important;
        overflow-wrap: anywhere !important;
        word-break: break-word !important;
        padding: 4px !important;
        {css_gen('table_head')}
    }}

    .meta-table td,
    .block-inner td {{
        border: 1px solid #555 !important;
        padding: 3px !important;
        color: #111 !important;
        text-align: left !important;
        white-space: normal !important;
        word-break: normal !important;
        overflow-wrap: normal !important;
        hyphens: none !important;
        {css_gen('table_body')}
    }}

    .meta-table td.key {{ font-weight: 700 !important; }}

    /* Delil içi tablo konteyneri: hardcode font-size KALDIRILDI */
    .block-inner table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed !important;
      border: 1px solid #555;
      box-sizing: border-box;
    }}

    /* Son kolon taşmasın */
    .block-inner th:last-child,
    .block-inner td:last-child{{
        white-space: normal !important;
        overflow-wrap: anywhere !important;
        word-break: break-word !important;
    }}

    .block-inner td.cell-break-anywhere {{
        overflow-wrap: anywhere !important;
    }}

    @media print {{
      thead {{ display: table-header-group; break-inside: avoid !important; }}
      tfoot {{ display: table-footer-group; break-inside: avoid !important; }}
      tr {{ break-inside: avoid !important; page-break-inside: avoid !important; }}
      thead tr {{ break-after: avoid !important; page-break-after: avoid !important; }}
      tbody tr:first-child {{ break-before: avoid !important; page-break-before: avoid !important; }}
    }}

    /* -------------------------------------------------------
       ✅ GÖVDE METNİ STANDARDI
       QTextEdit’in inline font-size/span stilleri bile ezilsin:
       - Sabit metin + kullanıcı metni = body ayarları
    ------------------------------------------------------- */
    .justify {{
        width: 100%;
        text-align: justify !important;
        white-space: normal !important;
    }}

    /* p/li ana kural */
    .justify p,
    .justify li {{
        font-size: {justify_size}pt !important;
        font-weight: {justify_bold} !important;
        font-style: {justify_italic} !important;
        text-decoration: {justify_under} !important;
        color: #111 !important;

        text-align: justify !important;
        white-space: normal !important;

        line-height: 1.5 !important;   /* ✅ 1,5 SATIR ARALIĞI */

        text-indent: 1.25cm !important;

        margin-top: 0 !important;
        margin-bottom: 3mm !important; /* satır aralığına uygun */
    }}

    /* QTextEdit’in ürettiği span/div/font gibi inline stilleri de gövdeye kilitle */
    .justify span,
    .justify div,
    .justify font {{
        font-size: {justify_size}pt !important;
        color: inherit !important;
    }}

    /* QTextEdit inline stilleri KORUNSUN */
    .justify b,
    .justify strong {{
        font-weight: bold !important;
    }}

    .justify i,
    .justify em {{
        font-style: italic !important;
    }}

    .justify u {{
        text-decoration: underline !important;
    }}

    .justify b, .justify strong {{ font-weight: bold !important; }}
    .justify i, .justify em {{ font-style: italic !important; }}
    .justify u {{ text-decoration: underline !important; }}

    /* --- DELİL BLOKLARI --- */
    .block {{
        margin: 4mm auto;
        text-indent: 0 !important;
        page-break-inside: avoid;
        break-inside: avoid;
    }}
    .block-splittable {{
        page-break-inside: auto !important;
        break-inside: auto !important;
    }}
    .block-title {{
        margin-bottom: 1.5mm;
        color: #2c3e50;
        display: block;
        width: 100%;
        {css_gen('evidence_title')}
        text-align: left !important;
        text-indent: 0 !important;
        page-break-after: avoid;
    }}
    .block-desc {{
        margin-top: 1.2mm;
        text-align: justify;
        {css_gen('evidence_desc')}
        text-indent: 0 !important;
        white-space: pre-wrap !important;
    }}
    .block-desc-full{{
        width: 100% !important;
        margin: 1.6mm 0 0 0 !important;
    }}
    .block img {{
        max-width: 100% !important;
        height: auto !important;
        display: block !important;
        margin: 0 auto !important;
    }}

    .map-wrap{{ width:100%; text-align:center; }}
    .block img.map-img,
    .block img.geo-map-img {{
        width: 100% !important;
        height: auto !important;
        max-height: 150mm !important;
        object-fit: contain !important;
        display: block !important;
        margin: 0 auto !important;
        border: none !important;
        border-radius: 0 !important;
        background: transparent !important;
        box-shadow: none !important;
    }}

    h2 {{
        page-break-after: avoid;
        break-after: avoid-page;
        page-break-inside: avoid;
        break-inside: avoid;
    }}
    h2 + div, h2 + p, h2 + table, h3 + div, h3 + p, h3 + table {{
        page-break-before: avoid;
        break-before: avoid-page;
    }}
    h2, h3 {{
        page-break-after: avoid !important;
        break-after: avoid-page !important;
    }}

    .cover {{
        position: relative;
        height: calc(297mm - {mt}mm - {mb}mm);
        background: #fff;
        overflow: hidden;
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        padding-top: 0mm;
    }}
    .cover-content {{
        width: 100%;
        max-width: 100%;
        padding: 0 6mm;
        text-align: center;
    }}
    .cover-kv {{
        width: 100%;
        margin: 0 auto;
        table-layout: fixed;
        border-collapse: collapse;
    }}
    .cover-kv td {{
        padding: 1.2mm 0 !important;
        vertical-align: middle !important;
    }}
    .cover-kv .k {{
        width: 52mm !important;
        white-space: nowrap !important;
        font-weight: 700 !important;
    }}
    .cover-title {{
        {css_gen('main_title')}
        text-align: center;
        margin-bottom: 8mm;
        text-decoration: underline;
        text-underline-offset: 2mm;
    }}
    </style>
    </head>
    <body>
    """]

        # --- KAPAK ---
        html_out.append(f"""
        <div class="cover">
          <div class="cover-content">
            <div class="cover-title">BİLİRKİŞİ RAPORU</div>

            <table class="kv-table cover-kv">
              <tr><td class='k'>Birim Adı</td><td class='c'>:</td><td class='v'>{html.escape((birim or '').strip())}</td></tr>
              <tr><td class='k'>Dosya Numarası</td><td class='c'>:</td><td class='v'>{html.escape((str(dosya_numarasi) or '').strip())}</td></tr>
              <tr><td class='k'>Bilirkişi</td><td class='c'>:</td><td class='v'>{html.escape((bilirkisi_adi or '').strip())}</td></tr>
              <tr><td class='k'>Ünvanı ve Sicili</td><td class='c'>:</td><td class='v'>{html.escape((bilirkisi_unvan or '').strip())}</td></tr>
              <tr><td class='k'>Görevlendirme Tarihi</td><td class='c'>:</td><td class='v'>{html.escape((gorev_t or '').strip())}</td></tr>

              <tr><td colspan="3" style="padding:4mm 0 0 0;"></td></tr>

              <tr>
                <td colspan="3" style="text-decoration: underline; text-underline-offset: 2mm; font-size: 18pt; font-weight: 800; padding:0 !important; margin:0 !important;">
                  TARAF BİLGİLERİ
                </td>
              </tr>

              {cover_parties_rows if cover_parties_rows else ""}

              <tr><td colspan="3" style="padding:2mm 0 0 0;"></td></tr>

              <tr><td class='k'>Suç</td><td class='c'>:</td><td class='v'>{html.escape((suc or '').strip())}</td></tr>
              <tr><td class='k'>Suç Tarihi</td><td class='c'>:</td><td class='v'>{html.escape((suc_t or '').strip())}</td></tr>
            </table>

          </div>
        </div>
        <div class="page-break"></div>
        """)

        section_no = 1

        if "gorev" not in disabled_sections:
            html_out.append(f"<h2>{section_no}. GÖREVLENDİRME</h2><div class='justify'>{gorev_metin}</div>")
            section_no += 1

        if "dosya" not in disabled_sections:
            html_out.append(f"<h2>{section_no}. DOSYA HAKKINDA</h2><div class='justify'>{dosya_metin}</div>")
            section_no += 1

        if "genel" not in disabled_sections:
            auto = self.build_auto_hts_dosya_abone_details()
            if isinstance(auto, str) and "GSM Hattı:" in auto:
                auto = re.sub(
                    r"(<h3[^>]*>\s*GSM\s*Hattı:\s*[^<]+</h3>)",
                    r"<div class='keep-with-next'>\1",
                    auto,
                    count=1
                )
                auto = auto.replace("<div class='block", "</div><div class='block", 1)

            html_out.append(f"<h2>{section_no}. HTS DOSYA VE ABONE BİLGİLERİ</h2>")
            html_out.append(f"<div class='justify'>{genel_metin}</div>")
            html_out.append(auto)
            section_no += 1

        # ✅ DELİLLER BURAYA ALINDI (4. BÖLÜM)
        if "deliller" not in disabled_sections:
            html_out.append(f"<h2>{section_no}. DELİLLER</h2>")

            for idx, (bid, sira, baslik, tur, gen, acik, htm, base_html, hidden_json, fmt_json, img, raw) in enumerate(blocks or []):
                row_char = chr(97 + idx)
                unique_html_id = f"evidence-{bid}"

                w = int(gen or 100)
                btitle = baslik or tur or "Delil"
                src_probe = (img or raw or "")

                block_cls = "block"
                if tur in ("HTML", "TABLE"):
                    block_cls = "block block-splittable"

                html_out.append(f"<div id='{unique_html_id}' class='{block_cls}'>")
                html_out.append(f"<div class='block-title'>{row_char}. {html.escape(btitle)}</div>")

                # ✅ sadece içerik (resim/tablo/html) width ile hareket edecek
                html_out.append(f"<div class='block-inner' style='width:{w}%; margin:0 auto;'>")

                if tur == "IMAGE":
                    src = img or raw
                    if src:
                        if not src.startswith("data:") and not src.startswith("http"):
                            src = self._file_uri(src)

                        is_event_map = ("event_map_" in (img or "") or "event_map_" in (raw or ""))
                        is_geo_map = ("geo_map_" in (img or "") or "geo_map_" in (raw or ""))
                        is_map = is_event_map or is_geo_map

                        if is_geo_map:
                            cls = "block-img geo-map-img"
                        elif is_event_map:
                            cls = "block-img map-img"
                        else:
                            cls = "block-img"

                        if is_map:
                            html_out.append("<div class='map-wrap'>")
                            html_out.append(f"<img class='{cls}' src='{src}'>")
                            html_out.append("</div>")
                        else:
                            html_out.append(f"<img class='{cls}' src='{src}'>")

                elif tur in ("HTML", "TABLE"):
                    # ✅ her zaman tam tabloyu kaynak al ki hem gizleme hem geri getirme çalışsın
                    content = (base_html or htm or raw)

                    if content:
                        try:
                            # hidden list
                            hidden_list = []
                            if hidden_json:
                                hidden_list = json.loads(hidden_json) if isinstance(hidden_json, str) else (hidden_json or [])

                            # fmt dict
                            fmt = {}
                            if fmt_json:
                                fmt = json.loads(fmt_json) if isinstance(fmt_json, str) else (fmt_json or {})

                            # ✅ sadece TABLE için uygula
                            if tur == "TABLE":
                                if hidden_list:
                                    content = _apply_hidden_cols_to_table_html(content, hidden_list)
                                if isinstance(fmt, dict) and fmt:
                                    content = _apply_fmt_to_table_html(content, fmt)

                        except Exception:
                            pass

                        if hasattr(self, "_postprocess_report_tables"):
                            content = self._postprocess_report_tables(content, btitle)

                        html_out.append(content)

                html_out.append("</div>")  # ✅ block-inner kapanır (açıklama artık etkilenmez)

                # ✅ AÇIKLAMA FULL-WIDTH: resimle birlikte daralıp genişlemez
                if acik:
                    html_out.append(f"<div class='block-desc block-desc-full'>{acik}</div>")

                html_out.append("</div>")  # block

            section_no += 1

        # ✅ DEĞERLENDİRME 5
        if "deg" not in disabled_sections:
            html_out.append(f"<h2>{section_no}. DEĞERLENDİRME</h2><div class='justify'>{deger_metin}</div>")
            section_no += 1

        # ✅ SONUÇ 6
        if "sonuc" not in disabled_sections:
            html_out.append(f"<h2>{section_no}. SONUÇ</h2><div class='justify'>{sonuc_metin}</div>")
            section_no += 1
        html_out.append("</body></html>")
        return "\n".join(html_out)


class PDFExporter:

    @staticmethod
    def export_pdf(
        html_string: str,
        out_path: str,
        margin_top_mm: int = 20,
        margin_right_mm: int = 20,
        margin_bottom_mm: int = 20,
        margin_left_mm: int = 20,
        zoom: float = 0.96,
        footer_font_size: int = 10,
        footer_right_text: str = "",
        # ÇERÇEVE
        draw_frame: bool = True,
        frame_inset_mm: float = 8.0,
        frame_radius_pt: float = 8.0,
        frame_stroke_rgb=(0.90, 0.90, 0.90),
        frame_line_width: float = 1.3,
        # KAPAK WATERMARK
        draw_cover_watermark: bool = True,
        cover_watermark_size_mm: float = 210.0,
        cover_watermark_opacity: float = 0.10,
        cover_watermark_blur_px: float = 2.2,
        # DAMGA (kapak hariç sol üst)
        draw_stamp: bool = True,
        stamp_size_mm: float = 20.0,          # büyüttük (12 -> 14)
        stamp_opacity: float = 0.18,
        stamp_offset_mm: float = 1.5,         # çerçeveye daha yakın (yazıya değmesin diye kontrol bizde)
        **_ignored_kwargs,
    ):
        import os, io, tempfile, shutil

        status_cb = _ignored_kwargs.pop("status_cb", None)      # callable(str)
        progress_cb = _ignored_kwargs.pop("progress_cb", None)  # callable(int) 0-100

        def _status(msg: str):
            try:
                if callable(status_cb):
                    status_cb(str(msg))
            except Exception:
                pass

        def _progress(pct: int):
            try:
                if callable(progress_cb):
                    progress_cb(int(pct))
            except Exception:
                pass

        try:
            out_path = os.path.abspath(out_path)
            out_dir = os.path.dirname(out_path)
            os.makedirs(out_dir, exist_ok=True)

            # PDF açık/kitli mi?
            if os.path.exists(out_path):
                lock_test = out_path + ".~locktest"
                try:
                    os.replace(out_path, lock_test)
                    os.replace(lock_test, out_path)
                except Exception:
                    raise RuntimeError(
                        "PDF dosyası şu anda AÇIK veya KİLİTLİ görünüyor.\n\n"
                        "Lütfen PDF'i kapatın ve tekrar deneyin."
                    )

            app_dir = APP_DIR
            logo_path = os.path.join(app_dir, "assets", "logo.png")
            if not os.path.exists(logo_path):
                logo_path = ""

            tmp_pdf_path = out_path + ".tmp.pdf"

            # ========== 1) CHROMIUM (Playwright) ile PDF render ==========
            _status("PDF oluşturuluyor... (Chromium render)")
            _progress(5)

            try:
                from playwright.sync_api import sync_playwright
            except Exception as e:
                raise RuntimeError(
                    "Chromium PDF motoru için Playwright gerekli.\n"
                    "Kurulum:\n"
                    "  pip install playwright\n"
                    "  playwright install chromium\n\n"
                    f"Detay: {e}"
                )

            with tempfile.TemporaryDirectory() as td:
                html_path = os.path.join(td, "report.html")
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(html_string)
                _progress(15)

                file_url = "file:///" + os.path.abspath(html_path).replace("\\", "/")

                with sync_playwright() as p:
                    browser = p.chromium.launch()
                    page = browser.new_page()
                    page.goto(file_url, wait_until="networkidle")
                    _progress(35)

                    page.pdf(
                        path=tmp_pdf_path,
                        format="A4",
                        print_background=True,
                        scale=float(zoom),
                        margin={
                            "top": f"{int(margin_top_mm)}mm",
                            "right": f"{int(margin_right_mm)}mm",
                            "bottom": f"{int(margin_bottom_mm)}mm",
                            "left": f"{int(margin_left_mm)}mm",
                        },
                    )
                    browser.close()

                _status("PDF düzenleniyor... (overlay/çerçeve/damga)")
                _progress(65)

                # ========== 2) POST-PROCESS OVERLAY ==========
                try:
                    try:
                        from pypdf import PdfReader, PdfWriter
                        _HAS_OVER_PARAM = True
                    except Exception:
                        from PyPDF2 import PdfReader, PdfWriter
                        _HAS_OVER_PARAM = False

                    from reportlab.pdfgen import canvas
                    from reportlab.lib.utils import ImageReader
                    from reportlab.pdfbase import pdfmetrics
                    from reportlab.pdfbase.ttfonts import TTFont

                    mm_to_pt = 2.83464567
                    inset_pt = float(frame_inset_mm) * mm_to_pt
                    wm_size_pt = float(cover_watermark_size_mm) * mm_to_pt
                    stamp_size_pt = float(stamp_size_mm) * mm_to_pt
                    stamp_pad_pt = float(stamp_offset_mm) * mm_to_pt

                    # Footer font: Segoe UI varsa yakala (yoksa Helvetica)
                    footer_font_name = "Helvetica"
                    font_candidates = [
                        os.path.join(app_dir, "assets", "fonts", "segoeui.ttf"),
                        os.path.join(app_dir, "assets", "fonts", "SegoeUI.ttf"),
                        os.path.join(app_dir, "assets", "segoeui.ttf"),
                        os.path.join(app_dir, "assets", "SegoeUI.ttf"),
                        os.path.join(app_dir, "assets", "fonts", "DejaVuSans.ttf"),
                        os.path.join(app_dir, "assets", "DejaVuSans.ttf"),
                    ]
                    for fp in font_candidates:
                        if os.path.exists(fp):
                            try:
                                pdfmetrics.registerFont(TTFont("HTSReportFont", fp))
                                footer_font_name = "HTSReportFont"
                                break
                            except Exception:
                                pass

                    wm_img_path = None
                    if draw_cover_watermark and logo_path:
                        try:
                            from PIL import Image, ImageFilter, ImageEnhance
                            im = Image.open(logo_path).convert("RGBA")
                            gray = im.convert("L")
                            im = Image.merge("RGBA", (gray, gray, gray, im.split()[-1]))
                            im = im.filter(ImageFilter.GaussianBlur(radius=float(cover_watermark_blur_px)))
                            r, g, b, a = im.split()
                            a = ImageEnhance.Brightness(a).enhance(float(cover_watermark_opacity))
                            im = Image.merge("RGBA", (r, g, b, a))
                            wm_img_path = os.path.join(td, "wm_cover.png")
                            im.save(wm_img_path, "PNG")
                        except Exception:
                            wm_img_path = None

                    stamp_img_path = None
                    if draw_stamp and logo_path:
                        try:
                            from PIL import Image, ImageEnhance
                            im = Image.open(logo_path).convert("RGBA")
                            gray = im.convert("L")
                            im = Image.merge("RGBA", (gray, gray, gray, im.split()[-1]))
                            r, g, b, a = im.split()
                            a = ImageEnhance.Brightness(a).enhance(float(stamp_opacity))
                            im = Image.merge("RGBA", (r, g, b, a))
                            stamp_img_path = os.path.join(td, "stamp.png")
                            im.save(stamp_img_path, "PNG")
                        except Exception:
                            stamp_img_path = None

                    reader = PdfReader(tmp_pdf_path)
                    writer = PdfWriter()

                    footer_y = inset_pt * 0.55

                    total_pages = len(reader.pages) if hasattr(reader, "pages") else 0
                    for i, page in enumerate(reader.pages):
                        if total_pages > 0:
                            _progress(65 + int(25 * (i / max(1, total_pages - 1))))

                        w = float(page.mediabox.width)
                        h = float(page.mediabox.height)

                        # 2.1 Kapak watermark (1. sayfa)
                        if draw_cover_watermark and i == 0 and wm_img_path and os.path.exists(wm_img_path):
                            packet = io.BytesIO()
                            c = canvas.Canvas(packet, pagesize=(w, h))
                            x = (w - wm_size_pt) / 2.0
                            y = (h - wm_size_pt) / 2.0
                            c.drawImage(
                                ImageReader(wm_img_path),
                                x, y,
                                width=wm_size_pt, height=wm_size_pt,
                                preserveAspectRatio=True,
                                mask="auto",
                            )
                            c.showPage()
                            c.save()
                            packet.seek(0)
                            ov = PdfReader(packet).pages[0]
                            if _HAS_OVER_PARAM:
                                page.merge_page(ov, over=True)
                            else:
                                page.merge_page(ov)

                        # 2.2 Damga (kapak hariç)
                        if draw_stamp and i >= 1 and stamp_img_path and os.path.exists(stamp_img_path):
                            packet = io.BytesIO()
                            c = canvas.Canvas(packet, pagesize=(w, h))
                            x = inset_pt + stamp_pad_pt
                            y = h - inset_pt - stamp_pad_pt - stamp_size_pt
                            c.drawImage(
                                ImageReader(stamp_img_path),
                                x, y,
                                width=stamp_size_pt, height=stamp_size_pt,
                                preserveAspectRatio=True,
                                mask="auto",
                            )
                            c.showPage()
                            c.save()
                            packet.seek(0)
                            ov = PdfReader(packet).pages[0]
                            if _HAS_OVER_PARAM:
                                page.merge_page(ov, over=True)
                            else:
                                page.merge_page(ov)

                        # 2.3 Footer
                        packet = io.BytesIO()
                        c = canvas.Canvas(packet, pagesize=(w, h))
                        c.setFont(footer_font_name, int(footer_font_size))
                        c.setFillColorRGB(0.60, 0.63, 0.66)

                        pno = str(i + 1)
                        c.drawCentredString(w / 2.0, footer_y, pno)

                        if footer_right_text:
                            c.drawRightString(w - inset_pt, footer_y, str(footer_right_text))

                        c.showPage()
                        c.save()
                        packet.seek(0)
                        ov = PdfReader(packet).pages[0]
                        if _HAS_OVER_PARAM:
                            page.merge_page(ov, over=True)
                        else:
                            page.merge_page(ov)

                        # 2.4 Çerçeve (en son)
                        if draw_frame:
                            packet = io.BytesIO()
                            c = canvas.Canvas(packet, pagesize=(w, h))
                            c.setLineWidth(float(frame_line_width))
                            rr, gg, bb = frame_stroke_rgb
                            c.setStrokeColorRGB(float(rr), float(gg), float(bb))
                            c.roundRect(
                                inset_pt, inset_pt,
                                w - 2 * inset_pt, h - 2 * inset_pt,
                                float(frame_radius_pt),
                                stroke=1, fill=0
                            )
                            c.showPage()
                            c.save()
                            packet.seek(0)
                            ov = PdfReader(packet).pages[0]
                            if _HAS_OVER_PARAM:
                                page.merge_page(ov, over=True)
                            else:
                                page.merge_page(ov)

                        writer.add_page(page)

                    post_path = tmp_pdf_path + ".post.pdf"
                    with open(post_path, "wb") as f:
                        writer.write(f)
                    os.replace(post_path, tmp_pdf_path)

                except Exception as e:
                    raise RuntimeError(f"PDF post-process başarısız: {e}")

                shutil.move(tmp_pdf_path, out_path)
                _progress(100)
                _status("PDF hazır.")

            try:
                if os.path.exists(tmp_pdf_path):
                    os.remove(tmp_pdf_path)
            except Exception:
                pass

        finally:
            _status("")


class StyleConfig:
    """
    Rapor stillerini veritabanından (sütunlu yapıdan) okur.
    JSON kullanmaz, doğrudan SQL sütunlarını sözlüğe çevirir.
    """

    DEFAULTS = {
        "main_title": {"label": "Rapor Ana Başlığı", "size": 18, "bold": True, "italic": False, "under": False},
        "section_title": {"label": "Bölüm Başlıkları", "size": 14, "bold": True, "italic": False, "under": False},
        "kv_keys": {"label": "Kapak Bilgi Başlıkları", "size": 12, "bold": True, "italic": False, "under": False},
        "kv_values": {"label": "Kapak Bilgi Değerleri", "size": 12, "bold": False, "italic": False, "under": False},
        "body": {"label": "Gövde Metinleri", "size": 12, "bold": False, "italic": False, "under": False},
        "table_head": {"label": "Tablo Başlıkları", "size": 14, "bold": True, "italic": False, "under": False},
        "table_body": {"label": "Tablo İçeriği", "size": 9, "bold": False, "italic": False, "under": False},
        "evidence_title": {"label": "Delil Başlıkları", "size": 14, "bold": True, "italic": False, "under": True},
        "evidence_desc": {"label": "Delil Açıklamaları", "size": 12, "bold": False, "italic": True, "under": False},
        "margins": {"top": 25, "right": 20, "bottom": 20, "left": 25}
    }

    @staticmethod
    def _init_tables(conn):
        """Tablolar yoksa oluşturur + eksik kolonları migrasyonla ekler."""
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS rapor_font_ayarlari (
                    Anahtar TEXT PRIMARY KEY,
                    Etiket TEXT,
                    Boyut INTEGER,
                    Kalin INTEGER DEFAULT 0,
                    Italik INTEGER DEFAULT 0,
                    AltiCizili INTEGER DEFAULT 0
                )
            """)

            # Esas tablo (TR kolonlar)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS rapor_sayfa_duzeni (
                    id INTEGER PRIMARY KEY CHECK (id = 1),
                    MarginUst INTEGER DEFAULT 25,
                    MarginSag INTEGER DEFAULT 20,
                    MarginAlt INTEGER DEFAULT 20,
                    MarginSol INTEGER DEFAULT 25
                )
            """)

            cols = [r[1] for r in conn.execute("PRAGMA table_info(rapor_sayfa_duzeni)").fetchall()]

            def _add_col(col_name: str, default_val: int):
                nonlocal cols
                if col_name not in cols:
                    conn.execute(f"ALTER TABLE rapor_sayfa_duzeni ADD COLUMN {col_name} INTEGER DEFAULT {int(default_val)}")
                    cols.append(col_name)

            # TR + EN kolonları garanti etmeye çalış
            _add_col("MarginUst", 25)
            _add_col("MarginSag", 20)
            _add_col("MarginAlt", 20)
            _add_col("MarginSol", 25)

            _add_col("TopMargin", 25)
            _add_col("RightMargin", 20)
            _add_col("BottomMargin", 20)
            _add_col("LeftMargin", 25)

            # id=1 satırı garanti
            conn.execute("INSERT OR IGNORE INTO rapor_sayfa_duzeni (id) VALUES (1)")

            # ✅ Güvenli NULL->default: sadece mevcut kolonlar için UPDATE üret
            cols = [r[1] for r in conn.execute("PRAGMA table_info(rapor_sayfa_duzeni)").fetchall()]
            sets = []
            if "MarginUst" in cols:    sets.append("MarginUst    = COALESCE(MarginUst, 25)")
            if "MarginSag" in cols:    sets.append("MarginSag    = COALESCE(MarginSag, 20)")
            if "MarginAlt" in cols:    sets.append("MarginAlt    = COALESCE(MarginAlt, 20)")
            if "MarginSol" in cols:    sets.append("MarginSol    = COALESCE(MarginSol, 25)")
            if "TopMargin" in cols:    sets.append("TopMargin    = COALESCE(TopMargin, 25)")
            if "RightMargin" in cols:  sets.append("RightMargin  = COALESCE(RightMargin, 20)")
            if "BottomMargin" in cols: sets.append("BottomMargin = COALESCE(BottomMargin, 20)")
            if "LeftMargin" in cols:   sets.append("LeftMargin   = COALESCE(LeftMargin, 25)")

            if sets:
                conn.execute(f"UPDATE rapor_sayfa_duzeni SET {', '.join(sets)} WHERE id=1")

            conn.commit()

        except Exception as e:
            print(f"Tablo kontrol hatası: {e}")

    @staticmethod
    def load():
        """Veritabanından ayarları okur ve Dict döndürür."""
        try:
            config = copy.deepcopy(StyleConfig.DEFAULTS)
        except Exception:
            config = StyleConfig.DEFAULTS.copy()

        try:
            with DB() as conn:
                StyleConfig._init_tables(conn)

                cols = [r[1] for r in conn.execute("PRAGMA table_info(rapor_sayfa_duzeni)").fetchall()]

                # Öncelik: TR kolonlar; yoksa EN kolonlar
                if all(c in cols for c in ("MarginUst", "MarginSag", "MarginAlt", "MarginSol")):
                    row = conn.execute(
                        "SELECT MarginUst, MarginSag, MarginAlt, MarginSol FROM rapor_sayfa_duzeni WHERE id=1"
                    ).fetchone()
                    if row:
                        config["margins"] = {"top": int(row[0]), "right": int(row[1]), "bottom": int(row[2]), "left": int(row[3])}

                elif all(c in cols for c in ("TopMargin", "RightMargin", "BottomMargin", "LeftMargin")):
                    row = conn.execute(
                        "SELECT TopMargin, RightMargin, BottomMargin, LeftMargin FROM rapor_sayfa_duzeni WHERE id=1"
                    ).fetchone()
                    if row:
                        config["margins"] = {"top": int(row[0]), "right": int(row[1]), "bottom": int(row[2]), "left": int(row[3])}

                rows = conn.execute(
                    "SELECT Anahtar, Etiket, Boyut, Kalin, Italik, AltiCizili FROM rapor_font_ayarlari"
                ).fetchall()

                for r in rows:
                    key = r[0]
                    if key in config and key != "margins":
                        config[key] = {
                            "label": str(r[1]),
                            "size": int(r[2]),
                            "bold": bool(r[3]),
                            "italic": bool(r[4]),
                            "under": bool(r[5])
                        }

        except Exception as e:
            print(f"Stil yükleme hatası: {e}")
            return config

        return config

    @staticmethod
    def save(data):
        """Verileri DB'ye yazar."""
        try:
            with DB() as conn:
                StyleConfig._init_tables(conn)

                for key, val in data.items():
                    if key == "margins":
                        continue
                    if isinstance(val, dict):
                        conn.execute("""
                            INSERT OR REPLACE INTO rapor_font_ayarlari
                            (Anahtar, Etiket, Boyut, Kalin, Italik, AltiCizili)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (
                            key,
                            val.get("label", ""),
                            int(val.get("size", 12)),
                            1 if val.get("bold") else 0,
                            1 if val.get("italic") else 0,
                            1 if val.get("under") else 0
                        ))

                margins = data.get("margins", {}) if isinstance(data.get("margins", {}), dict) else {}
                top = int(margins.get("top", 25))
                right = int(margins.get("right", 20))
                bottom = int(margins.get("bottom", 20))
                left = int(margins.get("left", 25))

                cols = [r[1] for r in conn.execute("PRAGMA table_info(rapor_sayfa_duzeni)").fetchall()]
                conn.execute("INSERT OR IGNORE INTO rapor_sayfa_duzeni (id) VALUES (1)")

                sets = []
                params = []

                # TR kolonlara yaz
                if "MarginUst" in cols:  sets.append("MarginUst=?");  params.append(top)
                if "MarginSag" in cols:  sets.append("MarginSag=?");  params.append(right)
                if "MarginAlt" in cols:  sets.append("MarginAlt=?");  params.append(bottom)
                if "MarginSol" in cols:  sets.append("MarginSol=?");  params.append(left)

                # EN kolonlara da yaz (TopMargin kullanan sürümlerle uyum)
                if "TopMargin" in cols:   sets.append("TopMargin=?");   params.append(top)
                if "RightMargin" in cols: sets.append("RightMargin=?"); params.append(right)
                if "BottomMargin" in cols:sets.append("BottomMargin=?");params.append(bottom)
                if "LeftMargin" in cols:  sets.append("LeftMargin=?");  params.append(left)

                if sets:
                    conn.execute(f"UPDATE rapor_sayfa_duzeni SET {', '.join(sets)} WHERE id=1", params)

                conn.commit()

        except Exception as e:
            print(f"Stil Kayıt Hatası: {e}")


class StyleEditorDialog(QDialog):
    """Gelişmiş Stil ve Düzen Editörü (Tablo Görünümlü - Başlıklı)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Rapor Biçim ve Düzen Ayarları")
        self.resize(700, 750)

        self.current_styles = StyleConfig.load()
        if isinstance(self.current_styles, str):
            try: self.current_styles = json.loads(self.current_styles)
            except: self.current_styles = None
        if not isinstance(self.current_styles, dict):
            self.current_styles = StyleConfig.DEFAULTS.copy()

        self.style_inputs = {}

        layout = QVBoxLayout(self)

        tabs = QTabWidget()

        tab_fonts = QWidget()
        l_fonts = QVBoxLayout(tab_fonts)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        content_widget = QWidget()

        grid = QGridLayout(content_widget)
        grid.setSpacing(15)
        grid.setContentsMargins(15, 20, 15, 20)

        headers = ["Düzenlenecek Alan", "Boyut (pt)", "Kalın", "İtalik", "Altı Çizili"]
        for col, text in enumerate(headers):
            lbl = QLabel(text)
            lbl.setStyleSheet("font-weight: bold; color: #555; text-transform: uppercase; font-size: 11px;")
            if col > 0:
                lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            grid.addWidget(lbl, 0, col)

        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("color: #ccc;")
        grid.addWidget(line, 1, 0, 1, 5)

        style_keys = [k for k in StyleConfig.DEFAULTS.keys() if k != "margins"]

        btn_style = """
            QPushButton { 
                background-color: #f9f9f9; 
                border: 1px solid #ccc; 
                border-radius: 4px; 
                font-size: 13px; font-weight: bold; color: #444;
            }
            QPushButton:checked { 
                background-color: #3498db; 
                color: white; 
                border: 1px solid #2980b9; 
            }
            QPushButton:hover { background-color: #e6e6e6; }
        """

        row = 2
        for key in style_keys:
            def_val = StyleConfig.DEFAULTS[key]
            curr_val = self.current_styles.get(key, def_val)
            if not isinstance(curr_val, dict): curr_val = def_val

            label_text = curr_val.get("label", def_val["label"])

            lbl_name = QLabel(label_text)
            lbl_name.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
            grid.addWidget(lbl_name, row, 0)

            sb = QSpinBox()
            sb.setRange(6, 72)
            sb.setValue(int(curr_val.get("size", 12)))
            sb.setSuffix(" pt")
            sb.setFixedWidth(90)
            sb.setAlignment(Qt.AlignmentFlag.AlignCenter)
            grid.addWidget(sb, row, 1, alignment=Qt.AlignmentFlag.AlignCenter)

            btn_bold = QPushButton("K")
            btn_bold.setCheckable(True)
            btn_bold.setChecked(bool(curr_val.get("bold", False)))
            btn_bold.setFixedSize(40, 30)
            btn_bold.setStyleSheet(btn_style)
            grid.addWidget(btn_bold, row, 2, alignment=Qt.AlignmentFlag.AlignCenter)

            btn_italic = QPushButton("I")
            btn_italic.setCheckable(True)
            btn_italic.setChecked(bool(curr_val.get("italic", False)))
            btn_italic.setFixedSize(40, 30)
            btn_italic.setStyleSheet(btn_style + "QPushButton { font-style: italic; font-family: serif; }")
            grid.addWidget(btn_italic, row, 3, alignment=Qt.AlignmentFlag.AlignCenter)

            btn_under = QPushButton("A")
            btn_under.setCheckable(True)
            btn_under.setChecked(bool(curr_val.get("under", False)))
            btn_under.setFixedSize(40, 30)
            btn_under.setStyleSheet(btn_style + "QPushButton { text-decoration: underline; }")
            grid.addWidget(btn_under, row, 4, alignment=Qt.AlignmentFlag.AlignCenter)

            self.style_inputs[key] = (sb, btn_bold, btn_italic, btn_under)
            row += 1

        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        grid.addWidget(spacer, row, 0)

        scroll.setWidget(content_widget)
        l_fonts.addWidget(scroll)
        tabs.addTab(tab_fonts, "🔤 Yazı Tipi ve Stiller")

        tab_page = QWidget()
        l_page = QVBoxLayout(tab_page)

        grp_margin = QGroupBox("Sayfa Kenar Boşlukları (mm)")
        grid_margin = QGridLayout(grp_margin)

        margins_curr = self.current_styles.get("margins", StyleConfig.DEFAULTS["margins"])

        self.sb_top = self._create_margin_spin(margins_curr.get("top", 25))
        self.sb_bottom = self._create_margin_spin(margins_curr.get("bottom", 20))
        self.sb_left = self._create_margin_spin(margins_curr.get("left", 25))
        self.sb_right = self._create_margin_spin(margins_curr.get("right", 20))

        grid_margin.setHorizontalSpacing(25)
        grid_margin.setVerticalSpacing(15)

        grid_margin.addWidget(QLabel("Üst"), 0, 1, Qt.AlignmentFlag.AlignCenter)
        grid_margin.addWidget(self.sb_top, 1, 1, Qt.AlignmentFlag.AlignCenter)

        grid_margin.addWidget(QLabel("Sol"), 2, 0, Qt.AlignmentFlag.AlignRight)
        grid_margin.addWidget(self.sb_left, 3, 0, Qt.AlignmentFlag.AlignRight)

        page_icon = QLabel("📄")
        page_icon.setStyleSheet("font-size: 48px; color: #bdc3c7;")
        page_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        grid_margin.addWidget(page_icon, 2, 1, 2, 1, Qt.AlignmentFlag.AlignCenter)

        grid_margin.addWidget(QLabel("Sağ"), 2, 2, Qt.AlignmentFlag.AlignLeft)
        grid_margin.addWidget(self.sb_right, 3, 2, Qt.AlignmentFlag.AlignLeft)

        grid_margin.addWidget(self.sb_bottom, 4, 1, Qt.AlignmentFlag.AlignCenter)
        grid_margin.addWidget(QLabel("Alt"), 5, 1, Qt.AlignmentFlag.AlignCenter)

        l_page.addWidget(grp_margin)
        l_page.addStretch()
        tabs.addTab(tab_page, "📐 Sayfa Düzeni")

        layout.addWidget(tabs)

        btn_box = QHBoxLayout()
        btn_reset = QPushButton("Varsayılanlara Dön")
        btn_reset.setStyleSheet("background-color: #95a5a6; color: white; padding: 8px;")
        btn_reset.clicked.connect(self.reset_defaults)

        btn_save = QPushButton("💾 Kaydet ve Uygula")
        btn_save.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 10px; font-size: 14px;")
        btn_save.clicked.connect(self.save_settings)

        btn_box.addWidget(btn_reset)
        btn_box.addStretch()
        btn_box.addWidget(btn_save)
        layout.addLayout(btn_box)

    def _create_margin_spin(self, val):
        sb = QSpinBox()
        sb.setRange(0, 100)
        sb.setValue(int(val))
        sb.setSuffix(" mm")
        sb.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sb.setFixedWidth(80)
        return sb

    def save_settings(self):
        new_data = {}
        for key, (sb, b, i, u) in self.style_inputs.items():
            orig_label = self.current_styles.get(key, {}).get("label", "")
            if not orig_label: orig_label = StyleConfig.DEFAULTS[key]["label"]

            new_data[key] = {
                "label": orig_label,
                "size": sb.value(),
                "bold": b.isChecked(),
                "italic": i.isChecked(),
                "under": u.isChecked()
            }

        new_data["margins"] = {
            "top": self.sb_top.value(),
            "bottom": self.sb_bottom.value(),
            "left": self.sb_left.value(),
            "right": self.sb_right.value()
        }

        if "body" not in new_data: new_data["body"] = StyleConfig.DEFAULTS["body"]
        StyleConfig.save(new_data)
        ModernDialog.show_success(self, "Kaydedildi", "Ayarlar kaydedildi.")
        self.accept()

    def reset_defaults(self):
        if ModernDialog.show_question(self, "Sıfırla", "Varsayılanlara dönülsün mü?"):
            StyleConfig.save(StyleConfig.DEFAULTS)
            self.accept()


class RichDescriptionDialog(QDialog):
    """Açıklama düzenlemek için özel Zengin Metin (Kalın/İtalik/Altı Çizili) penceresi."""
    def __init__(self, parent=None, text="", title="Açıklama Düzenle"):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(800, 450)

        layout = QVBoxLayout(self)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(5)

        btn_style = """
            QPushButton { 
                background-color: #f0f0f0; border: 1px solid #ccc; 
                border-radius: 3px; font-weight: bold; width: 28px; height: 28px;
                font-family: 'Times New Roman'; font-size: 14px; color: #333;
            }
            QPushButton:hover { background-color: #e0e0e0; }
            QPushButton:checked { background-color: #3498db; color: white; border: 1px solid #2980b9; }
        """

        def set_fmt(fmt_type):
            cursor = self.editor.textCursor()
            fmt = QTextCharFormat()
            if fmt_type == 'bold':
                fmt.setFontWeight(QFont.Weight.Bold if cursor.charFormat().fontWeight() != QFont.Weight.Bold else QFont.Weight.Normal)
            elif fmt_type == 'italic':
                fmt.setFontItalic(not cursor.charFormat().fontItalic())
            elif fmt_type == 'under':
                fmt.setFontUnderline(not cursor.charFormat().fontUnderline())
            self.editor.mergeCurrentCharFormat(fmt)
            self.editor.setFocus()

        btn_b = QPushButton("K"); btn_b.setToolTip("Kalın"); btn_b.setStyleSheet(btn_style); btn_b.clicked.connect(lambda: set_fmt('bold'))
        btn_i = QPushButton("I"); btn_i.setToolTip("İtalik"); btn_i.setStyleSheet(btn_style + "QPushButton { font-style: italic; }"); btn_i.clicked.connect(lambda: set_fmt('italic'))
        btn_u = QPushButton("A"); btn_u.setToolTip("Altı Çizili"); btn_u.setStyleSheet(btn_style + "QPushButton { text-decoration: underline; }"); btn_u.clicked.connect(lambda: set_fmt('under'))

        toolbar.addWidget(btn_b); toolbar.addWidget(btn_i); toolbar.addWidget(btn_u); toolbar.addStretch()
        layout.addLayout(toolbar)

        self.editor = QTextEdit()
        if "<" in text and ">" in text:
            self.editor.setHtml(text)
        else:
            self.editor.setPlainText(text)

        layout.addWidget(self.editor)

        btn_box = QHBoxLayout()
        btn_cancel = QPushButton("İptal"); btn_cancel.clicked.connect(self.reject)
        btn_save = QPushButton("Kaydet"); btn_save.setStyleSheet("background-color:#27ae60; color:white; font-weight:bold;"); btn_save.clicked.connect(self.accept)
        btn_box.addStretch(); btn_box.addWidget(btn_cancel); btn_box.addWidget(btn_save)
        layout.addLayout(btn_box)

    def get_html(self):
        return self.editor.toHtml()


class MetaEkRowWidget(QWidget):
    clicked = pyqtSignal(int)          # db_id
    edit_desc = pyqtSignal(int)        # db_id
    edit_props = pyqtSignal(int)       # db_id
    deleted = pyqtSignal(int)          # db_id

    def __init__(self, db_id: int, title: str, has_desc: bool, parent=None):
        super().__init__(parent)
        self.db_id = int(db_id)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(8, 2, 8, 2)
        lay.setSpacing(8)

        self.lbl = QLabel(title)
        self.lbl.setStyleSheet("font-weight:600;")
        self.lbl.setCursor(Qt.CursorShape.PointingHandCursor)
        self.lbl.mousePressEvent = lambda e: self.clicked.emit(self.db_id)
        lay.addWidget(self.lbl, 1)

        self.btn_desc = QPushButton("Açıklama" + (" ✅" if has_desc else ""))
        self.btn_desc.clicked.connect(lambda: self.edit_desc.emit(self.db_id))
        lay.addWidget(self.btn_desc)

        self.btn_props = QPushButton("Ayarlar")
        self.btn_props.clicked.connect(lambda: self.edit_props.emit(self.db_id))
        lay.addWidget(self.btn_props)

        self.btn_del = QPushButton("Sil")
        self.btn_del.clicked.connect(lambda: self.deleted.emit(self.db_id))
        lay.addWidget(self.btn_del)


class EvidenceRowWidget(QWidget):
    clicked = pyqtSignal(int)       # ✅ eid (önizlemede delile git)
    edit_desc = pyqtSignal(int)     # eid
    edit_props = pyqtSignal(int)    # eid
    move_up = pyqtSignal(int)       # eid
    move_down = pyqtSignal(int)     # eid
    deleted = pyqtSignal(int)

    def __init__(self, eid: int, title: str, has_desc: bool, parent=None):
        super().__init__(parent)
        self.eid = int(eid)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(8, 2, 8, 2)
        lay.setSpacing(8)

        lbl = QLabel(title)
        lbl.setStyleSheet("font-weight:600;")
        lbl.setCursor(Qt.CursorShape.PointingHandCursor)  # ✅
        lbl.mousePressEvent = lambda e: self.clicked.emit(self.eid)  # ✅
        lay.addWidget(lbl, 1)

        btn_up = QPushButton("↑")
        btn_up.setFixedWidth(36)
        btn_up.clicked.connect(lambda: self.move_up.emit(self.eid))
        lay.addWidget(btn_up)

        btn_down = QPushButton("↓")
        btn_down.setFixedWidth(36)
        btn_down.clicked.connect(lambda: self.move_down.emit(self.eid))
        lay.addWidget(btn_down)

        btn_desc = QPushButton("Açıklama" + (" ✅" if has_desc else ""))
        btn_desc.clicked.connect(lambda: self.edit_desc.emit(self.eid))
        lay.addWidget(btn_desc)

        btn_props = QPushButton("Ayarlar")
        btn_props.clicked.connect(lambda: self.edit_props.emit(self.eid))
        lay.addWidget(btn_props)

        btn_del = QPushButton("Sil")
        btn_del.clicked.connect(lambda: self.deleted.emit(self.eid))
        lay.addWidget(btn_del)


class TableFormatEditorDialog(QDialog):
    """
    Hücre/Satır/Sütun renklendirme editörü (delile özel).
    Değişiklikler fmt dict olarak döner.
    fmt:
      {"rows": {"0":"#.."}, "cols":{"1":"#.."}, "cells":{"2,3":"#.."}}
    """
    def __init__(self, parent, headers: list[str], rows: list[list[str]], initial_fmt: dict | None = None):
        super().__init__(parent)
        self.setWindowTitle("Tablo Düzenle (Renklendir)")
        self.setMinimumSize(860, 540)

        self._fmt = initial_fmt if isinstance(initial_fmt, dict) else {"rows": {}, "cols": {}, "cells": {}}
        for k in ("rows", "cols", "cells"):
            if k not in self._fmt or not isinstance(self._fmt[k], dict):
                self._fmt[k] = {}

        # history
        self._history = [copy.deepcopy(self._fmt)]
        self._hist_i = 0

        lay = QVBoxLayout(self)

        # üst bar
        tools = QHBoxLayout()
        btn_color = QPushButton("Renk Seç")
        btn_cells = QPushButton("Seçili Hücrelere Uygula")
        btn_row = QPushButton("Seçili Satıra Uygula")
        btn_col = QPushButton("Seçili Sütuna Uygula")
        btn_clear = QPushButton("Temizle")
        btn_undo = QPushButton("↩ Geri Al")
        btn_redo = QPushButton("↪ İleri Al")

        tools.addWidget(btn_color)
        tools.addWidget(btn_cells)
        tools.addWidget(btn_row)
        tools.addWidget(btn_col)
        tools.addWidget(btn_clear)
        tools.addWidget(btn_undo)
        tools.addWidget(btn_redo)
        tools.addStretch(1)
        lay.addLayout(tools)

        # tablo
        self.tbl = QTableWidget()
        self.table = self.tbl  # eski çağrılar için alias (güvenlik)
        self.tbl.setColumnCount(len(headers))
        self.tbl.setRowCount(len(rows))
        self.tbl.setHorizontalHeaderLabels([str(h) for h in headers])

        # ✅ EDIT KAPAT: çift tıkla düzenleme bitsin
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)

        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                it = QTableWidgetItem(str(val))
                it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)  # ekstra garanti
                self.tbl.setItem(r, c, it)

        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tbl.verticalHeader().setVisible(True)
        lay.addWidget(self.tbl, 1)

        # alt butonlar
        row_btns = QHBoxLayout()
        row_btns.addStretch(1)
        btn_cancel = QPushButton("Vazgeç")
        btn_ok = QPushButton("Kaydet")
        row_btns.addWidget(btn_cancel)
        row_btns.addWidget(btn_ok)
        lay.addLayout(row_btns)

        self._selected_color = "#ff0000"

        def _push_history():
            # aynı state tekrar eklenmesin
            cur = copy.deepcopy(self._fmt)
            if self._history and self._history[self._hist_i] == cur:
                return
            self._history = self._history[: self._hist_i + 1]
            self._history.append(cur)
            self._hist_i += 1

        def pick_color():
            c = QColorDialog.getColor(QColor(self._selected_color), self, "Renk Seç")
            if c.isValid():
                self._selected_color = c.name()

        def apply_cells():
            sels = self.tbl.selectedIndexes()
            if not sels:
                return
            for ix in sels:
                self._fmt["cells"][f"{ix.row()},{ix.column()}"] = self._selected_color
            _push_history()
            self._apply_preview_colors()

        def apply_row():
            sels = self.tbl.selectedIndexes()
            if not sels:
                return
            r = sels[0].row()
            self._fmt["rows"][str(r)] = self._selected_color
            _push_history()
            self._apply_preview_colors()

        def apply_col():
            sels = self.tbl.selectedIndexes()
            if not sels:
                return
            c = sels[0].column()
            self._fmt["cols"][str(c)] = self._selected_color
            _push_history()
            self._apply_preview_colors()

        def clear_all():
            self._fmt = {"rows": {}, "cols": {}, "cells": {}}
            _push_history()
            self._apply_preview_colors()

        def undo():
            if self._hist_i <= 0:
                return
            self._hist_i -= 1
            self._fmt = copy.deepcopy(self._history[self._hist_i])
            self._apply_preview_colors()

        def redo():
            if self._hist_i >= len(self._history) - 1:
                return
            self._hist_i += 1
            self._fmt = copy.deepcopy(self._history[self._hist_i])
            self._apply_preview_colors()

        btn_color.clicked.connect(pick_color)
        btn_cells.clicked.connect(apply_cells)
        btn_row.clicked.connect(apply_row)
        btn_col.clicked.connect(apply_col)
        btn_clear.clicked.connect(clear_all)
        btn_undo.clicked.connect(undo)
        btn_redo.clicked.connect(redo)

        btn_cancel.clicked.connect(self.reject)
        btn_ok.clicked.connect(self.accept)

        # ilk boya
        self._apply_preview_colors()

    def _apply_preview_colors(self):
        """✅ Daha önce çağrılıp bulunamadığı için eklendi."""
        # önce temizle
        for r in range(self.tbl.rowCount()):
            for c in range(self.tbl.columnCount()):
                it = self.tbl.item(r, c)
                if it:
                    it.setBackground(QColor("#ffffff"))

        rows_map = self._fmt.get("rows", {}) or {}
        cols_map = self._fmt.get("cols", {}) or {}
        cells_map = self._fmt.get("cells", {}) or {}

        # row/col/cell (cell en baskın)
        for r in range(self.tbl.rowCount()):
            row_color = rows_map.get(str(r))
            for c in range(self.tbl.columnCount()):
                color = cells_map.get(f"{r},{c}") or row_color or cols_map.get(str(c))
                if not color:
                    continue
                it = self.tbl.item(r, c)
                if it:
                    it.setBackground(QColor(color))

    def get_fmt(self) -> dict:
        return copy.deepcopy(self._fmt)

    def _push_history(self):
        cur = copy.deepcopy(self.fmt)
        # aynı state’i iki kere basma
        if self._history and json.dumps(self._history[self._hist_i], sort_keys=True, ensure_ascii=False) == json.dumps(cur, sort_keys=True, ensure_ascii=False):
            return
        # ileri geçmişi kırp
        self._history = self._history[: self._hist_i + 1]
        self._history.append(cur)
        self._hist_i += 1

    def undo(self):
        if self._hist_i <= 0:
            return
        self._hist_i -= 1
        self.fmt = copy.deepcopy(self._history[self._hist_i])
        self._repaint_from_fmt()

    def redo(self):
        if self._hist_i >= len(self._history) - 1:
            return
        self._hist_i += 1
        self.fmt = copy.deepcopy(self._history[self._hist_i])
        self._repaint_from_fmt()

    def _repaint_from_fmt(self):
        # tablo widget’ındaki tüm hücrelerin arkaplanını temizle
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                if item:
                    item.setBackground(QColor("white"))

        cols  = self.fmt.get("cols", {})  if isinstance(self.fmt.get("cols"), dict)  else {}
        rows  = self.fmt.get("rows", {})  if isinstance(self.fmt.get("rows"), dict)  else {}
        cells = self.fmt.get("cells", {}) if isinstance(self.fmt.get("cells"), dict) else {}

        # row + col
        for r in range(self.table.rowCount()):
            row_color = rows.get(str(r))
            for c in range(self.table.columnCount()):
                color = cells.get(f"{r},{c}") or row_color or cols.get(str(c))
                if not color:
                    continue
                item = self.table.item(r, c)
                if item:
                    item.setBackground(QColor(color))


class EvidencePropsDialog(QDialog):
    """
    Delil (rapor_taslagi) için: Başlık + Genişlik(%) + (TABLE ise) Kolon seçimi + Tablo renklendirme.
    """
    def __init__(
        self,
        parent,
        title: str,
        width_pct: int,
        is_table: bool = False,
        table_html: str = "",
        base_html: str = "",
        hidden_json: str = "",
        fmt_json: str = "",
    ):
        super().__init__(parent)
        self.setWindowTitle("Delil Ayarları")
        self.setModal(True)
        self.resize(520, 420)

        lay = QVBoxLayout(self)

        # --- başlık
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Başlık:"))
        self.edt_title = QLineEdit(title or "")
        row1.addWidget(self.edt_title)
        lay.addLayout(row1)

        # --- genişlik
        row2 = QHBoxLayout()
        w0 = int(width_pct) if width_pct else 100
        w0 = max(10, min(100, w0))

        self.lbl_w = QLabel(f"Genişlik: %{w0}")
        row2.addWidget(self.lbl_w)

        self.sld = QSlider(Qt.Orientation.Horizontal)
        self.sld.setMinimum(10)
        self.sld.setMaximum(100)
        self.sld.setValue(w0)
        self.sld.valueChanged.connect(lambda v: self.lbl_w.setText(f"Genişlik: %{v}"))
        row2.addWidget(self.sld)
        lay.addLayout(row2)

        # --- TABLE özel
        self._is_table = bool(is_table)
        self._table_html_in = str(table_html or "")
        self.base_html = (str(base_html or "").strip() or self._table_html_in)  # ✅ base öncelikli

        # hidden_json -> list
        try:
            hidden_list = json.loads(hidden_json) if hidden_json else []
            if not isinstance(hidden_list, list):
                hidden_list = []
        except Exception:
            hidden_list = []
        self._hidden_init = set(str(x) for x in hidden_list)

        # fmt_json -> dict
        try:
            fmt_obj = json.loads(fmt_json) if fmt_json else {}
            if not isinstance(fmt_obj, dict):
                fmt_obj = {}
        except Exception:
            fmt_obj = {}

        self._fmt = fmt_obj if isinstance(fmt_obj, dict) else {"rows": {}, "cols": {}, "cells": {}}
        if "rows" not in self._fmt or not isinstance(self._fmt["rows"], dict): self._fmt["rows"] = {}
        if "cols" not in self._fmt or not isinstance(self._fmt["cols"], dict): self._fmt["cols"] = {}
        if "cells" not in self._fmt or not isinstance(self._fmt["cells"], dict): self._fmt["cells"] = {}

        if self._is_table and self.base_html.strip():
            # ✅ kolonları base_html’den çıkar (kaldırılan kolon geri gelebilsin)
            headers, _rows = _extract_table_headers_rows(self.base_html)

            grp = QGroupBox("Tablo Kolonları (Bu delile özel)")
            g_lay = QVBoxLayout(grp)

            info = QLabel("Önizlemede görünen kolonlar işaretli gelir. İstemediğinizi kaldırın.")
            info.setStyleSheet("color:#444;")
            g_lay.addWidget(info)

            wrap = QWidget()
            v = QVBoxLayout(wrap)
            v.setContentsMargins(6, 6, 6, 6)

            self._col_checks: list[tuple[str, QCheckBox]] = []
            for h in headers:
                h_str = str(h)
                cb = QCheckBox(h_str)
                # ✅ daha önce gizlenmişse unchecked başlasın
                cb.setChecked(h_str not in self._hidden_init)
                self._col_checks.append((h_str, cb))
                v.addWidget(cb)

            scr = QScrollArea()
            scr.setWidgetResizable(True)
            scr.setMinimumHeight(160)
            scr.setWidget(wrap)
            g_lay.addWidget(scr)

            btn_edit = QPushButton("Tablo Düzenle (Renklendir)")
            g_lay.addWidget(btn_edit)

            def open_editor():
                # editor her zaman "güncel görünür kolonlar" üzerinden açılsın
                hidden = set(self.hidden_cols_list())
                html1 = _apply_hidden_cols_to_table_html(self.base_html, list(hidden))
                hdr2, rows2 = _extract_table_headers_rows(html1)

                dlg2 = TableFormatEditorDialog(self, hdr2, rows2, self._fmt)
                if dlg2.exec() == 1:
                    self._fmt = dlg2.get_fmt()

            btn_edit.clicked.connect(open_editor)

            lay.addWidget(grp)

        # --- butonlar
        btns = QHBoxLayout()
        btns.addStretch(1)
        b_no = QPushButton("Vazgeç")
        b_ok = QPushButton("Kaydet")
        b_ok.clicked.connect(self.accept)
        b_no.clicked.connect(self.reject)
        btns.addWidget(b_no)
        btns.addWidget(b_ok)
        lay.addLayout(btns)

    def hidden_cols_list(self) -> list[str]:
        """
        Checkbox’ta unchecked olan kolonlar = hidden.
        """
        if not getattr(self, "_is_table", False):
            return []
        if not hasattr(self, "_col_checks"):
            return []
        hidden = []
        for h, cb in self._col_checks:
            if not cb.isChecked():
                hidden.append(h)
        return hidden

    def get_fmt(self) -> dict:
        return self._fmt

    def values(self):
        """
        Dönüş: (title, width, hidden_list_or_None, fmt_or_None)
        """
        title = self.edt_title.text().strip()
        w = int(self.sld.value())

        if not self._is_table:
            return (title, w, None, None)

        return (title, w, self.hidden_cols_list(), self.get_fmt())


class MetaEkPropsDialog(QDialog):
    def __init__(self, parent, file_name: str, hide_name: bool, width_pct: int):
        super().__init__(parent)
        self.setWindowTitle("Ek Ayarları")
        self.setModal(True)
        self.resize(420, 180)

        lay = QVBoxLayout(self)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Dosya Adı:"))
        self.edt_name = QLineEdit(file_name or "")
        row1.addWidget(self.edt_name)
        lay.addLayout(row1)

        self.chk_hide = QCheckBox("Raporda dosya adını gizle")
        self.chk_hide.setChecked(bool(hide_name))
        lay.addWidget(self.chk_hide)

        row2 = QHBoxLayout()
        self.lbl_w = QLabel(f"Genişlik: %{int(width_pct) if width_pct else 80}")
        row2.addWidget(self.lbl_w)

        self.sld = QSlider(Qt.Orientation.Horizontal)
        self.sld.setMinimum(30)
        self.sld.setMaximum(100)
        self.sld.setValue(int(width_pct) if width_pct else 80)
        self.sld.valueChanged.connect(lambda v: self.lbl_w.setText(f"Genişlik: %{v}"))
        row2.addWidget(self.sld)
        lay.addLayout(row2)

        btns = QHBoxLayout()
        btns.addStretch(1)
        b_ok = QPushButton("Kaydet")
        b_no = QPushButton("Vazgeç")
        b_ok.clicked.connect(self.accept)
        b_no.clicked.connect(self.reject)
        btns.addWidget(b_no)
        btns.addWidget(b_ok)
        lay.addLayout(btns)

    def values(self):
        return (self.edt_name.text().strip(), bool(self.chk_hide.isChecked()), int(self.sld.value()))


class ReorderableListWidget(QListWidget):
    def __init__(self, on_reorder, parent=None):
        super().__init__(parent)
        self._on_reorder = on_reorder

    def dropEvent(self, event):
        super().dropEvent(event)
        try:
            if callable(self._on_reorder):
                self._on_reorder()
        except Exception:
            pass


class MetaEkPanel(QWidget):
    changed_any = pyqtSignal()

    def __init__(self, project_id: int, bolum: str, parent=None):
        super().__init__(parent)
        try:
            self.project_id = int(project_id)
        except Exception:
            self.project_id = 0
        self.bolum = str(bolum)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        # Üst bar
        top = QHBoxLayout()

        self.btn_add = QPushButton("Resim Ekle (Yerel Disk)")
        self.btn_add.clicked.connect(self.add_from_disk)
        top.addWidget(self.btn_add)

        # delil_cekmece: swap KALDIRILDI (tek liste)
        self.btn_swap = None

        top.addStretch(1)
        root.addLayout(top)

        if self.bolum == "delil_cekmece":
            # ✅ TEK LİSTE: rapor_taslagi tablosundaki tüm deliller (resim/tablo/görsel/html)
            self.lbl_evi = QLabel("Deliller (Resim / Tablo / Görsel / HTML)")
            self.lbl_evi.setStyleSheet("font-weight:700; padding:4px 0;")
            root.addWidget(self.lbl_evi)

            self.listw_evidence = QListWidget()
            self.listw_evidence.setSpacing(6)
            self.listw_evidence.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
            root.addWidget(self.listw_evidence, 1)

            # ✅ Eski sürümden kalan meta_ekler/delil_cekmece kayıtlarını bir kere taşımak için
            self._migrate_delil_cekmece_meta_to_taslak()

        else:
            # ✅ Diğer bölümler: mevcut davranış (rapor_meta_ekler listesi)
            self.lbl_meta = QLabel("Çekmeceye Eklenen Resimler")
            self.lbl_meta.setStyleSheet("font-weight:700; padding:4px 0;")
            root.addWidget(self.lbl_meta)

            self.listw = QListWidget()
            self.listw.setSpacing(6)
            root.addWidget(self.listw, 1)

        self.reload()

        try:
            if hasattr(self.parent(), "refresh_preview"):
                self.parent().refresh_preview()
        except Exception:
            pass

    def _migrate_delil_cekmece_meta_to_taslak(self):
        """
        Eski yapıda delil_cekmece resimleri rapor_meta_ekler(Bolum='delil_cekmece') içinde duruyordu.
        Yeni yapıda hepsi rapor_taslagi içinde olmalı (tek sıra/tek liste için).
        Bu fonksiyon varsa eski kayıtları rapor_taslagi'na taşır ve meta_ekler tarafını temizler.
        """
        if self.bolum != "delil_cekmece":
            return

        try:
            with DB() as conn:
                # meta_ekler tablosu yoksa/kolonlar farklıysa sessiz geç
                try:
                    cols = {r[1] for r in conn.execute("PRAGMA table_info(rapor_meta_ekler)").fetchall()}
                    if not cols:
                        return
                except Exception:
                    return

                # delil_cekmece kayıtlarını çek
                try:
                    rows = conn.execute("""
                        SELECT id, DosyaAdi, DosyaYolu, Aciklama
                        FROM rapor_meta_ekler
                        WHERE ProjeID=? AND Bolum='delil_cekmece'
                        ORDER BY id ASC
                    """, (int(self.project_id),)).fetchall()
                except Exception:
                    return

                if not rows:
                    return

                # rapor_taslagi'nda en sona eklemek için max sıra
                cur = conn.cursor()
                last_order = cur.execute(
                    "SELECT MAX(Sira) FROM rapor_taslagi WHERE ProjeID=?",
                    (int(self.project_id),)
                ).fetchone()[0]
                new_order = (last_order or 0)

                # rapor_taslagi kolonlarını dinamik oku (eski db sürümleriyle uyumlu)
                tcols = [r[1] for r in cur.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]

                for mid, fname, fpath, mdesc in rows:
                    if not fpath:
                        continue

                    new_order += 1
                    title = str(fname or os.path.basename(str(fpath)) or "Delil")

                    # rapor_taslagi insert (add_evidence_to_report mantığına paralel)
                    keys = ["ProjeID", "GSMNo", "Baslik", "Icerik", "Tur", "Tarih", "Sira"]
                    vals = [
                        int(self.project_id), "", title, str(fpath), "IMAGE",
                        datetime.now().strftime("%d.%m.%Y %H:%M"), int(new_order)
                    ]

                    # genişlik default 80 (delil resimleri için)
                    if "GenislikYuzde" in tcols:
                        keys.append("GenislikYuzde"); vals.append(100)
                    if "YukseklikMm" in tcols:
                        keys.append("YukseklikMm"); vals.append(0)
                    if "Hizalama" in tcols:
                        keys.append("Hizalama"); vals.append("center")
                    if "Aciklama" in tcols:
                        keys.append("Aciklama"); vals.append(str(mdesc or ""))
                    if "HtmlIcerik" in tcols:
                        keys.append("HtmlIcerik"); vals.append(None)
                    if "ImagePath" in tcols:
                        keys.append("ImagePath"); vals.append(str(fpath))

                    q_marks = ",".join(["?"] * len(keys))
                    col_names = ",".join(keys)
                    cur.execute(f"INSERT INTO rapor_taslagi ({col_names}) VALUES ({q_marks})", vals)

                    # meta_ekler kaydını sil
                    cur.execute("DELETE FROM rapor_meta_ekler WHERE id=?", (int(mid),))

                conn.commit()

        except Exception:
            # migration hata verirse UI'yı kilitlemeyelim
            return

    def _ensure_evidence_dir(self) -> str:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        evidence_dir = os.path.join(base_dir, "evidence_images")
        if not os.path.exists(evidence_dir):
            os.makedirs(evidence_dir)
        return evidence_dir

    def _edit_evidence_props(self, eid: int):
        """
        rapor_taslagi için delil başlığı + genişlik ayarı.
        UI: MetaEkPropsDialog ile aynı stile çekildi (slider + Kaydet/Vazgeç).
        """
        try:
            eid = int(eid)
        except Exception:
            return

        # Mevcut değerleri oku
        with DB() as conn:
            row = conn.execute(
                "SELECT id, Baslik, COALESCE(GenislikYuzde, 100), COALESCE(Tur,''), "
                "COALESCE(HtmlIcerik,''), COALESCE(BaseHtmlIcerik,''), "
                "COALESCE(HiddenColsJson,''), COALESCE(FmtJson,'') "
                "FROM rapor_taslagi WHERE id=? AND ProjeID=?",
                (eid, int(self.project_id))
            ).fetchone()

        if not row:
            return

        rid, cur_title, cur_w, cur_tur, cur_html, base_html, hidden_json, fmt_json = row

        cur_tur = str(cur_tur or "")
        is_table = (cur_tur.upper() == "TABLE")

        try:
            cur_w = int(cur_w or 100)
        except Exception:
            cur_w = 100

        # ✅ base boşsa backfill (senin örnek verdiğin “base boşsa backfill” tam burası)
        if (not base_html) and cur_html:
            base_html = cur_html
        is_table = (cur_tur.upper() == "TABLE")
        try:
            cur_w = int(cur_w or 100)
        except Exception:
            cur_w = 100

        # ✅ Ekler ile aynı UI yaklaşımı
        dlg = EvidencePropsDialog(
            self,
            str(cur_title or ""),
            int(cur_w),
            is_table=is_table,
            table_html=str(cur_html or ""),
            base_html=str(base_html or ""),
            hidden_json=str(hidden_json or ""),
            fmt_json=str(fmt_json or ""),
        )
        if dlg.exec() != 1:
            return

        new_title, new_w, hidden_cols, fmt = dlg.values()

        # DB güncelle
        with DB() as conn:
            if is_table and isinstance(cur_html, str) and cur_html.strip():
                # 1) kolon gizleme (checkbox’a göre)
                hidden = hidden_cols or set()
                html1 = _apply_hidden_cols_to_table_html(cur_html, hidden)

                # 2) renklendirme (hücre/satır/sütun)
                if isinstance(fmt, dict) and fmt:
                    html2 = _apply_fmt_to_table_html(html1, fmt)
                else:
                    html2 = html1

                hidden_list = dlg.hidden_cols_list()
                hidden_json_out = json.dumps(hidden_list, ensure_ascii=False)

                # BaseHtmlIcerik'i sakla; HtmlIcerik'i artık base olarak tut (istersen aynı bırak)
                conn.execute(
                    "UPDATE rapor_taslagi SET Baslik=?, GenislikYuzde=?, "
                    "BaseHtmlIcerik=?, HtmlIcerik=?, HiddenColsJson=?, FmtJson=? "
                    "WHERE id=? AND ProjeID=?",
                    (
                        new_title, int(new_w),
                        dlg.base_html,            # base sabit kalsın
                        html2,                    # ✅ preview'e uygulanmış sonuç HtmlIcerik'e yazılsın
                        hidden_json_out,
                        (json.dumps(fmt, ensure_ascii=False) if isinstance(fmt, dict) else ""),
                        rid, int(self.project_id)
                    )
                )
            else:
                conn.execute(
                    "UPDATE rapor_taslagi SET Baslik=?, GenislikYuzde=? WHERE id=? AND ProjeID=?",
                    (new_title, int(new_w), eid, int(self.project_id))
                )
            conn.commit()

        # Yenile
        self.reload()
        self.changed_any.emit()
        try:
            if hasattr(self.parent(), "refresh_preview"):
                self.parent().refresh_preview()
        except Exception:
            pass

    def _edit_props(self, db_id: int):
        with DB() as conn:
            row = conn.execute(
                "SELECT DosyaAdi, COALESCE(DosyaAdiGizle,0), COALESCE(GenislikYuzde,80) "
                "FROM rapor_meta_ekler WHERE id=? AND ProjeID=?",
                (int(db_id), self.project_id)
            ).fetchone()

        if not row:
            return

        cur_name, hide_name, width_pct = row
        dlg = MetaEkPropsDialog(self, cur_name or "", bool(hide_name), int(width_pct) if width_pct else 80)
        if dlg.exec() != 1:
            return

        new_name, new_hide, new_width = dlg.values()

        with DB() as conn:
            conn.execute(
                "UPDATE rapor_meta_ekler SET DosyaAdi=?, DosyaAdiGizle=?, GenislikYuzde=? "
                "WHERE id=? AND ProjeID=?",
                (new_name, 1 if new_hide else 0, int(new_width), int(db_id), self.project_id)
            )
            conn.commit()

        self.reload()
        self.changed_any.emit()
        # rapor önizleme yenilemen hangi fonksiyonsa onu çağır:
        try:
            if hasattr(self.parent(), "refresh_preview"):
                self.parent().refresh_preview()
        except Exception:
            pass

    def reload(self):
        if self.bolum == "delil_cekmece":
            # ✅ Tek liste: rapor_taslagi
            if not hasattr(self, "listw_evidence") or self.listw_evidence is None:
                return

            self.listw_evidence.clear()

            with DB() as conn:
                rows = conn.execute("""
                    SELECT id, COALESCE(Sira, 999999) AS S, Baslik, Tur, Aciklama
                    FROM rapor_taslagi
                    WHERE ProjeID=?
                    ORDER BY COALESCE(Sira, 999999) ASC, id ASC
                """, (int(self.project_id),)).fetchall()

            ek_no = 1
            for (eid, sira, baslik, tur, acik) in rows:
                title = f"Ek-{ek_no} - {str(baslik or '').strip()}"
                if not str(baslik or "").strip():
                    # Baslik boşsa en azından türü göster
                    title = f"Ek-{ek_no} - {str(tur or 'DELIL')}"

                has_desc = bool(str(acik or "").strip())

                item = QListWidgetItem()
                roww = EvidenceRowWidget(int(eid), title, has_desc, self)

                # sinyaller (mevcut fonksiyonlarını kullanıyoruz)
                roww.clicked.connect(lambda eid: self._focus_item(eid, "delil_cekmece"))
                roww.edit_desc.connect(self._edit_evidence_desc)
                roww.edit_props.connect(self._edit_evidence_props)
                roww.move_up.connect(lambda _eid, self=self: self._move_evidence(_eid, -1))
                roww.move_down.connect(lambda _eid, self=self: self._move_evidence(_eid, +1))
                roww.deleted.connect(self._delete_evidence)

                item.setSizeHint(QSize(10, 44))
                self.listw_evidence.addItem(item)
                self.listw_evidence.setItemWidget(item, roww)

                ek_no += 1

            return

        # ✅ Diğer bölümler: mevcut davranış (rapor_meta_ekler)
        self.listw.clear()

        with DB() as conn:
            cols = {r[1] for r in conn.execute("PRAGMA table_info(rapor_meta_ekler)").fetchall()}

            select_cols = ["id", "DosyaAdi", "DosyaYolu", "Aciklama"]
            if "DosyaAdiGizle" in cols:
                select_cols.append("DosyaAdiGizle")
            else:
                select_cols.append("0 AS DosyaAdiGizle")

            if "GenislikYuzde" in cols:
                select_cols.append("GenislikYuzde")
            else:
                select_cols.append("100 AS GenislikYuzde")

            sql = f"""
                SELECT {",".join(select_cols)}
                FROM rapor_meta_ekler
                WHERE ProjeID=? AND Bolum=?
                ORDER BY id ASC
            """
            rows = conn.execute(sql, (int(self.project_id), str(self.bolum))).fetchall()

        ek_no = 1
        for row in rows:
            # row: id, DosyaAdi, DosyaYolu, Aciklama, DosyaAdiGizle, GenislikYuzde
            db_id, fname, fpath, desc, gizle, w = row

            has_desc = bool(str(desc or "").strip())
            if int(gizle or 0) == 1:
                title = f"Ek-{ek_no}"
            else:
                title = f"Ek-{ek_no} - {fname}"

            item = QListWidgetItem()
            roww = MetaEkRowWidget(int(db_id), title, has_desc, self)

            roww.clicked.connect(lambda mid: self._focus_item(mid, "dosya_ek"))
            roww.edit_desc.connect(self._edit_desc)
            roww.edit_props.connect(self._edit_props)
            roww.deleted.connect(self._delete_item)

            item.setSizeHint(QSize(10, 44))
            self.listw.addItem(item)
            self.listw.setItemWidget(item, roww)

            ek_no += 1

    def _delete_evidence(self, eid: int):
        title_txt = ""
        try:
            with DB() as conn:
                row = conn.execute(
                    "SELECT Baslik FROM rapor_taslagi WHERE id=? AND ProjeID=?",
                    (eid, self.project_id)
                ).fetchone()
                if row and row[0]:
                    title_txt = str(row[0]).strip()
        except Exception:
            title_txt = ""

        msg = f"Bu delili ({title_txt}) silmek istiyor musunuz?" if title_txt else "Bu delili silmek istiyor musunuz?"

        ok = ModernDialog.show_question(
            self,
            "Sil",
            msg,
            "Evet",
            "Hayır"
        )
        if not ok:
            return

        with DB() as conn:
            cur = conn.cursor()
            cur.execute(
                "DELETE FROM rapor_taslagi WHERE id=? AND ProjeID=?",
                (eid, self.project_id)
            )
            conn.commit()

        self.reload()
        self.changed_any.emit()

    def _persist_delil_order(self):
        if self.bolum != "delil_cekmece":
            return

        with DB() as conn:
            s = 10
            for i in range(self.listw.count()):
                it = self.listw.item(i)
                kind, _id = it.data(Qt.ItemDataRole.UserRole) or (None, None)
                if kind == "meta":
                    conn.execute("UPDATE rapor_meta_ekler SET Sira=? WHERE id=? AND ProjeID=?",
                                 (s, int(_id), int(self.project_id)))
                elif kind == "taslak":
                    conn.execute("UPDATE rapor_taslagi SET Sira=? WHERE id=? AND ProjeID=?",
                                 (s, int(_id), int(self.project_id)))
                s += 10
            conn.commit()

        self.changed_any.emit()
        try:
            if hasattr(self.parent(), "refresh_preview"):
                self.parent().refresh_preview()
        except Exception:
            pass

    def _move_evidence(self, eid: int, direction: int):
        """
        rapor_taslagi.Sira swap ile yukarı/aşağı taşır.
        direction: -1 (yukarı), +1 (aşağı)
        """
        try:
            eid = int(eid)
            direction = -1 if int(direction) < 0 else 1
        except Exception:
            return

        with DB() as conn:
            rows = conn.execute("""
                SELECT id, COALESCE(Sira, 999999) AS S
                FROM rapor_taslagi
                WHERE ProjeID=?
                ORDER BY S ASC, id ASC
            """, (int(self.project_id),)).fetchall()

            ids = [r[0] for r in rows]
            sira_map = {r[0]: r[1] for r in rows}

            if eid not in ids:
                return

            idx = ids.index(eid)
            j = idx + direction
            if j < 0 or j >= len(ids):
                return

            other_id = ids[j]

            s1 = sira_map.get(eid, 999999)
            s2 = sira_map.get(other_id, 999999)

            conn.execute("UPDATE rapor_taslagi SET Sira=? WHERE id=?", (int(s2), int(eid)))
            conn.execute("UPDATE rapor_taslagi SET Sira=? WHERE id=?", (int(s1), int(other_id)))
            conn.commit()

        self.reload()
        self.changed_any.emit()
        try:
            if hasattr(self.parent(), "refresh_preview"):
                self.parent().refresh_preview()
        except Exception:
            pass

    def _edit_evidence_desc(self, eid: int):
        try:
            eid = int(eid)
        except Exception:
            return

        with DB() as conn:
            row = conn.execute(
                "SELECT Aciklama FROM rapor_taslagi WHERE id=? AND ProjeID=?",
                (eid, int(self.project_id))
            ).fetchone()
            current_text = row[0] if row else ""

        # ✅ Ekler bölümündekiyle aynı: Zengin metin editörü
        dlg = RichDescriptionDialog(self, current_text, "Delil Açıklaması Düzenle")
        if not dlg.exec():
            return

        new_text = dlg.get_html()

        with DB() as conn:
            conn.execute(
                "UPDATE rapor_taslagi SET Aciklama=? WHERE id=? AND ProjeID=?",
                (new_text, eid, int(self.project_id))
            )
            conn.commit()

        self.reload()
        self.changed_any.emit()
        try:
            if hasattr(self.parent(), "refresh_preview"):
                self.parent().refresh_preview()
        except Exception:
            pass


    def _focus_item(self, db_id: int, bolum: str):
        dlg = self.parent()
        while dlg and not hasattr(dlg, "scroll_to_meta_ek"):
            dlg = dlg.parent()

        if not dlg:
            return

        # ✅ DELİL ÇEKMECESİ
        if bolum == "delil_cekmece":
            dlg.scroll_to_meta_ek(db_id)

        # ✅ DOSYA HAKKINDAKİ EKLER (SADECE META-EK)
        else:
            dlg.scroll_to_metaek_only(db_id)

    def _edit_desc(self, db_id: int):
        with DB() as conn:
            row = conn.execute(
                "SELECT Aciklama FROM rapor_meta_ekler WHERE id=? AND ProjeID=?",
                (db_id, self.project_id)
            ).fetchone()
        current_text = row[0] if row else ""

        dlg = RichDescriptionDialog(self, current_text, "Ek Açıklaması Düzenle")
        if dlg.exec():
            new_text = dlg.get_html()
            with DB() as conn:
                conn.execute(
                    "UPDATE rapor_meta_ekler SET Aciklama=? WHERE id=? AND ProjeID=?",
                    (new_text, db_id, self.project_id)
                )
                conn.commit()
            self.reload()
            self.changed_any.emit()

    def add_from_disk(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Delil Resmi Seç",
            "",
            "Resimler (*.png *.jpg *.jpeg *.bmp *.webp)"
        )
        if not file_path:
            return

        # sadece delil_cekmece için: resimleri evidence_images altına kopyala ve rapor_taslagi'na ekle
        if self.bolum == "delil_cekmece":
            base_dir = os.path.dirname(os.path.abspath(__file__))
            evidence_dir = os.path.join(base_dir, "evidence_images")
            if not os.path.exists(evidence_dir):
                os.makedirs(evidence_dir)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            ext = os.path.splitext(file_path)[1].lower() or ".png"
            safe_name = f"delil_{ts}{ext}"
            dest = os.path.join(evidence_dir, safe_name)

            try:
                shutil.copy2(file_path, dest)
            except Exception as e:
                ModernDialog.show_error(self, "Hata", f"Resim kopyalanamadı:\n{e}")
                return

            try:
                with DB() as conn:
                    cur = conn.cursor()

                    last_order = cur.execute(
                        "SELECT MAX(Sira) FROM rapor_taslagi WHERE ProjeID=?",
                        (int(self.project_id),)
                    ).fetchone()[0]
                    new_order = (last_order or 0) + 1

                    # başlık: dosya adı (istersen burada başka kural koyabilirsin)
                    title = os.path.basename(dest)

                    cols = [r[1] for r in cur.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]

                    keys = ["ProjeID", "GSMNo", "Baslik", "Icerik", "Tur", "Tarih", "Sira"]
                    vals = [
                        int(self.project_id), "", title, dest, "IMAGE",
                        datetime.now().strftime("%d.%m.%Y %H:%M"),
                        int(new_order)
                    ]

                    # delil resimleri için varsayılanlar
                    if "GenislikYuzde" in cols: keys.append("GenislikYuzde"); vals.append(80)
                    if "YukseklikMm" in cols:   keys.append("YukseklikMm");   vals.append(0)
                    if "Hizalama" in cols:      keys.append("Hizalama");      vals.append("center")
                    if "Aciklama" in cols:      keys.append("Aciklama");      vals.append("")
                    if "HtmlIcerik" in cols:    keys.append("HtmlIcerik");    vals.append(None)
                    if "ImagePath" in cols:     keys.append("ImagePath");     vals.append(dest)

                    q_marks = ",".join(["?"] * len(keys))
                    col_names = ",".join(keys)
                    cur.execute(f"INSERT INTO rapor_taslagi ({col_names}) VALUES ({q_marks})", vals)

                    conn.commit()

                self.reload()
                self.changed_any.emit()
                try:
                    if hasattr(self.parent(), "refresh_preview"):
                        self.parent().refresh_preview()
                except Exception:
                    pass

                ModernDialog.show_success(self, "Eklendi", "Delil resim dosyası delil listesine eklendi.")
                return

            except Exception as e:
                ModernDialog.show_error(self, "Hata", str(e))
                return

        # diğer bölümler: mevcut davranış (rapor_meta_ekler)
        try:
            self._ensure_evidence_dir()

            base_dir = os.path.dirname(os.path.abspath(__file__))
            evidence_dir = os.path.join(base_dir, "evidence_images")
            if not os.path.exists(evidence_dir):
                os.makedirs(evidence_dir)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            ext = os.path.splitext(file_path)[1].lower() or ".png"
            safe_name = f"ek_{ts}{ext}"
            dest = os.path.join(evidence_dir, safe_name)

            shutil.copy2(file_path, dest)

            with DB() as conn:
                cols = {r[1] for r in conn.execute("PRAGMA table_info(rapor_meta_ekler)").fetchall()}
                has_hide = "DosyaAdiGizle" in cols
                has_w = "GenislikYuzde" in cols

                if has_hide and has_w:
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama, DosyaAdiGizle, GenislikYuzde) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (int(self.project_id), str(self.bolum), os.path.basename(dest), dest, "", 0, 100),
                    )
                elif has_hide and (not has_w):
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama, DosyaAdiGizle) "
                        "VALUES (?,?,?,?,?,?)",
                        (int(self.project_id), str(self.bolum), os.path.basename(dest), dest, "", 0),
                    )
                elif (not has_hide) and has_w:
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama, GenislikYuzde) "
                        "VALUES (?,?,?,?,?,?)",
                        (int(self.project_id), str(self.bolum), os.path.basename(dest), dest, "", 100),
                    )
                else:
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama) "
                        "VALUES (?,?,?,?,?)",
                        (int(self.project_id), str(self.bolum), os.path.basename(dest), dest, ""),
                    )
                conn.commit()

            self.reload()
            self.changed_any.emit()

            try:
                if hasattr(self.parent(), "refresh_preview"):
                    self.parent().refresh_preview()
            except Exception:
                pass

        except Exception as e:
            ModernDialog.show_error(self, "Hata", str(e))


    def _delete_item(self, db_id: int):
        # ✅ Silmeden önce program tarzı onay sor (dosya adını da göster)
        try:
            bolum = (getattr(self, "bolum", "") or "").strip()
            if not bolum:
                bolum = "Ek"

            dosya_adi = ""
            dosya_yolu = ""

            with DB() as conn:
                row = conn.execute(
                    "SELECT COALESCE(DosyaAdi,''), COALESCE(DosyaYolu,''), COALESCE(Bolum,'') "
                    "FROM rapor_meta_ekler WHERE id=?",
                    (db_id,)
                ).fetchone()

            if row:
                dosya_adi = (row[0] or "").strip()
                dosya_yolu = (row[1] or "").strip()
                if not bolum:
                    bolum = (row[2] or "").strip() or "Ek"

            # DosyaAdi boşsa dosya yolundan türet
            if not dosya_adi and dosya_yolu:
                try:
                    dosya_adi = os.path.basename(dosya_yolu)
                except Exception:
                    dosya_adi = ""

            # Kullanıcıya gösterilecek isim
            shown = dosya_adi if dosya_adi else bolum

            ok = ModernDialog.show_question(
                self,
                f"{bolum} Sil",
                f"Seçili kayıt ({shown}) silinecek. Emin misiniz?",
                yes_btn="Evet",
                no_btn="Hayır"
            )
            if not ok:
                return

        except Exception:
            return

        # ✅ Silme işlemi
        row = None
        with DB() as conn:
            row = conn.execute(
                "SELECT DosyaYolu FROM rapor_meta_ekler WHERE id=?",
                (db_id,)
            ).fetchone()
            conn.execute("DELETE FROM rapor_meta_ekler WHERE id=?", (db_id,))

        try:
            if row and row[0] and os.path.exists(row[0]):
                os.remove(row[0])
        except Exception:
            pass

        self.reload()
        self.changed_any.emit()


    def _update_desc(self, db_id: int, desc: str):
        with DB() as conn:
            conn.execute("UPDATE rapor_meta_ekler SET Aciklama=? WHERE id=?", (desc, db_id))
        self.changed_any.emit()


class ReportCenterDialog(QDialog):
    """
    Rapor ve Tasarım Merkezi
    - Sol: metin editörleri + delil çekmecesi + devre dışı checkbox’lar + marginler
    - Sağ: QWebEngineView canlı önizleme
    """

    def __init__(self, parent, project_id: int):
        super().__init__(parent)
        self.setWindowTitle("Rapor ve Tasarım Merkezi")
        self.setWindowState(Qt.WindowState.WindowMaximized)

        project_id = int(project_id)  # burada patlaması doğru
        self.project_id = project_id
        self.builder = ReportHtmlBuilder(project_id)

        self.preview = None
        self.ed_gorev = None
        self.ed_dosya = None
        self.ed_genel = None
        self.ed_deg = None
        self.ed_sonuc = None
        self.blocks_table = None

        self.cb_gorev = None
        self.cb_dosya = None
        self.cb_genel = None
        self.cb_delillers = None
        self.cb_deg = None
        self.cb_sonuc = None
        self.cb_taraflar = None

        # ✅ KRİTİK: açılışta DB’ye yazmayı kilitle
        self._loading_meta = True
        self._meta_ready = False

        self._build_ui()
        self._load_meta_to_ui()          # editörler burada dolar
        self.load_blocks_into_table()

        # autosave bağlantılarını meta yüklendikten SONRA bağla
        self._wire_meta_autosave()

        self._loading_meta = False
        self._meta_ready = True

        self.refresh_preview()           # artık güvenli
        self.loader = LoadingOverlay(self)

    def _wire_meta_autosave(self):
        if getattr(self, "_meta_autosave_wired", False):
            return
        self._meta_autosave_wired = True

        for ed in (self.ed_gorev, self.ed_dosya, self.ed_genel, self.ed_deg, self.ed_sonuc):
            if ed is None:
                continue
            ed.textChanged.connect(self.save_meta_silent)

    def closeEvent(self, event):
        try:
            self.save_meta_silent()   # son değişiklik DB’ye gitsin
        except Exception:
            pass
        super().closeEvent(event)

    def _norm_pid(self):
        pid = getattr(self, "project_id", None)
        if pid is None:
            return None
        # her yerde aynı: önce int dene
        try:
            return int(str(pid).strip())
        except Exception:
            return str(pid).strip()

    def _build_ui(self):
        # --- Ana yerleşim ---
        main = QHBoxLayout(self)
        main.setContentsMargins(10, 10, 10, 10)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        main.addWidget(splitter)

        # --- SOL PANEL ---
        left_panel = QWidget()
        left_panel.setMinimumWidth(520)

        left_outer = QVBoxLayout(left_panel)
        left_outer.setContentsMargins(0, 0, 0, 0)
        left_outer.setSpacing(10)

        # Sol içerik scroll içinde olsun
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        left_container = QWidget()
        left = QVBoxLayout(left_container)              # ✅ left artık TANIMLI
        left.setContentsMargins(0, 0, 0, 0)
        left.setSpacing(10)

        scroll.setWidget(left_container)
        left_outer.addWidget(scroll)

        # --- Sağ panel (önizleme) ---
        right_panel = QWidget()
        right = QVBoxLayout(right_panel)
        right.setContentsMargins(0, 0, 0, 0)
        right.setSpacing(10)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        try:
            splitter.setSizes([560, 1200])
        except Exception:
            pass

        # --- Meta editörleri (veri deposu) ---
        self.ed_gorev = QTextEdit()
        self.ed_dosya = QTextEdit()
        self.ed_genel = QTextEdit()
        self.ed_deg = QTextEdit()
        self.ed_sonuc = QTextEdit()

        # --- Kompakt kart üreticisi ---
        def _make_card(title: str, cb_attr_name: str, on_edit, on_reset=None):
            card = QFrame()
            card.setStyleSheet(
                "QFrame{background:#ffffff;border:1px solid #dcdcdc;border-radius:8px;}"
            )
            lay = QVBoxLayout(card)
            lay.setContentsMargins(10, 8, 10, 8)
            lay.setSpacing(6)

            header = QHBoxLayout()
            header.setContentsMargins(0, 0, 0, 0)
            header.setSpacing(8)

            lbl = QLabel(title)
            lbl.setStyleSheet("font-weight:800; color:#2c3e50; font-size:12px; text-transform:uppercase;")
            header.addWidget(lbl)
            header.addStretch(1)

            btn_edit = QPushButton("📝 Düzenle")
            btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_edit.setStyleSheet(
                "QPushButton{background:#ecf0f1;border:1px solid #bdc3c7;border-radius:6px;padding:6px 10px;font-weight:700;color:#7f8c8d;}"
                "QPushButton:hover{background:#dfe6e9;}"
            )
            btn_edit.clicked.connect(on_edit)
            header.addWidget(btn_edit)

            if on_reset is not None:
                btn_reset = QPushButton("↩ Varsayılan")
                btn_reset.setCursor(Qt.CursorShape.PointingHandCursor)
                btn_reset.setStyleSheet(
                    "QPushButton{background:transparent;border:none;color:#7f8c8d;font-weight:700;}"
                    "QPushButton:hover{color:#2c3e50;}"
                )
                btn_reset.clicked.connect(on_reset)
                header.addWidget(btn_reset)

            cb = QCheckBox("Gizle")
            setattr(self, cb_attr_name, cb)
            cb.stateChanged.connect(self.refresh_preview)
            header.addWidget(cb)

            lay.addLayout(header)

            summary = QLabel("-")
            summary.setWordWrap(True)
            summary.setStyleSheet("color:#7f8c8d; font-size:11px;")
            lay.addWidget(summary)
            return card, summary

        # --- Taraflar ---
        card_taraf = QFrame()
        card_taraf.setStyleSheet("QFrame{background:#ffffff;border:1px solid #dcdcdc;border-radius:8px;}")
        l_tar = QHBoxLayout(card_taraf)
        l_tar.setContentsMargins(10, 8, 10, 8)
        lbl_tar = QLabel("TARAFLAR")
        lbl_tar.setStyleSheet("font-weight:800; color:#2c3e50; font-size:12px; text-transform:uppercase;")
        l_tar.addWidget(lbl_tar)
        l_tar.addStretch(1)
        self.cb_taraflar = QCheckBox("Gizle")
        self.cb_taraflar.stateChanged.connect(self.refresh_preview)
        l_tar.addWidget(self.cb_taraflar)
        left.addWidget(card_taraf)

        # --- Bölüm kartları ---
        self.lbl_sum_gorev = None
        self.lbl_sum_dosya = None
        self.lbl_sum_genel = None
        self.lbl_sum_deg = None
        self.lbl_sum_sonuc = None

        card_g, self.lbl_sum_gorev = _make_card(
            "1. GÖREVLENDİRME", "cb_gorev",
            on_edit=lambda: self._open_meta_editor("gorev"),
            on_reset=self.reset_default_gorev,
        )
        left.addWidget(card_g)

        card_d, self.lbl_sum_dosya = _make_card(
            "2. DOSYA HAKKINDA", "cb_dosya",
            on_edit=lambda: self._open_meta_editor("dosya"),
            on_reset=self.reset_default_dosya,
        )
        left.addWidget(card_d)

        # --- 2. DOSYA HAKKINDA EKLER ---
        dosya_ek_frame = QFrame()
        dosya_ek_frame.setStyleSheet("QFrame{background:#ffffff;border:1px solid #dcdcdc;border-radius:8px;}")
        dosya_ek_lay = QVBoxLayout(dosya_ek_frame)
        dosya_ek_lay.setContentsMargins(10, 8, 10, 8)
        dosya_ek_lay.setSpacing(6)

        lbl_ek = QLabel("DOSYA HAKKINDA EKLER")
        lbl_ek.setStyleSheet("font-weight:800; color:#2c3e50; font-size:12px; text-transform:uppercase;")
        dosya_ek_lay.addWidget(lbl_ek)

        self.pnl_dosya_ekler = MetaEkPanel(self.project_id, "dosya_hakkinda", self)
        self.pnl_dosya_ekler.changed_any.connect(self.refresh_preview)
        dosya_ek_lay.addWidget(self.pnl_dosya_ekler, 1)

        left.addWidget(dosya_ek_frame)

        card_genel, self.lbl_sum_genel = _make_card(
            "3. HTS DOSYA VE ABONE BİLGİLERİ", "cb_genel",
            on_edit=lambda: self._open_meta_editor("genel"),
            on_reset=self.reset_default_genel,
        )
        left.addWidget(card_genel)

        # --- 4. DELİL ÇEKMECESİ (MetaEkPanel) ---
        card_delil = QFrame()
        card_delil.setStyleSheet("QFrame{background:#ffffff;border:1px solid #dcdcdc;border-radius:8px;}")
        l_delil = QVBoxLayout(card_delil)
        l_delil.setContentsMargins(10, 8, 10, 8)
        l_delil.setSpacing(6)

        head_delil = QHBoxLayout()
        head_delil.setContentsMargins(0, 0, 0, 0)
        lbl_delil = QLabel("4. DELİL ÇEKMECESİ")
        lbl_delil.setStyleSheet("font-weight:800; color:#2c3e50; font-size:12px; text-transform:uppercase;")
        head_delil.addWidget(lbl_delil)
        head_delil.addStretch(1)

        self.cb_delillers = QCheckBox("Gizle")
        self.cb_delillers.stateChanged.connect(self.refresh_preview)
        head_delil.addWidget(self.cb_delillers)

        l_delil.addLayout(head_delil)

        # ✅ TEK panel: delil_cekmece (deliller diye ikinci kez ezme yok)
        self.pnl_delil_ekler = MetaEkPanel(self.project_id, "delil_cekmece", self)
        self.pnl_delil_ekler.changed_any.connect(self.refresh_preview)
        l_delil.addWidget(self.pnl_delil_ekler)

        left.addWidget(card_delil)

        card_deg, self.lbl_sum_deg = _make_card(
            "5. DEĞERLENDİRME", "cb_deg",
            on_edit=lambda: self._open_meta_editor("deg"),
        )
        left.addWidget(card_deg)

        card_s, self.lbl_sum_sonuc = _make_card(
            "6. SONUÇ", "cb_sonuc",
            on_edit=lambda: self._open_meta_editor("sonuc"),
        )
        left.addWidget(card_s)

        left.addStretch(1)

        # --- Sol alt aksiyon barı (scroll DIŞI) ---
        action_layout = QHBoxLayout()
        action_layout.setSpacing(10)
        action_layout.setContentsMargins(0, 0, 0, 0)

        btn_reset_all = QPushButton("Sıfırla")
        btn_reset_all.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_reset_all.setStyleSheet("background-color:#c0392b; color:white; font-weight:bold; border-radius:6px; padding:10px;")
        btn_reset_all.clicked.connect(self.reset_all_defaults)
        action_layout.addWidget(btn_reset_all)

        btn_save_text = QPushButton("💾 Kaydet")
        btn_save_text.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_save_text.setToolTip("Manuel Kaydet")
        btn_save_text.setStyleSheet("background-color:#2980b9; color:white; font-weight:bold; border-radius:6px; padding:10px;")
        btn_save_text.clicked.connect(self.save_meta)
        action_layout.addWidget(btn_save_text)

        btn_style = QPushButton("🎨 Ayarlar")
        btn_style.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_style.setStyleSheet("background-color:#8e44ad; color: white; font-weight: bold; border-radius: 6px; padding: 10px;")
        btn_style.clicked.connect(self.open_style_editor)
        action_layout.addWidget(btn_style)

        left_outer.addLayout(action_layout)

        # --- SAĞ: Önizleme ---
        if getattr(self, "preview", None) is None:
            self.preview = QWebEngineView()
        self.preview.setStyleSheet("border: 1px solid #bdc3c7; background:#ffffff;")
        right.addWidget(self.preview, 1)
        self.web = self.preview  # uyumluluk

        right_btn_layout = QHBoxLayout()
        self.btn_refresh_preview = QPushButton("🔄 Önizlemeyi Güncelle")
        self.btn_refresh_preview.setStyleSheet("font-weight:bold; padding:10px; border-radius:6px;")
        self.btn_refresh_preview.clicked.connect(self.refresh_preview)
        right_btn_layout.addWidget(self.btn_refresh_preview)

        btn_pdf = QPushButton("📄 PDF Olarak Kaydet")
        btn_pdf.setStyleSheet("background-color:#27ae60; color:white; font-weight:bold; padding:10px; border-radius:6px;")
        btn_pdf.clicked.connect(self.export_pdf)
        right_btn_layout.addWidget(btn_pdf)

        right.addLayout(right_btn_layout)

    def scroll_to_metaek_only(self, metaek_id: int):
        js = f"""
        var el = document.getElementById('metaek-{metaek_id}');
        if (el) {{
            el.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
            el.style.outline = '3px solid #4CAF50';
            setTimeout(() => el.style.outline = '', 1200);
        }}
        """
        self.web.page().runJavaScript(js)

    def add_evidence_image_from_disk(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Delil Resmi Seç", "", "Resimler (*.png *.jpg *.jpeg *.bmp *.webp)"
        )
        if not file_path:
            return

        # kopyala
        base_dir = os.path.dirname(os.path.abspath(__file__))
        evidence_dir = os.path.join(base_dir, "evidence_images")
        if not os.path.exists(evidence_dir):
            os.makedirs(evidence_dir)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        ext = os.path.splitext(file_path)[1].lower() or ".png"
        safe_name = f"delil_{ts}{ext}"
        dest = os.path.join(evidence_dir, safe_name)

        try:
            shutil.copy2(file_path, dest)
        except Exception as e:
            ModernDialog.show_error(self, "Hata", f"Resim kopyalanamadı:\n{e}")
            return

        # rapor_taslagi insert
        try:
            with DB() as conn:
                cols = {r[1] for r in conn.execute("PRAGMA table_info(rapor_meta_ekler)").fetchall()}

                # Kolonlar eski sürümde yoksa yine de insert edebilmek için
                has_hide = "DosyaAdiGizle" in cols
                has_w    = "GenislikYuzde" in cols

                if has_hide and has_w:
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama, DosyaAdiGizle, GenislikYuzde) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (int(self.project_id), "delil_cekmece", os.path.basename(dest), dest, "", 0, 80),
                    )
                elif has_hide and (not has_w):
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama, DosyaAdiGizle) "
                        "VALUES (?,?,?,?,?,?)",
                        (int(self.project_id), "delil_cekmece", os.path.basename(dest), dest, "", 0),
                    )
                elif (not has_hide) and has_w:
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama, GenislikYuzde) "
                        "VALUES (?,?,?,?,?,?)",
                        (int(self.project_id), "delil_cekmece", os.path.basename(dest), dest, "", 80),
                    )
                else:
                    conn.execute(
                        "INSERT INTO rapor_meta_ekler (ProjeID, Bolum, DosyaAdi, DosyaYolu, Aciklama) "
                        "VALUES (?,?,?,?,?)",
                        (int(self.project_id), "delil_cekmece", os.path.basename(dest), dest, ""),
                    )

                conn.commit()

            # Delil çekmecesi panelini yenile (varsa)
            try:
                if hasattr(self, "meta_ek_panel") and self.meta_ek_panel:
                    self.meta_ek_panel.reload()
            except Exception:
                pass

            self.refresh_preview()
            ModernDialog.show_success(self, "Eklendi", "Delil dosyası delil çekmecesine eklendi.")
        except Exception as e:
            ModernDialog.show_error(self, "Hata", str(e))

    def scroll_to_meta_ek(self, db_id: int):
        try:
            target = getattr(self, "web", None) or getattr(self, "preview", None)
            if target is None:
                return

            js = f"""
            (function() {{
                var id = {int(db_id)};
                var el = document.getElementById('evidence-' + id);
                if(!el) el = document.getElementById('metaek-' + id);  // ✅ fallback
                if(!el) return;
                el.scrollIntoView({{behavior:'smooth', block:'center'}});
                el.style.outline='2px solid #e74c3c';
                setTimeout(function(){{ el.style.outline=''; }}, 1200);
            }})();"""

            target.page().runJavaScript(js)
        except Exception as e:
            print("scroll_to_meta_ek hata:", e)

    def create_rich_editor(self, min_height=80):
        """Üzerinde Kalın/İtalik/AltıÇizili butonları olan editör döndürür."""
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(5)
        toolbar.setContentsMargins(0, 0, 0, 0)

        btn_style = """
            QPushButton { 
                background-color: #ecf0f1; border: 1px solid #bdc3c7; 
                border-radius: 3px; font-weight: bold; width: 24px; height: 24px;
                font-family: 'Times New Roman'; font-size: 14px;
            }
            QPushButton:hover { background-color: #dfe6e9; }
            QPushButton:checked { background-color: #3498db; color: white; border: 1px solid #2980b9; }
        """

        editor = QTextEdit()
        editor.setMinimumHeight(min_height)
        editor.setStyleSheet("background-color: white; border: 1px solid #bdc3c7;")

        def set_bold():
            fmt = editor.currentCharFormat()
            fmt.setFontWeight(QFont.Weight.Bold if fmt.fontWeight() != QFont.Weight.Bold else QFont.Weight.Normal)
            editor.mergeCurrentCharFormat(fmt)
            editor.setFocus()

        def set_italic():
            fmt = editor.currentCharFormat()
            fmt.setFontItalic(not fmt.fontItalic())
            editor.mergeCurrentCharFormat(fmt)
            editor.setFocus()

        def set_underline():
            fmt = editor.currentCharFormat()
            fmt.setFontUnderline(not fmt.fontUnderline())
            editor.mergeCurrentCharFormat(fmt)
            editor.setFocus()

        btn_b = QPushButton("K")
        btn_b.setToolTip("Kalın (Bold)")
        btn_b.setStyleSheet(btn_style)
        btn_b.clicked.connect(set_bold)

        btn_i = QPushButton("I")
        btn_i.setToolTip("İtalik")
        btn_i.setStyleSheet(btn_style + "QPushButton { font-style: italic; }")
        btn_i.clicked.connect(set_italic)

        btn_u = QPushButton("A")
        btn_u.setToolTip("Altı Çizili")
        btn_u.setStyleSheet(btn_style + "QPushButton { text-decoration: underline; }")
        btn_u.clicked.connect(set_underline)

        toolbar.addWidget(btn_b)
        toolbar.addWidget(btn_i)
        toolbar.addWidget(btn_u)
        toolbar.addStretch()

        layout.addLayout(toolbar)
        layout.addWidget(editor)

        return container, editor

    def _build_blocks_table(self, left_layout: QVBoxLayout):
        self.blocks_table = QTableWidget()
        self.blocks_table.setColumnCount(5)
        self.blocks_table.setHorizontalHeaderLabels(["SIRA", "BAŞLIK", "TÜR", "GENİŞLİK %", "AÇIKLAMA"])

        h = self.blocks_table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed); self.blocks_table.setColumnWidth(0, 50)
        h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        h.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed); self.blocks_table.setColumnWidth(2, 70)
        h.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed); self.blocks_table.setColumnWidth(3, 100)
        h.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed); self.blocks_table.setColumnWidth(4, 110)

        self.blocks_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.blocks_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.blocks_table.setMinimumHeight(180)
        self.blocks_table.itemClicked.connect(self.scroll_to_evidence)
        left_layout.addWidget(self.blocks_table)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(5)

        mini_btn_style = """
            QPushButton {
                background-color: #ecf0f1; border: 1px solid #bdc3c7; border-radius: 4px;
                padding: 5px 10px; font-weight: bold; color: #2c3e50; font-size: 12px;
            }
            QPushButton:hover { background-color: #d5dbdb; }
        """

        btn_up = QPushButton("▲"); btn_up.setToolTip("Yukarı Taşı"); btn_up.setFixedWidth(40)
        btn_up.setStyleSheet(mini_btn_style); btn_up.clicked.connect(lambda: self.move_block(-1))
        btn_row.addWidget(btn_up)

        btn_down = QPushButton("▼"); btn_down.setToolTip("Aşağı Taşı"); btn_down.setFixedWidth(40)
        btn_down.setStyleSheet(mini_btn_style); btn_down.clicked.connect(lambda: self.move_block(1))
        btn_row.addWidget(btn_down)

        btn_save = QPushButton("Kaydet")
        btn_save.setStyleSheet(mini_btn_style + "QPushButton { color: #2980b9; }")
        btn_save.clicked.connect(self.save_block_edits)
        btn_row.addWidget(btn_save)

        self.btn_delete_block = QPushButton("Sil")
        self.btn_delete_block.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_delete_block.setMinimumHeight(30)
        self.btn_delete_block.setStyleSheet(mini_btn_style + "QPushButton { color: #c0392b; }")
        self.btn_delete_block.clicked.connect(self.delete_selected_block)
        btn_row.addWidget(self.btn_delete_block)

        btn_add_img = QPushButton("Delil Resmi Ekle (Yerel Disk)")
        btn_add_img.clicked.connect(self.add_evidence_image_from_disk)
        left_layout.addLayout(btn_row)

    def reset_default_dosya(self):
        self.ed_dosya.setHtml("")
        self.refresh_preview()

    def reset_default_genel(self):
        self.ed_genel.setHtml(self.builder.build_default_genel_bilgi())
        self.refresh_preview()

    def reset_all_defaults(self):
        self.ed_gorev.setHtml(self.builder.build_default_gorevlendirme())
        self.ed_dosya.setHtml("")
        self.ed_genel.setHtml(self.builder.build_default_genel_bilgi())
        self.ed_deg.setHtml("")
        self.ed_sonuc.setHtml("")
        self._refresh_meta_summaries()
        self.refresh_preview()

    def _html_to_summary(self, html_text: str, max_len: int = 180) -> str:
        try:
            doc = QTextDocument()
            doc.setHtml(html_text or "")
            plain = doc.toPlainText()
        except Exception:
            plain = re.sub(r"<[^>]+>", " ", html_text or "")

        plain = re.sub(r"\s+", " ", plain).strip()
        if not plain:
            return "(Metin boş)"
        if len(plain) <= max_len:
            return plain
        return plain[:max_len].rstrip() + "…"

    def _refresh_meta_summaries(self):
        if getattr(self, "lbl_sum_gorev", None):
            self.lbl_sum_gorev.setText(self._html_to_summary(self.ed_gorev.toHtml()))
        if getattr(self, "lbl_sum_dosya", None):
            self.lbl_sum_dosya.setText(self._html_to_summary(self.ed_dosya.toHtml()))
        if getattr(self, "lbl_sum_genel", None):
            self.lbl_sum_genel.setText(self._html_to_summary(self.ed_genel.toHtml()))
        if getattr(self, "lbl_sum_deg", None):
            self.lbl_sum_deg.setText(self._html_to_summary(self.ed_deg.toHtml()))
        if getattr(self, "lbl_sum_sonuc", None):
            self.lbl_sum_sonuc.setText(self._html_to_summary(self.ed_sonuc.toHtml()))

    def _open_meta_editor(self, key: str):
        mapping = {
            "gorev": ("1. GÖREVLENDİRME", self.ed_gorev),
            "dosya": ("2. DOSYA HAKKINDA", self.ed_dosya),
            "genel": ("3. HTS DOSYA VE ABONE BİLGİLERİ", self.ed_genel),
            "deg": ("5. DEĞERLENDİRME", self.ed_deg),
            "sonuc": ("6. SONUÇ", self.ed_sonuc),
        }
        if key not in mapping:
            return

        title, editor = mapping[key]
        dlg = RichDescriptionDialog(self, editor.toHtml(), f"{title} Metni Düzenle")
        if dlg.exec():
            editor.setHtml(dlg.get_html())
            self.save_meta_silent()
            self._refresh_meta_summaries()
            self.refresh_preview()

    def _normalize_html_for_store(self, html: str) -> str:
        """
        QTextEdit.toHtml() boşken bile 600+ karakterlik HTML iskeleti üretir.
        DB'ye bunu yazarsan açılışta 'boş' gibi davranıp standart metinlere dönme / kaybolma olur.
        Bu fonksiyon boş içeriği gerçekten "" yapar.
        """
        try:
            doc = QTextDocument()
            doc.setHtml(html or "")
            if doc.toPlainText().strip() == "":
                return ""
        except Exception:
            if (html or "").strip() == "":
                return ""
        return html or ""

    def _load_meta_to_ui(self):
        self._loading_meta = True
        try:
            meta = self.builder._fetch_meta()
            if not meta:
                self.builder._ensure_meta_row()
                meta = self.builder._fetch_meta()

            def _html_is_effectively_empty(h: str) -> bool:
                try:
                    doc = QTextDocument()
                    doc.setHtml(h or "")
                    return doc.toPlainText().strip() == ""
                except Exception:
                    return (h or "").strip() == ""

            def mget(idx, default=""):
                if not meta or idx >= len(meta):
                    return default
                v = meta[idx]
                return default if v is None else v

            gorev_db = mget(1, "")
            dosya_db = mget(2, "")
            genel_db = mget(3, "")
            deg_db   = mget(4, "")
            sonuc_db = mget(5, "")

            gorev_html = self.builder.build_default_gorevlendirme() if _html_is_effectively_empty(gorev_db) else gorev_db
            genel_html = self.builder.build_default_genel_bilgi()     if _html_is_effectively_empty(genel_db) else genel_db

            # setHtml sırasında autosave tetiklenmesin
            for ed in (self.ed_gorev, self.ed_dosya, self.ed_genel, self.ed_deg, self.ed_sonuc):
                ed.blockSignals(True)
            try:
                self.ed_gorev.setHtml(gorev_html)
                self.ed_dosya.setHtml(dosya_db or "")
                self.ed_genel.setHtml(genel_html)
                self.ed_deg.setHtml(deg_db or "")
                self.ed_sonuc.setHtml(sonuc_db or "")
            finally:
                for ed in (self.ed_gorev, self.ed_dosya, self.ed_genel, self.ed_deg, self.ed_sonuc):
                    ed.blockSignals(False)

            self._refresh_meta_summaries()
            try:
                self.pnl_dosya_ekler.reload()
            except Exception:
                pass

        finally:
            self._loading_meta = False

    def save_meta(self):
        pid = self._norm_pid()
        if pid is None:
            return

        try:
            self.builder._ensure_meta_row()

            g  = self._normalize_html_for_store(self.ed_gorev.toHtml())
            d  = self._normalize_html_for_store(self.ed_dosya.toHtml())
            ge = self._normalize_html_for_store(self.ed_genel.toHtml())
            de = self._normalize_html_for_store(self.ed_deg.toHtml())
            s  = self._normalize_html_for_store(self.ed_sonuc.toHtml())

            with DB() as conn:
                conn.execute("""
                    UPDATE rapor_meta SET
                        GorevlendirmeMetni=?,
                        DosyaHakkindaMetni=?,
                        GenelBilgilendirmeMetni=?,
                        DegerlendirmeMetni=?,
                        SonucMetni=?,
                        GuncellemeTarihi=CURRENT_TIMESTAMP
                    WHERE ProjeID=?
                """, (g, d, ge, de, s, pid))
                conn.commit()

            ModernDialog.show_success(self, "Kaydedildi", "Rapor metinleri kaydedildi.")
            self.refresh_preview()

        except Exception as e:
            ModernDialog.show_error(self, "Hata", str(e))

    def _on_dosya_ek_changed(self):
        # metin değişmiyor ama rapor HTML değişecek
        self.refresh_preview()

    def reset_default_gorev(self):
        self.ed_gorev.setHtml(self.builder.build_default_gorevlendirme())
        self.refresh_preview()

    def _rapor_taslagi_pk_col(self, conn):
        cols = [r[1] for r in conn.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]
        return "id" if "id" in cols else "rowid"

    def _rt_select_id_expr(self, conn):
        return "id" if self._rapor_taslagi_pk_col(conn) == "id" else "rowid AS id"

    def load_blocks_into_table(self):
        if getattr(self, "blocks_table", None) is None:
            return
        # Tabloyu temizle
        self.blocks_table.setRowCount(0)

        # Verileri çek
        with DB() as conn:
            # Sira'ya göre sıralı çekiyoruz
            rows = conn.execute(
                "SELECT id, Sira, Baslik, Tur, GenislikYuzde, Aciklama FROM rapor_taslagi WHERE ProjeID=? ORDER BY Sira ASC",
                (self.project_id,)
            ).fetchall()

        self.blocks_table.setRowCount(len(rows))

        for i, r in enumerate(rows):
            bid, sira, baslik, tur, gen, acik = r

            # Harf sıralaması (Görsel)
            harf = chr(97 + i)

            # --- KRİTİK DÜZELTME: ID'yi gizli bir veriye (UserRole) saklıyoruz ---
            # 0. Kolon (Sıra) hücresini oluştururken ID'yi içine gömüyoruz.
            item_sira = QTableWidgetItem(f"{harf}.")
            item_sira.setData(Qt.ItemDataRole.UserRole, bid) # <-- GİZLİ ID BURADA
            # --------------------------------------------------------------------

            self.blocks_table.setItem(i, 0, item_sira)
            self.blocks_table.setItem(i, 1, QTableWidgetItem(str(baslik)))
            self.blocks_table.setItem(i, 2, QTableWidgetItem(str(tur)))
            self.blocks_table.setItem(i, 3, QTableWidgetItem(f"%{gen}"))
            self.blocks_table.setItem(i, 4, QTableWidgetItem(str(acik)[:50]))

    def save_meta_silent(self):
        """Kullanıcı yazı yazdıkça arka planda sessizce kaydeder."""
        if getattr(self, "_loading_meta", False):
            return
        if not getattr(self, "_meta_ready", False):
            return

        pid = self._norm_pid()
        if pid is None:
            return

        try:
            # satır yoksa defaultlarla oluştur
            self.builder._ensure_meta_row()

            with DB() as conn:
                cur = conn.cursor()

                # ✅ AÇILIŞ KORUMASI:
                # Editörler tamamen boşsa, default satırı "boş" ile ezmeyelim.
                # (Kullanıcı zaten yazınca dolacak ve kaydolacak.)
                gorev_html = self.ed_gorev.toHtml() or ""
                genel_html = self.ed_genel.toHtml() or ""

                # QTextEdit boşken "<p></p>" gibi dönebilir; düz metin kontrolü yapalım
                def _plain_len(ed: QTextEdit) -> int:
                    return len((ed.toPlainText() or "").strip())

                if _plain_len(self.ed_gorev) == 0 and _plain_len(self.ed_genel) == 0 and \
                   _plain_len(self.ed_dosya) == 0 and _plain_len(self.ed_deg) == 0 and _plain_len(self.ed_sonuc) == 0:
                    return

                cur.execute("""
                    UPDATE rapor_meta SET
                        GorevlendirmeMetni=?,
                        DosyaHakkindaMetni=?,
                        GenelBilgilendirmeMetni=?,
                        DegerlendirmeMetni=?,
                        SonucMetni=?,
                        GuncellemeTarihi=CURRENT_TIMESTAMP
                    WHERE ProjeID=?
                """, (
                    self.ed_gorev.toHtml(),
                    self.ed_dosya.toHtml(),
                    self.ed_genel.toHtml(),
                    self.ed_deg.toHtml(),
                    self.ed_sonuc.toHtml(),
                    pid
                ))

                if cur.rowcount == 0:
                    raise RuntimeError(f"rapor_meta UPDATE 0 satır etkiledi. ProjeID eşleşmiyor olabilir. pid={pid!r}")

                conn.commit()

        except Exception:
            import traceback
            print("save_meta_silent HATA:")
            print(traceback.format_exc())

    def open_aciklama_editor(self, bid, current_text):
        """Delil açıklamasını düzenlemek için ZENGİN METİN editörünü açar."""

        dlg = RichDescriptionDialog(self, current_text, "Delil Açıklaması Düzenle")

        if dlg.exec():
            new_text = dlg.get_html()

            try:
                with DB() as conn:
                    pk = self._rapor_taslagi_pk_col(conn)
                    conn.execute(f"UPDATE rapor_taslagi SET Aciklama=? WHERE {pk}=? AND ProjeID=?",
                                 (new_text, bid, self.project_id))
                    conn.commit()

                self.load_blocks_into_table()
                self.refresh_preview()
                ModernDialog.show_success(self, "Güncellendi", "Açıklama kaydedildi.")

            except Exception as e:
                ModernDialog.show_error(self, "Hata", str(e))

    def scroll_to_evidence(self, item):
        """Tablodan tıklanan satırın ID'sini alıp HTML'de o ID'ye scroll eder."""
        row = item.row()

        # 0. Kolondaki (Sıra) hücreye erişip gizli ID'yi (UserRole) alıyoruz
        sira_item = self.blocks_table.item(row, 0)
        if not sira_item:
            return

        # Gizli ID'yi çek
        db_id = sira_item.data(Qt.ItemDataRole.UserRole)

        if not db_id:
            return

        # JavaScript ile o ID'ye git
        # Hem 'evidence-ID' hem 'metaek-ID' dener (garanti olsun diye)
        js = f"""
        (function() {{
            var el = document.getElementById('evidence-{db_id}');
            if (!el) el = document.getElementById('metaek-{db_id}');
            
            if (el) {{
                el.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                
                // Vurgu efekti (kırmızı çerçeve)
                var oldOutline = el.style.outline;
                el.style.outline = '3px solid #e74c3c';
                el.style.transition = 'outline 0.3s';
                
                setTimeout(function(){{
                    el.style.outline = oldOutline;
                }}, 1500);
            }} else {{
                console.log('Element bulunamadı: evidence-{db_id}');
            }}
        }})();
        """

        # Web görünümü nesnesini bul (self.preview veya self.web olabilir)
        target_web = getattr(self, "preview", None) or getattr(self, "web", None)
        if target_web:
            target_web.page().runJavaScript(js)

    def _rapor_taslagi_pk_col(self, conn):
        cols = [r[1] for r in conn.execute("PRAGMA table_info(rapor_taslagi)").fetchall()]
        return "id" if "id" in cols else "rowid"

    def _rt_select_id_expr(self, conn):
        return "id" if self._rapor_taslagi_pk_col(conn) == "id" else "rowid AS id"

    def save_block_edits(self):
        try:
            with DB() as conn:
                pk = self._rapor_taslagi_pk_col(conn)

                for r in range(self.blocks_table.rowCount()):
                    item_sira = self.blocks_table.item(r, 0)
                    if not item_sira:
                        continue

                    bid = item_sira.data(Qt.ItemDataRole.UserRole)
                    if bid is None:
                        continue

                    sira = int(item_sira.text() or (r + 1))
                    baslik = self.blocks_table.item(r, 1).text() if self.blocks_table.item(r, 1) else ""
                    gen = int(self.blocks_table.item(r, 3).text() or 100) if self.blocks_table.item(r, 3) else 100
                    conn.execute(f"""
                        UPDATE rapor_taslagi
                        SET Sira=?, Baslik=?, GenislikYuzde=?
                        WHERE {pk}=? AND ProjeID=?
                    """, (sira, baslik, gen, bid, self.project_id))

                conn.commit()

            ModernDialog.show_success(self, "Kaydedildi", "Blok ayarları güncellendi.")
            self.load_blocks_into_table()
            self.refresh_preview()

        except Exception as e:
            ModernDialog.show_error(self, "Rapor Hatası", str(e))

    def _normalize_block_order(self):
        """Delil silme/taşıma sonrası Sira’yı 1..n yap."""
        with DB() as conn:
            id_expr = self._rt_select_id_expr(conn)
            pk = self._rapor_taslagi_pk_col(conn)

            rows = conn.execute(f"""
                SELECT {id_expr}
                FROM rapor_taslagi
                WHERE ProjeID=?
                ORDER BY Sira ASC, id ASC
            """, (self.project_id,)).fetchall()

            for idx, (bid,) in enumerate(rows, start=1):
                conn.execute(f"""
                    UPDATE rapor_taslagi SET Sira=?
                    WHERE {pk}=? AND ProjeID=?
                """, (idx, bid, self.project_id))

            conn.commit()

    def move_block(self, delta):
        row = self.blocks_table.currentRow()
        if row < 0: return

        target_row = row + delta
        if target_row < 0 or target_row >= self.blocks_table.rowCount(): return

        item_current = self.blocks_table.item(row, 0)
        item_target = self.blocks_table.item(target_row, 0)
        if not item_current or not item_target: return

        bid_current = item_current.data(Qt.ItemDataRole.UserRole)
        bid_target = item_target.data(Qt.ItemDataRole.UserRole)

        try:
            sira_current = int(item_current.text())
            sira_target = int(item_target.text())
        except: return

        try:
            with DB() as conn:
                pk = self._rapor_taslagi_pk_col(conn)

                conn.execute(f"UPDATE rapor_taslagi SET Sira=-999 WHERE {pk}=?", (bid_current,))
                conn.execute(f"UPDATE rapor_taslagi SET Sira=? WHERE {pk}=?", (sira_current, bid_target))
                conn.execute(f"UPDATE rapor_taslagi SET Sira=? WHERE {pk}=?", (sira_target, bid_current))
                conn.execute(f"UPDATE rapor_taslagi SET Sira=? WHERE {pk}=? AND Sira=-999", (sira_target, bid_current))

                conn.commit()

            self.load_blocks_into_table()

            self.blocks_table.selectRow(target_row)

        except Exception as e:
            ModernDialog.show_error(self, "Hata", f"Sıralama hatası: {e}")

    def delete_selected_block(self):
        try:
            row = self.blocks_table.currentRow()
            if row < 0:
                ModernDialog.show_warning(self, "Delil Sil", "Silinecek delili seçiniz.")
                return

            ok = ModernDialog.show_question(
                self, "Delili Sil",
                "Seçili delil rapor taslağından silinecek. Emin misiniz?",
                yes_btn="Evet", no_btn="Hayır"
            )
            if not ok:
                return

            item_sira = self.blocks_table.item(row, 0)
            bid = item_sira.data(Qt.ItemDataRole.UserRole)

            with DB() as conn:
                pk = self._rapor_taslagi_pk_col(conn)
                conn.execute(f"DELETE FROM rapor_taslagi WHERE {pk}=? AND ProjeID=?",
                             (bid, self.project_id))
                conn.commit()

            self._normalize_block_order()
            self.load_blocks_into_table()
            self.refresh_preview()

        except Exception as e:
            ModernDialog.show_error(self, "Rapor Hatası", str(e))

    def _disabled_sections_set(self):
        s = set()
        if self.cb_taraflar.isChecked(): s.add("taraflar")
        if self.cb_gorev.isChecked(): s.add("gorev")
        if self.cb_dosya.isChecked(): s.add("dosya")
        if self.cb_genel.isChecked(): s.add("genel")
        if self.cb_delillers.isChecked(): s.add("deliller")
        if self.cb_deg.isChecked(): s.add("deg")
        if self.cb_sonuc.isChecked(): s.add("sonuc")
        return s

    def refresh_preview(self):
        # ✅ meta hazır değilken DB'ye yazma
        if getattr(self, "_meta_ready", False):
            self.save_meta_silent()

        disabled = self._disabled_sections_set()
        try:
            html_doc = self.builder.build_html(disabled_sections=disabled)
            base_url = QUrl.fromLocalFile(os.path.join(APP_DIR, ""))
            self.preview.setHtml(html_doc, base_url)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.preview.setHtml(f"<h3 style='color:red'>Rapor Oluşturma Hatası: {e}</h3>")

        try:
            self.preview.setZoomFactor(0.92)
        except Exception:
            pass

    def open_style_editor(self):
        dlg = StyleEditorDialog(self)
        if dlg.exec():
            self.refresh_preview()

    def export_pdf(self):
        try:
            if hasattr(self, "loader"):
                self.loader.start("PDF oluşturuluyor.")
                QApplication.processEvents()

            styles = StyleConfig.load()
            if not isinstance(styles, dict):
                styles = StyleConfig.DEFAULTS.copy()

            m = styles.get("margins", {"top": 25, "right": 20, "bottom": 20, "left": 25})

            disabled = self._disabled_sections_set()
            html_doc = self.builder.build_html(disabled_sections=disabled)

            out_path, _ = QFileDialog.getSaveFileName(self, "PDF Kaydet", "", "PDF (*.pdf)")
            if not out_path:
                return
            if not out_path.lower().endswith(".pdf"):
                out_path += ".pdf"

            def _status_cb(msg: str):
                try:
                    if hasattr(self, "loader"):
                        self.loader.setText(str(msg))
                        QApplication.processEvents()
                except Exception:
                    pass

            def _progress_cb(p: int):
                try:
                    if hasattr(self, "loader"):
                        self.loader.setProgress(int(p))
                        QApplication.processEvents()
                except Exception:
                    pass

            PDFExporter.export_pdf(
                html_string=html_doc,
                out_path=out_path,
                margin_top_mm=int(m.get("top", 25)),
                margin_right_mm=int(m.get("right", 20)),
                margin_bottom_mm=int(m.get("bottom", 20)),
                margin_left_mm=int(m.get("left", 25)),
                zoom=1.0,
                footer_font_size=10,
                status_cb=_status_cb,
                progress_cb=_progress_cb,
            )

            ModernDialog.show_success(self, "Başarılı", "PDF oluşturuldu.")

        except Exception as e:
            ModernDialog.show_error(self, "Hata", str(e))

        finally:
            try:
                if hasattr(self, "loader"):
                    self.loader.stop()
                    QApplication.processEvents()
            except Exception:
                pass


def _quit_app():
    """sys.exit yerine: Qt event loop'tan temiz çıkış."""
    try:
        app = QApplication.instance()
        if app is not None:
            QTimer.singleShot(0, app.quit)
            return
    except Exception:
        pass
    try:
        os._exit(0)
    except Exception:
        pass


def restart_application():
    """
    Uygulamayı yeniden başlatır.
    - Önce detached başlatmayı dener (daha temiz).
    - Olmazsa execv ile aynı process'i replace eder.
    """
    try:
        # PyQt6 varsa genelde mevcut (QtCore importların içinde)
        try:
            from PyQt6.QtCore import QProcess
            ok = QProcess.startDetached(sys.executable, sys.argv)
            if ok:
                _quit_app()
                return
        except Exception:
            pass

        # Fallback: process replace
        try:
            os.execv(sys.executable, [sys.executable] + sys.argv[1:])
        except Exception:
            pass

    finally:
        _quit_app()
